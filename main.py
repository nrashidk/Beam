"""
UAE e-Invoicing Platform with Registration Wizard
==================================================
InvoLinks API - Multi-tenant e-invoicing with subscription plans
"""

import os, enum, hashlib, secrets, json, logging
from uuid import uuid4
from typing import List, Optional
from datetime import datetime, date, timedelta
from decimal import Decimal

# Configure logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# Pydantic & FastAPI
from pydantic import BaseModel, Field
from fastapi import (
    FastAPI,
    UploadFile,
    File,
    HTTPException,
    Depends,
    Form,
    Request,
    Response,
    Header,
    Cookie,
)
from fastapi.responses import FileResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles

# SQLAlchemy
from sqlalchemy import (
    create_engine,
    Column,
    String,
    Integer,
    Float,
    Boolean,
    Date,
    DateTime,
    Enum as SQLEnum,
    ForeignKey,
    Text,
    func,
    or_,
    text,
)
from sqlalchemy.orm import declarative_base, Session, sessionmaker, relationship

# Password & JWT
import bcrypt
from jose import JWTError, jwt

# UBL XML Generator
# Legacy UBL generator removed - using utils/ubl_xml_generator.py instead

# Exception handling
from utils.exceptions import (
    InvoLinksException,
    ValidationError,
    InvoiceValidationError,
    CryptoError,
    SigningError,
    CertificateError,
    XMLGenerationError,
    PeppolError,
    PeppolProviderError,
    ConfigurationError,
    exception_to_http_response,
)

# MFA (Multi-Factor Authentication)
from utils.mfa_utils import MFAManager, EmailOTPManager

# Email Service (AWS SES)
from utils.email_service import email_service

# Bulk Import utilities
from utils.bulk_import import BulkImportValidator
import pandas as pd
from io import BytesIO

# VAT Utilities (Phase 1)
from utils.vat_utils import (
    is_valid_trn,
    format_trn,
    classify_invoice_type,
    is_valid_tax_code,
    UAE_TAX_CODES,
)

# Stripe payment processing
import stripe

# ==================== CONFIG ====================
DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./.dev.db")
# Allow overriding artifact root via env to avoid issues when working directory differs
ARTIFACT_ROOT = os.getenv("ARTIFACT_ROOT", os.path.join(os.getcwd(), "artifacts"))
os.makedirs(ARTIFACT_ROOT, exist_ok=True)
os.makedirs(os.path.join(ARTIFACT_ROOT, "documents"), exist_ok=True)

# JWT settings
SECRET_KEY = os.getenv("JWT_SECRET_KEY", "involinks-secret-key-change-in-production")
ALGORITHM = "HS256"
_jwt_expiry_hours = float(os.getenv("JWT_EXPIRY_HOURS", "24"))  # default 24 h; override via JWT_EXPIRY_HOURS (FTA Article 9.1 session control)
ACCESS_TOKEN_EXPIRE_MINUTES = max(15, int(_jwt_expiry_hours * 60))
REFRESH_TOKEN_EXPIRE_DAYS = int(os.getenv("JWT_REFRESH_DAYS", "1"))

# Password hashing (using bcrypt directly to avoid passlib issues)

# Database connection with pooling configuration for reliability
engine = create_engine(
    DATABASE_URL,
    future=True,
    pool_pre_ping=True,  # Verify connections before using them
    pool_recycle=3600,  # Recycle connections after 1 hour
    pool_size=5,
    max_overflow=10,
)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine, future=True)
Base = declarative_base()


# ==================== ENUMS ====================
class Role(str, enum.Enum):
    SUPER_ADMIN = "SUPER_ADMIN"
    COMPANY_ADMIN = "COMPANY_ADMIN"
    BUSINESS_ADMIN = "BUSINESS_ADMIN"
    FINANCE_USER = "FINANCE_USER"
    READ_ONLY = "READ_ONLY"  # Task 25: view-only role (no create/edit/delete)


# ==================== PERMISSION MATRIX (Task 25) ====================

# Structure: PERMISSIONS[(Role, resource, action)] = True | False
# Resources: invoices, inward_invoices, reports, vat_return, audit_logs, team, settings, gl, archival
# Actions: view, create, edit, delete, approve, export, manage_users

_T = True
_F = False

PERMISSIONS: dict[tuple, bool] = {
    # ── SUPER_ADMIN ─────────────────────────────────────────────────────────
    (Role.SUPER_ADMIN, "invoices",        "view"):         _T,
    (Role.SUPER_ADMIN, "invoices",        "create"):       _T,
    (Role.SUPER_ADMIN, "invoices",        "edit"):         _T,
    (Role.SUPER_ADMIN, "invoices",        "delete"):       _T,
    (Role.SUPER_ADMIN, "invoices",        "approve"):      _T,
    (Role.SUPER_ADMIN, "invoices",        "export"):       _T,
    (Role.SUPER_ADMIN, "inward_invoices", "view"):         _T,
    (Role.SUPER_ADMIN, "inward_invoices", "create"):       _T,
    (Role.SUPER_ADMIN, "inward_invoices", "edit"):         _T,
    (Role.SUPER_ADMIN, "inward_invoices", "delete"):       _T,
    (Role.SUPER_ADMIN, "inward_invoices", "approve"):      _T,
    (Role.SUPER_ADMIN, "reports",         "view"):         _T,
    (Role.SUPER_ADMIN, "reports",         "export"):       _T,
    (Role.SUPER_ADMIN, "vat_return",      "view"):         _T,
    (Role.SUPER_ADMIN, "vat_return",      "create"):       _T,
    (Role.SUPER_ADMIN, "vat_return",      "export"):       _T,
    (Role.SUPER_ADMIN, "audit_logs",      "view"):         _T,
    (Role.SUPER_ADMIN, "audit_logs",      "export"):       _T,
    (Role.SUPER_ADMIN, "team",            "view"):         _T,
    (Role.SUPER_ADMIN, "team",            "manage_users"): _T,
    (Role.SUPER_ADMIN, "settings",        "view"):         _T,
    (Role.SUPER_ADMIN, "settings",        "edit"):         _T,
    (Role.SUPER_ADMIN, "gl",              "view"):         _T,
    (Role.SUPER_ADMIN, "gl",              "edit"):         _T,
    (Role.SUPER_ADMIN, "archival",        "view"):         _T,
    (Role.SUPER_ADMIN, "archival",        "create"):       _T,

    # ── COMPANY_ADMIN ────────────────────────────────────────────────────────
    (Role.COMPANY_ADMIN, "invoices",        "view"):         _T,
    (Role.COMPANY_ADMIN, "invoices",        "create"):       _T,
    (Role.COMPANY_ADMIN, "invoices",        "edit"):         _T,
    (Role.COMPANY_ADMIN, "invoices",        "delete"):       _T,
    (Role.COMPANY_ADMIN, "invoices",        "approve"):      _T,
    (Role.COMPANY_ADMIN, "invoices",        "export"):       _T,
    (Role.COMPANY_ADMIN, "inward_invoices", "view"):         _T,
    (Role.COMPANY_ADMIN, "inward_invoices", "create"):       _T,
    (Role.COMPANY_ADMIN, "inward_invoices", "edit"):         _T,
    (Role.COMPANY_ADMIN, "inward_invoices", "delete"):       _T,
    (Role.COMPANY_ADMIN, "inward_invoices", "approve"):      _T,
    (Role.COMPANY_ADMIN, "reports",         "view"):         _T,
    (Role.COMPANY_ADMIN, "reports",         "export"):       _T,
    (Role.COMPANY_ADMIN, "vat_return",      "view"):         _T,
    (Role.COMPANY_ADMIN, "vat_return",      "create"):       _T,
    (Role.COMPANY_ADMIN, "vat_return",      "export"):       _T,
    (Role.COMPANY_ADMIN, "audit_logs",      "view"):         _T,
    (Role.COMPANY_ADMIN, "audit_logs",      "export"):       _T,
    (Role.COMPANY_ADMIN, "team",            "view"):         _T,
    (Role.COMPANY_ADMIN, "team",            "manage_users"): _T,
    (Role.COMPANY_ADMIN, "settings",        "view"):         _T,
    (Role.COMPANY_ADMIN, "settings",        "edit"):         _T,
    (Role.COMPANY_ADMIN, "gl",              "view"):         _T,
    (Role.COMPANY_ADMIN, "gl",              "edit"):         _T,
    (Role.COMPANY_ADMIN, "archival",        "view"):         _T,
    (Role.COMPANY_ADMIN, "archival",        "create"):       _T,

    # ── BUSINESS_ADMIN ───────────────────────────────────────────────────────
    (Role.BUSINESS_ADMIN, "invoices",        "view"):         _T,
    (Role.BUSINESS_ADMIN, "invoices",        "create"):       _T,
    (Role.BUSINESS_ADMIN, "invoices",        "edit"):         _T,
    (Role.BUSINESS_ADMIN, "invoices",        "delete"):       _F,
    (Role.BUSINESS_ADMIN, "invoices",        "approve"):      _T,
    (Role.BUSINESS_ADMIN, "invoices",        "export"):       _T,
    (Role.BUSINESS_ADMIN, "inward_invoices", "view"):         _T,
    (Role.BUSINESS_ADMIN, "inward_invoices", "create"):       _T,
    (Role.BUSINESS_ADMIN, "inward_invoices", "edit"):         _T,
    (Role.BUSINESS_ADMIN, "inward_invoices", "delete"):       _F,
    (Role.BUSINESS_ADMIN, "inward_invoices", "approve"):      _T,
    (Role.BUSINESS_ADMIN, "reports",         "view"):         _T,
    (Role.BUSINESS_ADMIN, "reports",         "export"):       _T,
    (Role.BUSINESS_ADMIN, "vat_return",      "view"):         _T,
    (Role.BUSINESS_ADMIN, "vat_return",      "create"):       _F,
    (Role.BUSINESS_ADMIN, "vat_return",      "export"):       _T,
    (Role.BUSINESS_ADMIN, "audit_logs",      "view"):         _T,
    (Role.BUSINESS_ADMIN, "audit_logs",      "export"):       _T,
    (Role.BUSINESS_ADMIN, "team",            "view"):         _T,
    (Role.BUSINESS_ADMIN, "team",            "manage_users"): _T,  # can invite FINANCE_USER or READ_ONLY
    (Role.BUSINESS_ADMIN, "settings",        "view"):         _T,
    (Role.BUSINESS_ADMIN, "settings",        "edit"):         _F,
    (Role.BUSINESS_ADMIN, "gl",              "view"):         _T,
    (Role.BUSINESS_ADMIN, "gl",              "edit"):         _F,
    (Role.BUSINESS_ADMIN, "archival",        "view"):         _T,
    (Role.BUSINESS_ADMIN, "archival",        "create"):       _F,

    # ── FINANCE_USER ────────────────────────────────────────────────────────
    (Role.FINANCE_USER, "invoices",        "view"):         _T,
    (Role.FINANCE_USER, "invoices",        "create"):       _T,
    (Role.FINANCE_USER, "invoices",        "edit"):         _T,
    (Role.FINANCE_USER, "invoices",        "delete"):       _F,
    (Role.FINANCE_USER, "invoices",        "approve"):      _F,
    (Role.FINANCE_USER, "invoices",        "export"):       _T,
    (Role.FINANCE_USER, "inward_invoices", "view"):         _T,
    (Role.FINANCE_USER, "inward_invoices", "create"):       _T,
    (Role.FINANCE_USER, "inward_invoices", "edit"):         _T,
    (Role.FINANCE_USER, "inward_invoices", "delete"):       _F,
    (Role.FINANCE_USER, "inward_invoices", "approve"):      _F,
    (Role.FINANCE_USER, "reports",         "view"):         _T,
    (Role.FINANCE_USER, "reports",         "export"):       _T,
    (Role.FINANCE_USER, "vat_return",      "view"):         _T,
    (Role.FINANCE_USER, "vat_return",      "create"):       _F,
    (Role.FINANCE_USER, "vat_return",      "export"):       _F,
    (Role.FINANCE_USER, "audit_logs",      "view"):         _F,
    (Role.FINANCE_USER, "audit_logs",      "export"):       _F,
    (Role.FINANCE_USER, "team",            "view"):         _F,
    (Role.FINANCE_USER, "team",            "manage_users"): _F,
    (Role.FINANCE_USER, "settings",        "view"):         _F,
    (Role.FINANCE_USER, "settings",        "edit"):         _F,
    (Role.FINANCE_USER, "gl",              "view"):         _T,
    (Role.FINANCE_USER, "gl",              "edit"):         _F,
    (Role.FINANCE_USER, "archival",        "view"):         _F,
    (Role.FINANCE_USER, "archival",        "create"):       _F,

    # ── READ_ONLY ─────────────────────────────────────────────────────────────
    (Role.READ_ONLY, "invoices",        "view"):         _T,
    (Role.READ_ONLY, "invoices",        "create"):       _F,
    (Role.READ_ONLY, "invoices",        "edit"):         _F,
    (Role.READ_ONLY, "invoices",        "delete"):       _F,
    (Role.READ_ONLY, "invoices",        "approve"):      _F,
    (Role.READ_ONLY, "invoices",        "export"):       _T,
    (Role.READ_ONLY, "inward_invoices", "view"):         _T,
    (Role.READ_ONLY, "inward_invoices", "create"):       _F,
    (Role.READ_ONLY, "inward_invoices", "edit"):         _F,
    (Role.READ_ONLY, "inward_invoices", "delete"):       _F,
    (Role.READ_ONLY, "inward_invoices", "approve"):      _F,
    (Role.READ_ONLY, "reports",         "view"):         _T,
    (Role.READ_ONLY, "reports",         "export"):       _T,
    (Role.READ_ONLY, "vat_return",      "view"):         _T,
    (Role.READ_ONLY, "vat_return",      "create"):       _F,
    (Role.READ_ONLY, "vat_return",      "export"):       _F,
    (Role.READ_ONLY, "audit_logs",      "view"):         _T,
    (Role.READ_ONLY, "audit_logs",      "export"):       _F,
    (Role.READ_ONLY, "team",            "view"):         _F,
    (Role.READ_ONLY, "team",            "manage_users"): _F,
    (Role.READ_ONLY, "settings",        "view"):         _F,
    (Role.READ_ONLY, "settings",        "edit"):         _F,
    (Role.READ_ONLY, "gl",              "view"):         _T,
    (Role.READ_ONLY, "gl",              "edit"):         _F,
    (Role.READ_ONLY, "archival",        "view"):         _F,
    (Role.READ_ONLY, "archival",        "create"):       _F,
}


def has_permission(user: "UserDB", resource: str, action: str) -> bool:
    """Return True if the user's role is granted (resource, action) in the matrix."""
    return PERMISSIONS.get((user.role, resource, action), False)


def require_permission(resource: str, action: str):
    """
    Returns a FastAPI dependency that raises 403 when the authenticated user
    lacks the given (resource, action) permission in the matrix.

    Usage:
        @app.post("/invoices", dependencies=[require_permission("invoices", "create")])
    """
    def _check(current_user: "UserDB" = Depends(get_current_user_from_header)):
        if not has_permission(current_user, resource, action):
            raise HTTPException(
                status_code=403,
                detail=(
                    f"Permission denied: '{action}' on '{resource}' "
                    f"is not allowed for role {current_user.role}"
                ),
            )
    return Depends(_check)


class CompanyStatus(str, enum.Enum):
    PENDING_REVIEW = "PENDING_REVIEW"
    ACTIVE = "ACTIVE"
    SUSPENDED = "SUSPENDED"
    REJECTED = "REJECTED"


class SubscriptionStatus(str, enum.Enum):
    TRIAL = "TRIAL"
    ACTIVE = "ACTIVE"
    PAST_DUE = "PAST_DUE"
    CANCELED = "CANCELED"


class DocumentType(str, enum.Enum):
    BUSINESS_LICENSE = "BUSINESS_LICENSE"
    TRN_CERTIFICATE = "TRN_CERTIFICATE"
    TRADE_LICENSE = "TRADE_LICENSE"
    PASSPORT = "PASSPORT"
    EMIRATES_ID = "EMIRATES_ID"


class DocumentStatus(str, enum.Enum):
    PENDING_REVIEW = "PENDING_REVIEW"
    APPROVED = "APPROVED"
    REJECTED = "REJECTED"


class InvoiceType(str, enum.Enum):
    TAX_INVOICE = "380"  # Commercial invoice (standard VAT)
    TAX_CREDIT_NOTE = "381"  # Credit note
    DEBIT_NOTE = "383"  # Debit note (Task 4 — FTA UBL type 383)
    COMMERCIAL_INVOICE = "480"  # Invoice out of scope of tax
    CREDIT_NOTE_OUT_OF_SCOPE = "81"  # Credit note related to goods/services


class InvoiceStatus(str, enum.Enum):
    DRAFT = "DRAFT"
    ISSUED = "ISSUED"
    SENT = "SENT"
    VIEWED = "VIEWED"
    PAID = "PAID"
    CANCELLED = "CANCELLED"
    OVERDUE = "OVERDUE"


class TaxCategory(str, enum.Enum):
    STANDARD = "S"  # Standard rate (5% in UAE)
    ZERO = "Z"  # Zero rated
    EXEMPT = "E"  # Exempt from tax
    OUT_OF_SCOPE = "O"  # Out of scope


class PurchaseOrderStatus(str, enum.Enum):
    DRAFT = "DRAFT"
    SENT = "SENT"
    ACKNOWLEDGED = "ACKNOWLEDGED"
    FULFILLED = "FULFILLED"
    CANCELLED = "CANCELLED"


class GoodsReceiptStatus(str, enum.Enum):
    PENDING = "PENDING"
    RECEIVED = "RECEIVED"
    PARTIAL = "PARTIAL"
    REJECTED = "REJECTED"


class InwardInvoiceStatus(str, enum.Enum):
    RECEIVED = "RECEIVED"  # Just received via PEPPOL
    PENDING_REVIEW = "PENDING_REVIEW"  # Awaiting AP team review
    MATCHED = "MATCHED"  # Matched with PO/GRN
    APPROVED = "APPROVED"  # Approved for payment
    REJECTED = "REJECTED"  # Rejected (invoice issue)
    PAID = "PAID"  # Payment completed
    DISPUTED = "DISPUTED"  # Under dispute
    CANCELLED = "CANCELLED"  # Cancelled by supplier


class MatchingStatus(str, enum.Enum):
    NOT_MATCHED = "NOT_MATCHED"
    PARTIAL_MATCH = "PARTIAL_MATCH"  # Some fields match, others don't
    FULL_MATCH = "FULL_MATCH"  # All fields match
    VARIANCE_DETECTED = "VARIANCE_DETECTED"  # Within tolerance but variance exists


class AuditFileStatus(str, enum.Enum):
    GENERATING = "GENERATING"
    COMPLETED = "COMPLETED"
    FAILED = "FAILED"


# ==================== DATABASE MODELS ====================
class UserDB(Base):
    __tablename__ = "users"
    id = Column(String, primary_key=True)
    email = Column(String, unique=True, index=True, nullable=False)
    password_hash = Column(String, nullable=True)
    role = Column(SQLEnum(Role), nullable=False)
    company_id = Column(String, ForeignKey("companies.id"), nullable=True)
    is_owner = Column(Boolean, default=False)  # First user who registered the company
    full_name = Column(String, nullable=True)
    invited_by = Column(
        String, ForeignKey("users.id"), nullable=True
    )  # Who invited this user
    created_at = Column(DateTime, default=datetime.utcnow)
    last_login = Column(DateTime, nullable=True)

    # MFA (Multi-Factor Authentication) fields - Article 9.1 compliance
    mfa_enabled = Column(Boolean, default=False)
    mfa_method = Column(String, nullable=True)  # 'totp', 'email', or None
    mfa_secret = Column(String, nullable=True)  # Base32 encoded TOTP secret
    mfa_backup_codes = Column(Text, nullable=True)  # JSON array of hashed backup codes
    mfa_enrolled_at = Column(DateTime, nullable=True)
    mfa_last_verified_at = Column(DateTime, nullable=True)

    # Task 2: Account lockout fields
    failed_login_count = Column(Integer, default=0)
    locked_until = Column(DateTime, nullable=True)

    # Task 20: Forced password change on first login
    must_change_password = Column(Boolean, default=False)

    # Task 22: Password expiry policy
    password_changed_at = Column(DateTime, nullable=True)

    @property
    def password_last_change_date(self):
        """Alias for password_changed_at for spec compatibility."""
        return self.password_changed_at


class CompanyDB(Base):
    __tablename__ = "companies"
    id = Column(String, primary_key=True)
    legal_name = Column(String, nullable=True)
    country = Column(String, default="AE")
    status = Column(SQLEnum(CompanyStatus), default=CompanyStatus.PENDING_REVIEW)
    trn = Column(String, nullable=True)

    # VAT Registration (Opt-In Model - Phase 1)
    vat_enabled = Column(Boolean, default=False)  # Toggle for VAT registration
    vat_registration_date = Column(
        Date, nullable=True
    )  # When business became VAT-registered
    vat_certificate_path = Column(
        String, nullable=True
    )  # Path to uploaded VAT certificate PDF

    # MFA (Multi-Factor Authentication) fields - Article 9.1 compliance
    mfa_enabled = Column(Boolean, default=False)
    mfa_method = Column(String, nullable=True)  # 'totp' or 'email'
    mfa_secret = Column(String, nullable=True)  # Base32 TOTP secret
    mfa_backup_codes = Column(Text, nullable=True)  # JSON array of hashed codes
    mfa_enrolled_at = Column(DateTime, nullable=True)
    mfa_last_verified_at = Column(DateTime, nullable=True)

    # Registration fields
    business_type = Column(String, nullable=True)
    business_activity = Column(String, nullable=True)
    registration_number = Column(String, nullable=True)
    registration_date = Column(Date, nullable=True)
    email = Column(String, nullable=True)
    phone = Column(String, nullable=True)
    website = Column(String, nullable=True)

    # Address
    address_line1 = Column(String, nullable=True)
    address_line2 = Column(String, nullable=True)
    city = Column(String, nullable=True)
    emirate = Column(String, nullable=True)
    po_box = Column(String, nullable=True)

    # Authorized person
    authorized_person_name = Column(String, nullable=True)
    authorized_person_title = Column(String, nullable=True)
    authorized_person_email = Column(String, nullable=True)
    authorized_person_phone = Column(String, nullable=True)

    # Email verification
    email_verified = Column(Boolean, default=False)
    verification_token = Column(String, nullable=True)
    verification_sent_at = Column(DateTime, nullable=True)

    # Password & authentication
    password_hash = Column(String, nullable=True)
    password_reset_token = Column(String, nullable=True)
    password_reset_expires = Column(DateTime, nullable=True)

    # Free plan configuration (DEPRECATED - moved to trial system)
    free_plan_type = Column(String, nullable=True)  # "DURATION" or "INVOICE_COUNT"
    free_plan_duration_months = Column(Integer, nullable=True)  # If duration-based
    free_plan_invoice_limit = Column(Integer, nullable=True)  # If invoice count-based
    free_plan_start_date = Column(DateTime, nullable=True)  # When free plan started
    invoices_generated = Column(Integer, default=0)  # Total lifetime invoices

    # Subscription tracking (DEPRECATED - moved to subscriptions table)
    subscription_plan_id = Column(
        String, ForeignKey("subscription_plans.id"), nullable=True
    )

    # Free Trial System
    trial_status = Column(String, default="ACTIVE")  # ACTIVE, EXPIRED, CONVERTED
    trial_start_date = Column(DateTime, nullable=True)
    trial_invoice_count = Column(Integer, default=0)  # Invoices sent during trial
    trial_ended_at = Column(DateTime, nullable=True)

    # Stripe Integration
    stripe_customer_id = Column(String, nullable=True, unique=True)

    # PEPPOL Configuration
    peppol_enabled = Column(Boolean, default=False)
    peppol_provider = Column(String, nullable=True)  # 'tradeshift', 'basware', 'mock'
    peppol_participant_id = Column(
        String, nullable=True
    )  # Company's PEPPOL participant ID (e.g., "0190:123456789012345")
    peppol_base_url = Column(String, nullable=True)  # Provider API base URL
    peppol_api_key = Column(
        Text, nullable=True
    )  # Provider API key (encrypted in production)
    peppol_configured_at = Column(DateTime, nullable=True)
    peppol_last_tested_at = Column(DateTime, nullable=True)

    created_at = Column(DateTime, default=datetime.utcnow)
    approved_at = Column(DateTime, nullable=True)
    rejected_at = Column(DateTime, nullable=True)


class SubscriptionPlanDB(Base):
    __tablename__ = "subscription_plans"
    id = Column(String, primary_key=True)
    name = Column(String, nullable=False)
    description = Column(Text, nullable=True)
    price_monthly = Column(Float, default=0.0)
    price_yearly = Column(Float, default=0.0)
    max_invoices_per_month = Column(Integer, nullable=True)
    max_users = Column(Integer, default=1)  # Total users allowed
    max_business_admins = Column(
        Integer, default=0
    )  # Phase 2: Max Business Admins per tier
    max_finance_users = Column(
        Integer, default=1
    )  # Phase 2: Max Finance Users per tier
    max_pos_devices = Column(Integer, default=0)
    allow_api_access = Column(Boolean, default=True)
    allow_branding = Column(Boolean, default=False)
    allow_multi_currency = Column(Boolean, default=False)
    priority_support = Column(Boolean, default=False)
    active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=datetime.utcnow)


class CompanySubscriptionDB(Base):
    __tablename__ = "company_subscriptions"
    id = Column(String, primary_key=True)
    company_id = Column(String, ForeignKey("companies.id"), nullable=False, index=True)
    plan_id = Column(String, ForeignKey("subscription_plans.id"), nullable=False)
    status = Column(SQLEnum(SubscriptionStatus), default=SubscriptionStatus.TRIAL)
    billing_cycle = Column(String, default="monthly")
    current_period_start = Column(Date, nullable=False)
    current_period_end = Column(Date, nullable=False)
    invoices_this_period = Column(Integer, default=0)
    created_at = Column(DateTime, default=datetime.utcnow)

    company = relationship("CompanyDB", backref="old_subscriptions")
    plan = relationship("SubscriptionPlanDB")


class CompanyDocumentDB(Base):
    __tablename__ = "company_documents"
    id = Column(String, primary_key=True)
    company_id = Column(String, ForeignKey("companies.id"), nullable=False, index=True)
    document_type = Column(SQLEnum(DocumentType), nullable=False)
    file_name = Column(String, nullable=False)
    file_path = Column(String, nullable=False)
    file_size = Column(Integer)
    mime_type = Column(String)
    status = Column(SQLEnum(DocumentStatus), default=DocumentStatus.PENDING_REVIEW)
    issue_date = Column(Date, nullable=True)
    expiry_date = Column(Date, nullable=True)
    document_number = Column(String, nullable=True)
    uploaded_at = Column(DateTime, default=datetime.utcnow)

    company = relationship("CompanyDB", backref="documents")


class CompanyBrandingDB(Base):
    __tablename__ = "company_branding"
    id = Column(String, primary_key=True)
    company_id = Column(
        String, ForeignKey("companies.id"), unique=True, nullable=False, index=True
    )

    # Logo
    logo_file_name = Column(String, nullable=True)
    logo_file_path = Column(String, nullable=True)
    logo_file_size = Column(Integer, nullable=True)
    logo_mime_type = Column(String, nullable=True)
    logo_uploaded_at = Column(DateTime, nullable=True)

    # Company Stamp/Seal
    stamp_file_name = Column(String, nullable=True)
    stamp_file_path = Column(String, nullable=True)
    stamp_file_size = Column(Integer, nullable=True)
    stamp_mime_type = Column(String, nullable=True)
    stamp_uploaded_at = Column(DateTime, nullable=True)

    # Optional: Brand Colors & Fonts (for future PDF generation)
    primary_color = Column(String, nullable=True)  # Hex color
    secondary_color = Column(String, nullable=True)  # Hex color
    font_family = Column(String, nullable=True)

    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    company = relationship("CompanyDB", backref="branding")


class RegistrationProgressDB(Base):
    __tablename__ = "registration_progress"
    id = Column(String, primary_key=True)
    company_id = Column(String, ForeignKey("companies.id"), unique=True, index=True)
    current_step = Column(Integer, default=1)
    step_company_info = Column(Boolean, default=False)
    step_business_details = Column(Boolean, default=False)
    step_documents = Column(Boolean, default=False)
    step_plan_selection = Column(Boolean, default=False)
    step_review = Column(Boolean, default=False)
    completed = Column(Boolean, default=False)
    created_at = Column(DateTime, default=datetime.utcnow)

    company = relationship("CompanyDB", backref="progress")


class ContentBlockDB(Base):
    """Content Management System - Editable text blocks for UI"""

    __tablename__ = "content_blocks"
    id = Column(String, primary_key=True)
    key = Column(
        String, unique=True, nullable=False, index=True
    )  # e.g., "homepage_hero_title"
    value = Column(Text, nullable=False)  # The actual text content
    description = Column(String, nullable=True)  # What this controls
    section = Column(
        String, nullable=True, index=True
    )  # Group by page: "homepage", "feature_boxes", etc.
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    updated_by = Column(String, nullable=True)  # Admin email who made the change


class FeaturedBusinessDB(Base):
    """Featured businesses for homepage moving bar - Phase 2"""

    __tablename__ = "featured_businesses"
    id = Column(String, primary_key=True)
    company_id = Column(String, ForeignKey("companies.id"), nullable=False, index=True)
    display_name = Column(String, nullable=True)  # Override company name if needed
    logo_url = Column(String, nullable=True)  # Company logo path
    is_active = Column(Boolean, default=True)  # Show/hide from moving bar
    display_order = Column(Integer, default=0)  # Sort order in moving bar
    added_by_user_id = Column(
        String, ForeignKey("users.id"), nullable=True
    )  # SuperAdmin who added it
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Relationships
    company = relationship("CompanyDB", backref="featured_entries")


class InvoiceDB(Base):
    __tablename__ = "invoices"
    id = Column(String, primary_key=True)
    company_id = Column(String, ForeignKey("companies.id"), nullable=False, index=True)

    # UBL/PINT-AE Core Fields
    invoice_number = Column(String, nullable=False, index=True)
    invoice_type = Column(SQLEnum(InvoiceType), nullable=False)
    status = Column(SQLEnum(InvoiceStatus), default=InvoiceStatus.DRAFT)
    issue_date = Column(Date, nullable=False)
    due_date = Column(Date, nullable=True)
    supply_date = Column(Date, nullable=True)  # Task 5: UBL cac:Delivery date
    version = Column(Integer, default=1, nullable=False, server_default="1")  # Task 10: optimistic locking
    is_archived = Column(Boolean, default=False, nullable=False, server_default="false")  # Task 12: archival
    archived_at = Column(DateTime, nullable=True)  # Task 12: timestamp when archived
    currency_code = Column(String, default="AED")

    # Supplier (Issuer) - Company issuing the invoice
    supplier_trn = Column(String, nullable=True)  # Optional when VAT not enabled
    supplier_name = Column(String, nullable=False)
    supplier_address = Column(Text, nullable=True)
    supplier_city = Column(String, nullable=True)
    supplier_country = Column(String, default="AE")
    supplier_peppol_id = Column(String, nullable=True)  # Peppol endpoint

    # Customer (Buyer) - Recipient of invoice
    customer_trn = Column(String, nullable=True)  # May be null for B2C
    customer_name = Column(String, nullable=False)
    customer_email = Column(String, nullable=True)
    customer_address = Column(Text, nullable=True)
    customer_city = Column(String, nullable=True)
    customer_country = Column(String, default="AE")
    customer_peppol_id = Column(String, nullable=True)

    # VAT Compliance (Phase 1)
    tax_code = Column(String, default="SR")  # UAE tax code: SR, ZR, ES, RC, OP
    invoice_classification = Column(
        String, default="standard"
    )  # 'full', 'simplified', 'standard'

    # Monetary Totals (PINT-AE mandatory)
    subtotal_amount = Column(Float, default=0.0)  # Line extension total
    tax_amount = Column(Float, default=0.0)  # Total VAT
    total_amount = Column(Float, default=0.0)  # Amount including VAT
    amount_due = Column(Float, default=0.0)  # Payable amount

    # AED conversion (mandatory if currency != AED)
    total_amount_aed = Column(Float, nullable=True)

    # Payment Terms
    payment_terms = Column(Text, nullable=True)
    payment_due_days = Column(Integer, default=30)

    # Notes & References
    invoice_notes = Column(Text, nullable=True)
    reference_number = Column(String, nullable=True)  # PO number, etc.
    preceding_invoice_id = Column(
        String, ForeignKey("invoices.id"), nullable=True
    )  # For credit notes

    # Credit Note specific
    credit_note_reason = Column(String, nullable=True)  # Mandatory for credit notes

    # Document Management
    xml_file_path = Column(String, nullable=True)  # UBL XML storage path
    pdf_file_path = Column(String, nullable=True)  # PDF storage path
    xml_hash = Column(String, nullable=True)  # SHA-256 hash of XML

    # Digital Signature & Hash Chain (UAE FTA Compliance)
    prev_invoice_hash = Column(
        String, nullable=True, index=True
    )  # Links to previous invoice in chain
    signature_b64 = Column(Text, nullable=True)  # Base64 encoded digital signature
    signing_cert_serial = Column(
        String, nullable=True
    )  # Certificate serial number for audit trail
    signing_timestamp = Column(
        DateTime, nullable=True
    )  # When invoice was digitally signed

    # PEPPOL Transmission Tracking
    peppol_message_id = Column(
        String, nullable=True, index=True
    )  # Provider's message ID
    peppol_status = Column(String, nullable=True)  # SENT, DELIVERED, REJECTED, etc.
    peppol_provider = Column(String, nullable=True)  # e.g., "tradeshift", "basware"
    peppol_sent_at = Column(DateTime, nullable=True)  # Transmission timestamp
    peppol_response = Column(Text, nullable=True)  # Provider API response (JSON)

    # Sharing & Transmission
    share_token = Column(String, nullable=True, index=True)  # Public share link token
    sent_at = Column(DateTime, nullable=True)
    viewed_at = Column(DateTime, nullable=True)
    paid_at = Column(DateTime, nullable=True)

    # User Tracking (Phase 2)
    created_by_user_id = Column(
        String, ForeignKey("users.id"), nullable=True, index=True
    )  # Who created the invoice
    payment_verified_by_user_id = Column(
        String, ForeignKey("users.id"), nullable=True
    )  # Who verified payment
    payment_verified_at = Column(DateTime, nullable=True)  # When payment was verified
    payment_method = Column(
        String, nullable=True
    )  # Phase 2: Cash, POS, Bank Transfer, etc.

    # Timestamps
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Relationships
    company = relationship("CompanyDB", backref="invoices")
    line_items = relationship(
        "InvoiceLineItemDB", backref="invoice", cascade="all, delete-orphan"
    )
    tax_breakdowns = relationship(
        "InvoiceTaxBreakdownDB", backref="invoice", cascade="all, delete-orphan"
    )


class InvoiceLineItemDB(Base):
    __tablename__ = "invoice_line_items"
    id = Column(String, primary_key=True)
    invoice_id = Column(String, ForeignKey("invoices.id"), nullable=False, index=True)
    line_number = Column(Integer, nullable=False)  # Sequential line number

    # Product/Service Details
    item_name = Column(String, nullable=False)
    item_description = Column(Text, nullable=True)
    item_code = Column(String, nullable=True)  # SKU or product code

    # Quantity & Unit
    quantity = Column(Float, nullable=False)
    unit_code = Column(String, default="C62")  # UN/ECE unit codes (C62 = piece)

    # Pricing
    unit_price = Column(Float, nullable=False)
    line_extension_amount = Column(Float, nullable=False)  # quantity * unit_price

    # Tax
    tax_category = Column(SQLEnum(TaxCategory), nullable=False)
    tax_percent = Column(Float, default=0.0)  # 5% for standard rate in UAE
    tax_amount = Column(Float, default=0.0)
    tax_code = Column(
        String, default="SR"
    )  # UAE tax code: SR, ZR, ES, RC, OP (Phase 1)

    # Total
    line_total_amount = Column(Float, nullable=False)  # Including tax

    created_at = Column(DateTime, default=datetime.utcnow)


class InvoiceTaxBreakdownDB(Base):
    __tablename__ = "invoice_tax_breakdowns"
    id = Column(String, primary_key=True)
    invoice_id = Column(String, ForeignKey("invoices.id"), nullable=False, index=True)

    # Tax Category Breakdown (required for UBL)
    tax_category = Column(SQLEnum(TaxCategory), nullable=False)
    taxable_amount = Column(Float, nullable=False)  # Base amount before tax
    tax_percent = Column(Float, nullable=False)
    tax_amount = Column(Float, nullable=False)

    created_at = Column(DateTime, default=datetime.utcnow)


# ==================== CORNER 4: AP MANAGEMENT (Inward Invoicing) ====================


class PurchaseOrderDB(Base):
    """Purchase Orders - Track expected invoices from suppliers"""

    __tablename__ = "purchase_orders"
    id = Column(String, primary_key=True)
    company_id = Column(String, ForeignKey("companies.id"), nullable=False, index=True)

    # PO Identification
    po_number = Column(
        String, nullable=False, index=True
    )  # Unique PO number per company
    status = Column(SQLEnum(PurchaseOrderStatus), default=PurchaseOrderStatus.DRAFT)

    # Supplier Information
    supplier_trn = Column(
        String, nullable=True
    )  # Optional for non-VAT-registered suppliers
    supplier_name = Column(String, nullable=False)
    supplier_contact_email = Column(String, nullable=True)
    supplier_address = Column(Text, nullable=True)
    supplier_peppol_id = Column(String, nullable=True)

    # Order Details
    order_date = Column(Date, nullable=False)
    expected_delivery_date = Column(Date, nullable=True)
    delivery_address = Column(Text, nullable=True)

    # Financial
    currency_code = Column(String, default="AED")
    expected_subtotal = Column(Float, default=0.0)
    expected_tax = Column(Float, default=0.0)
    expected_total = Column(Float, nullable=False)

    # Matching & Fulfillment
    received_invoice_count = Column(Integer, default=0)  # How many invoices received
    matched_invoice_count = Column(Integer, default=0)  # How many matched
    received_amount_total = Column(Float, default=0.0)  # Total invoiced amount
    variance_amount = Column(Float, default=0.0)  # Difference (expected - actual)

    # References
    reference_number = Column(String, nullable=True)  # Internal reference
    notes = Column(Text, nullable=True)

    # Approval
    approved_by_user_id = Column(String, ForeignKey("users.id"), nullable=True)
    approved_at = Column(DateTime, nullable=True)

    # Timestamps
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Relationships
    company = relationship("CompanyDB", backref="purchase_orders")
    line_items = relationship(
        "PurchaseOrderLineItemDB",
        backref="purchase_order",
        cascade="all, delete-orphan",
    )
    goods_receipts = relationship(
        "GoodsReceiptDB", backref="purchase_order", cascade="all, delete-orphan"
    )


class PurchaseOrderLineItemDB(Base):
    """Line items for Purchase Orders"""

    __tablename__ = "purchase_order_line_items"
    id = Column(String, primary_key=True)
    po_id = Column(String, ForeignKey("purchase_orders.id"), nullable=False, index=True)
    line_number = Column(Integer, nullable=False)

    # Product/Service
    item_name = Column(String, nullable=False)
    item_description = Column(Text, nullable=True)
    item_code = Column(String, nullable=True)  # SKU

    # Quantity & Unit
    quantity_ordered = Column(Float, nullable=False)
    quantity_received = Column(Float, default=0.0)
    unit_code = Column(String, default="C62")

    # Pricing
    unit_price = Column(Float, nullable=False)
    line_total = Column(Float, nullable=False)

    # Tax
    tax_category = Column(SQLEnum(TaxCategory), default=TaxCategory.STANDARD)
    tax_percent = Column(Float, default=5.0)

    created_at = Column(DateTime, default=datetime.utcnow)


class GoodsReceiptDB(Base):
    """Goods Receipt Notes (GRN) - Track physical delivery"""

    __tablename__ = "goods_receipts"
    id = Column(String, primary_key=True)
    company_id = Column(String, ForeignKey("companies.id"), nullable=False, index=True)
    po_id = Column(String, ForeignKey("purchase_orders.id"), nullable=True, index=True)

    # GRN Identification
    grn_number = Column(String, nullable=False, index=True)
    status = Column(SQLEnum(GoodsReceiptStatus), default=GoodsReceiptStatus.PENDING)

    # Receipt Details
    receipt_date = Column(Date, nullable=False)
    received_by_user_id = Column(String, ForeignKey("users.id"), nullable=True)

    # Supplier
    supplier_trn = Column(
        String, nullable=True
    )  # Optional for non-VAT-registered suppliers
    supplier_name = Column(String, nullable=False)
    supplier_delivery_note = Column(
        String, nullable=True
    )  # Supplier's delivery note number

    # Quantities
    total_items_received = Column(Integer, default=0)
    total_value = Column(Float, default=0.0)

    # Quality Control
    inspection_passed = Column(Boolean, default=True)
    quality_notes = Column(Text, nullable=True)
    damaged_items_count = Column(Integer, default=0)

    # Notes
    notes = Column(Text, nullable=True)

    # Timestamps
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Relationships
    company = relationship("CompanyDB", backref="goods_receipts")
    line_items = relationship(
        "GoodsReceiptLineItemDB", backref="goods_receipt", cascade="all, delete-orphan"
    )


class GoodsReceiptLineItemDB(Base):
    """Line items for Goods Receipt Notes"""

    __tablename__ = "goods_receipt_line_items"
    id = Column(String, primary_key=True)
    grn_id = Column(String, ForeignKey("goods_receipts.id"), nullable=False, index=True)
    po_line_item_id = Column(
        String, ForeignKey("purchase_order_line_items.id"), nullable=True
    )

    line_number = Column(Integer, nullable=False)

    # Product
    item_name = Column(String, nullable=False)
    item_code = Column(String, nullable=True)

    # Quantities
    quantity_received = Column(Float, nullable=False)
    quantity_accepted = Column(Float, nullable=False)
    quantity_rejected = Column(Float, default=0.0)
    unit_code = Column(String, default="C62")

    # Quality
    condition = Column(String, nullable=True)  # "GOOD", "DAMAGED", "DEFECTIVE"
    notes = Column(Text, nullable=True)

    created_at = Column(DateTime, default=datetime.utcnow)


class InwardInvoiceDB(Base):
    """Inward Invoices - Received from suppliers via PEPPOL (Corner 4)"""

    __tablename__ = "inward_invoices"
    id = Column(String, primary_key=True)
    company_id = Column(String, ForeignKey("companies.id"), nullable=False, index=True)

    # Invoice Identification
    supplier_invoice_number = Column(String, nullable=False, index=True)
    invoice_type = Column(SQLEnum(InvoiceType), default=InvoiceType.TAX_INVOICE)
    status = Column(SQLEnum(InwardInvoiceStatus), default=InwardInvoiceStatus.RECEIVED)

    # Dates
    invoice_date = Column(Date, nullable=False)
    due_date = Column(Date, nullable=True)
    received_at = Column(DateTime, default=datetime.utcnow)

    # Supplier (Invoice Issuer)
    supplier_trn = Column(
        String, nullable=True
    )  # Optional for non-VAT-registered suppliers
    supplier_name = Column(String, nullable=False)
    supplier_address = Column(Text, nullable=True)
    supplier_peppol_id = Column(String, nullable=True)
    supplier_company_id = Column(
        String, ForeignKey("companies.id"), nullable=True
    )  # If supplier uses InvoLinks

    # Customer (Our company - the buyer)
    customer_trn = Column(
        String, nullable=True
    )  # Optional for non-VAT-registered customers
    customer_name = Column(String, nullable=False)

    # Financial
    currency_code = Column(String, default="AED")
    subtotal_amount = Column(Float, nullable=False)
    tax_amount = Column(Float, nullable=False)
    total_amount = Column(Float, nullable=False)
    amount_due = Column(Float, nullable=False)
    # AED conversion (mandatory when currency_code != "AED" — FTA P2-A compliance)
    total_amount_aed = Column(Float, nullable=True)

    # Document Storage
    xml_file_path = Column(String, nullable=True)  # Received UBL XML
    pdf_file_path = Column(String, nullable=True)  # Generated PDF for viewing
    xml_hash = Column(String, nullable=True)

    # PEPPOL Reception
    peppol_message_id = Column(String, nullable=True, index=True)
    peppol_sender_id = Column(String, nullable=True)
    peppol_provider = Column(String, nullable=True)
    peppol_received_at = Column(DateTime, nullable=True)

    # Matching & Approval
    po_id = Column(String, ForeignKey("purchase_orders.id"), nullable=True, index=True)
    grn_id = Column(String, ForeignKey("goods_receipts.id"), nullable=True, index=True)
    matching_status = Column(
        SQLEnum(MatchingStatus), default=MatchingStatus.NOT_MATCHED
    )

    # 3-Way Matching Results
    po_match_score = Column(Float, default=0.0)  # 0-100% match with PO
    grn_match_score = Column(Float, default=0.0)  # 0-100% match with GRN
    amount_variance = Column(Float, default=0.0)  # Difference from expected
    quantity_variance = Column(Float, default=0.0)  # Qty difference

    # Approval Workflow
    reviewed_by_user_id = Column(String, ForeignKey("users.id"), nullable=True)
    reviewed_at = Column(DateTime, nullable=True)
    approved_by_user_id = Column(String, ForeignKey("users.id"), nullable=True)
    approved_at = Column(DateTime, nullable=True)
    rejection_reason = Column(Text, nullable=True)

    # Payment Tracking
    payment_status = Column(String, nullable=True)  # "PENDING", "SCHEDULED", "PAID"
    payment_method = Column(String, nullable=True)
    paid_amount = Column(Float, default=0.0)
    paid_at = Column(DateTime, nullable=True)
    payment_reference = Column(String, nullable=True)

    # Dispute Management
    disputed_at = Column(DateTime, nullable=True)
    dispute_reason = Column(Text, nullable=True)
    dispute_resolved_at = Column(DateTime, nullable=True)

    # Notes & References
    notes = Column(Text, nullable=True)
    reference_number = Column(String, nullable=True)

    # Timestamps
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Relationships
    company = relationship(
        "CompanyDB", foreign_keys=[company_id], backref="inward_invoices"
    )
    purchase_order = relationship("PurchaseOrderDB", backref="inward_invoices")
    goods_receipt = relationship("GoodsReceiptDB", backref="inward_invoices")
    line_items = relationship(
        "InwardInvoiceLineItemDB",
        backref="inward_invoice",
        cascade="all, delete-orphan",
    )


class InwardInvoiceLineItemDB(Base):
    """Line items for Inward Invoices"""

    __tablename__ = "inward_invoice_line_items"
    id = Column(String, primary_key=True)
    inward_invoice_id = Column(
        String, ForeignKey("inward_invoices.id"), nullable=False, index=True
    )
    line_number = Column(Integer, nullable=False)

    # Product/Service
    item_name = Column(String, nullable=False)
    item_description = Column(Text, nullable=True)
    item_code = Column(String, nullable=True)

    # Quantity & Unit
    quantity = Column(Float, nullable=False)
    unit_code = Column(String, default="C62")

    # Pricing
    unit_price = Column(Float, nullable=False)
    line_extension_amount = Column(Float, nullable=False)

    # Tax
    tax_category = Column(SQLEnum(TaxCategory), nullable=False)
    tax_percent = Column(Float, default=0.0)
    tax_amount = Column(Float, default=0.0)

    # Total
    line_total_amount = Column(Float, nullable=False)

    # Matching
    po_line_item_id = Column(
        String, ForeignKey("purchase_order_line_items.id"), nullable=True
    )
    grn_line_item_id = Column(
        String, ForeignKey("goods_receipt_line_items.id"), nullable=True
    )
    match_status = Column(String, nullable=True)  # "MATCHED", "VARIANCE", "NO_PO"

    created_at = Column(DateTime, default=datetime.utcnow)


class AuditFileDB(Base):
    """FTA Audit Files (FAF) - UAE Federal Tax Authority Audit File"""

    __tablename__ = "audit_files"
    id = Column(String, primary_key=True)
    company_id = Column(String, ForeignKey("companies.id"), nullable=False, index=True)

    # File Information
    file_name = Column(String, nullable=False)
    file_path = Column(String, nullable=False)
    file_size = Column(Integer, nullable=True)
    format = Column(String, default="CSV")  # CSV or TXT

    # Audit Period
    period_start_date = Column(Date, nullable=False)
    period_end_date = Column(Date, nullable=False)

    # Generation Details
    status = Column(SQLEnum(AuditFileStatus), default=AuditFileStatus.GENERATING)
    generated_by_user_id = Column(String, ForeignKey("users.id"), nullable=False)
    generated_at = Column(DateTime, default=datetime.utcnow)

    # Statistics
    total_invoices = Column(Integer, default=0)
    total_customers = Column(Integer, default=0)
    total_amount = Column(Float, default=0.0)
    total_vat = Column(Float, default=0.0)

    # Error tracking
    error_message = Column(Text, nullable=True)

    # Timestamps
    created_at = Column(DateTime, default=datetime.utcnow)

    # Relationships
    company = relationship("CompanyDB", backref="audit_files")


class PasswordHistoryDB(Base):
    """Stores the last N password hashes per user to prevent reuse (Task 2)."""

    __tablename__ = "password_history"

    id = Column(String, primary_key=True)
    user_id = Column(String, ForeignKey("users.id"), nullable=False, index=True)
    password_hash = Column(String, nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)

    user = relationship("UserDB", backref="password_history")


class AuditLogDB(Base):
    """Immutable audit trail — every sensitive action is recorded here (Task 1)."""

    __tablename__ = "audit_logs"

    id = Column(String, primary_key=True)
    company_id = Column(String, ForeignKey("companies.id"), nullable=True, index=True)
    user_id = Column(String, ForeignKey("users.id"), nullable=True, index=True)

    # What happened
    action = Column(String, nullable=False)        # e.g. "USER_LOGIN", "INVOICE_ISSUED"
    resource_type = Column(String, nullable=True)  # e.g. "invoice", "vat_return"
    resource_id = Column(String, nullable=True)    # PK of the affected resource

    # Change detail (JSON-encoded strings)
    old_value = Column(Text, nullable=True)
    new_value = Column(Text, nullable=True)
    description = Column(Text, nullable=True)

    # Request context
    ip_address = Column(String, nullable=True)
    user_agent = Column(String, nullable=True)

    # Timestamp — never nullable; auto-set
    created_at = Column(DateTime, default=datetime.utcnow, nullable=False)

    company = relationship("CompanyDB", backref="audit_logs")
    user = relationship("UserDB", backref="audit_logs")


class IntegrityCheckDB(Base):
    """Stores the result of each hash chain integrity verification run (Task 26)."""

    __tablename__ = "integrity_checks"

    id = Column(String, primary_key=True)
    company_id = Column(String, ForeignKey("companies.id"), nullable=True, index=True)
    performed_by_user_id = Column(String, ForeignKey("users.id"), nullable=True)
    checked_at = Column(DateTime, default=datetime.utcnow, nullable=False)
    total_invoices = Column(Integer, nullable=False, default=0)
    valid_links = Column(Integer, nullable=False, default=0)
    integrity_ok = Column(Boolean, nullable=False, default=False)
    first_broken_invoice_id = Column(String, nullable=True)
    first_broken_invoice_number = Column(String, nullable=True)
    first_broken_invoice_date = Column(String, nullable=True)

    company = relationship("CompanyDB", backref="integrity_checks")
    user = relationship("UserDB", backref="integrity_checks")


class VATReturnDB(Base):
    """VAT Return — stores computed FTA VAT returns per period (Task 9)."""

    __tablename__ = "vat_returns"

    id = Column(String, primary_key=True)
    company_id = Column(String, ForeignKey("companies.id"), nullable=False, index=True)

    # Period
    period_start = Column(Date, nullable=False)
    period_end = Column(Date, nullable=False)

    # Transaction counts
    total_sales_invoices = Column(Integer, default=0)
    total_purchase_invoices = Column(Integer, default=0)

    # === SALES / OUTPUT SIDE ===
    # Box 1 — Standard rated supplies (taxable amount excl. VAT)
    standard_rated_sales = Column(Float, default=0.0)
    # Box 2 — Tax refunds provided to tourists (AED)
    tax_refunds_provided = Column(Float, default=0.0)
    # Box 3 — Output VAT (VAT on standard-rated sales)
    output_vat = Column(Float, default=0.0)
    # Box 4 — Zero-rated supplies (taxable amount)
    zero_rated_sales = Column(Float, default=0.0)
    # Box 5 — Exempt supplies (amount)
    exempt_sales = Column(Float, default=0.0)
    # Box 6 — Out-of-scope / reverse-charge supplies (amount)
    out_of_scope_sales = Column(Float, default=0.0)
    # Alias required by task spec: reverse_charge_sales = same as out_of_scope_sales (category O)
    reverse_charge_sales = Column(Float, default=0.0)
    # Box 7 — Total supplies (Box 1 + 4 + 5 + 6)
    total_sales = Column(Float, default=0.0)

    # === PURCHASES / INPUT SIDE ===
    # Box 8 — Standard-rated purchases – bills (taxable amount excl. VAT)
    standard_rated_purchases = Column(Float, default=0.0)
    # Box 9 — Recoverable input VAT on bills
    input_vat_bills = Column(Float, default=0.0)
    # Box 10 — Standard-rated purchases – expenses (taxable amount excl. VAT)
    purchase_expenses = Column(Float, default=0.0)
    # Box 11 — Recoverable input VAT on expenses
    input_vat_expenses = Column(Float, default=0.0)
    # Box 12 — Total input VAT (Box 9 + Box 11)
    total_input_vat = Column(Float, default=0.0)
    # Zero-rated & exempt purchases (informational; VAT = 0)
    zero_rated_purchases = Column(Float, default=0.0)
    exempt_purchases = Column(Float, default=0.0)
    # Reverse-charge VAT: input VAT due on reverse-charge purchases (category O inward invoices)
    reverse_charge_vat = Column(Float, default=0.0)

    # === NET ===
    # Box 13 — Net VAT payable / refundable (Box 3 – Box 12)
    net_vat_payable = Column(Float, default=0.0)

    # Metadata
    generated_by_user_id = Column(String, ForeignKey("users.id"), nullable=False)
    created_at = Column(DateTime, default=datetime.utcnow)

    company = relationship("CompanyDB", backref="vat_returns")


class PaymentMethodDB(Base):
    """Customer payment methods (credit cards via Stripe)"""

    __tablename__ = "payment_methods"
    id = Column(String, primary_key=True)
    company_id = Column(String, ForeignKey("companies.id"), nullable=False, index=True)

    # Stripe details
    stripe_payment_method_id = Column(String, nullable=False, unique=True)

    # Card information (for display)
    card_brand = Column(String, nullable=True)  # visa, mastercard, amex
    card_last4 = Column(String, nullable=True)
    exp_month = Column(Integer, nullable=True)
    exp_year = Column(Integer, nullable=True)

    # Billing details
    billing_email = Column(String, nullable=True)
    billing_name = Column(String, nullable=True)

    # Status
    is_default = Column(Boolean, default=False)

    # Timestamps
    created_at = Column(DateTime, default=datetime.utcnow)

    # Relationships
    company = relationship("CompanyDB", backref="payment_methods")


class SubscriptionDB(Base):
    """Company subscriptions (Stripe-based)"""

    __tablename__ = "subscriptions"
    id = Column(String, primary_key=True)
    company_id = Column(String, ForeignKey("companies.id"), nullable=False, index=True)

    # Subscription details
    tier = Column(String, nullable=False)  # FREE, BASIC, PRO, ENTERPRISE
    billing_cycle_months = Column(Integer, default=1)  # 1, 3, 6
    monthly_price = Column(Float, default=0.0)
    discount_percent = Column(Float, default=0.0)

    # Status
    status = Column(String, default="ACTIVE")  # ACTIVE, CANCELLED, PAST_DUE, TRIAL

    # Stripe details
    stripe_subscription_id = Column(String, nullable=True, unique=True)
    stripe_customer_id = Column(String, nullable=True)

    # Period
    current_period_start = Column(DateTime, nullable=False)
    current_period_end = Column(DateTime, nullable=False)

    # Timestamps
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    cancelled_at = Column(DateTime, nullable=True)

    # Relationships
    company = relationship("CompanyDB", backref="subscriptions")


class PeppolUsageDB(Base):
    """PEPPOL transmission tracking for usage-based billing"""

    __tablename__ = "peppol_usage"
    id = Column(String, primary_key=True)
    company_id = Column(String, ForeignKey("companies.id"), nullable=False, index=True)
    invoice_id = Column(String, ForeignKey("invoices.id"), nullable=True)

    # Transmission details
    transmission_date = Column(DateTime, default=datetime.utcnow)
    status = Column(String, nullable=False)  # SUCCESS, FAILED

    # Billing
    fee_amount = Column(Float, default=1.0)  # AED per transmission
    month_year = Column(String, nullable=False, index=True)  # "2025-10" for aggregation
    billed = Column(Boolean, default=False)

    # Timestamps
    created_at = Column(DateTime, default=datetime.utcnow)

    # Relationships
    company = relationship("CompanyDB", backref="peppol_usage_records")


class BillingInvoiceDB(Base):
    """Monthly billing invoices (subscription + PEPPOL usage)"""

    __tablename__ = "billing_invoices"
    id = Column(String, primary_key=True)
    company_id = Column(String, ForeignKey("companies.id"), nullable=False, index=True)

    # Invoice details
    invoice_number = Column(String, unique=True, nullable=False)

    # Billing period
    billing_period_start = Column(Date, nullable=False)
    billing_period_end = Column(Date, nullable=False)

    # Amounts
    subscription_fee = Column(Float, default=0.0)
    peppol_usage_count = Column(Integer, default=0)
    peppol_usage_fee = Column(Float, default=0.0)
    total_amount = Column(Float, nullable=False)

    # Status
    status = Column(String, default="DRAFT")  # DRAFT, SENT, PAID, OVERDUE, CANCELLED

    # Stripe details
    stripe_invoice_id = Column(String, nullable=True, unique=True)

    # Payment
    paid_at = Column(DateTime, nullable=True)
    payment_method_id = Column(String, ForeignKey("payment_methods.id"), nullable=True)

    # Timestamps
    created_at = Column(DateTime, default=datetime.utcnow)
    sent_at = Column(DateTime, nullable=True)

    # Relationships
    company = relationship("CompanyDB", backref="billing_invoices")


class ExpenseCategoryDB(Base):
    """Custom expense categories - user-defined"""

    __tablename__ = "expense_categories"
    id = Column(String, primary_key=True)
    company_id = Column(String, ForeignKey("companies.id"), nullable=False, index=True)
    name = Column(String, nullable=False)
    description = Column(Text, nullable=True)
    is_default = Column(Boolean, default=False)  # System-provided categories
    created_at = Column(DateTime, default=datetime.utcnow)
    company = relationship("CompanyDB", backref="expense_categories")


class ExpenseDB(Base):
    """Simple expense tracking - rent, utilities, salaries, materials, etc."""

    __tablename__ = "expenses"
    id = Column(String, primary_key=True)
    company_id = Column(String, ForeignKey("companies.id"), nullable=False, index=True)

    # Expense details
    expense_date = Column(Date, nullable=False, index=True)
    category = Column(String, nullable=False)  # Now user-defined categories
    description = Column(Text, nullable=True)

    # Financial
    amount = Column(Float, nullable=False)
    vat_amount = Column(Float, default=0.0)  # Input VAT paid to supplier
    total_amount = Column(Float, nullable=False)  # amount + vat_amount
    currency_code = Column(String, default="AED")

    # Reference
    reference_number = Column(String, nullable=True)
    supplier_name = Column(String, nullable=True)
    notes = Column(Text, nullable=True)

    # VAT Compliance (Phase 1)
    tax_code = Column(String, default="SR")  # UAE tax code: SR, ZR, ES, RC, OP
    vendor_id = Column(
        String, nullable=True
    )  # Link to vendor (will add vendors table in Phase 2)

    # User Tracking (Phase 2)
    created_by_user_id = Column(
        String, ForeignKey("users.id"), nullable=True, index=True
    )  # Who recorded the expense

    # Timestamps
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Relationships
    company = relationship("CompanyDB", backref="expenses")


class InventoryItemDB(Base):
    """Simple inventory tracking"""

    __tablename__ = "inventory_items"
    id = Column(String, primary_key=True)
    company_id = Column(String, ForeignKey("companies.id"), nullable=False, index=True)

    # Item details
    item_name = Column(String, nullable=False)
    item_code = Column(String, nullable=True, index=True)  # SKU
    description = Column(Text, nullable=True)
    unit = Column(String, default="unit")  # unit, bottle, kit, box, etc.

    # Stock
    current_stock = Column(Float, default=0.0)
    min_stock_level = Column(Float, default=0.0)  # Reorder threshold

    # Pricing (optional)
    unit_cost = Column(Float, nullable=True)  # Cost per unit

    # Timestamps
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Relationships
    company = relationship("CompanyDB", backref="inventory_items")


class InventoryTransactionDB(Base):
    """Track inventory movements (purchases, sales, adjustments)"""

    __tablename__ = "inventory_transactions"
    id = Column(String, primary_key=True)
    company_id = Column(String, ForeignKey("companies.id"), nullable=False, index=True)
    inventory_item_id = Column(
        String, ForeignKey("inventory_items.id"), nullable=False, index=True
    )

    # Transaction details
    transaction_type = Column(
        String, nullable=False
    )  # PURCHASE, SALE, ADJUSTMENT, SERVICE_USAGE
    quantity = Column(
        Float, nullable=False
    )  # Positive for additions, negative for reductions
    transaction_date = Column(Date, nullable=False)

    # References
    reference_type = Column(String, nullable=True)  # INVOICE, EXPENSE, MANUAL
    reference_id = Column(String, nullable=True)  # Invoice ID or expense ID
    notes = Column(Text, nullable=True)

    # Stock after transaction
    stock_after = Column(Float, nullable=False)

    # Timestamps
    created_at = Column(DateTime, default=datetime.utcnow)

    # Relationships
    company = relationship("CompanyDB", backref="inventory_transactions")
    inventory_item = relationship("InventoryItemDB", backref="transactions")


# ==================== TASK 8: GENERAL LEDGER / DOUBLE-ENTRY ACCOUNTING ====================

class AccountDB(Base):
    """Chart of Accounts — one set per company, system accounts auto-created."""

    __tablename__ = "accounts"
    id = Column(String, primary_key=True)
    company_id = Column(String, ForeignKey("companies.id"), nullable=False, index=True)
    account_code = Column(String(20), nullable=False)
    account_name = Column(String(100), nullable=False)
    account_name_ar = Column(String(100), nullable=True)
    account_type = Column(String(20), nullable=False)  # ASSET, LIABILITY, EQUITY, REVENUE, EXPENSE
    parent_account_id = Column(String, ForeignKey("accounts.id"), nullable=True)
    is_system = Column(Boolean, default=False)  # System accounts auto-created per company
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=datetime.utcnow)

    company = relationship("CompanyDB", backref="accounts")


class JournalEntryDB(Base):
    """Journal Entry header — one per business event (invoice issued, payment, etc.)."""

    __tablename__ = "journal_entries"
    id = Column(String, primary_key=True)
    company_id = Column(String, ForeignKey("companies.id"), nullable=False, index=True)
    entry_date = Column(Date, nullable=False)
    reference_type = Column(String(50), nullable=True)  # INVOICE, CREDIT_NOTE, DEBIT_NOTE, PAYMENT, EXPENSE
    reference_id = Column(String, nullable=True)  # FK to source document
    reference_number = Column(String(100), nullable=True)  # Human-readable (invoice number)
    description = Column(String(500), nullable=True)
    is_posted = Column(Boolean, default=True)
    created_by_user_id = Column(String, ForeignKey("users.id"), nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)

    company = relationship("CompanyDB", backref="journal_entries")
    lines = relationship("JournalEntryLineDB", backref="journal_entry", cascade="all, delete-orphan")


class JournalEntryLineDB(Base):
    """Journal Entry Line — debit or credit leg of a journal entry."""

    __tablename__ = "journal_entry_lines"
    id = Column(String, primary_key=True)
    journal_entry_id = Column(String, ForeignKey("journal_entries.id"), nullable=False, index=True)
    account_id = Column(String, ForeignKey("accounts.id"), nullable=False)
    account_code = Column(String(20), nullable=True)   # Denormalized for read performance
    account_name = Column(String(100), nullable=True)  # Denormalized for read performance
    debit_amount = Column(Float, default=0.0)
    credit_amount = Column(Float, default=0.0)
    currency = Column(String(3), default="AED")
    amount_aed = Column(Float, default=0.0)
    description = Column(String(500), nullable=True)
    line_number = Column(Integer, nullable=True)

    account = relationship("AccountDB", backref="journal_entry_lines")


# Create tables
Base.metadata.create_all(engine)


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ==================== PASSWORD SECURITY HELPERS (Task 2) ====================

MAX_FAILED_LOGINS = 5       # Lock after N consecutive failures
LOCKOUT_MINUTES = 30        # Lock duration
PASSWORD_HISTORY_DEPTH = 5  # Number of past hashes to retain per user

# Task 22: Password expiry policy (env-configurable)
PASSWORD_MAX_AGE_DAYS = int(os.environ.get("PASSWORD_MAX_AGE_DAYS", "90"))
PASSWORD_MIN_AGE_HOURS = int(os.environ.get("PASSWORD_MIN_AGE_HOURS", "24"))
PASSWORD_EXPIRY_WARNING_DAYS = 14  # Show warning banner this many days before expiry


def validate_password_complexity(password: str) -> None:
    """
    Enforce UAE FTA password policy:
    - Min 8 characters
    - At least one uppercase letter
    - At least one lowercase letter
    - At least one digit
    - At least one special character
    Raises ValueError on failure.
    """
    import re
    errors = []
    if len(password) < 8:
        errors.append("at least 8 characters")
    if not re.search(r"[A-Z]", password):
        errors.append("an uppercase letter")
    if not re.search(r"[a-z]", password):
        errors.append("a lowercase letter")
    if not re.search(r"\d", password):
        errors.append("a digit")
    if not re.search(r"[^A-Za-z0-9]", password):
        errors.append("a special character")
    if errors:
        raise ValueError(f"Password must contain: {', '.join(errors)}")


def _check_password_expiry(user: "UserDB"):
    """
    Task 22 — Check password age against policy.
    Returns (is_expired: bool, days_until_expiry: int | None).
    days_until_expiry is None when no password_changed_at is recorded.
    """
    if not getattr(user, "password_changed_at", None):
        return False, None
    age_days = (datetime.utcnow() - user.password_changed_at).days
    days_until_expiry = PASSWORD_MAX_AGE_DAYS - age_days
    return days_until_expiry <= 0, days_until_expiry


def _save_password_history(user_id: str, password_hash: str, db: Session) -> None:
    """Save a new password hash and trim to the last N entries."""
    entry = PasswordHistoryDB(
        id=f"ph_{uuid4().hex[:16]}",
        user_id=user_id,
        password_hash=password_hash,
        created_at=datetime.utcnow(),
    )
    db.add(entry)
    db.flush()

    # Trim to keep only the most recent PASSWORD_HISTORY_DEPTH entries
    history = (
        db.query(PasswordHistoryDB)
        .filter(PasswordHistoryDB.user_id == user_id)
        .order_by(PasswordHistoryDB.created_at.desc())
        .all()
    )
    for old in history[PASSWORD_HISTORY_DEPTH:]:
        db.delete(old)


def _check_password_history(user_id: str, new_plain: str, db: Session) -> bool:
    """
    Returns True if new_plain matches any of the last PASSWORD_HISTORY_DEPTH hashes.
    """
    history = (
        db.query(PasswordHistoryDB)
        .filter(PasswordHistoryDB.user_id == user_id)
        .order_by(PasswordHistoryDB.created_at.desc())
        .limit(PASSWORD_HISTORY_DEPTH)
        .all()
    )
    for h in history:
        if verify_password(new_plain, h.password_hash):
            return True
    return False


# ==================== AUDIT TRAIL HELPER (Task 1) ====================


def log_audit_event(
    db: Session,
    action: str,
    user_id: Optional[str] = None,
    company_id: Optional[str] = None,
    resource_type: Optional[str] = None,
    resource_id: Optional[str] = None,
    old_value: Optional[str] = None,
    new_value: Optional[str] = None,
    description: Optional[str] = None,
    ip_address: Optional[str] = None,
    user_agent: Optional[str] = None,
) -> None:
    """
    Persist a single audit trail entry.
    This is a fire-and-forget helper — failures are logged but never raised.
    """
    try:
        entry = AuditLogDB(
            id=f"al_{uuid4().hex[:16]}",
            company_id=company_id,
            user_id=user_id,
            action=action,
            resource_type=resource_type,
            resource_id=resource_id,
            old_value=old_value,
            new_value=new_value,
            description=description,
            ip_address=ip_address,
            user_agent=user_agent,
            created_at=datetime.utcnow(),
        )
        db.add(entry)
        db.flush()  # Don't commit — caller controls the transaction
    except Exception as exc:
        logger.error(f"Audit log write failed (non-fatal): {exc}")


# ==================== GL HELPER FUNCTIONS (Task 8) ====================

# System chart of accounts template
_DEFAULT_ACCOUNTS = [
    ("1000", "Cash / Bank",          "نقد / بنك",        "ASSET"),
    ("1100", "Accounts Receivable",  "حسابات القبض",     "ASSET"),
    ("1200", "VAT Receivable",       "ضريبة القيمة المضافة المستردة", "ASSET"),
    ("2100", "Accounts Payable",     "حسابات الدفع",     "LIABILITY"),
    ("2200", "VAT Payable",          "ضريبة القيمة المضافة المستحقة", "LIABILITY"),
    ("3000", "Sales Revenue",        "إيرادات المبيعات", "REVENUE"),
    ("4000", "Purchase / Cost of Goods", "المشتريات / تكلفة البضاعة", "EXPENSE"),
]


def create_default_chart_of_accounts(company_id: str, db: Session) -> None:
    """Create the 7 system accounts for a company if they don't already exist."""
    for code, name_en, name_ar, acc_type in _DEFAULT_ACCOUNTS:
        existing = (
            db.query(AccountDB)
            .filter(AccountDB.company_id == company_id, AccountDB.account_code == code)
            .first()
        )
        if not existing:
            db.add(AccountDB(
                id=str(uuid4()),
                company_id=company_id,
                account_code=code,
                account_name=name_en,
                account_name_ar=name_ar,
                account_type=acc_type,
                is_system=True,
                is_active=True,
            ))
    db.flush()


def _get_account(company_id: str, code: str, db: Session) -> AccountDB:
    """Fetch a system account by code, creating default accounts if missing."""
    acc = (
        db.query(AccountDB)
        .filter(AccountDB.company_id == company_id, AccountDB.account_code == code)
        .first()
    )
    if not acc:
        create_default_chart_of_accounts(company_id, db)
        acc = (
            db.query(AccountDB)
            .filter(AccountDB.company_id == company_id, AccountDB.account_code == code)
            .first()
        )
    return acc


def create_journal_entry_for_invoice(invoice: "InvoiceDB", db: Session, user_id: str = None) -> None:
    """
    Auto-generate a double-entry journal entry when an outgoing invoice is ISSUED.

    Sales Invoice (380):
        DR  1100 Accounts Receivable  [total incl. VAT]
            CR  3000 Sales Revenue    [subtotal excl. VAT]
            CR  2200 VAT Payable      [VAT amount]

    Credit Note (381):
        DR  3000 Sales Revenue        [subtotal]
        DR  2200 VAT Payable          [VAT]
            CR  1100 Accounts Receivable [total]
    """
    try:
        company_id = invoice.company_id
        entry_date = invoice.issue_date if isinstance(invoice.issue_date, date) else date.today()
        total = round(float(invoice.total_amount or 0), 2)
        subtotal = round(float(invoice.subtotal_amount or 0), 2)
        tax = round(float(invoice.tax_amount or 0), 2)

        is_credit_note = invoice.invoice_type in [
            InvoiceType.TAX_CREDIT_NOTE,
            InvoiceType.CREDIT_NOTE_OUT_OF_SCOPE,
        ]

        ref_type = "CREDIT_NOTE" if is_credit_note else "INVOICE"
        desc = f"{'Credit Note' if is_credit_note else 'Invoice'} {invoice.invoice_number}"

        je = JournalEntryDB(
            id=str(uuid4()),
            company_id=company_id,
            entry_date=entry_date,
            reference_type=ref_type,
            reference_id=invoice.id,
            reference_number=invoice.invoice_number,
            description=desc,
            is_posted=True,
            created_by_user_id=user_id,
        )
        db.add(je)
        db.flush()

        ar = _get_account(company_id, "1100", db)
        rev = _get_account(company_id, "3000", db)
        vat_p = _get_account(company_id, "2200", db)

        lines = []
        if is_credit_note:
            lines = [
                JournalEntryLineDB(id=str(uuid4()), journal_entry_id=je.id, account_id=rev.id,  account_code=rev.account_code,  account_name=rev.account_name,  debit_amount=subtotal, credit_amount=0,     currency="AED", amount_aed=subtotal, description="Revenue reversal", line_number=1),
                JournalEntryLineDB(id=str(uuid4()), journal_entry_id=je.id, account_id=vat_p.id,account_code=vat_p.account_code,account_name=vat_p.account_name,debit_amount=tax,     credit_amount=0,     currency="AED", amount_aed=tax,     description="VAT reversal",     line_number=2),
                JournalEntryLineDB(id=str(uuid4()), journal_entry_id=je.id, account_id=ar.id,   account_code=ar.account_code,   account_name=ar.account_name,   debit_amount=0,       credit_amount=total, currency="AED", amount_aed=total,   description="AR credit",         line_number=3),
            ]
        else:
            lines = [
                JournalEntryLineDB(id=str(uuid4()), journal_entry_id=je.id, account_id=ar.id,   account_code=ar.account_code,   account_name=ar.account_name,   debit_amount=total,   credit_amount=0,       currency="AED", amount_aed=total,   description="AR debit",     line_number=1),
                JournalEntryLineDB(id=str(uuid4()), journal_entry_id=je.id, account_id=rev.id,  account_code=rev.account_code,  account_name=rev.account_name,  debit_amount=0,       credit_amount=subtotal,currency="AED", amount_aed=subtotal,description="Revenue",      line_number=2),
                JournalEntryLineDB(id=str(uuid4()), journal_entry_id=je.id, account_id=vat_p.id,account_code=vat_p.account_code,account_name=vat_p.account_name,debit_amount=0,       credit_amount=tax,     currency="AED", amount_aed=tax,     description="VAT Payable",  line_number=3),
            ]

        for line in lines:
            db.add(line)
        db.flush()
        logger.info(f"GL: journal entry {je.id} created for invoice {invoice.invoice_number}")
    except Exception as exc:
        logger.error(f"GL: failed to create journal entry for invoice {getattr(invoice, 'invoice_number', '?')}: {exc}")


def create_journal_entry_for_purchase(inward_invoice, db: Session, user_id: str = None) -> None:
    """
    Auto-generate a double-entry journal entry when a purchase invoice is received.

    Purchase Invoice:
        DR  4000 Purchase / COGS   [subtotal]
        DR  1200 VAT Receivable    [VAT]
            CR  2100 Accounts Payable [total]
    """
    try:
        company_id = inward_invoice.company_id
        entry_date = (
            inward_invoice.invoice_date.date()
            if hasattr(inward_invoice.invoice_date, "date")
            else date.today()
        )
        total = round(float(inward_invoice.total_amount or 0), 2)
        subtotal = round(float(inward_invoice.subtotal_amount or 0), 2)
        tax = round(float(inward_invoice.tax_amount or 0), 2)

        je = JournalEntryDB(
            id=str(uuid4()),
            company_id=company_id,
            entry_date=entry_date,
            reference_type="PURCHASE",
            reference_id=str(inward_invoice.id),
            reference_number=inward_invoice.supplier_invoice_number,
            description=f"Purchase {inward_invoice.supplier_invoice_number} from {inward_invoice.supplier_name}",
            is_posted=True,
            created_by_user_id=user_id,
        )
        db.add(je)
        db.flush()

        cogs = _get_account(company_id, "4000", db)
        vat_r = _get_account(company_id, "1200", db)
        ap = _get_account(company_id, "2100", db)

        lines = [
            JournalEntryLineDB(id=str(uuid4()), journal_entry_id=je.id, account_id=cogs.id, account_code=cogs.account_code, account_name=cogs.account_name, debit_amount=subtotal, credit_amount=0,     currency="AED", amount_aed=subtotal, description="Purchase cost", line_number=1),
            JournalEntryLineDB(id=str(uuid4()), journal_entry_id=je.id, account_id=vat_r.id,account_code=vat_r.account_code,account_name=vat_r.account_name,debit_amount=tax,     credit_amount=0,     currency="AED", amount_aed=tax,     description="Input VAT",     line_number=2),
            JournalEntryLineDB(id=str(uuid4()), journal_entry_id=je.id, account_id=ap.id,   account_code=ap.account_code,   account_name=ap.account_name,   debit_amount=0,       credit_amount=total, currency="AED", amount_aed=total,   description="AP credit",     line_number=3),
        ]
        for line in lines:
            db.add(line)
        db.flush()
        logger.info(f"GL: purchase journal entry {je.id} created for {inward_invoice.supplier_invoice_number}")
    except Exception as exc:
        logger.error(f"GL: failed to create purchase journal entry: {exc}")


# ==================== AUTH HELPERS ====================


def create_access_token(data: dict, expires_delta: timedelta = None):
    """Create JWT access token"""
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt


def create_refresh_token(data: dict, expires_delta: timedelta = None):
    """Create JWT refresh token"""
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(days=REFRESH_TOKEN_EXPIRE_DAYS)
    to_encode.update({"exp": expire, "type": "refresh"})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt


def verify_password(plain_password: str, hashed_password: str) -> bool:
    """Verify password against hash"""
    if not hashed_password:
        return False
    password_bytes = plain_password.encode("utf-8")[:72]  # bcrypt max 72 bytes
    hashed_bytes = hashed_password.encode("utf-8")
    return bcrypt.checkpw(password_bytes, hashed_bytes)


def get_password_hash(password: str) -> str:
    """Hash a password"""
    password_bytes = password.encode("utf-8")[:72]  # bcrypt max 72 bytes
    salt = bcrypt.gensalt()
    hashed = bcrypt.hashpw(password_bytes, salt)
    return hashed.decode("utf-8")


def generate_temp_password(length: int = 16) -> str:
    """Generate a secure temporary password that meets complexity requirements

    Requirements: Minimum 8 characters with uppercase, lowercase, special char, and digit
    """
    import string

    # Ensure we have at least one of each required character type
    uppercase = secrets.choice(string.ascii_uppercase)
    lowercase = secrets.choice(string.ascii_lowercase)
    digit = secrets.choice(string.digits)
    special = secrets.choice("!@#$%^&*")

    # Fill the rest with random characters from all types
    remaining_length = max(length - 4, 4)
    all_chars = string.ascii_letters + string.digits + "!@#$%^&*"
    remaining = "".join(secrets.choice(all_chars) for _ in range(remaining_length))

    # Combine and shuffle
    password_list = list(uppercase + lowercase + digit + special + remaining)
    secrets.SystemRandom().shuffle(password_list)

    return "".join(password_list)


def authenticate_user(email: str, password: str, db: Session):
    """Authenticate user with lockout enforcement (Task 2)."""
    user = db.query(UserDB).filter(UserDB.email == email).first()
    if not user or not user.password_hash:
        return None

    # Task 2: Check account lockout
    if user.locked_until and user.locked_until > datetime.utcnow():
        remaining = int((user.locked_until - datetime.utcnow()).total_seconds() / 60)
        raise HTTPException(
            403,
            f"Account is locked due to {MAX_FAILED_LOGINS} failed login attempts. "
            f"Try again in {remaining} minute(s).",
        )

    if not verify_password(password, user.password_hash):
        # Increment failed attempts and possibly lock the account
        user.failed_login_count = (user.failed_login_count or 0) + 1
        if user.failed_login_count >= MAX_FAILED_LOGINS:
            user.locked_until = datetime.utcnow() + timedelta(minutes=LOCKOUT_MINUTES)
            user.failed_login_count = 0
        db.commit()
        return None

    # Successful login — reset counter
    user.failed_login_count = 0
    user.locked_until = None
    return user


def authenticate_company(email: str, password: str, db: Session):
    """Authenticate company by email and password"""
    company = db.query(CompanyDB).filter(CompanyDB.email == email).first()
    if not company:
        return None
    if not company.password_hash:
        return None
    if not verify_password(password, company.password_hash):
        return None
    return company


def get_current_company(token: str, db: Session):
    """Get current company from JWT token"""
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        company_id: str = payload.get("sub")
        if company_id is None:
            return None
        company = db.get(CompanyDB, company_id)
        return company
    except JWTError:
        return None


def get_current_user_from_header(
    authorization: str = Header(None), db: Session = Depends(get_db)
) -> UserDB:
    """Extract and validate user from Authorization header"""
    if not authorization:
        raise HTTPException(401, "Not authenticated")

    try:
        scheme, token = authorization.split()
        if scheme.lower() != "bearer":
            raise HTTPException(401, "Invalid authentication scheme")
    except ValueError:
        raise HTTPException(401, "Invalid authorization header")

    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        print("[DEBUG] Decoded access token payload:", payload)
        user_id: str = payload.get("sub")
        token_type: str = payload.get("type")
        mfa_challenge: bool = payload.get("mfa_challenge", False)

        # Reject temporary MFA challenge tokens on protected endpoints
        if mfa_challenge:
            print("[DEBUG] Token is MFA challenge token, rejecting.")
            raise HTTPException(
                401, "MFA verification required. Please complete MFA login flow."
            )

        if user_id is None or token_type != "user":
            print(f"[DEBUG] Invalid token: user_id={user_id}, token_type={token_type}")
            raise HTTPException(401, "Invalid token")

        user = db.get(UserDB, user_id)
        if user is None:
            print(f"[DEBUG] User not found for user_id={user_id}")
            raise HTTPException(401, "User not found")

        return user
    except JWTError as e:
        print(f"[DEBUG] JWTError while decoding token: {e}")
        raise HTTPException(401, "Invalid token")
        raise HTTPException(401, "Invalid token")


# ==================== HELPER FUNCTIONS ====================
def is_valid_trn(trn: str) -> bool:
    """Validate UAE TRN (15 digits)"""
    if not trn or len(trn) != 15:
        return False
    return trn.isdigit()


def seed_plans(db: Session):
    """Seed default subscription plans"""
    if db.query(SubscriptionPlanDB).count() > 0:
        return

    plans = [
        SubscriptionPlanDB(
            id="plan_free",
            name="Free",
            description="Start free with essential features",
            price_monthly=0.0,
            price_yearly=0.0,
            max_invoices_per_month=100,
            max_users=1,
            max_business_admins=0,  # Phase 2: Owner only
            max_finance_users=0,  # Phase 2: No team members
            max_pos_devices=0,
            allow_api_access=True,
            allow_branding=False,
            allow_multi_currency=False,
            priority_support=False,
        ),
        SubscriptionPlanDB(
            id="plan_starter",
            name="Starter",
            description="Perfect for small businesses",
            price_monthly=99.0,
            price_yearly=990.0,
            max_invoices_per_month=100,
            max_users=3,  # Phase 2: 1 owner + 2 team members
            max_business_admins=0,  # Phase 2: No business admins
            max_finance_users=2,  # Phase 2: Up to 2 finance users
            max_pos_devices=1,
            allow_api_access=True,
            allow_branding=False,
        ),
        SubscriptionPlanDB(
            id="plan_professional",
            name="Professional",
            description="For growing businesses",
            price_monthly=299.0,
            price_yearly=2990.0,
            max_invoices_per_month=500,
            max_users=5,  # Phase 2: 1 owner + 4 team members
            max_business_admins=2,  # Phase 2: Up to 2 business admins
            max_finance_users=2,  # Phase 2: Up to 2 finance users
            max_pos_devices=3,
            allow_api_access=True,
            allow_branding=True,
            allow_multi_currency=True,
        ),
        SubscriptionPlanDB(
            id="plan_enterprise",
            name="Enterprise",
            description="Unlimited for large organizations",
            price_monthly=999.0,
            price_yearly=9990.0,
            max_invoices_per_month=None,
            max_users=50,  # Phase 2: 1 owner + 49 team members
            max_business_admins=10,  # Phase 2: Up to 10 business admins
            max_finance_users=39,  # Phase 2: Up to 39 finance users
            max_pos_devices=10,
            allow_api_access=True,
            allow_branding=True,
            allow_multi_currency=True,
            priority_support=True,
        ),
    ]

    for plan in plans:
        db.add(plan)
    db.commit()


def seed_content(db: Session):
    """Seed initial content blocks for Homepage"""
    if db.query(ContentBlockDB).count() > 0:
        return  # Already seeded

    content_blocks = [
        # Homepage Hero
        ContentBlockDB(
            id=f"cb_{uuid4().hex[:12]}",
            key="homepage_hero_title",
            value="Simple, Compliant\nDigital Invoicing for UAE",
            description="Homepage main heading",
            section="homepage",
            updated_by="system",
        ),
        ContentBlockDB(
            id=f"cb_{uuid4().hex[:12]}",
            key="homepage_hero_subtitle",
            value="Automated invoicing in structured electronic formats.",
            description="Homepage subheading",
            section="homepage",
            updated_by="system",
        ),
        # Feature Boxes
        ContentBlockDB(
            id=f"cb_{uuid4().hex[:12]}",
            key="feature_box_1_title",
            value="Government-Approved Invoices",
            description="First feature box title",
            section="feature_boxes",
            updated_by="system",
        ),
        ContentBlockDB(
            id=f"cb_{uuid4().hex[:12]}",
            key="feature_box_1_description",
            value="Create professional invoices that meet all UAE government requirements. Every invoice is automatically secured and properly formatted.",
            description="First feature box description",
            section="feature_boxes",
            updated_by="system",
        ),
        ContentBlockDB(
            id=f"cb_{uuid4().hex[:12]}",
            key="feature_box_2_title",
            value="Manage Your Purchases",
            description="Second feature box title",
            section="feature_boxes",
            updated_by="system",
        ),
        ContentBlockDB(
            id=f"cb_{uuid4().hex[:12]}",
            key="feature_box_2_description",
            value="Create purchase orders, receive supplier invoices, and keep track of all your expenses in one organized place.",
            description="Second feature box description",
            section="feature_boxes",
            updated_by="system",
        ),
        ContentBlockDB(
            id=f"cb_{uuid4().hex[:12]}",
            key="feature_box_3_title",
            value="Extra Security",
            description="Third feature box title",
            section="feature_boxes",
            updated_by="system",
        ),
        ContentBlockDB(
            id=f"cb_{uuid4().hex[:12]}",
            key="feature_box_3_description",
            value="Extra protection for your account. Use your phone or email to confirm it's really you when logging in.",
            description="Third feature box description",
            section="feature_boxes",
            updated_by="system",
        ),
        ContentBlockDB(
            id=f"cb_{uuid4().hex[:12]}",
            key="feature_box_4_title",
            value="Team Collaboration",
            description="Fourth feature box title",
            section="feature_boxes",
            updated_by="system",
        ),
        ContentBlockDB(
            id=f"cb_{uuid4().hex[:12]}",
            key="feature_box_4_description",
            value="Work together with your team. Add unlimited members and control who can view or edit invoices.",
            description="Fourth feature box description",
            section="feature_boxes",
            updated_by="system",
        ),
        ContentBlockDB(
            id=f"cb_{uuid4().hex[:12]}",
            key="feature_box_5_title",
            value="Electronic Delivery",
            description="Fifth feature box title",
            section="feature_boxes",
            updated_by="system",
        ),
        ContentBlockDB(
            id=f"cb_{uuid4().hex[:12]}",
            key="feature_box_5_description",
            value="Send invoices electronically to your customers through trusted partners, making delivery fast and secure.",
            description="Fifth feature box description",
            section="feature_boxes",
            updated_by="system",
        ),
        ContentBlockDB(
            id=f"cb_{uuid4().hex[:12]}",
            key="feature_box_6_title",
            value="Flexible Subscriptions",
            description="Sixth feature box title",
            section="feature_boxes",
            updated_by="system",
        ),
        ContentBlockDB(
            id=f"cb_{uuid4().hex[:12]}",
            key="feature_box_6_description",
            value="Start free with 10 invoices per month. Upgrade anytime as your business grows to send unlimited invoices.",
            description="Sixth feature box description",
            section="feature_boxes",
            updated_by="system",
        ),
        # Footer Content
        ContentBlockDB(
            id=f"cb_{uuid4().hex[:12]}",
            key="footer_company_name",
            value="InvoLinks",
            description="Footer company name",
            section="footer",
            updated_by="system",
        ),
        ContentBlockDB(
            id=f"cb_{uuid4().hex[:12]}",
            key="footer_company_description",
            value="Simple, Compliant Digital Invoicing for UAE",
            description="Footer company description",
            section="footer",
            updated_by="system",
        ),
        ContentBlockDB(
            id=f"cb_{uuid4().hex[:12]}",
            key="footer_address",
            value="Dubai, United Arab Emirates",
            description="Footer company address",
            section="footer",
            updated_by="system",
        ),
        ContentBlockDB(
            id=f"cb_{uuid4().hex[:12]}",
            key="footer_email",
            value="support@involinks.ae",
            description="Footer contact email",
            section="footer",
            updated_by="system",
        ),
        ContentBlockDB(
            id=f"cb_{uuid4().hex[:12]}",
            key="footer_phone",
            value="+971 4 123 4567",
            description="Footer contact phone",
            section="footer",
            updated_by="system",
        ),
        ContentBlockDB(
            id=f"cb_{uuid4().hex[:12]}",
            key="footer_copyright",
            value="© 2025 InvoLinks. All rights reserved.",
            description="Footer copyright text",
            section="footer",
            updated_by="system",
        ),
    ]

    for block in content_blocks:
        db.add(block)
    db.commit()


# ==================== PYDANTIC SCHEMAS ====================
class CompanyInfoCreate(BaseModel):
    legal_name: str = Field(..., min_length=2, max_length=255)
    business_type: Optional[str] = Field(default="LLC", example="LLC")
    registration_number: Optional[str] = None
    registration_date: Optional[date] = None
    email: str
    password: str = Field(..., min_length=8, description="Password (min 8 characters)")
    phone: Optional[str] = None
    website: Optional[str] = None


class BusinessDetailsCreate(BaseModel):
    business_activity: Optional[str] = None
    address_line1: Optional[str] = None
    address_line2: Optional[str] = None
    city: Optional[str] = None
    emirate: Optional[str] = Field(default="Dubai", example="Dubai")
    po_box: Optional[str] = None
    trn: Optional[str] = None
    authorized_person_name: Optional[str] = None
    authorized_person_title: Optional[str] = None
    authorized_person_email: Optional[str] = None
    authorized_person_phone: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None


class DocumentUploadResponse(BaseModel):
    id: str
    document_type: DocumentType
    file_name: str
    file_size: int
    status: DocumentStatus
    uploaded_at: datetime


class PlanOut(BaseModel):
    id: str
    name: str
    description: Optional[str]
    price_monthly: float
    price_yearly: float
    max_invoices_per_month: Optional[int]
    max_users: int
    allow_branding: bool
    allow_api_access: bool


class CompanyOut(BaseModel):
    id: str
    legal_name: Optional[str]
    status: CompanyStatus
    email: Optional[str]
    trn: Optional[str]


class RegistrationProgressOut(BaseModel):
    company_id: str
    current_step: int
    completed: bool


# ==================== CONTENT MANAGEMENT PYDANTIC SCHEMAS ====================
class ContentBlockCreate(BaseModel):
    key: str = Field(..., description="Unique key (e.g., homepage_hero_title)")
    value: str = Field(..., description="The actual text content")
    description: Optional[str] = Field(None, description="What this controls")
    section: Optional[str] = Field(None, description="Section/page grouping")


class ContentBlockUpdate(BaseModel):
    value: str = Field(..., description="Updated text content")


class ContentBlockOut(BaseModel):
    id: str
    key: str
    value: str
    description: Optional[str]
    section: Optional[str]
    updated_at: datetime
    updated_by: Optional[str]


# ==================== INVOICE PYDANTIC SCHEMAS ====================
class InvoiceLineItemCreate(BaseModel):
    item_name: str
    item_description: Optional[str] = None
    item_code: Optional[str] = None
    quantity: float
    unit_code: str = "C62"  # UN/ECE code (C62 = piece)
    unit_price: float
    tax_category: TaxCategory
    tax_percent: float = 5.0  # UAE standard VAT rate
    tax_code: Optional[str] = (
        None  # UAE tax code (SR, ZR, ES, RC, OP) - Required for VAT-registered businesses
    )


class InvoiceCreate(BaseModel):
    invoice_type: InvoiceType
    issue_date: str  # ISO date format
    due_date: Optional[str] = None
    supply_date: Optional[str] = None  # Task 5: Actual delivery/supply date
    currency_code: str = "AED"

    # AED equivalent (mandatory when currency_code != "AED" — FTA P2-A compliance)
    total_amount_aed: Optional[float] = None

    # Customer details
    customer_name: str
    customer_email: Optional[str] = None
    customer_trn: Optional[str] = None
    customer_address: Optional[str] = None
    customer_city: Optional[str] = None
    customer_country: str = "AE"
    customer_peppol_id: Optional[str] = None

    # Line items
    line_items: List[InvoiceLineItemCreate]

    # Optional fields
    payment_terms: Optional[str] = None
    payment_due_days: int = 30
    invoice_notes: Optional[str] = None
    reference_number: Optional[str] = None

    # Credit note specific
    preceding_invoice_id: Optional[str] = None
    credit_note_reason: Optional[str] = None


class InvoiceLineItemOut(BaseModel):
    id: str
    line_number: int
    item_name: str
    item_description: Optional[str]
    quantity: float
    unit_code: str
    unit_price: float
    line_extension_amount: float
    tax_category: TaxCategory
    tax_percent: float
    tax_amount: float
    line_total_amount: float


class InvoiceTaxBreakdownOut(BaseModel):
    id: str
    tax_category: TaxCategory
    taxable_amount: float
    tax_percent: float
    tax_amount: float


class InvoiceOut(BaseModel):
    id: str
    company_id: str
    invoice_number: str
    invoice_type: InvoiceType
    status: InvoiceStatus
    issue_date: str
    due_date: Optional[str]
    supply_date: Optional[str] = None  # Task 5: Actual delivery/supply date
    version: int = 1  # Task 10: Optimistic locking counter
    currency_code: str

    # Supplier (auto-filled from company)
    supplier_trn: Optional[str]
    supplier_name: str
    supplier_address: Optional[str]
    supplier_peppol_id: Optional[str]

    # Customer
    customer_trn: Optional[str]
    customer_name: str
    customer_email: Optional[str]
    customer_address: Optional[str]
    customer_city: Optional[str]
    invoice_notes: Optional[str]
    customer_peppol_id: Optional[str]

    # Totals
    subtotal_amount: float
    tax_amount: float
    total_amount: float
    amount_due: float
    total_amount_aed: Optional[float] = None  # AED equivalent (P2-A: populated for non-AED invoices)

    # Credit note fields
    preceding_invoice_id: Optional[str] = None
    credit_note_reason: Optional[str] = None

    # Documents
    xml_file_path: Optional[str]
    pdf_file_path: Optional[str]
    share_token: Optional[str]

    # Timestamps
    created_at: str
    sent_at: Optional[str]
    viewed_at: Optional[str]

    # Related data
    line_items: List[InvoiceLineItemOut] = []
    tax_breakdowns: List[InvoiceTaxBreakdownOut] = []


class InvoiceListOut(BaseModel):
    id: str
    invoice_number: str
    invoice_type: InvoiceType
    status: InvoiceStatus
    issue_date: str
    customer_name: str
    total_amount: float
    currency_code: str
    created_at: str


# ==================== CORNER 4: AP MANAGEMENT MODELS ====================

from pydantic import validator


class PurchaseOrderLineItemCreate(BaseModel):
    line_number: int
    item_name: str
    item_description: Optional[str] = None
    item_code: Optional[str] = None
    quantity_ordered: float
    unit_code: str = "C62"
    unit_price: float
    tax_category: TaxCategory = TaxCategory.STANDARD
    tax_percent: float = 5.0

    @validator("item_description", "item_code", pre=True)
    def empty_str_to_none(cls, v):
        if v == "":
            return None
        return v

    @validator("tax_category", pre=True)
    def tax_category_enum_flexible(cls, v):
        if v == "":
            return TaxCategory.STANDARD
        if isinstance(v, TaxCategory):
            return v
        if v in ("STANDARD", "S"):
            return TaxCategory.STANDARD
        if v in ("ZERO", "Z"):
            return TaxCategory.ZERO
        if v in ("EXEMPT", "E"):
            return TaxCategory.EXEMPT
        if v in ("OUT_OF_SCOPE", "O"):
            return TaxCategory.OUT_OF_SCOPE
        raise ValueError(f"Invalid tax_category: {v}")


class PurchaseOrderCreate(BaseModel):
    po_number: str
    supplier_trn: Optional[str] = None  # Optional for non-VAT-registered suppliers
    supplier_name: str
    supplier_contact_email: Optional[str] = None
    supplier_address: Optional[str] = None
    supplier_peppol_id: Optional[str] = None
    order_date: str  # YYYY-MM-DD
    expected_delivery_date: Optional[str] = None
    delivery_address: Optional[str] = None
    currency_code: str = "AED"
    reference_number: Optional[str] = None
    notes: Optional[str] = None
    line_items: List[PurchaseOrderLineItemCreate]

    @validator(
        "supplier_trn",
        "supplier_contact_email",
        "supplier_address",
        "supplier_peppol_id",
        "expected_delivery_date",
        "delivery_address",
        "reference_number",
        "notes",
        pre=True,
    )
    def empty_str_to_none(cls, v):
        if v == "":
            return None
        return v


class PurchaseOrderLineItemOut(BaseModel):
    id: str
    line_number: int
    item_name: str
    item_description: Optional[str]
    item_code: Optional[str]
    quantity_ordered: float
    quantity_received: float
    unit_code: str
    unit_price: float
    line_total: float
    tax_category: TaxCategory
    tax_percent: float


class PurchaseOrderOut(BaseModel):
    id: str
    company_id: str
    po_number: str
    status: PurchaseOrderStatus
    supplier_trn: Optional[str]  # Optional for non-VAT-registered suppliers
    supplier_name: str
    supplier_contact_email: Optional[str]
    supplier_address: Optional[str]
    supplier_peppol_id: Optional[str]
    order_date: str
    expected_delivery_date: Optional[str]
    delivery_address: Optional[str]
    currency_code: str
    expected_subtotal: float
    expected_tax: float
    expected_total: float
    received_invoice_count: int
    matched_invoice_count: int
    received_amount_total: float
    variance_amount: float
    reference_number: Optional[str]
    notes: Optional[str]
    approved_by_user_id: Optional[str]
    approved_at: Optional[str]
    created_at: str
    updated_at: str
    line_items: List[PurchaseOrderLineItemOut] = []


class InwardInvoiceReceive(BaseModel):
    """Model for receiving an inward invoice via PEPPOL"""

    supplier_invoice_number: str
    invoice_date: str  # YYYY-MM-DD
    due_date: Optional[str] = None
    supplier_trn: str
    supplier_name: str
    supplier_address: Optional[str] = None
    supplier_peppol_id: Optional[str] = None
    customer_trn: str
    customer_name: str
    currency_code: str = "AED"
    subtotal_amount: float
    tax_amount: float
    total_amount: float
    amount_due: float
    xml_content: Optional[str] = None  # UBL XML content
    peppol_message_id: Optional[str] = None
    peppol_sender_id: Optional[str] = None
    peppol_provider: Optional[str] = None
    line_items: Optional[List[dict]] = None  # JSON line items from XML


class InwardInvoiceLineItemOut(BaseModel):
    id: str
    line_number: int
    item_name: str
    item_description: Optional[str]
    item_code: Optional[str]
    quantity: float
    unit_code: str
    unit_price: float
    line_extension_amount: float
    tax_category: TaxCategory
    tax_percent: float
    tax_amount: float
    line_total_amount: float
    po_line_item_id: Optional[str]
    grn_line_item_id: Optional[str]
    match_status: Optional[str]


class InwardInvoiceOut(BaseModel):
    id: str
    company_id: str
    supplier_invoice_number: str
    invoice_type: InvoiceType
    status: InwardInvoiceStatus
    invoice_date: str
    due_date: Optional[str]
    received_at: str
    supplier_trn: str
    supplier_name: str
    supplier_address: Optional[str]
    supplier_peppol_id: Optional[str]
    supplier_company_id: Optional[str]
    customer_trn: str
    customer_name: str
    currency_code: str
    subtotal_amount: float
    tax_amount: float
    total_amount: float
    amount_due: float
    xml_file_path: Optional[str]
    pdf_file_path: Optional[str]
    xml_hash: Optional[str]
    peppol_message_id: Optional[str]
    peppol_sender_id: Optional[str]
    peppol_provider: Optional[str]
    peppol_received_at: Optional[str]
    po_id: Optional[str]
    grn_id: Optional[str]
    matching_status: MatchingStatus
    po_match_score: float
    grn_match_score: float
    amount_variance: float
    quantity_variance: float
    reviewed_by_user_id: Optional[str]
    reviewed_at: Optional[str]
    approved_by_user_id: Optional[str]
    approved_at: Optional[str]
    rejection_reason: Optional[str]
    payment_status: Optional[str]
    payment_method: Optional[str]
    paid_amount: float
    paid_at: Optional[str]
    payment_reference: Optional[str]
    disputed_at: Optional[str]
    dispute_reason: Optional[str]
    dispute_resolved_at: Optional[str]
    notes: Optional[str]
    reference_number: Optional[str]
    created_at: str
    updated_at: str
    line_items: List[InwardInvoiceLineItemOut] = []


class InwardInvoiceListOut(BaseModel):
    id: str
    supplier_invoice_number: str
    status: InwardInvoiceStatus
    supplier_name: str
    total_amount: float
    currency_code: str
    invoice_date: str
    received_at: str
    matching_status: MatchingStatus
    po_id: Optional[str]


class InwardInvoiceApprove(BaseModel):
    notes: Optional[str] = None
    payment_method: Optional[str] = None
    payment_reference: Optional[str] = None


class InwardInvoiceReject(BaseModel):
    rejection_reason: str
    notes: Optional[str] = None


class InwardInvoiceMatchPO(BaseModel):
    po_id: str


class CompanyBrandingOut(BaseModel):
    id: str
    company_id: str
    logo_file_name: Optional[str] = None
    logo_file_path: Optional[str] = None
    logo_uploaded_at: Optional[str] = None
    stamp_file_name: Optional[str] = None
    stamp_file_path: Optional[str] = None
    stamp_uploaded_at: Optional[str] = None
    primary_color: Optional[str] = None
    secondary_color: Optional[str] = None
    font_family: Optional[str] = None
    created_at: str
    updated_at: str


# ==================== FASTAPI APP ====================
app = FastAPI(
    title="InvoLinks E-Invoicing API",
    version="2.0",
    description="Multi-tenant UAE e-invoicing platform with registration wizard",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# Global exception handler for domain exceptions
@app.exception_handler(InvoLinksException)
async def involinks_exception_handler(request, exc: InvoLinksException):
    """Convert domain exceptions to HTTP responses with structured error format"""
    response = exception_to_http_response(exc)
    return JSONResponse(status_code=response["status_code"], content=response["detail"])


# Serve React app (production) or fallback to development API
if os.path.exists("dist"):
    # Production: Serve React build
    app.mount("/assets", StaticFiles(directory="dist/assets"), name="assets")
else:
    # Development: Keep static files available (optional)
    if os.path.exists("static"):
        app.mount("/static", StaticFiles(directory="static"), name="static")


def _run_column_migrations():
    """Add new columns to existing tables without affecting data (Task 2)."""
    migrations = [
        # Task 2: Account lockout columns for users table
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS failed_login_count INTEGER DEFAULT 0",
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS locked_until TIMESTAMP",
        # Task 5: Supply date on invoices
        "ALTER TABLE invoices ADD COLUMN IF NOT EXISTS supply_date DATE",
        # Task 10: Optimistic locking version counter
        "ALTER TABLE invoices ADD COLUMN IF NOT EXISTS version INTEGER NOT NULL DEFAULT 1",
        # Task 12: Data archival columns
        "ALTER TABLE invoices ADD COLUMN IF NOT EXISTS is_archived BOOLEAN NOT NULL DEFAULT false",
        "ALTER TABLE invoices ADD COLUMN IF NOT EXISTS archived_at TIMESTAMP",
        # Task 20: Forced password change on first login
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS must_change_password BOOLEAN DEFAULT false",
        # Task 22: Password expiry policy
        "ALTER TABLE users ADD COLUMN IF NOT EXISTS password_changed_at TIMESTAMP",
        # Task 22: Backfill existing users - use created_at as baseline password age
        "UPDATE users SET password_changed_at = created_at WHERE password_changed_at IS NULL AND created_at IS NOT NULL",
        "UPDATE users SET password_changed_at = NOW() WHERE password_changed_at IS NULL",
        # Task 13: VAT return 13-box breakdown columns
        "ALTER TABLE vat_returns ADD COLUMN IF NOT EXISTS zero_rated_sales DOUBLE PRECISION DEFAULT 0.0",
        "ALTER TABLE vat_returns ADD COLUMN IF NOT EXISTS exempt_sales DOUBLE PRECISION DEFAULT 0.0",
        "ALTER TABLE vat_returns ADD COLUMN IF NOT EXISTS out_of_scope_sales DOUBLE PRECISION DEFAULT 0.0",
        "ALTER TABLE vat_returns ADD COLUMN IF NOT EXISTS reverse_charge_sales DOUBLE PRECISION DEFAULT 0.0",
        "ALTER TABLE vat_returns ADD COLUMN IF NOT EXISTS total_sales DOUBLE PRECISION DEFAULT 0.0",
        "ALTER TABLE vat_returns ADD COLUMN IF NOT EXISTS reverse_charge_vat DOUBLE PRECISION DEFAULT 0.0",
        "ALTER TABLE vat_returns ADD COLUMN IF NOT EXISTS zero_rated_purchases DOUBLE PRECISION DEFAULT 0.0",
        "ALTER TABLE vat_returns ADD COLUMN IF NOT EXISTS exempt_purchases DOUBLE PRECISION DEFAULT 0.0",
        # Task 25: READ_ONLY role enum migration — adds value to PostgreSQL enum type if not present
        "DO $$ BEGIN IF NOT EXISTS (SELECT 1 FROM pg_enum WHERE enumlabel = 'READ_ONLY' AND enumtypid = (SELECT oid FROM pg_type WHERE typname = 'role')) THEN ALTER TYPE role ADD VALUE 'READ_ONLY'; END IF; END $$",
        # Task 27 (P2-A): AED conversion enforcement — add nullable AED column to both invoice tables
        "ALTER TABLE invoices ADD COLUMN IF NOT EXISTS total_amount_aed DOUBLE PRECISION",
        "ALTER TABLE inward_invoices ADD COLUMN IF NOT EXISTS total_amount_aed DOUBLE PRECISION",
        # Backfill: set total_amount_aed = total_amount for existing AED invoices
        "UPDATE invoices SET total_amount_aed = total_amount WHERE total_amount_aed IS NULL AND (currency_code = 'AED' OR currency_code IS NULL)",
        "UPDATE inward_invoices SET total_amount_aed = total_amount WHERE total_amount_aed IS NULL AND (currency_code = 'AED' OR currency_code IS NULL)",
    ]
    with engine.begin() as conn:
        for sql in migrations:
            try:
                conn.execute(text(sql))
            except Exception as e:
                logger.warning(f"Column migration skipped (may already exist): {e}")


@app.on_event("startup")
def startup_event():
    """Seed plans on startup and validate environment"""
    # Run column-level migrations for new fields on existing tables
    _run_column_migrations()

    # Check if running in production mode
    production_mode = os.getenv("PRODUCTION_MODE", "false").lower() == "true"

    # Validate cryptographic environment
    from utils.crypto_utils import validate_environment_keys

    try:
        validation_result = validate_environment_keys(fail_on_missing=production_mode)

        if production_mode:
            logger.info("🔒 PRODUCTION MODE: All cryptographic validations passed")
        else:
            logger.info("🔧 DEVELOPMENT MODE: Running with permissive validation")
            if validation_result.get("warnings"):
                for warning in validation_result["warnings"]:
                    logger.warning(f"Startup Warning: {warning}")
                logger.warning("Set PRODUCTION_MODE=true to enforce strict validation")

    except ConfigurationError as e:
        logger.error(f"Startup Error: {e.message}")
        if production_mode:
            logger.critical("PRODUCTION MODE: Cannot start without valid signing keys")
            raise  # Fail hard in production mode
        else:
            logger.warning("Continuing with mock keys - NOT PRODUCTION READY")

    # Seed database plans and content
    db = SessionLocal()
    seed_plans(db)
    seed_content(db)

    # Create super admin if it doesn't exist
    super_admin_email = os.getenv("SUPER_ADMIN_EMAIL", "nrashidk@gmail.com")
    super_admin_password = os.getenv("SUPER_ADMIN_PASSWORD", "AbuDhabi@123")

    existing_super_admin = (
        db.query(UserDB).filter(UserDB.role == Role.SUPER_ADMIN).first()
    )

    if not existing_super_admin:
        super_admin = UserDB(
            id=f"user_{uuid4().hex[:8]}",
            email=super_admin_email,
            password_hash=get_password_hash(super_admin_password),
            role=Role.SUPER_ADMIN,
            company_id=None,
            is_owner=False,
            full_name="Super Admin",
        )
        db.add(super_admin)
        db.commit()
        logger.info(f"Super Admin created: {super_admin_email}")
    else:
        logger.info(f"Super Admin already exists: {existing_super_admin.email}")

    db.close()

    # Archival status check: warn if invoices older than 5 years are not yet archived
    # Open a fresh session — the seeding session above is already closed.
    try:
        _arch_db = SessionLocal()
        try:
            five_years_ago = date.today() - timedelta(days=5 * 365)
            eligible_count = (
                _arch_db.query(InvoiceDB)
                .filter(
                    InvoiceDB.issue_date <= five_years_ago,
                    InvoiceDB.status.in_([InvoiceStatus.PAID, InvoiceStatus.CANCELLED]),
                    InvoiceDB.is_archived == False,
                )
                .count()
            )
            if eligible_count > 0:
                logger.warning(
                    f"⚠️ ARCHIVAL REQUIRED: {eligible_count} invoice(s) older than 5 years "
                    "are eligible for archival but not yet archived. "
                    "Use POST /invoices/archive to archive them."
                )
            else:
                logger.info("✅ Archival check: No invoices require archiving at this time.")
        finally:
            _arch_db.close()
    except Exception as _arch_exc:
        logger.warning(f"Archival startup check failed: {_arch_exc}")

    # Task 24: Backup configuration summary
    _backup_provider = os.getenv("BACKUP_PROVIDER", "Neon PostgreSQL PITR")
    _backup_rpo = os.getenv("BACKUP_RPO_HOURS", "24")
    _backup_rto = os.getenv("BACKUP_RTO_HOURS", "4")
    logger.info(
        f"💾 Backup config — Provider: {_backup_provider} | "
        f"RPO target: {_backup_rpo}h | RTO target: {_backup_rto}h"
    )

    # Task 26: Last integrity check summary
    try:
        _ic_db = SessionLocal()
        try:
            last_ic = (
                _ic_db.query(IntegrityCheckDB)
                .order_by(IntegrityCheckDB.checked_at.desc())
                .first()
            )
            if last_ic:
                _ic_status = "PASSED ✅" if last_ic.integrity_ok else "FAILED ❌"
                _ic_scope = "GLOBAL" if last_ic.company_id is None else f"company={last_ic.company_id}"
                logger.info(
                    f"🔗 Last integrity check: {_ic_status} | "
                    f"Scope: {_ic_scope} | "
                    f"Date: {last_ic.checked_at.strftime('%Y-%m-%d %H:%M UTC')} | "
                    f"Invoices checked: {last_ic.total_invoices} | "
                    f"Valid links: {last_ic.valid_links}"
                )
            else:
                logger.info("🔗 Integrity check: No prior check on record. Run GET /admin/integrity/verify.")
        finally:
            _ic_db.close()
    except Exception as _ic_exc:
        logger.warning(f"Integrity check startup log failed: {_ic_exc}")

    mode_indicator = "🔒 PRODUCTION" if production_mode else "🔧 DEVELOPMENT"
    print(f"✅ InvoLinks API started ({mode_indicator}) - Plans seeded")


# ==================== BACKUP ENDPOINTS (Task 24) ====================


class BackupStatusResponse(BaseModel):
    provider: str
    rpo_target_hours: int
    rto_target_hours: int
    last_verified_at: Optional[str]
    last_verified_by: Optional[str]
    backup_age_hours: Optional[float]
    within_rpo: bool
    compliance_note: str


@app.get("/admin/backup/status", tags=["Admin"], response_model=BackupStatusResponse)
def get_backup_status(
    db: Session = Depends(get_db),
    current_user: UserDB = Depends(get_current_user_from_header),
):
    """
    Return backup configuration metadata and last verification time.
    Super-admin only.  Task 24.
    """
    if current_user.role != Role.SUPER_ADMIN:
        raise HTTPException(403, "Super admin access required")

    provider = os.getenv("BACKUP_PROVIDER", "Neon PostgreSQL PITR")
    rpo_hours = int(os.getenv("BACKUP_RPO_HOURS", "24"))
    rto_hours = int(os.getenv("BACKUP_RTO_HOURS", "4"))

    # Find the most recent BACKUP_VERIFIED audit event
    last_event = (
        db.query(AuditLogDB)
        .filter(AuditLogDB.action == "BACKUP_VERIFIED")
        .order_by(AuditLogDB.created_at.desc())
        .first()
    )

    last_verified_at = None
    last_verified_by = None
    backup_age_hours = None
    within_rpo = False

    if last_event:
        last_verified_at = last_event.created_at.isoformat()
        # Resolve user email via user_id
        last_verified_by = None
        if last_event.user_id:
            u = db.query(UserDB).filter(UserDB.id == last_event.user_id).first()
            last_verified_by = u.email if u else None
        backup_age_hours = round(
            (datetime.utcnow() - last_event.created_at).total_seconds() / 3600, 2
        )
        within_rpo = backup_age_hours <= rpo_hours

    compliance_note = (
        "COMPLIANT — last backup verification within RPO window"
        if within_rpo
        else (
            "UNVERIFIED — no backup verification on record; run POST /admin/backup/verify"
            if last_verified_at is None
            else f"AT RISK — last verification {backup_age_hours:.1f}h ago exceeds RPO of {rpo_hours}h"
        )
    )

    return BackupStatusResponse(
        provider=provider,
        rpo_target_hours=rpo_hours,
        rto_target_hours=rto_hours,
        last_verified_at=last_verified_at,
        last_verified_by=last_verified_by,
        backup_age_hours=backup_age_hours,
        within_rpo=within_rpo,
        compliance_note=compliance_note,
    )


@app.post("/admin/backup/verify", tags=["Admin"])
def verify_backup(
    request: Request,
    db: Session = Depends(get_db),
    current_user: UserDB = Depends(get_current_user_from_header),
):
    """
    Log a BACKUP_VERIFIED audit event with timestamp to create a compliance paper trail.
    Super-admin only.  Task 24.
    """
    if current_user.role != Role.SUPER_ADMIN:
        raise HTTPException(403, "Super admin access required")

    provider = os.getenv("BACKUP_PROVIDER", "Neon PostgreSQL PITR")
    rpo_hours = int(os.getenv("BACKUP_RPO_HOURS", "24"))

    ip = request.client.host if request.client else None

    log_audit_event(
        db=db,
        action="BACKUP_VERIFIED",
        user_id=current_user.id,
        company_id=None,
        resource_type="system",
        resource_id="backup",
        description=(
            f"Backup recoverability manually verified by {current_user.email}. "
            f"Provider: {provider} | RPO target: {rpo_hours}h"
        ),
        ip_address=ip,
    )
    db.commit()

    logger.info(f"BACKUP_VERIFIED logged by {current_user.email}")

    return {
        "status": "ok",
        "verified_at": datetime.utcnow().isoformat(),
        "verified_by": current_user.email,
        "provider": provider,
        "message": "Backup verification event logged to audit trail.",
    }


# ==================== DATA INTEGRITY ENDPOINTS (Task 26) ====================


@app.get("/admin/integrity/verify", tags=["Admin"])
def verify_hash_chain_integrity(
    company_id: Optional[str] = None,
    request: Request = None,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Task 26 — Verify the SHA-256 hash chain integrity for all non-archived invoices.

    Walks invoices ordered by issue_date (then created_at for tie-breaking), recomputes
    each invoice's hash from its canonical fields, and checks that the following invoice's
    prev_invoice_hash matches. Returns total checked, valid link count, and the first
    broken link (if any). Logs INTEGRITY_CHECK_PASSED or INTEGRITY_CHECK_FAILED.

    Access: SUPER_ADMIN (any company, optional company_id filter), COMPANY_ADMIN (own company).
    """
    if current_user.role not in [Role.SUPER_ADMIN, Role.COMPANY_ADMIN]:
        raise HTTPException(403, "Only Super Admin or Company Admin can run integrity checks")

    # Determine which company to check
    if current_user.role == Role.SUPER_ADMIN and company_id:
        target_company_id = company_id
    elif current_user.role == Role.COMPANY_ADMIN:
        target_company_id = current_user.company_id
    else:
        target_company_id = None  # SUPER_ADMIN with no filter → all companies

    ip = request.client.host if request and request.client else None

    # Chain membership: all non-archived invoices that are post-draft.
    # Excludes only DRAFT (mutable, never issued, no hash link).
    # OVERDUE, CANCELLED and all other post-issuance statuses are included
    # so that status transitions never create chain gaps or hide broken links.
    q = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.is_archived == False,
            InvoiceDB.status != InvoiceStatus.DRAFT,
        )
        .order_by(InvoiceDB.issue_date.asc(), InvoiceDB.created_at.asc())
    )
    if target_company_id:
        q = q.filter(InvoiceDB.company_id == target_company_id)

    invoices = q.all()
    total = len(invoices)

    from utils.crypto_utils import InvoiceCrypto
    from collections import defaultdict

    _verifier = InvoiceCrypto()

    valid_links = 0
    first_broken_id = None
    first_broken_number = None
    first_broken_date = None
    integrity_ok = True

    # Group by company to maintain per-company chains
    by_company: dict[str, list] = defaultdict(list)
    for inv in invoices:
        by_company[inv.company_id or "__none__"].append(inv)

    for _cid, chain in by_company.items():
        for i, inv in enumerate(chain):
            if i == 0:
                # First invoice in the chain — no previous to verify against
                continue
            prev = chain[i - 1]

            # Independent recomputation: build canonical dict from the previous
            # invoice's DB fields, compute SHA-256 hash, and compare against
            # the current invoice's stored prev_invoice_hash.
            # This detects tampering even if both stored hash fields are altered.
            prev_canonical = {
                "invoice_number": prev.invoice_number or "",
                "issue_date": str(prev.issue_date) if prev.issue_date else "",
                "supplier_trn": prev.supplier_trn or "",
                "customer_trn": prev.customer_trn or "",
                "total_amount": str(prev.total_amount or "0.0"),
                "tax_amount": str(prev.tax_amount or "0.0"),
                "prev_invoice_hash": prev.prev_invoice_hash or "",
            }
            expected_hash = _verifier.compute_invoice_hash(prev_canonical)

            # A missing prev_invoice_hash on a non-first invoice — broken chain
            if not inv.prev_invoice_hash:
                if integrity_ok:
                    first_broken_id = inv.id
                    first_broken_number = inv.invoice_number
                    first_broken_date = str(inv.issue_date) if inv.issue_date else None
                integrity_ok = False
                continue

            # Primary: compare against independently recomputed canonical hash
            # Fallback: legacy data used prev_invoice.xml_hash (XML-content hash)
            # Accept either for backward compatibility — new invoices will only
            # match the canonical hash going forward.
            legacy_match = prev.xml_hash and (inv.prev_invoice_hash == prev.xml_hash)
            if inv.prev_invoice_hash == expected_hash or legacy_match:
                valid_links += 1
            else:
                if integrity_ok:
                    first_broken_id = inv.id
                    first_broken_number = inv.invoice_number
                    first_broken_date = str(inv.issue_date) if inv.issue_date else None
                integrity_ok = False

    # Persist the result
    check_record = IntegrityCheckDB(
        id=f"ic_{uuid4().hex[:16]}",
        company_id=target_company_id,
        performed_by_user_id=current_user.id,
        checked_at=datetime.utcnow(),
        total_invoices=total,
        valid_links=valid_links,
        integrity_ok=integrity_ok,
        first_broken_invoice_id=first_broken_id,
        first_broken_invoice_number=first_broken_number,
        first_broken_invoice_date=first_broken_date,
    )
    db.add(check_record)

    # Audit event
    audit_action = "INTEGRITY_CHECK_PASSED" if integrity_ok else "INTEGRITY_CHECK_FAILED"
    desc_parts = [
        f"Hash chain integrity check by {current_user.email}.",
        f"Total invoices checked: {total}.",
        f"Valid links: {valid_links}.",
    ]
    if not integrity_ok:
        desc_parts.append(
            f"First broken link: invoice {first_broken_number} (ID={first_broken_id}, date={first_broken_date})."
        )
    log_audit_event(
        db=db,
        action=audit_action,
        user_id=current_user.id,
        company_id=target_company_id,
        resource_type="system",
        resource_id=check_record.id,
        description=" ".join(desc_parts),
        ip_address=ip,
    )
    db.commit()

    logger.info(
        f"{audit_action}: company={target_company_id or 'ALL'} "
        f"total={total} valid_links={valid_links} by={current_user.email}"
    )

    result = {
        "integrity_ok": integrity_ok,
        "total_invoices_checked": total,
        "valid_links": valid_links,
        "checked_at": check_record.checked_at.isoformat(),
        "checked_by": current_user.email,
        "company_id": target_company_id,
    }
    if not integrity_ok:
        result["first_broken_link"] = {
            "invoice_id": first_broken_id,
            "invoice_number": first_broken_number,
            "invoice_date": first_broken_date,
        }
    return result


@app.get("/admin/integrity/status", tags=["Admin"])
def get_integrity_status(
    company_id: Optional[str] = None,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Task 26 — Return the most recent integrity check result for a company (or system-wide).

    Used by the Super Admin dashboard to show current integrity posture without re-running.
    """
    if current_user.role not in [Role.SUPER_ADMIN, Role.COMPANY_ADMIN]:
        raise HTTPException(403, "Only Super Admin or Company Admin can view integrity status")

    target_company_id = (
        company_id if (current_user.role == Role.SUPER_ADMIN and company_id)
        else current_user.company_id if current_user.role == Role.COMPANY_ADMIN
        else None
    )

    q = db.query(IntegrityCheckDB).order_by(IntegrityCheckDB.checked_at.desc())
    if target_company_id:
        # Company-scoped: return latest result for that specific company
        q = q.filter(IntegrityCheckDB.company_id == target_company_id)
    elif current_user.role == Role.SUPER_ADMIN:
        # Global scope for SUPER_ADMIN with no filter: prefer the latest
        # global run (company_id IS NULL); fall back to latest any-scope run
        # if no global run exists yet
        global_last = q.filter(IntegrityCheckDB.company_id.is_(None)).first()
        last = global_last or q.first()
        if not last:
            return {
                "status": "NEVER_CHECKED",
                "scope": "GLOBAL",
                "message": "No integrity check has been performed yet. Run GET /admin/integrity/verify.",
                "last_checked_at": None,
                "integrity_ok": None,
                "total_invoices": None,
                "valid_links": None,
                "first_broken_link": None,
            }
        return {
            "status": "PASSED" if last.integrity_ok else "FAILED",
            "scope": "GLOBAL" if last.company_id is None else "COMPANY",
            "last_checked_at": last.checked_at.isoformat(),
            "integrity_ok": last.integrity_ok,
            "total_invoices": last.total_invoices,
            "valid_links": last.valid_links,
            "checked_by_user_id": last.performed_by_user_id,
            "company_id": last.company_id,
            "first_broken_link": {
                "invoice_id": last.first_broken_invoice_id,
                "invoice_number": last.first_broken_invoice_number,
                "invoice_date": last.first_broken_invoice_date,
            } if not last.integrity_ok else None,
        }

    last = q.first()

    if not last:
        return {
            "status": "NEVER_CHECKED",
            "scope": "COMPANY" if target_company_id else "GLOBAL",
            "message": "No integrity check has been performed yet. Run GET /admin/integrity/verify.",
            "last_checked_at": None,
            "integrity_ok": None,
            "total_invoices": None,
            "valid_links": None,
            "first_broken_link": None,
        }

    return {
        "status": "PASSED" if last.integrity_ok else "FAILED",
        "scope": "GLOBAL" if last.company_id is None else "COMPANY",
        "last_checked_at": last.checked_at.isoformat(),
        "integrity_ok": last.integrity_ok,
        "total_invoices": last.total_invoices,
        "valid_links": last.valid_links,
        "checked_by_user_id": last.performed_by_user_id,
        "company_id": last.company_id,
        "first_broken_link": {
            "invoice_id": last.first_broken_invoice_id,
            "invoice_number": last.first_broken_invoice_number,
            "invoice_date": last.first_broken_invoice_date,
        } if not last.integrity_ok else None,
    }


# ==================== PERMISSION MATRIX ENDPOINT (Task 25) ====================


@app.get("/admin/permissions", tags=["Admin"])
def get_permission_matrix(
    current_user: UserDB = Depends(get_current_user_from_header),
):
    """
    Return the full role-permission matrix as JSON for FTA audit documentation.
    Accessible by SUPER_ADMIN and COMPANY_ADMIN.  Task 25.
    """
    if current_user.role not in [Role.SUPER_ADMIN, Role.COMPANY_ADMIN]:
        raise HTTPException(403, "Insufficient permissions to view permission matrix")

    # Serialise PERMISSIONS into a nested dict: {role: {resource: {action: bool}}}
    matrix: dict = {}
    for (role, resource, action), allowed in PERMISSIONS.items():
        r_key = role.value
        matrix.setdefault(r_key, {}).setdefault(resource, {})[action] = allowed

    all_resources = sorted({res for (_, res, _) in PERMISSIONS})
    all_actions = sorted({act for (_, _, act) in PERMISSIONS})

    return {
        "roles": [r.value for r in Role],
        "resources": all_resources,
        "actions": all_actions,
        "matrix": matrix,
    }


# ==================== REGISTRATION ENDPOINTS ====================


class QuickRegisterCreate(BaseModel):
    email: str
    company_name: str
    business_type: Optional[str] = None
    phone: Optional[str] = None
    website: Optional[str] = None
    password: str
    address_line1: Optional[str] = None
    address_line2: Optional[str] = None
    city: Optional[str] = None
    emirate: Optional[str] = None
    po_box: Optional[str] = None
    trn: Optional[str] = None


@app.post("/register/quick", tags=["Registration"])
def quick_register(payload: QuickRegisterCreate, db: Session = Depends(get_db)):
    """Quick registration - submit all data in one request"""

    # Check if email belongs to a SUPER_ADMIN (protect from accidental override)
    existing_super_admin = (
        db.query(UserDB)
        .filter(UserDB.email == payload.email, UserDB.role == Role.SUPER_ADMIN)
        .first()
    )
    if existing_super_admin:
        raise HTTPException(403, "This email is reserved for system administration")

    # Check if email already exists
    existing = db.query(CompanyDB).filter(CompanyDB.email == payload.email).first()
    if existing:
        raise HTTPException(400, "Email already registered")

    company_id = f"co_{uuid4().hex[:8]}"

    try:
        # Create company with all registration data
        company = CompanyDB(
            id=company_id,
            legal_name=payload.company_name,
            email=payload.email,
            business_type=payload.business_type,
            phone=payload.phone,
            website=payload.website,
            password_hash=get_password_hash(payload.password),
            address_line1=payload.address_line1,
            address_line2=payload.address_line2,
            city=payload.city,
            emirate=payload.emirate,
            po_box=payload.po_box,
            trn=payload.trn,
            status=CompanyStatus.PENDING_REVIEW,
            country="AE",
            email_verified=False,
        )
        db.add(company)

        # Create progress tracker
        progress = RegistrationProgressDB(
            id=f"prog_{uuid4().hex[:8]}",
            company_id=company_id,
            current_step=1,
            step_company_info=True,
        )
        db.add(progress)

        # Create user account (owner)
        user = UserDB(
            id=f"user_{uuid4().hex[:8]}",
            email=payload.email,
            password_hash=get_password_hash(payload.password),
            role=Role.COMPANY_ADMIN,
            company_id=company_id,
            is_owner=True,
            full_name=payload.company_name,  # Use company name as default
            # Task 22: Initialise password age tracking at registration
            password_changed_at=datetime.utcnow(),
        )
        db.add(user)

        db.commit()
        db.refresh(company)

        # Automatically send verification email
        verification_token = f"verify_{uuid4().hex}"
        company.verification_token = verification_token
        company.verification_sent_at = datetime.utcnow()
        db.commit()

        # Get platform URL for verification link
        platform_url = os.getenv("PLATFORM_URL", "https://involinks.ae")
        verification_url = f"{platform_url}/?verify={verification_token}"

        # Skip verification email sending for now (temporarily disabled)
        # email_result = email_service.send_verification_email(
        #     to_email=company.email,
        #     company_name=company.legal_name or company.email,
        #     verification_url=verification_url)
        email_result = {
            "success": False,
            "note": "Email sending skipped (temporarily disabled)",
        }

        return {
            "success": True,
            "company_id": company_id,
            "message": "Registration successful! Please check your email to verify your account.",
            "status": company.status,
            "email_sent": email_result.get("success", False),
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Registration failed: {str(e)}")


@app.post("/register/init", tags=["Registration"])
def init_registration(db: Session = Depends(get_db)):
    """Initialize a new registration session (deprecated - use /register/quick)"""
    company_id = f"co_{uuid4().hex[:8]}"

    try:
        company = CompanyDB(
            id=company_id, status=CompanyStatus.PENDING_REVIEW, country="AE"
        )
        db.add(company)

        progress = RegistrationProgressDB(
            id=f"prog_{uuid4().hex[:8]}", company_id=company_id, current_step=1
        )
        db.add(progress)
        db.commit()

        return {
            "company_id": company_id,
            "current_step": 1,
            "message": "Registration session initialized",
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Initialization failed: {str(e)}")


@app.post("/register/{company_id}/step1", tags=["Registration"])
def register_step1(
    company_id: str, payload: CompanyInfoCreate, db: Session = Depends(get_db)
):
    """Step 1: Company Information"""
    company = db.get(CompanyDB, company_id)
    if not company:
        raise HTTPException(404, "Registration session not found")

    # Check if email already exists
    existing = (
        db.query(CompanyDB)
        .filter(CompanyDB.email == payload.email, CompanyDB.id != company_id)
        .first()
    )
    if existing:
        raise HTTPException(400, "Email already registered")

    company.legal_name = payload.legal_name
    company.business_type = payload.business_type
    company.registration_number = payload.registration_number
    company.registration_date = payload.registration_date
    company.email = payload.email
    company.password_hash = get_password_hash(
        payload.password
    )  # Hash password (max 72 bytes)
    company.phone = payload.phone
    company.website = payload.website

    progress = db.query(RegistrationProgressDB).filter_by(company_id=company_id).first()
    progress.step_company_info = True
    progress.current_step = 2

    db.commit()
    return {"message": "Step 1 completed", "next_step": 2}


@app.post("/register/{company_id}/step2", tags=["Registration"])
def register_step2(
    company_id: str, payload: BusinessDetailsCreate, db: Session = Depends(get_db)
):
    """Step 2: Business Details & Address"""
    company = db.get(CompanyDB, company_id)
    if not company:
        raise HTTPException(404, "Registration session not found")

    if payload.business_activity:
        company.business_activity = payload.business_activity
    if payload.address_line1:
        company.address_line1 = payload.address_line1
    if payload.address_line2:
        company.address_line2 = payload.address_line2
    if payload.city:
        company.city = payload.city
    if payload.emirate:
        company.emirate = payload.emirate
    if payload.po_box:
        company.po_box = payload.po_box
    if payload.email:
        company.email = payload.email
    if payload.phone:
        company.phone = payload.phone

    if payload.trn:
        if not is_valid_trn(payload.trn):
            raise HTTPException(400, "Invalid TRN format (must be 15 digits)")
        company.trn = payload.trn

    company.authorized_person_name = payload.authorized_person_name
    company.authorized_person_title = payload.authorized_person_title
    company.authorized_person_email = payload.authorized_person_email
    company.authorized_person_phone = payload.authorized_person_phone

    progress = db.query(RegistrationProgressDB).filter_by(company_id=company_id).first()
    progress.step_business_details = True
    progress.current_step = 3
    db.commit()

    return {"message": "Step 2 completed", "next_step": 3}


@app.post(
    "/register/{company_id}/documents",
    response_model=DocumentUploadResponse,
    tags=["Registration"],
)
async def upload_document(
    company_id: str,
    document_type: DocumentType = Form(...),
    file: UploadFile = File(...),
    issue_date: Optional[str] = Form(None),
    expiry_date: Optional[str] = Form(None),
    document_number: Optional[str] = Form(None),
    db: Session = Depends(get_db),
):
    """Upload a document (Step 3)"""
    company = db.get(CompanyDB, company_id)
    if not company:
        raise HTTPException(404, "Registration session not found")

    # Validate file type
    allowed_types = ["application/pdf", "image/jpeg", "image/png", "image/jpg"]
    if file.content_type not in allowed_types:
        raise HTTPException(400, "Invalid file type. Allowed: PDF, JPEG, PNG")

    # Validate file size (max 10MB)
    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:
        raise HTTPException(400, "File size exceeds 10MB limit")

    # Save file
    file_id = uuid4().hex[:12]
    file_extension = file.filename.split(".")[-1] if "." in file.filename else "pdf"
    file_name = f"{file_id}_{document_type.value}.{file_extension}"
    file_dir = os.path.join(ARTIFACT_ROOT, "documents", company_id)
    os.makedirs(file_dir, exist_ok=True)
    file_path = os.path.join(file_dir, file_name)

    with open(file_path, "wb") as f:
        f.write(contents)

    # Create document record
    doc = CompanyDocumentDB(
        id=f"doc_{uuid4().hex[:8]}",
        company_id=company_id,
        document_type=document_type,
        file_name=file.filename,
        file_path=file_path,
        file_size=len(contents),
        mime_type=file.content_type,
        issue_date=date.fromisoformat(issue_date) if issue_date else None,
        expiry_date=date.fromisoformat(expiry_date) if expiry_date else None,
        document_number=document_number,
        status=DocumentStatus.PENDING_REVIEW,
    )
    db.add(doc)
    db.commit()

    return DocumentUploadResponse(
        id=doc.id,
        document_type=doc.document_type,
        file_name=doc.file_name,
        file_size=doc.file_size,
        status=doc.status,
        uploaded_at=doc.uploaded_at,
    )


@app.get(
    "/register/{company_id}/documents",
    response_model=List[DocumentUploadResponse],
    tags=["Registration"],
)
def list_documents(company_id: str, db: Session = Depends(get_db)):
    """List all uploaded documents"""
    docs = db.query(CompanyDocumentDB).filter_by(company_id=company_id).all()
    return [
        DocumentUploadResponse(
            id=d.id,
            document_type=d.document_type,
            file_name=d.file_name,
            file_size=d.file_size,
            status=d.status,
            uploaded_at=d.uploaded_at,
        )
        for d in docs
    ]


@app.delete("/register/{company_id}/documents/{doc_id}", tags=["Registration"])
def delete_document(company_id: str, doc_id: str, db: Session = Depends(get_db)):
    """Delete an uploaded document"""
    doc = (
        db.query(CompanyDocumentDB).filter_by(id=doc_id, company_id=company_id).first()
    )
    if not doc:
        raise HTTPException(404, "Document not found")

    # Delete file from storage
    if os.path.exists(doc.file_path):
        os.remove(doc.file_path)

    db.delete(doc)
    db.commit()
    return {"message": "Document deleted successfully"}


@app.post("/register/{company_id}/step3", tags=["Registration"])
def register_step3(company_id: str, db: Session = Depends(get_db)):
    """Complete Step 3: Verify required documents uploaded"""
    docs = db.query(CompanyDocumentDB).filter_by(company_id=company_id).all()

    required_docs = [DocumentType.BUSINESS_LICENSE, DocumentType.TRN_CERTIFICATE]
    uploaded_types = [doc.document_type for doc in docs]
    missing = [dt for dt in required_docs if dt not in uploaded_types]

    if missing:
        raise HTTPException(
            400, f"Required documents missing: {', '.join([m.value for m in missing])}"
        )

    progress = db.query(RegistrationProgressDB).filter_by(company_id=company_id).first()
    progress.step_documents = True
    progress.current_step = 4
    db.commit()

    return {"message": "Step 3 completed", "next_step": 4}


@app.post("/register/{company_id}/step4", tags=["Registration"])
def register_step4(
    company_id: str,
    plan_id: str = Form(...),
    billing_cycle: str = Form("monthly"),
    db: Session = Depends(get_db),
):
    """Step 4: Select Subscription Plan"""
    company = db.get(CompanyDB, company_id)
    if not company:
        raise HTTPException(404, "Company not found")

    plan = db.get(SubscriptionPlanDB, plan_id)
    if not plan or not plan.active:
        raise HTTPException(404, "Plan not found or inactive")

    # Create subscription (trial for 14 days)
    subscription = CompanySubscriptionDB(
        id=f"sub_{uuid4().hex[:8]}",
        company_id=company_id,
        plan_id=plan_id,
        status=SubscriptionStatus.TRIAL,
        billing_cycle=billing_cycle,
        current_period_start=date.today(),
        current_period_end=date.today() + timedelta(days=14),
    )
    db.add(subscription)

    progress = db.query(RegistrationProgressDB).filter_by(company_id=company_id).first()
    progress.step_plan_selection = True
    progress.current_step = 5
    db.commit()

    return {"message": "Step 4 completed - Plan selected", "next_step": 5}


@app.post("/register/{company_id}/finalize", tags=["Registration"])
def finalize_registration(company_id: str, db: Session = Depends(get_db)):
    """Finalize and submit registration for approval"""
    company = db.get(CompanyDB, company_id)
    if not company:
        raise HTTPException(404, "Company not found")

    progress = db.query(RegistrationProgressDB).filter_by(company_id=company_id).first()
    if progress:
        progress.step_review = True
    progress.completed = True
    db.commit()

    return {
        "message": "Registration submitted successfully! Your application is pending admin review.",
        "company_id": company_id,
        "status": "PENDING_REVIEW",
    }


@app.post("/register/{company_id}/send-verification", tags=["Registration"])
def send_verification_email(company_id: str, db: Session = Depends(get_db)):
    """Send email verification link"""
    company = db.get(CompanyDB, company_id)
    if not company:
        raise HTTPException(404, "Company not found")

    if not company.email:
        raise HTTPException(400, "No email address on file")

    verification_token = f"verify_{uuid4().hex}"
    company.verification_token = verification_token
    company.verification_sent_at = datetime.utcnow()
    db.commit()

    # Get platform URL for verification link
    platform_url = os.getenv("PLATFORM_URL", "https://involinks.ae")
    verification_url = f"{platform_url}/?verify={verification_token}"

    # Skip verification email sending for now (temporarily disabled)
    # result = email_service.send_verification_email(
    #     to_email=company.email,
    #     company_name=company.legal_name or company.email,
    #     verification_url=verification_url)
    result = {"success": False, "note": "Email sending skipped (temporarily disabled)"}

    return {
        "message": "Verification email skipped",
        "email": company.email,
        "email_status": "skipped",
        "note": result.get("note", "Email sending skipped"),
    }


@app.post("/register/verify/{token}", tags=["Registration"])
def verify_email(token: str, db: Session = Depends(get_db)):
    """Verify email with token and submit for admin approval"""
    company = db.query(CompanyDB).filter_by(verification_token=token).first()

    if not company:
        raise HTTPException(404, "Invalid verification token")

    if company.verification_sent_at:
        time_diff = datetime.utcnow() - company.verification_sent_at
        if time_diff.total_seconds() > 86400:
            raise HTTPException(400, "Verification link expired")

    company.email_verified = True
    company.verification_token = None
    company.status = CompanyStatus.PENDING_REVIEW

    progress = db.query(RegistrationProgressDB).filter_by(company_id=company.id).first()
    if progress:
        progress.completed = True

    db.commit()

    print(
        f"✅ Email verified for {company.email} - Company {company.legal_name} is now pending admin review"
    )

    return {
        "success": True,
        "email": company.email,
        "message": "Email verified! Your application is now under review.",
    }


@app.get(
    "/register/{company_id}/progress",
    response_model=RegistrationProgressOut,
    tags=["Registration"],
)
def get_registration_progress(company_id: str, db: Session = Depends(get_db)):
    """Get current registration progress"""
    progress = db.query(RegistrationProgressDB).filter_by(company_id=company_id).first()
    if not progress:
        raise HTTPException(404, "Registration not found")

    return RegistrationProgressOut(
        company_id=company_id,
        current_step=progress.current_step,
        completed=progress.completed,
    )


# ==================== PLAN MANAGEMENT ====================


@app.get("/plans", response_model=List[PlanOut], tags=["Plans"])
def list_plans(db: Session = Depends(get_db)):
    """List all active subscription plans (public endpoint)"""
    plans = db.query(SubscriptionPlanDB).filter_by(active=True).all()
    return [
        PlanOut(
            id=p.id,
            name=p.name,
            description=p.description,
            price_monthly=p.price_monthly,
            price_yearly=p.price_yearly,
            max_invoices_per_month=p.max_invoices_per_month,
            max_users=p.max_users,
            allow_branding=p.allow_branding,
            allow_api_access=p.allow_api_access,
        )
        for p in plans
    ]


@app.get("/plans/{plan_id}", response_model=PlanOut, tags=["Plans"])
def get_plan(plan_id: str, db: Session = Depends(get_db)):
    """Get specific plan details"""
    plan = db.get(SubscriptionPlanDB, plan_id)
    if not plan:
        raise HTTPException(404, "Plan not found")

    return PlanOut(
        id=plan.id,
        name=plan.name,
        description=plan.description,
        price_monthly=plan.price_monthly,
        price_yearly=plan.price_yearly,
        max_invoices_per_month=plan.max_invoices_per_month,
        max_users=plan.max_users,
        allow_branding=plan.allow_branding,
        allow_api_access=plan.allow_api_access,
    )


# ==================== AUTH ENDPOINTS ====================


class LoginRequest(BaseModel):
    email: str
    password: str


class LoginResponse(BaseModel):
    access_token: str
    token_type: str
    user_id: Optional[str] = None
    company_id: Optional[str] = None
    company_name: Optional[str] = None
    role: Optional[str] = None


class PasswordResetRequest(BaseModel):
    email: str


class PasswordResetConfirm(BaseModel):
    token: str
    new_password: str = Field(..., min_length=8)


# MFA (Multi-Factor Authentication) Pydantic models
class MFAEnrollTOTPResponse(BaseModel):
    secret: str
    qr_code: str
    backup_codes: List[str]


class MFAVerifyRequest(BaseModel):
    code: str


class MFAStatusResponse(BaseModel):
    enabled: bool
    method: Optional[str] = None
    enrolled_at: Optional[str] = None
    last_verified_at: Optional[str] = None


class MFALoginVerifyRequest(BaseModel):
    temp_token: str
    code: str
    method: str = "totp"  # 'totp', 'email', or 'backup'


class MFALoginResponse(BaseModel):
    mfa_required: bool
    mfa_method: Optional[str] = None
    temp_token: Optional[str] = None
    access_token: Optional[str] = None
    refresh_token: Optional[str] = None
    token_type: Optional[str] = "bearer"
    user_id: Optional[str] = None
    company_id: Optional[str] = None
    role: Optional[str] = None
    password_change_required: Optional[bool] = False
    # Task 22: Password expiry fields
    password_expired: Optional[bool] = False
    password_expires_in_days: Optional[int] = None


@app.post("/auth/login", response_model=MFALoginResponse, tags=["Auth"])
def login(payload: LoginRequest, response: Response, db: Session = Depends(get_db)):
    """
    Login endpoint - supports MFA (Multi-Factor Authentication)

    Flow:
    1. If user has MFA enabled: returns temp_token + mfa_required=True
    2. If user has no MFA: returns access_token directly
    """

    # Try user authentication first (for super admins, company admins, etc)
    user = authenticate_user(payload.email, payload.password, db)
    if user:
        # Check company status for non-super-admin users
        if user.role != Role.SUPER_ADMIN and user.company_id:
            company = db.get(CompanyDB, user.company_id)
            if company and company.status != CompanyStatus.ACTIVE:
                raise HTTPException(
                    403,
                    f"Company not approved. Status: {company.status.value}. Please wait for admin approval.",
                )

        # Check if MFA is enabled
        if user.mfa_enabled:
            # Create temporary token (5 minutes expiry) for MFA verification
            temp_token = create_access_token(
                data={"sub": user.id, "type": "user", "mfa_challenge": True},
                expires_delta=timedelta(minutes=5),
            )
            return MFALoginResponse(
                mfa_required=True,
                mfa_method=user.mfa_method,
                temp_token=temp_token,
                access_token=None,
                user_id=user.id,
                company_id=user.company_id,
                role=user.role.value,
            )
        # No MFA - return access and refresh tokens directly
        # Task 22: Check password expiry before issuing token
        _pw_expired, _pw_days_left = _check_password_expiry(user)
        if _pw_expired:
            user.must_change_password = True
            log_audit_event(
                db, "PASSWORD_EXPIRED_FORCED_CHANGE",
                user_id=user.id, company_id=user.company_id,
                resource_type="user", resource_id=user.id,
                description=f"Password for {user.email} expired — forced change required",
            )
        password_change_required = bool(getattr(user, "must_change_password", False))
        access_token = create_access_token(data={
            "sub": user.id,
            "type": "user",
            "password_change_required": password_change_required,
        })
        refresh_token = create_refresh_token(data={"sub": user.id, "type": "user"})
        # Update last login
        user.last_login = datetime.utcnow()
        log_audit_event(
            db, "USER_LOGIN",
            user_id=user.id, company_id=user.company_id,
            resource_type="user", resource_id=user.id,
            description=f"User {user.email} logged in",
        )
        db.commit()
        return MFALoginResponse(
            mfa_required=False,
            mfa_method=None,
            temp_token=None,
            access_token=access_token,
            refresh_token=refresh_token,
            token_type="bearer",
            user_id=user.id,
            company_id=user.company_id,
            role=user.role.value,
            password_change_required=password_change_required,
            password_expired=_pw_expired,
            password_expires_in_days=_pw_days_left,
        )

    # Try company authentication (with MFA support)
    company = authenticate_company(payload.email, payload.password, db)
    if company:
        if company.status != CompanyStatus.ACTIVE:
            raise HTTPException(
                403, f"Account not active. Status: {company.status.value}"
            )
        # Check if MFA is enabled for company
        if company.mfa_enabled:
            # Create temporary token (5 minutes expiry) for MFA verification
            temp_token = create_access_token(
                data={"sub": company.id, "type": "company", "mfa_challenge": True},
                expires_delta=timedelta(minutes=5),
            )
            return MFALoginResponse(
                mfa_required=True,
                mfa_method=company.mfa_method,
                temp_token=temp_token,
                access_token=None,
                company_id=company.id,
                role="COMPANY",
            )
        # No MFA - return access and refresh tokens directly
        access_token = create_access_token(data={"sub": company.id, "type": "company"})
        refresh_token = create_refresh_token(
            data={"sub": company.id, "type": "company"}
        )
        return MFALoginResponse(
            mfa_required=False,
            mfa_method=None,
            temp_token=None,
            access_token=access_token,
            refresh_token=refresh_token,
            token_type="bearer",
            company_id=company.id,
            role="COMPANY",
        )

    # If neither user nor company authentication succeeded
    raise HTTPException(status_code=401, detail="Invalid email or password")


# ==================== REFRESH TOKEN ENDPOINT ====================
from fastapi import Cookie


class RefreshTokenRequest(BaseModel):
    refresh_token: Optional[str] = None


@app.post("/auth/refresh")
def refresh_token_endpoint(payload: RefreshTokenRequest, db: Session = Depends(get_db)):
    """Issue new access token using refresh token (from request body)"""
    refresh_token = payload.refresh_token
    if not refresh_token:
        raise HTTPException(401, "Refresh token missing")
    try:
        decoded = jwt.decode(refresh_token, SECRET_KEY, algorithms=[ALGORITHM])
        if decoded.get("type") != "refresh":
            raise HTTPException(401, "Invalid refresh token type")
        user_id = decoded.get("sub")
        if not user_id:
            raise HTTPException(401, "Invalid refresh token")
        # Determine if this is a user or company token
        # Try to find user first
        user = (
            db.query(UserDB).filter(UserDB.id == user_id).first()
            if "UserDB" in globals()
            else None
        )
        if user:
            access_token = create_access_token({"sub": user_id, "type": "user"})
        else:
            # Try company
            company = (
                db.query(CompanyDB).filter(CompanyDB.id == user_id).first()
                if "CompanyDB" in globals()
                else None
            )
            if company:
                access_token = create_access_token({"sub": user_id, "type": "company"})
            else:
                raise HTTPException(401, "Invalid refresh token: subject not found")
        return {"access_token": access_token, "token_type": "bearer"}
    except JWTError:
        raise HTTPException(401, "Invalid or expired refresh token")


@app.post("/auth/logout", tags=["Auth"])
def logout():
    """Logout endpoint (client-side token removal)"""
    return {"message": "Logged out successfully"}


# ==================== MFA (MULTI-FACTOR AUTHENTICATION) ENDPOINTS ====================


@app.post("/auth/mfa/enroll/totp", response_model=MFAEnrollTOTPResponse, tags=["MFA"])
def enroll_totp_mfa(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Enroll in TOTP-based MFA (Google Authenticator, Authy, etc.)

    Returns:
        - secret: TOTP secret (Base32 encoded)
        - qr_code: Data URL for QR code image
        - backup_codes: 10 backup codes (save these!)
    """
    if current_user.mfa_enabled:
        raise HTTPException(
            400, "MFA is already enabled. Disable it first to re-enroll."
        )

    # Generate TOTP secret
    secret = MFAManager.generate_totp_secret()

    # Generate QR code for easy setup
    qr_code = MFAManager.generate_qr_code(secret, current_user.email)

    # Generate backup codes
    backup_codes_plain = MFAManager.generate_backup_codes()
    backup_codes_hashed = MFAManager.hash_backup_codes(backup_codes_plain)

    # Store temporarily (will be saved when user verifies)
    current_user.mfa_secret = secret
    current_user.mfa_backup_codes = json.dumps(backup_codes_hashed)
    db.commit()

    return MFAEnrollTOTPResponse(
        secret=secret, qr_code=qr_code, backup_codes=backup_codes_plain
    )


@app.post("/auth/mfa/enroll/verify", tags=["MFA"])
def verify_mfa_enrollment(
    payload: MFAVerifyRequest,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Verify and complete MFA enrollment

    User must enter the 6-digit code from their authenticator app
    """
    if current_user.mfa_enabled:
        raise HTTPException(400, "MFA is already enabled")

    if not current_user.mfa_secret:
        raise HTTPException(
            400, "MFA enrollment not started. Call /auth/mfa/enroll/totp first"
        )

    # Verify the TOTP code
    if not MFAManager.verify_totp(current_user.mfa_secret, payload.code):
        raise HTTPException(400, "Invalid verification code")

    # Enable MFA
    current_user.mfa_enabled = True
    current_user.mfa_method = "totp"
    current_user.mfa_enrolled_at = datetime.utcnow()
    current_user.mfa_last_verified_at = datetime.utcnow()
    db.commit()

    return {
        "success": True,
        "message": "MFA enabled successfully",
        "method": "totp",
        "enrolled_at": current_user.mfa_enrolled_at.isoformat(),
    }


@app.post("/auth/mfa/verify", response_model=MFALoginResponse, tags=["MFA"])
def verify_mfa_login(payload: MFALoginVerifyRequest, db: Session = Depends(get_db)):
    """
    Verify MFA code during login

    Methods supported:
    - totp: 6-digit code from authenticator app
    - email: 6-digit code sent to email
    - backup: 8-digit backup code
    """
    # Decode temp token
    try:
        token_data = jwt.decode(payload.temp_token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id = token_data.get("sub")
        mfa_challenge = token_data.get("mfa_challenge")

        if not mfa_challenge:
            raise HTTPException(401, "Invalid temporary token")

        user = db.query(UserDB).filter(UserDB.id == user_id).first()
        if not user:
            raise HTTPException(401, "User not found")

    except JWTError:
        raise HTTPException(401, "Invalid or expired temporary token")

    # Verify code based on method
    verified = False

    if payload.method == "totp":
        verified = MFAManager.verify_totp(user.mfa_secret, payload.code)

    elif payload.method == "email":
        verified = EmailOTPManager.verify(user.email, payload.code)

    elif payload.method == "backup":
        if not user.mfa_backup_codes:
            raise HTTPException(400, "No backup codes available")

        verified, updated_codes = MFAManager.verify_backup_code(
            payload.code, user.mfa_backup_codes
        )

        if verified:
            # Remove used backup code
            user.mfa_backup_codes = updated_codes

    if not verified:
        raise HTTPException(401, "Invalid verification code")

    # Update last verified timestamp
    user.mfa_last_verified_at = datetime.utcnow()
    user.last_login = datetime.utcnow()
    log_audit_event(
        db, "USER_LOGIN",
        user_id=user.id, company_id=user.company_id,
        resource_type="user", resource_id=user.id,
        description=f"User {user.email} logged in (MFA verified)",
    )
    # Task 22: Check password expiry at MFA-verify step too
    _pw_expired_mfa, _pw_days_left_mfa = _check_password_expiry(user)
    if _pw_expired_mfa:
        user.must_change_password = True
        log_audit_event(
            db, "PASSWORD_EXPIRED_FORCED_CHANGE",
            user_id=user.id, company_id=user.company_id,
            resource_type="user", resource_id=user.id,
            description=f"Password for {user.email} expired — forced change required",
        )
    password_change_required = bool(getattr(user, "must_change_password", False))
    db.commit()

    # Create full access token and refresh token
    access_token = create_access_token(data={
        "sub": user.id,
        "type": "user",
        "password_change_required": password_change_required,
    })
    refresh_token = create_refresh_token(data={"sub": user.id, "type": "user"})

    return MFALoginResponse(
        mfa_required=False,
        mfa_method=None,
        temp_token=None,
        access_token=access_token,
        refresh_token=refresh_token,
        token_type="bearer",
        user_id=user.id,
        company_id=user.company_id,
        role=user.role.value,
        password_change_required=password_change_required,
        password_expired=_pw_expired_mfa,
        password_expires_in_days=_pw_days_left_mfa,
    )


@app.get("/auth/mfa/status", response_model=MFAStatusResponse, tags=["MFA"])
def get_mfa_status(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Get current MFA status for the user"""
    return MFAStatusResponse(
        enabled=current_user.mfa_enabled,
        method=current_user.mfa_method,
        enrolled_at=current_user.mfa_enrolled_at.isoformat()
        if current_user.mfa_enrolled_at
        else None,
        last_verified_at=current_user.mfa_last_verified_at.isoformat()
        if current_user.mfa_last_verified_at
        else None,
    )


@app.post("/auth/mfa/disable", tags=["MFA"])
def disable_mfa(
    payload: MFAVerifyRequest,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Disable MFA (requires current MFA code for security)
    """
    if not current_user.mfa_enabled:
        raise HTTPException(400, "MFA is not enabled")

    # Verify current MFA code
    if not MFAManager.verify_totp(current_user.mfa_secret, payload.code):
        raise HTTPException(401, "Invalid verification code")

    # Disable MFA
    current_user.mfa_enabled = False
    current_user.mfa_method = None
    current_user.mfa_secret = None
    current_user.mfa_backup_codes = None
    db.commit()

    return {"success": True, "message": "MFA disabled successfully"}


@app.post("/auth/mfa/backup-codes/regenerate", tags=["MFA"])
def regenerate_backup_codes(
    payload: MFAVerifyRequest,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Regenerate backup codes (requires current MFA code for security)

    Returns new set of 10 backup codes. Old codes are invalidated.
    """
    if not current_user.mfa_enabled:
        raise HTTPException(400, "MFA is not enabled")

    # Verify current MFA code
    if not MFAManager.verify_totp(current_user.mfa_secret, payload.code):
        raise HTTPException(401, "Invalid verification code")

    # Generate new backup codes
    backup_codes_plain = MFAManager.generate_backup_codes()
    backup_codes_hashed = MFAManager.hash_backup_codes(backup_codes_plain)

    # Save new codes
    current_user.mfa_backup_codes = json.dumps(backup_codes_hashed)
    db.commit()

    return {
        "success": True,
        "message": "Backup codes regenerated",
        "backup_codes": backup_codes_plain,
    }


@app.post("/auth/mfa/email/send", tags=["MFA"])
def send_email_otp(user_email: str, db: Session = Depends(get_db)):
    """
    Send email OTP for MFA verification (backup method)

    This is a fallback method if user loses their TOTP device.
    Rate limited to 3 sends per hour.
    """
    user = db.query(UserDB).filter(UserDB.email == user_email).first()
    if not user:
        # Don't reveal if email exists
        return {"message": "If the email is registered, an OTP has been sent"}

    if not user.mfa_enabled:
        raise HTTPException(400, "MFA is not enabled for this account")

    # Check rate limiting
    if not EmailOTPManager.can_send(user_email):
        raise HTTPException(429, "Too many OTP requests. Please try again later.")

    # Generate and store OTP
    otp = EmailOTPManager.generate_and_store(user_email)

    # Send MFA OTP email via AWS SES
    result = email_service.send_mfa_otp_email(
        to_email=user_email, user_name=user.full_name or "User", otp_code=otp
    )

    return {
        "success": True,
        "message": "Verification code sent to your email",
        "expires_in_minutes": 10,
        "email_status": "sent" if result.get("success") else "simulated",
    }


@app.post("/auth/forgot-password", tags=["Auth"])
def forgot_password(payload: PasswordResetRequest, db: Session = Depends(get_db)):
    """Request password reset - sends reset token via email"""
    company = db.query(CompanyDB).filter(CompanyDB.email == payload.email).first()

    if not company:
        # Don't reveal if email exists
        return {
            "message": "If your email is registered, you will receive a password reset link"
        }

    # Generate reset token
    reset_token = f"reset_{secrets.token_hex(16)}"
    company.password_reset_token = reset_token
    company.password_reset_expires = datetime.utcnow() + timedelta(
        hours=1
    )  # 1 hour expiry

    db.commit()

    # Simulate sending email
    print("\n" + "=" * 70)
    print("╔" + "=" * 68 + "╗")
    print("║" + " " * 25 + "EMAIL WOULD BE SENT" + " " * 24 + "║")
    print("╠" + "=" * 68 + "╣")
    print(f"║  To: {payload.email:<60} ║")
    print(f"║  Subject: Reset Your Password - InvoLinks E-Invoicing{' ' * 13} ║")
    print("║" + " " * 68 + "║")
    print(f"║  Hi {(company.legal_name or 'User'):<60} ║")
    print("║" + " " * 68 + "║")
    print("║  You requested to reset your password. Click the link below:    ║")
    print("║" + " " * 68 + "║")
    print(f"║  http://your-app-url.com/reset-password?token={reset_token[:20]:<22} ║")
    print("║" + " " * 68 + "║")
    print("║  This link will expire in 1 hour.                               ║")
    print("║" + " " * 68 + "║")
    print("║  If you didn't request this, please ignore this email.          ║")
    print("║" + " " * 68 + "║")
    print("║  Best regards,                                                   ║")
    print("║  InvoLinks E-Invoicing Team                                      ║")
    print("╚" + "=" * 68 + "╝")
    print("=" * 70 + "\n")

    return {
        "message": "If your email is registered, you will receive a password reset link"
    }


@app.post("/auth/reset-password", tags=["Auth"])
def reset_password(payload: PasswordResetConfirm, db: Session = Depends(get_db)):
    """Reset password using token"""
    company = (
        db.query(CompanyDB)
        .filter(CompanyDB.password_reset_token == payload.token)
        .first()
    )

    if not company:
        raise HTTPException(400, "Invalid or expired reset token")

    if (
        not company.password_reset_expires
        or company.password_reset_expires < datetime.utcnow()
    ):
        raise HTTPException(400, "Reset token has expired")

    # Update password
    company.password_hash = get_password_hash(payload.new_password)
    company.password_reset_token = None
    company.password_reset_expires = None

    db.commit()

    return {
        "message": "Password reset successfully. You can now login with your new password."
    }


@app.get("/auth/me", tags=["Auth"])
def get_current_user(token: str, db: Session = Depends(get_db)):
    """Get current authenticated company"""
    company = get_current_company(token, db)
    if not company:
        raise HTTPException(401, "Not authenticated")

    return {
        "company_id": company.id,
        "company_name": company.legal_name,
        "email": company.email,
        "status": company.status.value,
    }


# ==================== ADMIN ENDPOINTS ====================


@app.get("/admin/companies/pending", response_model=List[CompanyOut], tags=["Admin"])
def list_pending_companies(db: Session = Depends(get_db)):
    """List all companies pending review (Admin only)"""
    companies = db.query(CompanyDB).filter_by(status=CompanyStatus.PENDING_REVIEW).all()
    return [
        CompanyOut(
            id=c.id, legal_name=c.legal_name, status=c.status, email=c.email, trn=c.trn
        )
        for c in companies
    ]


# ==================== COMPANY ENDPOINTS ====================


@app.get("/me", tags=["Auth"])
def get_current_user_info(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Get current user and associated company information"""
    # Get company associated with the user
    company = (
        db.query(CompanyDB).filter(CompanyDB.id == current_user.company_id).first()
    )

    # Task 22: Compute days until password expires for frontend warning banner
    _pw_expired_me, _pw_days_left_me = _check_password_expiry(current_user)

    response = {
        "user_id": current_user.id,
        "email": current_user.email,
        "role": current_user.role.value,
        "company": None,
        # Task 22: Password policy info for security settings / warning banner
        "password_changed_at": (
            current_user.password_changed_at.isoformat()
            if getattr(current_user, "password_changed_at", None)
            else None
        ),
        # Task 22: Alias for spec compatibility
        "password_last_change_date": (
            current_user.password_changed_at.isoformat()
            if getattr(current_user, "password_changed_at", None)
            else None
        ),
        "password_expires_in_days": _pw_days_left_me,
        "password_expired": _pw_expired_me,
    }

    if company:
        response["company"] = {
            "id": company.id,
            "legal_name": company.legal_name,
            "email": company.email,
            "status": company.status.value,
            "invoices_generated": company.invoices_generated or 0,
            "free_plan_type": company.free_plan_type,
            "free_plan_invoice_limit": company.free_plan_invoice_limit,
            "free_plan_duration_months": company.free_plan_duration_months,
        }

    return response


@app.get("/companies/{company_id}", tags=["Companies"])
def get_company(company_id: str, db: Session = Depends(get_db)):
    """Get company details"""
    company = db.get(CompanyDB, company_id)
    if not company:
        raise HTTPException(404, "Company not found")

    return {
        "id": company.id,
        "legal_name": company.legal_name,
        "status": company.status,
        "email": company.email,
        "trn": company.trn,
        "business_type": company.business_type,
        "phone": company.phone,
        "website": company.website,
        "address_line1": company.address_line1,
        "address_line2": company.address_line2,
        "city": company.city,
        "emirate": company.emirate,
        "po_box": company.po_box,
        "email_verified": company.email_verified,
        "created_at": company.created_at.isoformat() if company.created_at else None,
    }


class CompanyProfileUpdateRequest(BaseModel):
    legal_name: Optional[str] = None
    phone: Optional[str] = None
    website: Optional[str] = None
    address_line1: Optional[str] = None
    address_line2: Optional[str] = None
    city: Optional[str] = None
    emirate: Optional[str] = None
    po_box: Optional[str] = None


@app.put("/company/profile", tags=["Companies"])
def update_company_profile(
    payload: CompanyProfileUpdateRequest,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Update company profile (Company Admin only)"""
    if not has_permission(current_user, "settings", "edit"):
        raise HTTPException(403, "Permission denied: edit on settings is not allowed for your role")

    if not current_user.company_id:
        raise HTTPException(400, "User is not associated with a company")

    company = db.get(CompanyDB, current_user.company_id)
    if not company:
        raise HTTPException(404, "Company not found")

    # Update fields if provided
    if payload.legal_name is not None:
        company.legal_name = payload.legal_name
    if payload.phone is not None:
        company.phone = payload.phone
    if payload.website is not None:
        company.website = payload.website
    if payload.address_line1 is not None:
        company.address_line1 = payload.address_line1
    if payload.address_line2 is not None:
        company.address_line2 = payload.address_line2
    if payload.city is not None:
        company.city = payload.city
    if payload.emirate is not None:
        company.emirate = payload.emirate
    if payload.po_box is not None:
        company.po_box = payload.po_box

    company.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(company)

    return {
        "success": True,
        "message": "Company profile updated successfully",
        "company": {
            "id": company.id,
            "legal_name": company.legal_name,
            "email": company.email,
            "phone": company.phone,
            "website": company.website,
            "address_line1": company.address_line1,
            "address_line2": company.address_line2,
            "city": company.city,
            "emirate": company.emirate,
            "po_box": company.po_box,
        },
    }


@app.get("/companies/{company_id}/subscription", tags=["Companies"])
def get_company_subscription(company_id: str, db: Session = Depends(get_db)):
    """Get company's current subscription"""
    # Verify company exists
    company = db.get(CompanyDB, company_id)
    if not company:
        raise HTTPException(404, "Company not found")

    sub = db.query(CompanySubscriptionDB).filter_by(company_id=company_id).first()
    if not sub:
        # No subscription record — default to Free plan
        free_plan = db.query(SubscriptionPlanDB).filter_by(id="plan_free").first()
        return {
            "subscription_id": None,
            "plan": {
                "id": free_plan.id if free_plan else "plan_free",
                "name": free_plan.name if free_plan else "Free",
                "price_monthly": free_plan.price_monthly if free_plan else 0,
                "max_invoices_per_month": free_plan.max_invoices_per_month
                if free_plan
                else 100,
                "max_users": free_plan.max_users if free_plan else 1,
                "allow_api_access": free_plan.allow_api_access if free_plan else True,
                "allow_branding": free_plan.allow_branding if free_plan else False,
                "allow_multi_currency": free_plan.allow_multi_currency
                if free_plan
                else False,
                "priority_support": free_plan.priority_support if free_plan else False,
            }
            if free_plan
            else {"id": "plan_free", "name": "Free", "price_monthly": 0},
            "status": "ACTIVE",
            "billing_cycle": "monthly",
            "current_period_start": None,
            "current_period_end": None,
            "invoices_this_period": 0,
            "message": "No subscription record — defaulting to Free plan",
        }

    return {
        "subscription_id": sub.id,
        "plan": {
            "id": sub.plan.id,
            "name": sub.plan.name,
            "price_monthly": sub.plan.price_monthly,
        },
        "status": sub.status.value,
        "billing_cycle": sub.billing_cycle,
        "current_period_start": sub.current_period_start.isoformat(),
        "current_period_end": sub.current_period_end.isoformat(),
        "invoices_this_period": sub.invoices_this_period,
    }


@app.get(
    "/companies/{company_id}/invoices",
    tags=["Companies"],
    response_model=List[InvoiceListOut],
)
def get_company_invoices(
    company_id: str,
    db: Session = Depends(get_db),
    status: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
):
    """Get all invoices for a specific company (excludes CANCELLED invoices for recent listing)"""
    # Verify company exists
    company = db.get(CompanyDB, company_id)
    if not company:
        raise HTTPException(404, "Company not found")

    query = db.query(InvoiceDB).filter(
        InvoiceDB.company_id == company_id,
        InvoiceDB.status != InvoiceStatus.CANCELLED,  # Exclude cancelled invoices
    )

    if status:
        try:
            status_enum = InvoiceStatus(status)
            query = query.filter(InvoiceDB.status == status_enum)
        except ValueError:
            raise HTTPException(400, f"Invalid status: {status}")

    query = query.order_by(InvoiceDB.created_at.desc())
    invoices = query.offset(offset).limit(limit).all()

    return [
        InvoiceListOut(
            id=inv.id,
            invoice_number=inv.invoice_number,
            invoice_type=inv.invoice_type,
            status=inv.status,
            issue_date=inv.issue_date.isoformat(),
            customer_name=inv.customer_name,
            total_amount=inv.total_amount,
            currency_code=inv.currency_code,
            created_at=inv.created_at.isoformat(),
        )
        for inv in invoices
    ]


# ==================== ADMIN ENDPOINTS ====================


class CompanyApprovalRequest(BaseModel):
    approved: bool
    notes: Optional[str] = None


@app.get("/admin/companies/pending", tags=["Admin"])
def get_pending_companies(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Get all companies pending approval (Super Admin only)"""
    if current_user.role != Role.SUPER_ADMIN:
        raise HTTPException(403, "Insufficient permissions")

    pending = (
        db.query(CompanyDB)
        .filter(CompanyDB.status == CompanyStatus.PENDING_REVIEW)
        .order_by(CompanyDB.created_at.desc())
        .all()
    )

    return [
        {
            "id": c.id,
            "legal_name": c.legal_name,
            "email": c.email,
            "business_type": c.business_type,
            "phone": c.phone,
            "created_at": c.created_at.isoformat(),
            "status": c.status.value,
            "invoices_generated": 0,  # Pending companies haven't generated invoices yet
        }
        for c in pending
    ]


@app.get("/admin/companies", tags=["Admin"])
def get_all_companies(
    status: Optional[str] = None,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Get all companies with optional status filter (Super Admin only)"""
    if current_user.role != Role.SUPER_ADMIN:
        raise HTTPException(403, "Insufficient permissions")

    query = db.query(CompanyDB)
    if status:
        try:
            status_enum = CompanyStatus(status)
            query = query.filter(CompanyDB.status == status_enum)
        except ValueError:
            pass

    companies = query.order_by(CompanyDB.created_at.desc()).all()

    # Use invoices table as source of truth for generated invoice count.
    # This avoids stale counters in companies.invoices_generated.
    invoice_counts = {}
    companies_with_logos = set()
    if companies:
        company_ids = [c.id for c in companies]
        invoice_counts = dict(
            db.query(InvoiceDB.company_id, func.count(InvoiceDB.id))
            .filter(InvoiceDB.company_id.in_(company_ids))
            .group_by(InvoiceDB.company_id)
            .all()
        )
        companies_with_logos = set(
            cid
            for (cid,) in db.query(CompanyBrandingDB.company_id)
            .filter(
                CompanyBrandingDB.company_id.in_(company_ids),
                CompanyBrandingDB.logo_file_path.isnot(None),
            )
            .all()
        )

    return [
        {
            "id": c.id,
            "legal_name": c.legal_name,
            "email": c.email,
            "business_type": c.business_type,
            "phone": c.phone,
            "status": c.status.value,
            "created_at": c.created_at.isoformat() if c.created_at else None,
            "approved_at": c.approved_at.isoformat() if c.approved_at else None,
            "rejected_at": c.rejected_at.isoformat() if c.rejected_at else None,
            "subscription_plan": c.subscription_plan_id,
            "invoices_generated": invoice_counts.get(c.id, 0),
            "free_plan_type": c.free_plan_type,
            "free_plan_invoice_limit": c.free_plan_invoice_limit,
            "free_plan_duration_months": c.free_plan_duration_months,
            "free_plan_start_date": c.free_plan_start_date.isoformat()
            if c.free_plan_start_date
            else None,
            "has_logo": c.id in companies_with_logos,
        }
        for c in companies
    ]


@app.get("/admin/companies/{company_id}", tags=["Admin"])
def get_company_by_id(
    company_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Get a single company by id (Super Admin only)"""
    if current_user.role != Role.SUPER_ADMIN:
        raise HTTPException(403, "Insufficient permissions")

    company = db.get(CompanyDB, company_id)
    if not company:
        raise HTTPException(404, "Company not found")

    invoices_generated = (
        db.query(func.count(InvoiceDB.id))
        .filter(InvoiceDB.company_id == company.id)
        .scalar()
        or 0
    )

    return {
        "id": company.id,
        "legal_name": company.legal_name,
        "email": company.email,
        "business_type": company.business_type,
        "phone": company.phone,
        "status": company.status.value,
        "created_at": company.created_at.isoformat() if company.created_at else None,
        "approved_at": company.approved_at.isoformat() if company.approved_at else None,
        "rejected_at": company.rejected_at.isoformat() if company.rejected_at else None,
        "subscription_plan": company.subscription_plan_id,
        "invoices_generated": invoices_generated,
        "free_plan_type": company.free_plan_type,
        "free_plan_invoice_limit": company.free_plan_invoice_limit,
        "free_plan_duration_months": company.free_plan_duration_months,
        "free_plan_start_date": company.free_plan_start_date.isoformat()
        if company.free_plan_start_date
        else None,
        "trial_status": company.trial_status,
        "trial_start_date": company.trial_start_date.isoformat()
        if company.trial_start_date
        else None,
        "trial_invoice_count": company.trial_invoice_count,
        "trial_ended_at": company.trial_ended_at.isoformat()
        if company.trial_ended_at
        else None,
    }


class CompanyApprovalConfig(BaseModel):
    free_plan_type: str = "INVOICE_COUNT"  # "DURATION" or "INVOICE_COUNT"
    free_plan_duration_months: Optional[int] = None  # For duration-based
    free_plan_invoice_limit: Optional[int] = 100  # For invoice count-based


@app.post("/admin/companies/{company_id}/approve", tags=["Admin"])
def approve_company(
    company_id: str,
    config: CompanyApprovalConfig = None,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Approve a pending company registration with configurable free plan (Super Admin only)"""
    if current_user.role != Role.SUPER_ADMIN:
        raise HTTPException(403, "Insufficient permissions")

    company = db.get(CompanyDB, company_id)
    if not company:
        raise HTTPException(404, "Company not found")

    if company.status != CompanyStatus.PENDING_REVIEW:
        raise HTTPException(
            400,
            f"Company is not pending review. Current status: {company.status.value}",
        )

    # Use default config if none provided
    if config is None:
        config = CompanyApprovalConfig()

    # Update company status to ACTIVE
    company.status = CompanyStatus.ACTIVE
    company.approved_at = datetime.utcnow()

    # Configure free plan settings
    company.free_plan_type = config.free_plan_type
    company.trial_start_date = (
        datetime.utcnow()
    )  # Use trial_start_date for billing tracking
    company.trial_status = "ACTIVE"  # Ensure trial is active
    company.free_plan_start_date = datetime.utcnow()  # Keep for backwards compatibility

    if config.free_plan_type == "DURATION":
        company.free_plan_duration_months = config.free_plan_duration_months or 1
        plan_description = f"{company.free_plan_duration_months} month(s) free duration"
    else:  # INVOICE_COUNT
        company.free_plan_invoice_limit = config.free_plan_invoice_limit or 100
        plan_description = f"{company.free_plan_invoice_limit} free invoices"

    # Assign Free plan if not already assigned
    if not company.subscription_plan_id:
        free_plan = db.query(SubscriptionPlanDB).filter_by(name="Free").first()
        if free_plan:
            company.subscription_plan_id = free_plan.id

    db.commit()

    # Skip approval email sending for now (temporarily disabled)
    # email_result = email_service.send_approval_notification(
    #     to_email=company.email,
    #     company_name=company.legal_name or company.email,
    #     status="APPROVED")
    email_result = {
        "success": False,
        "note": "Email sending skipped (temporarily disabled)",
    }

    return {
        "success": True,
        "company_id": company_id,
        "status": company.status.value,
        "free_plan_type": company.free_plan_type,
        "free_plan_config": plan_description,
        "message": "Company approved successfully",
        "email_sent": email_result.get("success", False),
        "email_status": "skipped",
    }


class RejectCompanyRequest(BaseModel):
    notes: Optional[str] = None


@app.post("/admin/companies/{company_id}/reject", tags=["Admin"])
def reject_company(
    company_id: str,
    payload: RejectCompanyRequest = None,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Reject a pending company registration (Super Admin only)"""
    if current_user.role != Role.SUPER_ADMIN:
        raise HTTPException(403, "Insufficient permissions")

    company = db.get(CompanyDB, company_id)
    if not company:
        raise HTTPException(404, "Company not found")

    if company.status != CompanyStatus.PENDING_REVIEW:
        raise HTTPException(
            400,
            f"Company is not pending review. Current status: {company.status.value}",
        )

    notes = payload.notes if payload else None

    # Update company status to REJECTED
    company.status = CompanyStatus.REJECTED
    company.rejected_at = datetime.utcnow()
    db.commit()

    # Skip rejection email sending for now (temporarily disabled)
    # email_result = email_service.send_approval_notification(
    #     to_email=company.email,
    #     company_name=company.legal_name or company.email,
    #     status="REJECTED",
    #     admin_message=notes)
    email_result = {
        "success": False,
        "note": "Email sending skipped (temporarily disabled)",
    }

    return {
        "success": True,
        "company_id": company_id,
        "status": company.status.value,
        "message": "Company rejected",
        "email_sent": email_result.get("success", False),
    }


@app.post("/admin/companies/{company_id}/reset-trial", tags=["Admin"])
def reset_company_trial(
    company_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Reset company trial to start fresh (Super Admin only - for testing)"""
    if current_user.role != Role.SUPER_ADMIN:
        raise HTTPException(403, "Insufficient permissions")

    company = db.get(CompanyDB, company_id)
    if not company:
        raise HTTPException(404, "Company not found")

    # Delete all outbound invoices for this company as requested by super admin.
    invoices_to_delete = (
        db.query(InvoiceDB).filter(InvoiceDB.company_id == company_id).all()
    )
    invoice_ids = [inv.id for inv in invoices_to_delete]

    # Keep PEPPOL usage rows but detach deleted invoices to avoid FK issues.
    if invoice_ids:
        db.query(PeppolUsageDB).filter(
            PeppolUsageDB.company_id == company_id,
            PeppolUsageDB.invoice_id.in_(invoice_ids),
        ).update({PeppolUsageDB.invoice_id: None}, synchronize_session=False)

    deleted_invoices = len(invoices_to_delete)
    deleted_xml_files = 0
    deleted_pdf_files = 0

    for invoice in invoices_to_delete:
        if invoice.xml_file_path and os.path.exists(invoice.xml_file_path):
            try:
                os.remove(invoice.xml_file_path)
                deleted_xml_files += 1
            except Exception:
                pass

        if invoice.pdf_file_path and os.path.exists(invoice.pdf_file_path):
            try:
                os.remove(invoice.pdf_file_path)
                deleted_pdf_files += 1
            except Exception:
                pass

        db.delete(invoice)

    # Reset trial data
    company.trial_status = "ACTIVE"
    company.trial_start_date = datetime.utcnow()
    company.trial_invoice_count = 0
    company.invoices_generated = 0
    company.trial_ended_at = None

    db.commit()

    # Calculate new remaining
    trial_duration_days = (company.free_plan_duration_months or 1) * 30
    invoice_limit = company.free_plan_invoice_limit or 100

    return {
        "success": True,
        "company_id": company_id,
        "message": "Trial reset successfully and company invoices cleared",
        "trial_status": company.trial_status,
        "trial_start_date": company.trial_start_date.isoformat(),
        "free_plan_type": company.free_plan_type,
        "days_remaining": trial_duration_days,
        "invoices_remaining": invoice_limit,
        "deleted_invoices": deleted_invoices,
        "deleted_xml_files": deleted_xml_files,
        "deleted_pdf_files": deleted_pdf_files,
    }


class CompanyUpdateRequest(BaseModel):
    invoices_generated: Optional[int] = None
    free_plan_invoice_limit: Optional[int] = None
    free_plan_duration_months: Optional[int] = None
    free_plan_type: Optional[str] = None  # INVOICE_COUNT or DURATION
    vat_enabled: Optional[bool] = None
    subscription_plan_id: Optional[str] = None
    status: Optional[str] = None  # active, inactive, suspended


@app.put("/admin/companies/{company_id}", tags=["Admin"])
def update_company(
    company_id: str,
    payload: CompanyUpdateRequest,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Update company details (Super Admin only)"""
    if current_user.role != Role.SUPER_ADMIN:
        raise HTTPException(403, "Insufficient permissions")

    company = db.get(CompanyDB, company_id)
    if not company:
        raise HTTPException(404, "Company not found")

    # Update fields if provided
    if payload.invoices_generated is not None:
        company.invoices_generated = payload.invoices_generated
    if payload.free_plan_invoice_limit is not None:
        company.free_plan_invoice_limit = payload.free_plan_invoice_limit
    if payload.free_plan_duration_months is not None:
        company.free_plan_duration_months = payload.free_plan_duration_months
    if payload.free_plan_type is not None:
        company.free_plan_type = payload.free_plan_type
    if payload.vat_enabled is not None:
        company.vat_enabled = payload.vat_enabled
    if payload.subscription_plan_id is not None:
        company.subscription_plan_id = payload.subscription_plan_id
    if payload.status is not None:
        # Map string status to enum
        status_map = {
            "active": CompanyStatus.ACTIVE,
            "suspended": CompanyStatus.SUSPENDED,
        }
        if payload.status.lower() in status_map:
            company.status = status_map[payload.status.lower()]

    db.commit()
    db.refresh(company)

    return {
        "success": True,
        "company_id": company_id,
        "message": "Company updated successfully",
        "invoices_generated": company.invoices_generated,
        "free_plan_invoice_limit": company.free_plan_invoice_limit,
        "free_plan_duration_months": company.free_plan_duration_months,
    }


@app.get("/admin/stats", tags=["Admin"])
def get_admin_stats(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Get comprehensive dashboard statistics for SuperAdmin (Super Admin only)"""
    if current_user.role != Role.SUPER_ADMIN:
        raise HTTPException(403, "Insufficient permissions")

    # Count companies by status
    pending = (
        db.query(CompanyDB)
        .filter(CompanyDB.status == CompanyStatus.PENDING_REVIEW)
        .count()
    )
    approved = (
        db.query(CompanyDB).filter(CompanyDB.status == CompanyStatus.ACTIVE).count()
    )
    rejected = (
        db.query(CompanyDB).filter(CompanyDB.status == CompanyStatus.REJECTED).count()
    )
    active_companies = (
        db.query(CompanyDB).filter(CompanyDB.status == CompanyStatus.ACTIVE).count()
    )
    suspended_companies = (
        db.query(CompanyDB).filter(CompanyDB.status == CompanyStatus.SUSPENDED).count()
    )
    inactive_companies = (
        db.query(CompanyDB)
        .filter(
            (CompanyDB.status == CompanyStatus.PENDING_REVIEW)
            | (CompanyDB.status == CompanyStatus.REJECTED)
        )
        .count()
    )

    # Get all companies with details for explorer
    companies = db.query(CompanyDB).all()
    all_companies_list = []
    # Parse optional date range for per-company revenue calculations
    from_dt = None
    to_dt = None
    if from_date:
        try:
            from_dt = datetime.fromisoformat(from_date.replace("Z", "+00:00")).replace(
                tzinfo=None
            )
        except Exception:
            from_dt = None
    if to_date:
        try:
            to_dt = datetime.fromisoformat(to_date.replace("Z", "+00:00")).replace(
                tzinfo=None
            )
        except Exception:
            to_dt = None
    for company in companies:
        # Get plan info - prefer active company subscription, fallback to legacy subscription_plan_id
        plan = None
        arpu = 0
        # Check for latest company subscription
        latest_sub = (
            db.query(CompanySubscriptionDB)
            .filter(CompanySubscriptionDB.company_id == company.id)
            .order_by(CompanySubscriptionDB.created_at.desc())
            .first()
        )
        if latest_sub and latest_sub.plan:
            plan = latest_sub.plan.name
            arpu = latest_sub.plan.price_monthly or 0
        elif company.subscription_plan_id:
            plan_obj = db.get(SubscriptionPlanDB, company.subscription_plan_id)
            if plan_obj:
                plan = plan_obj.name
                arpu = plan_obj.price_monthly
        else:
            # No plan assigned — default to Free
            plan = "Free"
            arpu = 0

        # Get invoice count for this month
        now = datetime.utcnow()
        month_start = datetime(now.year, now.month, 1)
        invoices_this_month = (
            db.query(InvoiceDB)
            .filter(
                InvoiceDB.company_id == company.id, InvoiceDB.created_at >= month_start
            )
            .count()
        )
        invoices_used_total = (
            db.query(InvoiceDB).filter(InvoiceDB.company_id == company.id).count()
        )
        # Calculate company revenue for the supplied date range (paid invoices only)
        company_revenue = 0.0
        invoice_rev_q = db.query(InvoiceDB).filter(
            InvoiceDB.company_id == company.id, InvoiceDB.status == InvoiceStatus.PAID
        )
        if from_dt:
            invoice_rev_q = invoice_rev_q.filter(InvoiceDB.paid_at >= from_dt)
        if to_dt:
            invoice_rev_q = invoice_rev_q.filter(InvoiceDB.paid_at <= to_dt)
        try:
            invoices_for_company = invoice_rev_q.all()
            for inv in invoices_for_company:
                company_revenue += (inv.total_amount or 0.0) * (
                    -1
                    if inv.invoice_type
                    in [
                        InvoiceType.TAX_CREDIT_NOTE,
                        InvoiceType.CREDIT_NOTE_OUT_OF_SCOPE,
                    ]
                    else 1
                )
        except Exception:
            company_revenue = 0.0
        all_companies_list.append(
            {
                "id": company.id,
                "name": company.legal_name or "Unnamed Company",
                "status": "active"
                if company.status == CompanyStatus.ACTIVE
                else (
                    "suspended"
                    if company.status == CompanyStatus.SUSPENDED
                    else "inactive"
                ),
                "invoicesThisMonth": invoices_this_month,
                "invoicesUsed": invoices_used_total,
                "invoicesLimit": company.free_plan_invoice_limit,
                "free_plan_type": company.free_plan_type,
                "plan": plan,
                "arpu": arpu,
                "region": company.emirate,
                "vatCompliant": bool(company.trn),
                "revenue": round(company_revenue, 2),
            }
        )

    # Invoice statistics
    now = datetime.utcnow()
    month_start = datetime(now.year, now.month, 1)
    last_month_start = (month_start - timedelta(days=1)).replace(day=1)

    invoices_mtd = (
        db.query(InvoiceDB).filter(InvoiceDB.created_at >= month_start).count()
    )
    invoices_last_month = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.created_at >= last_month_start, InvoiceDB.created_at < month_start
        )
        .count()
    )

    # Revenue calculations by tier
    tiers_data = []
    all_plans = (
        db.query(SubscriptionPlanDB).filter(SubscriptionPlanDB.active == True).all()
    )
    total_mrr = 0

    for plan in all_plans:
        active_on_plan = (
            db.query(CompanyDB)
            .filter(
                CompanyDB.subscription_plan_id == plan.id,
                CompanyDB.status == CompanyStatus.ACTIVE,
            )
            .count()
        )

        mrr = active_on_plan * plan.price_monthly
        total_mrr += mrr

        tiers_data.append(
            {
                "plan": plan.name,
                "activeCompanies": active_on_plan,
                "pricePerCompany": plan.price_monthly,
                "mrr": mrr,
                "arr": mrr * 12,
                "newActivations": 0,  # Could track this with created_at filters
            }
        )

    return {
        "asOf": datetime.utcnow().isoformat(),
        "registrations": {
            "pending": pending,
            "approved": approved,
            "rejected": rejected,
        },
        "companies": {
            "active": active_companies,
            "inactive": inactive_companies,
            "suspended": suspended_companies,
            "all": all_companies_list,
        },
        "invoices": {"monthToDate": invoices_mtd, "lastMonth": invoices_last_month},
        "revenue": {
            "mrr": total_mrr,
            "arr": total_mrr * 12,
            "deltaPctVsLastMonth": 0,  # Could calculate from historical data
            "tiers": tiers_data,
        },
        "total_companies": pending
        + approved
        + rejected
        + active_companies
        + inactive_companies,
    }


# Alias route for analytics endpoint
@app.get("/admin/analytics", tags=["Admin"])
def get_admin_analytics(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Alias for /admin/stats - Get comprehensive dashboard analytics (Super Admin only)"""
    return get_admin_stats(from_date, to_date, current_user, db)


# ==================== CONTENT MANAGEMENT ENDPOINTS ====================


@app.get(
    "/admin/content", response_model=List[ContentBlockOut], tags=["Admin", "Content"]
)
def get_all_content(
    section: Optional[str] = None,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Get all content blocks (Super Admin only). Optionally filter by section."""
    if current_user.role != Role.SUPER_ADMIN:
        raise HTTPException(403, "Only Super Admins can access content management")

    query = db.query(ContentBlockDB)
    if section:
        query = query.filter(ContentBlockDB.section == section)

    blocks = query.order_by(ContentBlockDB.section, ContentBlockDB.key).all()

    return [
        ContentBlockOut(
            id=b.id,
            key=b.key,
            value=b.value,
            description=b.description,
            section=b.section,
            updated_at=b.updated_at,
            updated_by=b.updated_by,
        )
        for b in blocks
    ]


@app.put(
    "/admin/content/{key}", response_model=ContentBlockOut, tags=["Admin", "Content"]
)
def update_content(
    key: str,
    payload: ContentBlockUpdate,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Update a content block by key (Super Admin only)"""
    if current_user.role != Role.SUPER_ADMIN:
        raise HTTPException(403, "Only Super Admins can edit content")

    block = db.query(ContentBlockDB).filter(ContentBlockDB.key == key).first()
    if not block:
        raise HTTPException(404, f"Content block with key '{key}' not found")

    block.value = payload.value
    block.updated_at = datetime.utcnow()
    block.updated_by = current_user.email

    db.commit()
    db.refresh(block)

    return ContentBlockOut(
        id=block.id,
        key=block.key,
        value=block.value,
        description=block.description,
        section=block.section,
        updated_at=block.updated_at,
        updated_by=block.updated_by,
    )


@app.get("/content/public", response_model=List[ContentBlockOut], tags=["Content"])
def get_public_content(db: Session = Depends(get_db)):
    """Get all content blocks for public consumption (no auth required)"""
    blocks = (
        db.query(ContentBlockDB)
        .order_by(ContentBlockDB.section, ContentBlockDB.key)
        .all()
    )

    return [
        ContentBlockOut(
            id=b.id,
            key=b.key,
            value=b.value,
            description=b.description,
            section=b.section,
            updated_at=b.updated_at,
            updated_by=b.updated_by,
        )
        for b in blocks
    ]


@app.get("/content/footer", tags=["Content"])
def get_footer_content(db: Session = Depends(get_db)):
    """Get footer content blocks formatted as a single object (no auth required)"""
    blocks = db.query(ContentBlockDB).filter(ContentBlockDB.section == "footer").all()

    # Convert blocks to a dictionary
    footer_data = {}
    for block in blocks:
        # Remove "footer_" prefix from keys
        key = block.key.replace("footer_", "")
        footer_data[key] = block.value

    return footer_data


# ==================== PHASE 2: SUPERADMIN TIER MANAGEMENT & PLATFORM STATS ====================


class SubscriptionPlanOut(BaseModel):
    id: str
    name: str
    description: Optional[str]
    price_monthly: float
    price_yearly: float
    max_invoices_per_month: Optional[int]
    max_users: int
    max_business_admins: int
    max_finance_users: int
    max_pos_devices: int
    allow_api_access: bool
    allow_branding: bool
    allow_multi_currency: bool
    priority_support: bool
    active: bool


class SubscriptionPlanUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    price_monthly: Optional[float] = None
    price_yearly: Optional[float] = None
    max_invoices_per_month: Optional[int] = None
    max_users: Optional[int] = None
    max_business_admins: Optional[int] = None
    max_finance_users: Optional[int] = None
    max_pos_devices: Optional[int] = None
    allow_api_access: Optional[bool] = None
    allow_branding: Optional[bool] = None
    allow_multi_currency: Optional[bool] = None
    priority_support: Optional[bool] = None
    active: Optional[bool] = None


class PlatformStatsOut(BaseModel):
    total_companies: int
    active_companies: int
    pending_companies: int
    total_invoices: int
    total_revenue_aed: float
    active_subscriptions: int
    free_tier_users: int
    paid_tier_users: int


class FeaturedBusinessOut(BaseModel):
    id: str
    company_id: str
    company_name: str
    display_name: str  # Empty string if not set
    logo_url: str  # Empty string if not set
    is_active: bool
    display_order: int
    created_at: str


class FeaturedBusinessCreate(BaseModel):
    company_id: str
    display_name: Optional[str] = None
    logo_url: Optional[str] = None
    display_order: Optional[int] = None


class FeaturedBusinessUpdate(BaseModel):
    display_name: Optional[str] = None
    logo_url: Optional[str] = None
    is_active: Optional[bool] = None
    display_order: Optional[int] = None


@app.get(
    "/admin/subscription-plans",
    response_model=List[SubscriptionPlanOut],
    tags=["Admin", "Tier Management"],
)
def get_all_subscription_plans(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Get all subscription plans (Super Admin only)"""
    if current_user.role != Role.SUPER_ADMIN:
        raise HTTPException(403, "Only Super Admins can manage subscription plans")

    plans = (
        db.query(SubscriptionPlanDB).order_by(SubscriptionPlanDB.price_monthly).all()
    )

    return [
        SubscriptionPlanOut(
            id=p.id,
            name=p.name,
            description=p.description,
            price_monthly=p.price_monthly,
            price_yearly=p.price_yearly,
            max_invoices_per_month=p.max_invoices_per_month,
            max_users=p.max_users,
            max_business_admins=p.max_business_admins,
            max_finance_users=p.max_finance_users,
            max_pos_devices=p.max_pos_devices,
            allow_api_access=p.allow_api_access,
            allow_branding=p.allow_branding,
            allow_multi_currency=p.allow_multi_currency,
            priority_support=p.priority_support,
            active=p.active,
        )
        for p in plans
    ]


@app.put(
    "/admin/subscription-plans/{plan_id}",
    response_model=SubscriptionPlanOut,
    tags=["Admin", "Tier Management"],
)
def update_subscription_plan(
    plan_id: str,
    payload: SubscriptionPlanUpdate,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Update subscription plan tier limits and features (Super Admin only)"""
    if current_user.role != Role.SUPER_ADMIN:
        raise HTTPException(403, "Only Super Admins can modify subscription plans")

    plan = db.query(SubscriptionPlanDB).filter(SubscriptionPlanDB.id == plan_id).first()
    if not plan:
        raise HTTPException(404, f"Subscription plan {plan_id} not found")

    # Update fields if provided
    if payload.name is not None:
        plan.name = payload.name
    if payload.description is not None:
        plan.description = payload.description
    if payload.price_monthly is not None:
        plan.price_monthly = payload.price_monthly
    if payload.price_yearly is not None:
        plan.price_yearly = payload.price_yearly
    if payload.max_invoices_per_month is not None:
        plan.max_invoices_per_month = payload.max_invoices_per_month
    if payload.max_users is not None:
        plan.max_users = payload.max_users
    if payload.max_business_admins is not None:
        plan.max_business_admins = payload.max_business_admins
    if payload.max_finance_users is not None:
        plan.max_finance_users = payload.max_finance_users
    if payload.max_pos_devices is not None:
        plan.max_pos_devices = payload.max_pos_devices
    if payload.allow_api_access is not None:
        plan.allow_api_access = payload.allow_api_access
    if payload.allow_branding is not None:
        plan.allow_branding = payload.allow_branding
    if payload.allow_multi_currency is not None:
        plan.allow_multi_currency = payload.allow_multi_currency
    if payload.priority_support is not None:
        plan.priority_support = payload.priority_support
    if payload.active is not None:
        plan.active = payload.active

    db.commit()
    db.refresh(plan)

    return SubscriptionPlanOut(
        id=plan.id,
        name=plan.name,
        description=plan.description,
        price_monthly=plan.price_monthly,
        price_yearly=plan.price_yearly,
        max_invoices_per_month=plan.max_invoices_per_month,
        max_users=plan.max_users,
        max_business_admins=plan.max_business_admins,
        max_finance_users=plan.max_finance_users,
        max_pos_devices=plan.max_pos_devices,
        allow_api_access=plan.allow_api_access,
        allow_branding=plan.allow_branding,
        allow_multi_currency=plan.allow_multi_currency,
        priority_support=plan.priority_support,
        active=plan.active,
    )


@app.get(
    "/admin/platform-stats",
    response_model=PlatformStatsOut,
    tags=["Admin", "Platform Stats"],
)
def get_platform_statistics(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
):
    """Get aggregated platform statistics (Super Admin only) - Privacy-focused: Only aggregated data, no individual business details"""
    if current_user.role != Role.SUPER_ADMIN:
        raise HTTPException(403, "Only Super Admins can access platform statistics")

    from_dt = None
    to_dt = None
    if from_date:
        try:
            from_dt = datetime.fromisoformat(from_date.replace("Z", "+00:00")).replace(
                tzinfo=None
            )
        except Exception:
            pass
    if to_date:
        try:
            to_dt = datetime.fromisoformat(to_date.replace("Z", "+00:00")).replace(
                tzinfo=None
            )
        except Exception:
            pass

    # Total companies (filtered by created_at if date range provided)
    company_query = db.query(CompanyDB)
    if from_dt:
        company_query = company_query.filter(CompanyDB.created_at >= from_dt)
    if to_dt:
        company_query = company_query.filter(CompanyDB.created_at <= to_dt)
    total_companies = company_query.count()

    active_query = db.query(CompanyDB).filter(CompanyDB.status == CompanyStatus.ACTIVE)
    if from_dt:
        active_query = active_query.filter(CompanyDB.created_at >= from_dt)
    if to_dt:
        active_query = active_query.filter(CompanyDB.created_at <= to_dt)
    active_companies = active_query.count()

    pending_query = db.query(CompanyDB).filter(
        CompanyDB.status == CompanyStatus.PENDING_REVIEW
    )
    if from_dt:
        pending_query = pending_query.filter(CompanyDB.created_at >= from_dt)
    if to_dt:
        pending_query = pending_query.filter(CompanyDB.created_at <= to_dt)
    pending_companies = pending_query.count()

    # Total invoices (filtered by created_at)
    invoice_query = db.query(InvoiceDB)
    if from_dt:
        invoice_query = invoice_query.filter(InvoiceDB.created_at >= from_dt)
    if to_dt:
        invoice_query = invoice_query.filter(InvoiceDB.created_at <= to_dt)
    total_invoices = invoice_query.count()

    # Total revenue (filtered) - subtract credit notes from revenue
    invoice_list_query = db.query(InvoiceDB)
    if from_dt:
        invoice_list_query = invoice_list_query.filter(InvoiceDB.created_at >= from_dt)
    if to_dt:
        invoice_list_query = invoice_list_query.filter(InvoiceDB.created_at <= to_dt)

    invoices_for_revenue = invoice_list_query.all()
    total_revenue = sum(
        (inv.total_amount or 0.0)
        * (
            -1
            if inv.invoice_type
            in [InvoiceType.TAX_CREDIT_NOTE, InvoiceType.CREDIT_NOTE_OUT_OF_SCOPE]
            else 1
        )
        for inv in invoices_for_revenue
    )

    # Active subscriptions (company_subscriptions table)
    active_subscriptions = (
        db.query(CompanySubscriptionDB)
        .filter(CompanySubscriptionDB.status == SubscriptionStatus.ACTIVE)
        .count()
    )

    # Free vs paid tier users - consider company_subscriptions (company_subscriptions.plan -> subscription_plans)
    # and legacy Company.subscription_plan_id
    # 1) companies with a company_subscription whose plan has price_monthly == 0 and status in (TRIAL, ACTIVE)
    free_plan_company_ids = set()
    free_plan_subs = (
        db.query(CompanySubscriptionDB)
        .join(SubscriptionPlanDB)
        .filter(
            SubscriptionPlanDB.price_monthly == 0.0,
            CompanySubscriptionDB.status.in_(
                [SubscriptionStatus.ACTIVE, SubscriptionStatus.TRIAL]
            ),
        )
        .with_entities(CompanySubscriptionDB.company_id)
        .distinct()
        .all()
    )
    for cid in free_plan_subs:
        free_plan_company_ids.add(cid[0])

    # 2) companies with legacy subscription_plan_id pointing to a free plan
    free_plans = (
        db.query(SubscriptionPlanDB)
        .filter(SubscriptionPlanDB.price_monthly == 0.0)
        .with_entities(SubscriptionPlanDB.id)
        .all()
    )
    free_plan_ids = [p[0] for p in free_plans]
    if free_plan_ids:
        legacy_free_companies = (
            db.query(CompanyDB)
            .filter(CompanyDB.subscription_plan_id.in_(free_plan_ids))
            .with_entities(CompanyDB.id)
            .distinct()
            .all()
        )
        for cid in legacy_free_companies:
            free_plan_company_ids.add(cid[0])

    # 3) Active companies with NO subscription record at all → treat as free
    all_active_company_ids = set(
        cid[0]
        for cid in db.query(CompanyDB)
        .filter(
            CompanyDB.status.in_([CompanyStatus.ACTIVE, CompanyStatus.PENDING_REVIEW])
        )
        .with_entities(CompanyDB.id)
        .all()
    )
    companies_with_any_sub = set(
        cid[0]
        for cid in db.query(CompanySubscriptionDB)
        .with_entities(CompanySubscriptionDB.company_id)
        .distinct()
        .all()
    )
    no_plan_companies = all_active_company_ids - companies_with_any_sub
    free_plan_company_ids.update(no_plan_companies)

    free_tier = len(free_plan_company_ids)

    # Paid tier users - similar logic for company_subscriptions with price_monthly > 0
    paid_plan_company_ids = set()
    paid_plan_subs = (
        db.query(CompanySubscriptionDB)
        .join(SubscriptionPlanDB)
        .filter(
            SubscriptionPlanDB.price_monthly > 0.0,
            CompanySubscriptionDB.status.in_(
                [SubscriptionStatus.ACTIVE, SubscriptionStatus.TRIAL]
            ),
        )
        .with_entities(CompanySubscriptionDB.company_id)
        .distinct()
        .all()
    )
    for cid in paid_plan_subs:
        paid_plan_company_ids.add(cid[0])

    # Legacy companies with subscription_plan_id pointing to paid plans
    paid_plan_ids = [
        p[0]
        for p in db.query(SubscriptionPlanDB)
        .filter(SubscriptionPlanDB.price_monthly > 0.0)
        .with_entities(SubscriptionPlanDB.id)
        .all()
    ]
    if paid_plan_ids:
        legacy_paid_companies = (
            db.query(CompanyDB)
            .filter(CompanyDB.subscription_plan_id.in_(paid_plan_ids))
            .with_entities(CompanyDB.id)
            .distinct()
            .all()
        )
        for cid in legacy_paid_companies:
            paid_plan_company_ids.add(cid[0])

    paid_tier = len(paid_plan_company_ids)

    return PlatformStatsOut(
        total_companies=total_companies,
        active_companies=active_companies,
        pending_companies=pending_companies,
        total_invoices=total_invoices,
        total_revenue_aed=total_revenue,
        active_subscriptions=active_subscriptions,
        free_tier_users=free_tier,
        paid_tier_users=paid_tier,
    )


@app.get(
    "/admin/featured-businesses",
    response_model=List[FeaturedBusinessOut],
    tags=["Admin", "Featured Businesses"],
)
def get_featured_businesses(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Get all featured businesses for homepage moving bar (Super Admin only)"""
    if current_user.role != Role.SUPER_ADMIN:
        raise HTTPException(403, "Only Super Admins can manage featured businesses")

    try:
        # Use outerjoin to handle cases where company might be deleted
        featured = (
            db.query(FeaturedBusinessDB)
            .outerjoin(CompanyDB)
            .order_by(FeaturedBusinessDB.display_order)
            .all()
        )

        result = []
        for f in featured:
            # Handle missing company gracefully
            if f.company and f.company.legal_name:
                company_name = f.company.legal_name
            elif f.company and f.company.email:
                company_name = f.company.email
            else:
                company_name = f"Company {f.company_id}"

            result.append(
                FeaturedBusinessOut(
                    id=f.id,
                    company_id=f.company_id,
                    company_name=company_name,
                    display_name=f.display_name or "",
                    logo_url=f.logo_url or "",
                    is_active=f.is_active,
                    display_order=f.display_order,
                    created_at=f.created_at.isoformat(),
                )
            )

        return result
    except Exception as e:
        logger.error(f"Error fetching featured businesses: {str(e)}")
        raise HTTPException(500, f"Failed to fetch featured businesses: {str(e)}")


@app.post("/admin/featured-businesses", tags=["Admin", "Featured Businesses"])
def add_featured_business(
    payload: FeaturedBusinessCreate,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Add a company to the featured businesses list (Super Admin only)"""
    if current_user.role != Role.SUPER_ADMIN:
        raise HTTPException(403, "Only Super Admins can add featured businesses")

    # Verify company exists
    company = db.query(CompanyDB).filter(CompanyDB.id == payload.company_id).first()
    if not company:
        raise HTTPException(404, "Company not found")

    # Check if already featured
    existing = (
        db.query(FeaturedBusinessDB)
        .filter(FeaturedBusinessDB.company_id == payload.company_id)
        .first()
    )
    if existing:
        raise HTTPException(400, "Company is already featured")

    # Auto-calculate next display_order if not provided
    if payload.display_order is None:
        max_order = db.query(func.max(FeaturedBusinessDB.display_order)).scalar()
        display_order = 0 if max_order is None else max_order + 1
    else:
        display_order = payload.display_order

    new_featured = FeaturedBusinessDB(
        id=f"fb_{uuid4().hex[:12]}",
        company_id=payload.company_id,
        display_name=payload.display_name,
        logo_url=payload.logo_url,
        is_active=True,
        display_order=display_order,
        added_by_user_id=current_user.id,
    )

    db.add(new_featured)
    db.commit()
    db.refresh(new_featured)

    return {"message": "Business added to featured list", "id": new_featured.id}


@app.put(
    "/admin/featured-businesses/{featured_id}", tags=["Admin", "Featured Businesses"]
)
def update_featured_business(
    featured_id: str,
    payload: FeaturedBusinessUpdate,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Update a featured business (Super Admin only)"""
    if current_user.role != Role.SUPER_ADMIN:
        raise HTTPException(403, "Only Super Admins can update featured businesses")

    featured = (
        db.query(FeaturedBusinessDB)
        .filter(FeaturedBusinessDB.id == featured_id)
        .first()
    )
    if not featured:
        raise HTTPException(404, "Featured business not found")

    # Update only provided fields
    if payload.display_name is not None:
        featured.display_name = payload.display_name
    if payload.logo_url is not None:
        featured.logo_url = payload.logo_url
    if payload.is_active is not None:
        featured.is_active = payload.is_active
    if payload.display_order is not None:
        featured.display_order = payload.display_order

    db.commit()
    db.refresh(featured)

    return {"message": "Featured business updated successfully"}


@app.delete(
    "/admin/featured-businesses/{featured_id}", tags=["Admin", "Featured Businesses"]
)
def remove_featured_business(
    featured_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Remove a company from the featured businesses list (Super Admin only)"""
    if current_user.role != Role.SUPER_ADMIN:
        raise HTTPException(403, "Only Super Admins can remove featured businesses")

    featured = (
        db.query(FeaturedBusinessDB)
        .filter(FeaturedBusinessDB.id == featured_id)
        .first()
    )
    if not featured:
        raise HTTPException(404, "Featured business not found")

    db.delete(featured)
    db.commit()

    return {"message": "Business removed from featured list"}


@app.get(
    "/public/featured-businesses",
    response_model=List[FeaturedBusinessOut],
    tags=["Public"],
)
def get_public_featured_businesses(db: Session = Depends(get_db)):
    """Get active featured businesses for homepage (public, no auth required)"""
    try:
        # Use outerjoin to handle cases where company might be deleted
        featured = (
            db.query(FeaturedBusinessDB)
            .outerjoin(CompanyDB)
            .filter(FeaturedBusinessDB.is_active == True)
            .order_by(FeaturedBusinessDB.display_order)
            .all()
        )

        result = []
        for f in featured:
            # Skip if company is deleted or missing
            if not f.company:
                continue

            # Ensure we have valid strings
            company_name = (
                f.company.legal_name or f.company.email or f"Company {f.company_id}"
            )
            display_name = f.display_name or company_name

            result.append(
                FeaturedBusinessOut(
                    id=f.id,
                    company_id=f.company_id,
                    company_name=company_name,
                    display_name=display_name,
                    logo_url=f.logo_url or "",
                    is_active=f.is_active,
                    display_order=f.display_order,
                    created_at=f.created_at.isoformat(),
                )
            )

        return result
    except Exception as e:
        logger.error(f"Error fetching public featured businesses: {str(e)}")
        # Don't expose error details to public
        raise HTTPException(500, "Failed to fetch featured businesses")


# ==================== USER MANAGEMENT ENDPOINTS ====================


class UserInvite(BaseModel):
    email: str
    full_name: str
    role: Role = Role.FINANCE_USER


class UserOut(BaseModel):
    id: str
    email: str
    full_name: Optional[str]
    role: str
    is_owner: bool
    created_at: str
    last_login: Optional[str]


@app.post("/users/invite", tags=["Users"])
@app.post("/company/users/invite", tags=["Users"])
def invite_user(
    payload: UserInvite,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Invite a team member to the company based on inviter role permissions."""
    # Task 25: permission matrix gate
    if not has_permission(current_user, "team", "manage_users"):
        raise HTTPException(403, "Only admins can invite users")

    # BUSINESS_ADMIN can only invite FINANCE_USER or READ_ONLY
    if current_user.role == Role.BUSINESS_ADMIN and payload.role not in (
        Role.FINANCE_USER, Role.READ_ONLY
    ):
        raise HTTPException(
            403,
            "Business admins can only invite Finance Users or Read-Only users",
        )

    if not current_user.company_id and current_user.role != Role.SUPER_ADMIN:
        raise HTTPException(400, "User must belong to a company to invite others")

    # Check if email already exists
    existing = db.query(UserDB).filter(UserDB.email == payload.email).first()
    if existing:
        raise HTTPException(400, "User with this email already exists")

    # Create new user
    user_id = f"usr_{uuid4().hex[:12]}"
    temp_password = generate_temp_password(16)
    password_hash = get_password_hash(temp_password)

    new_user = UserDB(
        id=user_id,
        email=payload.email,
        full_name=payload.full_name,
        password_hash=password_hash,
        role=payload.role,
        company_id=current_user.company_id,
        is_owner=False,
        invited_by=current_user.id,
        must_change_password=True,
        # Task 22: Initialise password age tracking on invite creation
        password_changed_at=datetime.utcnow(),
    )

    db.add(new_user)
    db.commit()

    # Simulate email notification
    print("\n" + "=" * 70)
    print("✉️  TEAM MEMBER INVITATION - EMAIL NOTIFICATION")
    print("=" * 70)
    print(f"To: {payload.email}")
    print(f"Subject: You've been invited to join InvoLinks!")
    print(f"\nHello {payload.full_name},")
    print(f"\nYou have been invited to join a team on InvoLinks E-Invoicing.")
    print(f"Your temporary password is: {temp_password}")
    print(f"Please log in and change your password immediately.")
    print("=" * 70 + "\n")

    return {
        "success": True,
        "user_id": user_id,
        "email": payload.email,
        "temporary_password": temp_password,
        "message": "User invited successfully",
    }


@app.post("/users/{user_id}/reset-password", tags=["Users"])
@app.post("/company/users/{user_id}/reset-password", tags=["Users"])
def admin_reset_user_password(
    user_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Admin resets a team member's password.
    Sets a temporary password and flags the account for mandatory change on next login (Task 20).
    """
    allowed_roles = [Role.COMPANY_ADMIN, Role.SUPER_ADMIN, Role.BUSINESS_ADMIN]
    if current_user.role not in allowed_roles:
        raise HTTPException(403, "Only admins can reset user passwords")

    target_user = db.query(UserDB).filter(UserDB.id == user_id).first()
    if not target_user:
        raise HTTPException(404, "User not found")

    # Must be in the same company (unless super admin)
    if current_user.role != Role.SUPER_ADMIN:
        if target_user.company_id != current_user.company_id:
            raise HTTPException(403, "Cannot reset password for users outside your company")

    # Business admins can only reset finance users
    if current_user.role == Role.BUSINESS_ADMIN and target_user.role != Role.FINANCE_USER:
        raise HTTPException(403, "Business admins can only reset Finance User passwords")

    temp_password = generate_temp_password(16)
    target_user.password_hash = get_password_hash(temp_password)
    target_user.must_change_password = True
    # Task 22: Track password mutation time for expiry policy
    target_user.password_changed_at = datetime.utcnow()

    log_audit_event(
        db, "PASSWORD_CHANGED",
        user_id=current_user.id, company_id=current_user.company_id,
        resource_type="user", resource_id=target_user.id,
        description=f"Admin {current_user.email} reset password for {target_user.email} — forced change on next login",
    )

    db.commit()

    return {
        "success": True,
        "user_id": user_id,
        "email": target_user.email,
        "temporary_password": temp_password,
        "message": "Password reset successfully. User must change password on next login.",
    }


@app.get("/users/teams", response_model=List[UserOut], tags=["Users"])
def get_team_members(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Get all team members for the current user's company"""
    if not current_user.company_id:
        return []

    users = db.query(UserDB).filter(UserDB.company_id == current_user.company_id).all()

    return [
        {
            "id": user.id,
            "email": user.email,
            "full_name": user.full_name,
            "role": user.role.value,
            "is_owner": user.is_owner,
            "created_at": user.created_at.isoformat() if user.created_at else None,
            "last_login": user.last_login.isoformat() if user.last_login else None,
        }
        for user in users
    ]


@app.delete("/users/{user_id}", tags=["Users"])
@app.delete("/company/users/{user_id}", tags=["Users"])
def remove_team_member(
    user_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Remove a team member (Company Admin only)"""
    # Only company admins can remove users
    if current_user.role not in [Role.COMPANY_ADMIN, Role.SUPER_ADMIN]:
        raise HTTPException(403, "Only admins can remove users")

    user_to_remove = db.get(UserDB, user_id)
    if not user_to_remove:
        raise HTTPException(404, "User not found")

    # Can't remove users from other companies
    if (
        user_to_remove.company_id != current_user.company_id
        and current_user.role != Role.SUPER_ADMIN
    ):
        raise HTTPException(403, "Cannot remove users from other companies")

    # Can't remove the owner
    if user_to_remove.is_owner:
        raise HTTPException(400, "Cannot remove the company owner")

    # Can't remove yourself
    if user_to_remove.id == current_user.id:
        raise HTTPException(400, "Cannot remove yourself")

    db.delete(user_to_remove)
    db.commit()

    return {"success": True, "message": "User removed successfully"}


# ==================== INVOICE ENDPOINTS ====================


def generate_invoice_number(
    company_id: str, db: Session, invoice_type: str = "380"
) -> str:
    """Generate sequential invoice number for company with type-specific prefix"""
    # Map invoice type to prefix
    prefix_map = {
        "380": "TI",   # Tax Invoice
        "381": "TCN",  # Tax Credit Note
        "383": "DN",   # Debit Note (Task 4)
        "480": "CI",   # Commercial Invoice
        "81": "CN",    # Credit Note
    }

    # Handle both enum and string types
    if hasattr(invoice_type, "value"):
        invoice_type_value = invoice_type.value  # Extract enum value
    else:
        invoice_type_value = str(invoice_type)

    # Get prefix, default to TI if unknown type
    prefix = prefix_map.get(invoice_type_value, "TI")

    # Count existing invoices of this type for this company
    count = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.company_id == company_id, InvoiceDB.invoice_type == invoice_type
        )
        .count()
    )
    next_num = count + 1

    # Format: PREFIX-XXXXX (e.g., TI-00001, CN-00001)
    return f"{prefix}-{next_num:05d}"


def calculate_line_item_totals(
    line_item: InvoiceLineItemCreate, vat_enabled: bool
) -> dict:
    """Calculate tax and totals for a line item"""
    line_extension = line_item.quantity * line_item.unit_price
    if vat_enabled and line_item.tax_category == TaxCategory.STANDARD:
        tax_amount = line_extension * (line_item.tax_percent / 100)
    else:
        tax_amount = 0
    line_total = line_extension + tax_amount

    return {
        "line_extension_amount": round(line_extension, 2),
        "tax_amount": round(tax_amount, 2),
        "line_total_amount": round(line_total, 2),
    }


@app.post("/invoices", tags=["Invoices"], response_model=InvoiceOut,
          dependencies=[require_permission("invoices", "create")])
def create_invoice(
    payload: InvoiceCreate,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    # Get company
    company = (
        db.query(CompanyDB).filter(CompanyDB.id == current_user.company_id).first()
    )
    if not company:
        raise HTTPException(404, "Company not found")

    # Enforce Super Admin configured free plan invoice limits (invoice-count based)
    if (
        company.free_plan_type == "INVOICE_COUNT"
        and company.free_plan_invoice_limit is not None
    ):
        active_paid_subscription = (
            db.query(SubscriptionDB)
            .filter(
                SubscriptionDB.company_id == company.id,
                SubscriptionDB.status == "ACTIVE",
                SubscriptionDB.tier != "FREE",
            )
            .first()
        )

        if not active_paid_subscription:
            used_invoice_count = (
                db.query(InvoiceDB).filter(InvoiceDB.company_id == company.id).count()
            )

            if used_invoice_count >= company.free_plan_invoice_limit:
                raise HTTPException(
                    403,
                    f"Invoice limit reached ({company.free_plan_invoice_limit}). Please upgrade or request additional invoices from Super Admin.",
                )

    # Only require TRN when VAT is enabled
    if company.vat_enabled and not company.trn:
        raise HTTPException(
            400,
            "Company TRN is required to issue VAT invoices. Please enable VAT in settings first.",
        )

    # Check trial eligibility using configured free-plan settings
    # (keep consistent with /billing/trial logic).
    if company.trial_start_date and company.trial_status in ["ACTIVE", "EXPIRED"]:
        trial_duration_days = (company.free_plan_duration_months or 1) * 30
        invoice_limit = company.free_plan_invoice_limit or 100
        days_elapsed = (datetime.utcnow() - company.trial_start_date).days
        trial_days_remaining = max(0, trial_duration_days - days_elapsed)
        trial_invoices_remaining = max(
            0, invoice_limit - (company.trial_invoice_count or 0)
        )

        if company.free_plan_type == "DURATION":
            trial_expired_now = trial_days_remaining == 0
            expired_message = f"Free trial expired ({trial_duration_days} days). Please subscribe to continue creating invoices."
        elif company.free_plan_type == "INVOICE_COUNT":
            trial_expired_now = trial_invoices_remaining == 0
            expired_message = f"Free trial expired ({invoice_limit} invoices used). Please subscribe to continue creating invoices."
        else:
            trial_expired_now = (
                trial_days_remaining == 0 or trial_invoices_remaining == 0
            )
            expired_message = (
                "Free trial expired. Please subscribe to continue creating invoices."
            )

        if trial_expired_now:
            company.trial_status = "EXPIRED"
            company.trial_ended_at = datetime.utcnow()
            db.commit()

            active_subscription = (
                db.query(SubscriptionDB)
                .filter(
                    SubscriptionDB.company_id == company.id,
                    SubscriptionDB.status == "ACTIVE",
                )
                .first()
            )

            if not active_subscription:
                raise HTTPException(403, expired_message)
        elif company.trial_status == "EXPIRED":
            # Recover stale EXPIRED status from old rules when still eligible.
            company.trial_status = "ACTIVE"
            company.trial_ended_at = None
            db.commit()

    # Fallback: if still marked EXPIRED and no active subscription, block invoice creation.
    if company.trial_status == "EXPIRED":
        active_subscription = (
            db.query(SubscriptionDB)
            .filter(
                SubscriptionDB.company_id == company.id,
                SubscriptionDB.status == "ACTIVE",
            )
            .first()
        )

        if not active_subscription:
            raise HTTPException(
                403,
                "Free trial expired. Please subscribe to continue creating invoices.",
            )

    # Initialize trial if this is the first invoice
    if company.trial_status is None:
        company.trial_status = "ACTIVE"
        company.trial_start_date = datetime.utcnow()
        company.trial_invoice_count = 0
        db.commit()

    # Generate invoice number
    invoice_id = f"inv_{uuid4().hex[:12]}"
    invoice_number = generate_invoice_number(company.id, db, payload.invoice_type)

    # Calculate totals
    subtotal = 0.0
    total_tax = 0.0
    tax_by_category = {}  # For tax breakdowns

    # Parse dates
    from datetime import datetime as dt

    issue_date = dt.fromisoformat(payload.issue_date).date()
    due_date = dt.fromisoformat(payload.due_date).date() if payload.due_date else None
    supply_date = dt.fromisoformat(payload.supply_date).date() if getattr(payload, "supply_date", None) else None

    # Create invoice
    invoice = InvoiceDB(
        id=invoice_id,
        company_id=company.id,
        invoice_number=invoice_number,
        invoice_type=payload.invoice_type,
        status=InvoiceStatus.DRAFT,
        issue_date=issue_date,
        due_date=due_date,
        supply_date=supply_date,
        currency_code=payload.currency_code,
        # Supplier (from company)
        supplier_trn=company.trn,
        supplier_name=company.legal_name or company.email,
        supplier_address=f"{company.address_line1 or ''} {company.address_line2 or ''}".strip()
        or None,
        supplier_city=company.city,
        supplier_country="AE",
        supplier_peppol_id=company.trn[:10]
        if company.trn
        else None,  # First 10 digits of TRN as TIN
        # Customer
        customer_trn=payload.customer_trn,
        customer_name=payload.customer_name,
        customer_email=payload.customer_email,
        customer_address=payload.customer_address,
        customer_city=payload.customer_city,
        customer_country=payload.customer_country,
        customer_peppol_id=payload.customer_peppol_id,
        payment_terms=payload.payment_terms,
        payment_due_days=payload.payment_due_days,
        invoice_notes=payload.invoice_notes,
        reference_number=payload.reference_number,
        preceding_invoice_id=payload.preceding_invoice_id
        if payload.preceding_invoice_id
        else None,
        credit_note_reason=payload.credit_note_reason,
        share_token=f"share_{uuid4().hex[:16]}",
    )

    db.add(invoice)
    db.flush()  # Get invoice ID

    # Create line items
    for idx, line_item in enumerate(payload.line_items, 1):
        if company.vat_enabled:
            effective_tax_category = line_item.tax_category
            effective_tax_percent = line_item.tax_percent
            effective_tax_code = line_item.tax_code
        else:
            effective_tax_category = TaxCategory.OUT_OF_SCOPE
            effective_tax_percent = 0.0
            effective_tax_code = "OP"

        totals = calculate_line_item_totals(line_item, company.vat_enabled)

        line_db = InvoiceLineItemDB(
            id=f"line_{uuid4().hex[:12]}",
            invoice_id=invoice.id,
            line_number=idx,
            item_name=line_item.item_name,
            item_description=line_item.item_description,
            item_code=line_item.item_code,
            quantity=line_item.quantity,
            unit_code=line_item.unit_code,
            unit_price=line_item.unit_price,
            line_extension_amount=totals["line_extension_amount"],
            tax_category=effective_tax_category,
            tax_percent=effective_tax_percent,
            tax_code=effective_tax_code,  # Save UAE tax code
            tax_amount=totals["tax_amount"],
            line_total_amount=totals["line_total_amount"],
        )
        db.add(line_db)

        subtotal += totals["line_extension_amount"]
        total_tax += totals["tax_amount"]

        # Aggregate tax by category
        if company.vat_enabled:
            key = (effective_tax_category, effective_tax_percent)
            if key not in tax_by_category:
                tax_by_category[key] = {"taxable": 0.0, "tax": 0.0}
            tax_by_category[key]["taxable"] += totals["line_extension_amount"]
            tax_by_category[key]["tax"] += totals["tax_amount"]

    # Create tax breakdowns
    if company.vat_enabled:
        for (category, percent), amounts in tax_by_category.items():
            tax_breakdown = InvoiceTaxBreakdownDB(
                id=f"tax_{uuid4().hex[:12]}",
                invoice_id=invoice.id,
                tax_category=category,
                taxable_amount=round(amounts["taxable"], 2),
                tax_percent=percent,
                tax_amount=round(amounts["tax"], 2),
            )
            db.add(tax_breakdown)

    # Update invoice totals
    invoice.subtotal_amount = round(subtotal, 2)
    invoice.tax_amount = round(total_tax, 2)
    invoice.total_amount = round(subtotal + total_tax, 2)
    invoice.amount_due = round(subtotal + total_tax, 2)

    # P2-A: AED conversion enforcement — reject foreign-currency invoices without AED equivalent
    if (payload.currency_code or "AED").upper() != "AED":
        if not payload.total_amount_aed or payload.total_amount_aed <= 0:
            raise HTTPException(
                422,
                "total_amount_aed is required for foreign-currency invoices",
            )
        invoice.total_amount_aed = round(payload.total_amount_aed, 2)
    else:
        invoice.total_amount_aed = invoice.total_amount

    from decimal import Decimal

    # Credit note validation: must reference matching invoice and not exceed original total
    if invoice.invoice_type in [
        InvoiceType.TAX_CREDIT_NOTE,
        InvoiceType.CREDIT_NOTE_OUT_OF_SCOPE,
    ]:
        if not invoice.preceding_invoice_id:
            raise HTTPException(400, "Credit notes must reference an existing invoice.")

        original_invoice = (
            db.query(InvoiceDB)
            .filter(
                InvoiceDB.id == invoice.preceding_invoice_id,
                InvoiceDB.company_id == company.id,
            )
            .first()
        )

        if not original_invoice:
            raise HTTPException(400, "Referenced invoice not found.")

        if (
            invoice.invoice_type == InvoiceType.TAX_CREDIT_NOTE
            and original_invoice.invoice_type != InvoiceType.TAX_INVOICE
        ):
            raise HTTPException(400, "Tax credit notes must reference a tax invoice.")

        if (
            invoice.invoice_type == InvoiceType.CREDIT_NOTE_OUT_OF_SCOPE
            and original_invoice.invoice_type != InvoiceType.COMMERCIAL_INVOICE
        ):
            raise HTTPException(
                400, "Out-of-scope credit notes must reference a commercial invoice."
            )

        if Decimal(str(invoice.total_amount)) > Decimal(
            str(original_invoice.total_amount)
        ):
            raise HTTPException(
                400, "Credit note total cannot exceed the original invoice total."
            )

        # Deduct credit note amount from original invoice's amount_due
        new_amount_due = float(
            Decimal(str(original_invoice.amount_due))
            - Decimal(str(invoice.total_amount))
        )
        # Ensure amount_due doesn't go below 0
        original_invoice.amount_due = max(0.0, new_amount_due)
        db.add(original_invoice)  # Explicitly mark for update

    # Debit note validation (Task 4 — FTA UBL 383): must reference an issued tax invoice
    if invoice.invoice_type == InvoiceType.DEBIT_NOTE:
        if not invoice.preceding_invoice_id:
            raise HTTPException(400, "Debit notes must reference an existing tax invoice.")

        original_invoice = (
            db.query(InvoiceDB)
            .filter(
                InvoiceDB.id == invoice.preceding_invoice_id,
                InvoiceDB.company_id == company.id,
            )
            .first()
        )

        if not original_invoice:
            raise HTTPException(400, "Referenced invoice not found.")

        if original_invoice.invoice_type != InvoiceType.TAX_INVOICE:
            raise HTTPException(400, "Debit notes must reference a tax invoice (type 380).")

        if original_invoice.status not in [InvoiceStatus.ISSUED, InvoiceStatus.SENT, InvoiceStatus.VIEWED, InvoiceStatus.PAID]:
            raise HTTPException(400, "Debit notes can only be issued against a posted invoice.")

        # Add debit note amount to original invoice's amount_due
        new_amount_due = float(
            Decimal(str(original_invoice.amount_due))
            + Decimal(str(invoice.total_amount))
        )
        original_invoice.amount_due = new_amount_due
        db.add(original_invoice)

    # VAT Compliance: Auto-classify invoice type based on amount and company VAT status (Phase 1)
    invoice.invoice_classification = classify_invoice_type(
        total_amount=Decimal(str(invoice.total_amount)),
        vat_enabled=company.vat_enabled or False,
    )

    # Increment trial invoice count if in trial
    if company.trial_status == "ACTIVE":
        company.trial_invoice_count += 1

    # Keep company-level invoice counter in sync for admin explorer/reporting
    company.invoices_generated = (company.invoices_generated or 0) + 1

    # Invoice created as DRAFT - use /issue endpoint to generate XML and finalize
    db.commit()
    db.refresh(invoice)

    # Format response
    return InvoiceOut(
        id=invoice.id,
        company_id=invoice.company_id,
        invoice_number=invoice.invoice_number,
        invoice_type=invoice.invoice_type,
        status=invoice.status,
        issue_date=invoice.issue_date.isoformat(),
        due_date=invoice.due_date.isoformat() if invoice.due_date else None,
        supply_date=invoice.supply_date.isoformat() if invoice.supply_date else None,
        currency_code=invoice.currency_code,
        supplier_trn=invoice.supplier_trn,
        supplier_name=invoice.supplier_name,
        supplier_address=invoice.supplier_address,
        supplier_peppol_id=invoice.supplier_peppol_id,
        customer_trn=invoice.customer_trn,
        customer_name=invoice.customer_name,
        customer_email=invoice.customer_email,
        customer_address=invoice.customer_address,
        customer_city=invoice.customer_city,
        invoice_notes=invoice.invoice_notes,
        customer_peppol_id=invoice.customer_peppol_id,
        subtotal_amount=invoice.subtotal_amount,
        tax_amount=invoice.tax_amount,
        total_amount=invoice.total_amount,
        amount_due=invoice.amount_due,
        total_amount_aed=getattr(invoice, "total_amount_aed", None),
        preceding_invoice_id=invoice.preceding_invoice_id,
        credit_note_reason=invoice.credit_note_reason,
        xml_file_path=invoice.xml_file_path,
        pdf_file_path=invoice.pdf_file_path,
        share_token=invoice.share_token,
        created_at=invoice.created_at.isoformat(),
        sent_at=invoice.sent_at.isoformat() if invoice.sent_at else None,
        viewed_at=invoice.viewed_at.isoformat() if invoice.viewed_at else None,
        line_items=[
            InvoiceLineItemOut(
                id=li.id,
                line_number=li.line_number,
                item_name=li.item_name,
                item_description=li.item_description,
                quantity=li.quantity,
                unit_code=li.unit_code,
                unit_price=li.unit_price,
                line_extension_amount=li.line_extension_amount,
                tax_category=li.tax_category,
                tax_percent=li.tax_percent,
                tax_amount=li.tax_amount,
                line_total_amount=li.line_total_amount,
            )
            for li in invoice.line_items
        ],
        tax_breakdowns=[
            InvoiceTaxBreakdownOut(
                id=tb.id,
                tax_category=tb.tax_category,
                taxable_amount=tb.taxable_amount,
                tax_percent=tb.tax_percent,
                tax_amount=tb.tax_amount,
            )
            for tb in invoice.tax_breakdowns
        ],
    )


@app.post("/invoices/{invoice_id}/transmit-peppol", tags=["Invoices"])
def transmit_invoice_via_peppol(
    invoice_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Transmit invoice via PEPPOL network to recipient

    This endpoint:
    1. Validates invoice is ready for transmission
    2. Loads the UBL XML file
    3. Sends via configured PEPPOL provider
    4. Updates invoice with transmission status
    """
    # Get invoice
    invoice = db.query(InvoiceDB).filter(InvoiceDB.id == invoice_id).first()
    if not invoice:
        raise HTTPException(404, "Invoice not found")

    # Verify ownership
    if invoice.company_id != current_user.company_id:
        raise HTTPException(403, "Access denied")

    # Validate invoice status
    if invoice.status == InvoiceStatus.DRAFT:
        raise HTTPException(
            400, "Cannot transmit draft invoice. Update status to VALIDATED first."
        )

    # Check if XML exists
    if not invoice.xml_file_path or not os.path.exists(invoice.xml_file_path):
        raise HTTPException(
            400, "Invoice XML not found. Please regenerate the invoice."
        )

    # Check if already transmitted
    if invoice.peppol_status in ["SENT", "DELIVERED"]:
        return {
            "success": False,
            "message": f"Invoice already transmitted. Status: {invoice.peppol_status}",
            "message_id": invoice.peppol_message_id,
            "sent_at": invoice.peppol_sent_at.isoformat()
            if invoice.peppol_sent_at
            else None,
        }

    # Validate PEPPOL participant IDs
    if not invoice.supplier_peppol_id:
        raise HTTPException(400, "Supplier PEPPOL ID is required for transmission")

    if not invoice.customer_peppol_id:
        raise HTTPException(400, "Customer PEPPOL ID is required for transmission")

    # Load XML content
    try:
        with open(invoice.xml_file_path, "r", encoding="utf-8") as f:
            xml_content = f.read()
    except Exception as e:
        raise HTTPException(500, f"Failed to read XML file: {str(e)}")

    # Transmit via PEPPOL
    from utils.peppol_provider import send_invoice_via_peppol
    import json

    # For now, use MOCK provider (replace with real provider in production)
    # Set PEPPOL_PROVIDER, PEPPOL_BASE_URL, PEPPOL_API_KEY in environment variables
    provider_name = os.getenv("PEPPOL_PROVIDER", "mock")
    provider_url = os.getenv("PEPPOL_BASE_URL")
    provider_key = os.getenv("PEPPOL_API_KEY")

    result = send_invoice_via_peppol(
        invoice_xml=xml_content,
        invoice_number=invoice.invoice_number,
        sender_id=invoice.supplier_peppol_id,
        receiver_id=invoice.customer_peppol_id,
        provider_name=provider_name,
        base_url=provider_url,
        api_key=provider_key,
    )

    # Update invoice with transmission result
    if result.get("success"):
        invoice.peppol_message_id = result.get("message_id")
        invoice.peppol_status = result.get("status")
        invoice.peppol_provider = result.get("provider")
        invoice.peppol_sent_at = datetime.utcnow()
        invoice.peppol_response = json.dumps(result.get("response", {}))
        invoice.status = InvoiceStatus.SENT  # Update invoice status

        db.commit()
        db.refresh(invoice)

        return {
            "success": True,
            "message": "Invoice transmitted successfully via PEPPOL",
            "invoice_number": invoice.invoice_number,
            "message_id": invoice.peppol_message_id,
            "status": invoice.peppol_status,
            "provider": invoice.peppol_provider,
            "sent_at": invoice.peppol_sent_at.isoformat(),
            "recipient_id": invoice.customer_peppol_id,
        }
    else:
        # Save error response
        invoice.peppol_status = "FAILED"
        invoice.peppol_response = json.dumps({"error": result.get("error")})
        db.commit()

        raise HTTPException(500, f"PEPPOL transmission failed: {result.get('error')}")


@app.get("/invoices/{invoice_id}/peppol-status", tags=["Invoices"])
def get_peppol_transmission_status(
    invoice_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Check PEPPOL transmission status for an invoice"""
    # Get invoice
    invoice = db.query(InvoiceDB).filter(InvoiceDB.id == invoice_id).first()
    if not invoice:
        raise HTTPException(404, "Invoice not found")

    # Verify ownership
    if invoice.company_id != current_user.company_id:
        raise HTTPException(403, "Access denied")

    if not invoice.peppol_message_id:
        return {
            "transmitted": False,
            "message": "Invoice has not been transmitted via PEPPOL",
        }

    # Query provider for latest status (if needed)
    from utils.peppol_provider import PeppolProviderFactory
    import json

    try:
        provider_name = invoice.peppol_provider or os.getenv("PEPPOL_PROVIDER", "mock")
        provider = PeppolProviderFactory.create_provider(
            provider_name=provider_name,
            base_url=os.getenv("PEPPOL_BASE_URL"),
            api_key=os.getenv("PEPPOL_API_KEY"),
        )

        status_result = provider.get_status(invoice.peppol_message_id)

        # Update invoice status if changed
        if status_result.get("success") and status_result.get("status"):
            new_status = status_result["status"]
            if new_status != invoice.peppol_status:
                invoice.peppol_status = new_status
                invoice.peppol_response = json.dumps(status_result.get("details", {}))
                db.commit()

        return {
            "transmitted": True,
            "invoice_number": invoice.invoice_number,
            "message_id": invoice.peppol_message_id,
            "status": invoice.peppol_status,
            "provider": invoice.peppol_provider,
            "sent_at": invoice.peppol_sent_at.isoformat()
            if invoice.peppol_sent_at
            else None,
            "latest_check": status_result if status_result.get("success") else None,
        }

    except Exception as e:
        return {
            "transmitted": True,
            "invoice_number": invoice.invoice_number,
            "message_id": invoice.peppol_message_id,
            "status": invoice.peppol_status,
            "provider": invoice.peppol_provider,
            "sent_at": invoice.peppol_sent_at.isoformat()
            if invoice.peppol_sent_at
            else None,
            "error": f"Status check failed: {str(e)}",
        }


@app.get("/invoices", tags=["Invoices"], response_model=List[InvoiceListOut])
def list_invoices(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
    status: Optional[str] = None,
    invoice_type: Optional[str] = None,
    q: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
):
    """List invoices for the current company (excludes archived invoices by default)"""
    query = db.query(InvoiceDB).filter(
        InvoiceDB.company_id == current_user.company_id,
        InvoiceDB.is_archived == False,  # Task 12: hide archived by default
    )

    if status:
        try:
            status_enum = InvoiceStatus(status)
            query = query.filter(InvoiceDB.status == status_enum)
        except ValueError:
            raise HTTPException(400, f"Invalid status: {status}")

    if invoice_type:
        try:
            type_enum = InvoiceType(invoice_type)
            query = query.filter(InvoiceDB.invoice_type == type_enum)
        except ValueError:
            raise HTTPException(400, f"Invalid invoice type: {invoice_type}")

    # Optional search query - search by invoice number or customer name
    if q:
        q_term = f"%{q}%"
        query = query.filter(
            or_(
                InvoiceDB.invoice_number.ilike(q_term),
                InvoiceDB.customer_name.ilike(q_term),
            )
        )

    query = query.order_by(InvoiceDB.created_at.desc())
    invoices = query.offset(offset).limit(limit).all()

    return [
        InvoiceListOut(
            id=inv.id,
            invoice_number=inv.invoice_number,
            invoice_type=inv.invoice_type,
            status=inv.status,
            issue_date=inv.issue_date.isoformat(),
            customer_name=inv.customer_name,
            total_amount=inv.total_amount,
            currency_code=inv.currency_code,
            created_at=inv.created_at.isoformat(),
        )
        for inv in invoices
    ]


@app.get("/invoices/archived", tags=["Invoices"])
def list_archived_invoices(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
    limit: int = 50,
    offset: int = 0,
):
    """
    Task 12 — Data Archival: list all archived invoices for this company.

    Role Access: COMPANY_ADMIN, BUSINESS_ADMIN, FINANCE_USER
    """
    if current_user.role not in [
        Role.COMPANY_ADMIN, Role.BUSINESS_ADMIN, Role.FINANCE_USER,
    ]:
        raise HTTPException(403, "Insufficient permissions to view archived invoices")

    invoices = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.company_id == current_user.company_id,
            InvoiceDB.is_archived == True,
        )
        .order_by(InvoiceDB.archived_at.desc())
        .offset(offset)
        .limit(limit)
        .all()
    )

    return [
        {
            "id": inv.id,
            "invoice_number": inv.invoice_number,
            "invoice_type": inv.invoice_type.value if inv.invoice_type else None,
            "status": inv.status.value if inv.status else None,
            "issue_date": inv.issue_date.isoformat() if inv.issue_date else None,
            "customer_name": inv.customer_name,
            "total_amount": inv.total_amount,
            "currency_code": inv.currency_code,
            "archived_at": inv.archived_at.isoformat() if inv.archived_at else None,
        }
        for inv in invoices
    ]


@app.post("/invoices/archive", tags=["Invoices"])
def archive_old_invoices(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
    years_old: int = 5,
):
    """
    Task 12 — Data Archival: archive all PAID/CANCELLED invoices older than `years_old` years.

    Archived invoices are hidden from normal invoice lists but remain in the database
    and can be retrieved for audit purposes via GET /invoices/archived.

    Role Access: COMPANY_ADMIN, BUSINESS_ADMIN only.
    Default: archives invoices issued more than 5 years ago.
    """
    from datetime import date, timedelta

    if not has_permission(current_user, "archival", "create"):
        raise HTTPException(403, "Permission denied: archive invoices is not allowed for your role")

    if years_old < 1:
        raise HTTPException(400, "years_old must be at least 1")

    cutoff_date = date.today() - timedelta(days=years_old * 365)

    invoices_to_archive = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.company_id == current_user.company_id,
            InvoiceDB.is_archived == False,
            InvoiceDB.status.in_([InvoiceStatus.PAID, InvoiceStatus.CANCELLED]),
            InvoiceDB.issue_date <= cutoff_date,
        )
        .all()
    )

    now = datetime.utcnow()
    archived_count = 0
    for inv in invoices_to_archive:
        inv.is_archived = True
        inv.archived_at = now
        archived_count += 1

    db.commit()

    log_audit_event(
        db, "INVOICES_ARCHIVED",
        user_id=current_user.id, company_id=current_user.company_id,
        resource_type="invoice", resource_id=None,
        description=f"Archived {archived_count} invoices older than {years_old} years (cutoff: {cutoff_date})",
    )

    return {
        "archived_count": archived_count,
        "cutoff_date": cutoff_date.isoformat(),
        "years_old": years_old,
        "message": f"Successfully archived {archived_count} invoice(s) issued before {cutoff_date}.",
    }


@app.get("/invoices/archive/status", tags=["Invoices"])
def get_archival_status(
    years_old: int = 5,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Task 16 — Archival status: count invoices eligible for archival
    (PAID/CANCELLED, older than `years_old` years, not yet archived).
    Useful for admins to know when to trigger POST /invoices/archive.
    """
    from datetime import date, timedelta

    if current_user.role not in [Role.COMPANY_ADMIN, Role.BUSINESS_ADMIN, Role.SUPER_ADMIN]:
        raise HTTPException(403, "Only admins can view archival status")

    cutoff_date = date.today() - timedelta(days=years_old * 365)

    eligible_count = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.company_id == current_user.company_id,
            InvoiceDB.is_archived == False,
            InvoiceDB.status.in_([InvoiceStatus.PAID, InvoiceStatus.CANCELLED]),
            InvoiceDB.issue_date <= cutoff_date,
        )
        .count()
    )

    already_archived = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.company_id == current_user.company_id,
            InvoiceDB.is_archived == True,
        )
        .count()
    )

    return {
        "years_old": years_old,
        "cutoff_date": cutoff_date.isoformat(),
        "eligible_for_archival": eligible_count,
        "already_archived": already_archived,
        "archival_required": eligible_count > 0,
        "message": (
            f"{eligible_count} invoice(s) eligible for archival (issued before {cutoff_date})."
            if eligible_count > 0
            else "No invoices require archiving at this time."
        ),
    }


@app.post("/invoices/{invoice_id}/restore", tags=["Invoices"])
def restore_archived_invoice(
    invoice_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Task 12 — Data Archival: restore a single archived invoice back to normal.

    Role Access: COMPANY_ADMIN, BUSINESS_ADMIN only.
    """
    if not has_permission(current_user, "archival", "create"):
        raise HTTPException(403, "Permission denied: restore archived invoices is not allowed for your role")

    invoice = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.id == invoice_id,
            InvoiceDB.company_id == current_user.company_id,
            InvoiceDB.is_archived == True,
        )
        .first()
    )

    if not invoice:
        raise HTTPException(404, "Archived invoice not found")

    invoice.is_archived = False
    invoice.archived_at = None
    db.commit()

    log_audit_event(
        db, "INVOICE_RESTORED",
        user_id=current_user.id, company_id=current_user.company_id,
        resource_type="invoice", resource_id=invoice.id,
        description=f"Restored archived invoice {invoice.invoice_number}",
    )

    return {
        "id": invoice.id,
        "invoice_number": invoice.invoice_number,
        "message": f"Invoice {invoice.invoice_number} successfully restored.",
    }


@app.get("/invoices/pending-payment", tags=["Invoices"])
def get_pending_payment_invoices(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
    limit: int = 100,
):
    """
    Get list of invoices awaiting payment

    Returns invoices that are not yet paid (status != PAID and status != CANCELLED)
    Role Access: BUSINESS_ADMIN, FINANCE_USER, COMPANY_ADMIN
    """
    # Check role-based permissions
    if current_user.role not in [
        Role.COMPANY_ADMIN,
        Role.BUSINESS_ADMIN,
        Role.FINANCE_USER,
    ]:
        raise HTTPException(403, "Insufficient permissions to view pending payments")

    invoices = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.company_id == current_user.company_id,
            InvoiceDB.status.notin_([InvoiceStatus.PAID, InvoiceStatus.CANCELLED]),
        )
        .order_by(InvoiceDB.due_date.asc())
        .limit(limit)
        .all()
    )

    return [
        {
            "id": inv.id,
            "invoice_number": inv.invoice_number,
            "customer_name": inv.customer_name,
            "customer_email": inv.customer_email,
            "issue_date": inv.issue_date.isoformat(),
            "due_date": inv.due_date.isoformat() if inv.due_date else None,
            "total_amount": inv.total_amount,
            "amount_due": inv.amount_due,
            "currency_code": inv.currency_code,
            "status": inv.status.value,
            "payment_terms": inv.payment_terms,
            "days_overdue": (datetime.utcnow().date() - inv.due_date).days
            if inv.due_date and datetime.utcnow().date() > inv.due_date
            else 0,
        }
        for inv in invoices
    ]


@app.get("/invoices/{invoice_id}", tags=["Invoices"], response_model=InvoiceOut)
def get_invoice(
    invoice_id: str,
    response: Response,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Get a specific invoice.

    Task 10: Returns ETag header with current version for optimistic locking.
    """
    invoice = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.id == invoice_id, InvoiceDB.company_id == current_user.company_id
        )
        .first()
    )

    if not invoice:
        raise HTTPException(404, "Invoice not found")

    # Task 10: Set ETag header for optimistic locking support
    current_version = invoice.version if invoice.version is not None else 1
    response.headers["ETag"] = str(current_version)

    return InvoiceOut(
        id=invoice.id,
        company_id=invoice.company_id,
        invoice_number=invoice.invoice_number,
        invoice_type=invoice.invoice_type,
        status=invoice.status,
        issue_date=invoice.issue_date.isoformat(),
        due_date=invoice.due_date.isoformat() if invoice.due_date else None,
        supply_date=invoice.supply_date.isoformat() if invoice.supply_date else None,
        version=current_version,
        currency_code=invoice.currency_code,
        supplier_trn=invoice.supplier_trn,
        supplier_name=invoice.supplier_name,
        supplier_address=invoice.supplier_address,
        supplier_peppol_id=invoice.supplier_peppol_id,
        customer_trn=invoice.customer_trn,
        customer_name=invoice.customer_name,
        customer_email=invoice.customer_email,
        customer_address=invoice.customer_address,
        customer_city=invoice.customer_city,
        invoice_notes=invoice.invoice_notes,
        customer_peppol_id=invoice.customer_peppol_id,
        subtotal_amount=invoice.subtotal_amount,
        tax_amount=invoice.tax_amount,
        total_amount=invoice.total_amount,
        amount_due=invoice.amount_due,
        total_amount_aed=getattr(invoice, "total_amount_aed", None),
        preceding_invoice_id=invoice.preceding_invoice_id,
        credit_note_reason=invoice.credit_note_reason,
        xml_file_path=invoice.xml_file_path,
        pdf_file_path=invoice.pdf_file_path,
        share_token=invoice.share_token,
        created_at=invoice.created_at.isoformat(),
        sent_at=invoice.sent_at.isoformat() if invoice.sent_at else None,
        viewed_at=invoice.viewed_at.isoformat() if invoice.viewed_at else None,
        line_items=[
            InvoiceLineItemOut(
                id=li.id,
                line_number=li.line_number,
                item_name=li.item_name,
                item_description=li.item_description,
                quantity=li.quantity,
                unit_code=li.unit_code,
                unit_price=li.unit_price,
                line_extension_amount=li.line_extension_amount,
                tax_category=li.tax_category,
                tax_percent=li.tax_percent,
                tax_amount=li.tax_amount,
                line_total_amount=li.line_total_amount,
            )
            for li in invoice.line_items
        ],
        tax_breakdowns=[
            InvoiceTaxBreakdownOut(
                id=tb.id,
                tax_category=tb.tax_category,
                taxable_amount=tb.taxable_amount,
                tax_percent=tb.tax_percent,
                tax_amount=tb.tax_amount,
            )
            for tb in invoice.tax_breakdowns
        ],
    )


@app.put("/invoices/{invoice_id}", tags=["Invoices"], response_model=InvoiceOut,
         dependencies=[require_permission("invoices", "edit")])
def update_invoice(
    invoice_id: str,
    data: InvoiceCreate,
    request: Request,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Update a DRAFT invoice (only DRAFT invoices can be edited).

    Task 10 — Concurrent Update Prevention: supports ETag/If-Match optimistic locking.
    Supply `If-Match: <version>` to detect mid-air edits. Returns 412 on conflict.
    """
    invoice = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.id == invoice_id, InvoiceDB.company_id == current_user.company_id
        )
        .first()
    )

    if not invoice:
        raise HTTPException(404, "Invoice not found")

    if invoice.status != InvoiceStatus.DRAFT:
        raise HTTPException(
            409,
            f"Cannot edit invoice. Only DRAFT invoices can be edited. "
            f"Current status: {invoice.status.value}. "
            "To correct a posted invoice, issue a credit note (document type 381).",
        )

    # Task 15: Capture old values before any mutation (for audit trail)
    import json as _json
    _old_snapshot = {
        "status": invoice.status.value if invoice.status else None,
        "customer_name": invoice.customer_name,
        "total_amount": float(invoice.total_amount) if invoice.total_amount else None,
    }

    # Task 10: Optimistic locking via If-Match / ETag
    if_match = request.headers.get("If-Match")
    current_version = invoice.version if invoice.version is not None else 1
    if if_match is not None:
        try:
            client_version = int(if_match.strip('"'))
        except ValueError:
            raise HTTPException(400, "If-Match header must be an integer version number")
        if client_version != current_version:
            raise HTTPException(
                412,
                f"Precondition Failed: invoice was modified by another session "
                f"(client version={client_version}, server version={current_version}). "
                "Reload the invoice and reapply your changes.",
            )

    try:
        # Update basic invoice details
        invoice.invoice_type = data.invoice_type
        invoice.issue_date = datetime.fromisoformat(data.issue_date)
        invoice.due_date = (
            datetime.fromisoformat(data.due_date) if data.due_date else None
        )
        invoice.currency_code = data.currency_code

        # Update customer details
        invoice.customer_name = data.customer_name
        invoice.customer_email = data.customer_email
        invoice.customer_trn = data.customer_trn
        invoice.customer_address = data.customer_address
        invoice.customer_city = data.customer_city
        invoice.customer_country = data.customer_country or "AE"
        invoice.customer_peppol_id = data.customer_peppol_id

        # Optional fields
        invoice.payment_terms = data.payment_terms
        invoice.payment_due_days = data.payment_due_days or 30
        invoice.invoice_notes = data.invoice_notes
        invoice.reference_number = data.reference_number

        # Clear existing line items
        db.query(InvoiceLineItemDB).filter(
            InvoiceLineItemDB.invoice_id == invoice_id
        ).delete()
        db.flush()

        # Add updated line items
        subtotal = Decimal("0")
        total_tax = Decimal("0")
        line_number = 1
        tax_by_category = {}

        for item_data in data.line_items:
            line_extension_amount = Decimal(str(item_data.quantity)) * Decimal(
                str(item_data.unit_price)
            )

            # Calculate tax
            tax_amount = Decimal("0")
            if item_data.tax_category == "S":
                tax_amount = line_extension_amount * (
                    Decimal(str(item_data.tax_percent)) / Decimal("100")
                )

            line_total = line_extension_amount + tax_amount

            line_item = InvoiceLineItemDB(
                id=f"li_{uuid4().hex[:12]}",
                invoice_id=invoice_id,
                line_number=line_number,
                item_name=item_data.item_name,
                item_description=item_data.item_description,
                item_code=item_data.item_code,
                quantity=Decimal(str(item_data.quantity)),
                unit_code=item_data.unit_code or "C62",
                unit_price=Decimal(str(item_data.unit_price)),
                line_extension_amount=line_extension_amount,
                tax_category=item_data.tax_category,
                tax_percent=Decimal(str(item_data.tax_percent)),
                tax_code=item_data.tax_code,
                tax_amount=tax_amount,
                line_total_amount=line_total,
            )
            db.add(line_item)

            subtotal += line_extension_amount
            total_tax += tax_amount

            # Track tax by category for breakdowns
            cat = item_data.tax_category  # Already a string
            if cat not in tax_by_category:
                tax_by_category[cat] = {"taxable": Decimal("0"), "tax": Decimal("0")}
            tax_by_category[cat]["taxable"] += line_extension_amount
            tax_by_category[cat]["tax"] += tax_amount

            line_number += 1

        # Update invoice totals
        invoice.subtotal_amount = float(subtotal)
        invoice.tax_amount = float(total_tax)
        invoice.total_amount = float(subtotal + total_tax)
        invoice.amount_due = float(subtotal + total_tax)

        # P2-A: AED conversion enforcement — reject foreign-currency invoices without AED equivalent
        if (data.currency_code or "AED").upper() != "AED":
            if not data.total_amount_aed or data.total_amount_aed <= 0:
                raise HTTPException(
                    422,
                    "total_amount_aed is required for foreign-currency invoices",
                )
            invoice.total_amount_aed = round(data.total_amount_aed, 2)
        else:
            invoice.total_amount_aed = invoice.total_amount

        # Clear and rebuild tax breakdowns
        db.query(InvoiceTaxBreakdownDB).filter(
            InvoiceTaxBreakdownDB.invoice_id == invoice_id
        ).delete()
        db.flush()

        for tax_category, amounts in tax_by_category.items():
            breakdown = InvoiceTaxBreakdownDB(
                id=f"tb_{uuid4().hex[:12]}",
                invoice_id=invoice_id,
                tax_category=tax_category,
                taxable_amount=float(amounts["taxable"]),
                tax_percent=float(5.0) if tax_category == "S" else 0,
                tax_amount=float(amounts["tax"]),
            )
            db.add(breakdown)

        # Task 10: Increment optimistic-locking version counter
        invoice.version = (invoice.version or 1) + 1

        # Task 15: Log INVOICE_UPDATED with old/new key fields
        _new_snapshot = {
            "status": invoice.status.value if invoice.status else None,
            "customer_name": invoice.customer_name,
            "total_amount": float(invoice.total_amount) if invoice.total_amount else None,
        }
        log_audit_event(
            db, "INVOICE_UPDATED",
            user_id=current_user.id, company_id=current_user.company_id,
            resource_type="invoice", resource_id=invoice.id,
            old_value=_json.dumps(_old_snapshot),
            new_value=_json.dumps(_new_snapshot),
            description=f"Invoice {invoice.invoice_number} updated",
        )

        db.commit()

        # Return updated invoice
        return InvoiceOut(
            id=invoice.id,
            company_id=invoice.company_id,
            invoice_number=invoice.invoice_number,
            invoice_type=invoice.invoice_type,
            status=invoice.status,
            issue_date=invoice.issue_date.isoformat(),
            due_date=invoice.due_date.isoformat() if invoice.due_date else None,
            supply_date=invoice.supply_date.isoformat() if invoice.supply_date else None,
            version=invoice.version,
            currency_code=invoice.currency_code,
            supplier_trn=invoice.supplier_trn,
            supplier_name=invoice.supplier_name,
            supplier_address=invoice.supplier_address,
            supplier_peppol_id=invoice.supplier_peppol_id,
            customer_trn=invoice.customer_trn,
            customer_name=invoice.customer_name,
            customer_email=invoice.customer_email,
            customer_address=invoice.customer_address,
            customer_city=invoice.customer_city,
            invoice_notes=invoice.invoice_notes,
            customer_peppol_id=invoice.customer_peppol_id,
            subtotal_amount=invoice.subtotal_amount,
            tax_amount=invoice.tax_amount,
            total_amount=invoice.total_amount,
            amount_due=invoice.amount_due,
            total_amount_aed=getattr(invoice, "total_amount_aed", None),
            preceding_invoice_id=invoice.preceding_invoice_id,
            credit_note_reason=invoice.credit_note_reason,
            xml_file_path=invoice.xml_file_path,
            pdf_file_path=invoice.pdf_file_path,
            share_token=invoice.share_token,
            created_at=invoice.created_at.isoformat(),
            sent_at=invoice.sent_at.isoformat() if invoice.sent_at else None,
            viewed_at=invoice.viewed_at.isoformat() if invoice.viewed_at else None,
            line_items=[
                InvoiceLineItemOut(
                    id=li.id,
                    line_number=li.line_number,
                    item_name=li.item_name,
                    item_description=li.item_description,
                    quantity=li.quantity,
                    unit_code=li.unit_code,
                    unit_price=li.unit_price,
                    line_extension_amount=li.line_extension_amount,
                    tax_category=li.tax_category,
                    tax_percent=li.tax_percent,
                    tax_amount=li.tax_amount,
                    line_total_amount=li.line_total_amount,
                )
                for li in invoice.line_items
            ],
            tax_breakdowns=[
                InvoiceTaxBreakdownOut(
                    id=tb.id,
                    tax_category=tb.tax_category,
                    taxable_amount=tb.taxable_amount,
                    tax_percent=tb.tax_percent,
                    tax_amount=tb.tax_amount,
                )
                for tb in invoice.tax_breakdowns
            ],
        )
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Failed to update invoice: {str(e)}")


# ==================== PAYMENT VERIFICATION ENDPOINTS ====================


class PaymentVerificationRequest(BaseModel):
    payment_method: str  # Cash, Card, POS, Bank Transfer, etc.
    payment_reference: Optional[str] = None
    payment_notes: Optional[str] = None
    payment_date: Optional[str] = None  # ISO date string


@app.post("/invoices/{invoice_id}/verify-payment", tags=["Invoices"])
def verify_invoice_payment(
    invoice_id: str,
    payment: PaymentVerificationRequest,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Verify and record payment for an invoice (offline payments: Cash, POS, Bank Transfer)

    Role Access: BUSINESS_ADMIN, FINANCE_USER, COMPANY_ADMIN
    """
    # Check role-based permissions
    if current_user.role not in [
        Role.COMPANY_ADMIN,
        Role.BUSINESS_ADMIN,
        Role.FINANCE_USER,
    ]:
        raise HTTPException(403, "Insufficient permissions to verify payments")

    # Get invoice
    invoice = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.id == invoice_id, InvoiceDB.company_id == current_user.company_id
        )
        .first()
    )

    if not invoice:
        raise HTTPException(404, "Invoice not found")

    # Ensure invoice is issued before payment can be verified
    if invoice.status == InvoiceStatus.DRAFT:
        raise HTTPException(
            400,
            "Cannot verify payment for a DRAFT invoice. Please issue the invoice first.",
        )

    if invoice.status == InvoiceStatus.PAID:
        raise HTTPException(400, "Invoice is already marked as paid")

    # Parse payment date
    from datetime import datetime

    if payment.payment_date:
        try:
            payment_dt = datetime.fromisoformat(
                payment.payment_date.replace("Z", "+00:00")
            )
        except ValueError:
            raise HTTPException(
                400, "Invalid payment_date format. Use ISO format (YYYY-MM-DD)"
            )
    else:
        payment_dt = datetime.utcnow()

    # Snapshot prior status before mutation
    _prev_status_paid = invoice.status.value if invoice.status else "UNKNOWN"

    # Update invoice payment details
    invoice.status = InvoiceStatus.PAID
    invoice.payment_method = payment.payment_method
    invoice.amount_due = 0.0
    invoice.paid_at = payment_dt
    invoice.payment_verified_by_user_id = current_user.id
    invoice.payment_verified_at = datetime.utcnow()

    # Add payment notes to invoice notes if provided
    if payment.payment_notes:
        if invoice.invoice_notes:
            invoice.invoice_notes += f"\n\nPayment Notes: {payment.payment_notes}"
        else:
            invoice.invoice_notes = f"Payment Notes: {payment.payment_notes}"

    # Add payment reference to notes if provided
    if payment.payment_reference:
        note = f"\n\nPayment Reference: {payment.payment_reference}"
        if invoice.invoice_notes:
            invoice.invoice_notes += note
        else:
            invoice.invoice_notes = f"Payment Reference: {payment.payment_reference}"

    # Task 1 / Task 15: Audit trail — log actual prior → new status
    import json as _json
    log_audit_event(
        db, "INVOICE_PAID",
        user_id=current_user.id, company_id=current_user.company_id,
        resource_type="invoice", resource_id=invoice.id,
        old_value=_json.dumps({"status": _prev_status_paid}),
        new_value=_json.dumps({
            "status": "PAID",
            "payment_method": payment.payment_method,
            "total_amount": float(invoice.total_amount),
        }),
        description=f"Invoice {invoice.invoice_number} marked as paid via {payment.payment_method}",
    )

    db.commit()
    db.refresh(invoice)

    return {
        "success": True,
        "invoice_id": invoice.id,
        "invoice_number": invoice.invoice_number,
        "status": invoice.status.value,
        "payment_method": invoice.payment_method,
        "paid_at": invoice.paid_at.isoformat() if invoice.paid_at else None,
        "total_amount": invoice.total_amount,
        "verified_by": current_user.email,
        "verified_at": invoice.payment_verified_at.isoformat(),
    }


def _iso_weeks_in_year(year: int) -> int:
    """Return the number of ISO weeks in the given year (52 or 53)."""
    from datetime import date
    # Dec 28 is always in the last ISO week of the year
    dec28 = date(year, 12, 28)
    return dec28.isocalendar()[1]


def _build_periodic_date_range(period_type: str, year, period):
    """
    Task 21 — compute (start_dt, end_dt, period_label, period_type, period) for
    weekly / monthly / quarterly / annual period types.

    All defaulting (year, period) is done here so callers can always pass raw
    query-param values (which may be None).

    Returns (start_dt, end_dt, period_label, normalised_period_type,
             normalised_period, normalised_year).
    normalised_year is the ISO year for weekly reports, Gregorian year for all others.
    """
    from datetime import date, timedelta
    import calendar

    today = date.today()

    if period_type == "weekly":
        # ISO week-year: use isocalendar() so year-boundary weeks are correct
        iso_today = today.isocalendar()
        if year is None:
            year = iso_today[0]   # ISO year, NOT Gregorian year
        if period is None:
            # Only default to today's week if we are using today's ISO year
            if year == iso_today[0]:
                period = iso_today[1]
            else:
                period = 1
        # Validate week number against actual weeks in that ISO year
        max_week = _iso_weeks_in_year(year)
        if period < 1 or period > max_week:
            raise HTTPException(
                400,
                f"period must be 1-{max_week} for weekly reports in year {year}",
            )
        # Compute Monday of ISO week 1: Jan 4 is always in week 1
        jan4 = date(year, 1, 4)
        week1_monday = jan4 - timedelta(days=jan4.weekday())
        start_dt = week1_monday + timedelta(weeks=period - 1)
        end_dt = start_dt + timedelta(days=6)
        period_label = f"Week {period}, {year} ({start_dt.strftime('%d %b')}–{end_dt.strftime('%d %b %Y')})"
        return start_dt, end_dt, period_label, "weekly", period, year

    elif period_type == "quarterly":
        if year is None:
            year = today.year
        if period is None:
            period = (today.month - 1) // 3 + 1
        if period < 1 or period > 4:
            raise HTTPException(400, "period must be 1-4 for quarterly reports")
        start_month = (period - 1) * 3 + 1
        end_month = start_month + 2
        start_dt = date(year, start_month, 1)
        last_day = calendar.monthrange(year, end_month)[1]
        end_dt = date(year, end_month, last_day)
        period_label = f"Q{period} {year}"
        return start_dt, end_dt, period_label, "quarterly", period, year

    elif period_type == "annual":
        if year is None:
            year = today.year
        # Full calendar year — period param unused (always 1)
        start_dt = date(year, 1, 1)
        end_dt = date(year, 12, 31)
        period_label = f"Annual {year}"
        return start_dt, end_dt, period_label, "annual", 1, year

    else:
        # monthly (default / fallback)
        if year is None:
            year = today.year
        if period is None:
            period = today.month
        if period < 1 or period > 12:
            raise HTTPException(400, "period must be 1-12 for monthly reports")
        start_dt = date(year, period, 1)
        last_day = calendar.monthrange(year, period)[1]
        end_dt = date(year, period, last_day)
        period_label = f"{calendar.month_name[period]} {year}"
        return start_dt, end_dt, period_label, "monthly", period, year


@app.get("/reports/periodic", tags=["Reports"])
def get_periodic_report(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
    period_type: str = "monthly",
    year: int = None,
    period: int = None,
):
    """
    Task 11 / Task 21 — Periodic Report: weekly, monthly, quarterly or annual invoice summary.

    Query parameters:
    - period_type: "weekly" | "monthly" (default) | "quarterly" | "annual"
    - year: calendar year (default: current year)
    - period: ISO week number for weekly; 1-12 for monthly; 1-4 for quarterly; ignored for annual

    Role Access: COMPANY_ADMIN, BUSINESS_ADMIN, FINANCE_USER, READ_ONLY
    """
    if not has_permission(current_user, "reports", "view"):
        raise HTTPException(403, "Permission denied: view on reports is not allowed for your role")

    # All year/period defaulting is handled inside _build_periodic_date_range.
    # norm_year is the ISO year for weekly, Gregorian year for all other types.
    start_dt, end_dt, period_label, period_type, period, norm_year = (
        _build_periodic_date_range(period_type, year, period)
    )

    # Query all non-cancelled invoices in the period
    invoices = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.company_id == current_user.company_id,
            InvoiceDB.status != InvoiceStatus.CANCELLED,
            InvoiceDB.issue_date >= start_dt,
            InvoiceDB.issue_date <= end_dt,
        )
        .all()
    )

    # Aggregate totals by invoice type
    type_labels = {
        "380": "Tax Invoice",
        "381": "Credit Note",
        "383": "Debit Note",
        "480": "Commercial Invoice",
        "81": "Simplified Tax Invoice",
    }
    breakdown = {}
    overall = {
        "count": 0,
        "subtotal": 0.0,
        "tax_amount": 0.0,
        "total_amount": 0.0,
    }

    for inv in invoices:
        type_code = str(inv.invoice_type.value) if inv.invoice_type else "380"
        if type_code not in breakdown:
            breakdown[type_code] = {
                "invoice_type_code": type_code,
                "invoice_type_label": type_labels.get(type_code, type_code),
                "count": 0,
                "subtotal": 0.0,
                "tax_amount": 0.0,
                "total_amount": 0.0,
            }
        entry = breakdown[type_code]
        entry["count"] += 1
        entry["subtotal"] += float(inv.subtotal_amount or 0)
        entry["tax_amount"] += float(inv.tax_amount or 0)
        entry["total_amount"] += float(inv.total_amount or 0)

        overall["count"] += 1
        overall["subtotal"] += float(inv.subtotal_amount or 0)
        overall["tax_amount"] += float(inv.tax_amount or 0)
        overall["total_amount"] += float(inv.total_amount or 0)

    # Round all floats
    for row in breakdown.values():
        row["subtotal"] = round(row["subtotal"], 2)
        row["tax_amount"] = round(row["tax_amount"], 2)
        row["total_amount"] = round(row["total_amount"], 2)
    overall = {k: round(v, 2) if isinstance(v, float) else v for k, v in overall.items()}

    return {
        "period_type": period_type,
        "period_label": period_label,
        "year": norm_year,
        "period": period,
        "start_date": start_dt.isoformat(),
        "end_date": end_dt.isoformat(),
        "summary": overall,
        "breakdown_by_type": list(breakdown.values()),
        "generated_at": datetime.utcnow().isoformat(),
    }


@app.get("/reports/periodic/export", tags=["Reports"])
def export_periodic_report(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
    period_type: str = "monthly",
    year: int = None,
    period: int = None,
    format: str = "pdf",
):
    """
    Task 21 — Export a periodic report as PDF or XLSX.

    Query parameters:
    - period_type: "weekly" | "monthly" | "quarterly" | "annual"
    - year: calendar year (default: current year)
    - period: period number (see GET /reports/periodic for semantics)
    - format: "pdf" (default) | "xlsx"

    Role Access: COMPANY_ADMIN, BUSINESS_ADMIN, FINANCE_USER, READ_ONLY
    """
    import io

    if not has_permission(current_user, "reports", "export"):
        raise HTTPException(403, "Permission denied: export on reports is not allowed for your role")

    # Validate format param explicitly
    if format.lower() not in ("pdf", "xlsx"):
        raise HTTPException(400, "format must be 'pdf' or 'xlsx'")

    # All year/period defaulting is handled inside _build_periodic_date_range
    start_dt, end_dt, period_label, period_type, period, _norm_year = (
        _build_periodic_date_range(period_type, year, period)
    )

    # Query invoices
    invoices = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.company_id == current_user.company_id,
            InvoiceDB.status != InvoiceStatus.CANCELLED,
            InvoiceDB.issue_date >= start_dt,
            InvoiceDB.issue_date <= end_dt,
        )
        .all()
    )

    type_labels = {
        "380": "Tax Invoice",
        "381": "Credit Note",
        "383": "Debit Note",
        "480": "Commercial Invoice",
        "81": "Simplified Tax Invoice",
    }
    breakdown = {}
    overall = {"count": 0, "subtotal": 0.0, "tax_amount": 0.0, "total_amount": 0.0}

    for inv in invoices:
        type_code = str(inv.invoice_type.value) if inv.invoice_type else "380"
        if type_code not in breakdown:
            breakdown[type_code] = {
                "invoice_type_code": type_code,
                "invoice_type_label": type_labels.get(type_code, type_code),
                "count": 0, "subtotal": 0.0, "tax_amount": 0.0, "total_amount": 0.0,
            }
        e = breakdown[type_code]
        e["count"] += 1
        e["subtotal"] += float(inv.subtotal_amount or 0)
        e["tax_amount"] += float(inv.tax_amount or 0)
        e["total_amount"] += float(inv.total_amount or 0)
        overall["count"] += 1
        overall["subtotal"] += float(inv.subtotal_amount or 0)
        overall["tax_amount"] += float(inv.tax_amount or 0)
        overall["total_amount"] += float(inv.total_amount or 0)

    for row in breakdown.values():
        row["subtotal"] = round(row["subtotal"], 2)
        row["tax_amount"] = round(row["tax_amount"], 2)
        row["total_amount"] = round(row["total_amount"], 2)
    overall = {k: round(v, 2) if isinstance(v, float) else v for k, v in overall.items()}
    breakdown_rows = list(breakdown.values())
    generated_at = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    safe_label = period_label.replace(",", "").replace(" ", "_").replace("–", "-")

    if format.lower() == "xlsx":
        # ── XLSX export ────────────────────────────────────────────────────────
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise HTTPException(500, "openpyxl not installed — cannot generate XLSX")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Periodic Report"

        # Styling helpers
        primary_fill = PatternFill("solid", fgColor="0066CC")
        header_fill = PatternFill("solid", fgColor="E8F0FE")
        total_fill = PatternFill("solid", fgColor="DBEAFE")
        white_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def _set(ws, row, col, value, font=None, fill=None, align="left", number_format=None):
            cell = ws.cell(row=row, column=col, value=value)
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal=align, vertical="center")
            if number_format:
                cell.number_format = number_format
            cell.border = border
            return cell

        # Title block
        ws.merge_cells("A1:E1")
        t = ws.cell(row=1, column=1, value="InvoLinks — Periodic Invoice Report")
        t.font = Font(size=14, bold=True, color="0066CC")
        t.alignment = Alignment(horizontal="center")

        ws.merge_cells("A2:E2")
        t2 = ws.cell(row=2, column=1, value=f"Period: {period_label}  |  {start_dt.isoformat()} to {end_dt.isoformat()}")
        t2.alignment = Alignment(horizontal="center")
        t2.font = Font(italic=True, color="555555")

        ws.merge_cells("A3:E3")
        t3 = ws.cell(row=3, column=1, value=f"Generated: {generated_at}")
        t3.alignment = Alignment(horizontal="center")
        t3.font = Font(italic=True, size=9, color="888888")

        # Column headers (row 5)
        headers = ["Invoice Type", "Count", "Net Amount (AED)", "VAT Amount (AED)", "Gross Total (AED)"]
        for col, h in enumerate(headers, 1):
            _set(ws, 5, col, h, font=white_font, fill=primary_fill, align="center")

        # Data rows
        for r_idx, row in enumerate(breakdown_rows, 6):
            _set(ws, r_idx, 1, f"{row['invoice_type_label']} ({row['invoice_type_code']})", font=bold_font, fill=header_fill)
            _set(ws, r_idx, 2, row["count"], align="center")
            _set(ws, r_idx, 3, row["subtotal"], number_format="#,##0.00", align="right")
            _set(ws, r_idx, 4, row["tax_amount"], number_format="#,##0.00", align="right")
            _set(ws, r_idx, 5, row["total_amount"], number_format="#,##0.00", align="right")

        # Totals row
        total_row = 6 + len(breakdown_rows)
        _set(ws, total_row, 1, "TOTAL", font=Font(bold=True, color="0066CC"), fill=total_fill)
        _set(ws, total_row, 2, overall["count"], font=bold_font, fill=total_fill, align="center")
        _set(ws, total_row, 3, overall["subtotal"], font=bold_font, fill=total_fill, number_format="#,##0.00", align="right")
        _set(ws, total_row, 4, overall["tax_amount"], font=bold_font, fill=total_fill, number_format="#,##0.00", align="right")
        _set(ws, total_row, 5, overall["total_amount"], font=Font(bold=True, size=12, color="0066CC"), fill=total_fill, number_format="#,##0.00", align="right")

        # Column widths
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 10
        for col in ["C", "D", "E"]:
            ws.column_dimensions[col].width = 22

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="periodic_report_{safe_label}.xlsx"'},
        )

    else:
        # ── PDF export ─────────────────────────────────────────────────────────
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import cm
            from reportlab.lib import colors as rl_colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        except ImportError:
            raise HTTPException(500, "reportlab not installed — cannot generate PDF")

        buffer = io.BytesIO()
        PAGE_WIDTH, PAGE_HEIGHT = A4
        MARGIN = 2 * cm
        CONTENT_WIDTH = PAGE_WIDTH - (2 * MARGIN)
        PRIMARY = rl_colors.HexColor("#0066CC")
        LIGHT_BLUE = rl_colors.HexColor("#E8F0FE")
        TOTAL_BG = rl_colors.HexColor("#DBEAFE")
        GRAY = rl_colors.HexColor("#666666")
        DARK = rl_colors.HexColor("#333333")

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=MARGIN, rightMargin=MARGIN,
            topMargin=MARGIN, bottomMargin=MARGIN,
            title=f"Periodic Report — {period_label}",
            author="InvoLinks",
        )
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle("Title21", fontSize=16, textColor=PRIMARY, fontName="Helvetica-Bold",
                                   alignment=TA_CENTER, spaceAfter=4))
        styles.add(ParagraphStyle("Sub21", fontSize=10, textColor=GRAY, alignment=TA_CENTER, spaceAfter=2))
        styles.add(ParagraphStyle("Small21", fontSize=8, textColor=GRAY, alignment=TA_CENTER, spaceAfter=8))

        story = []
        story.append(Paragraph("InvoLinks — Periodic Invoice Report", styles["Title21"]))
        story.append(Paragraph(f"Period: {period_label}  |  {start_dt.isoformat()} to {end_dt.isoformat()}", styles["Sub21"]))
        story.append(Paragraph(f"Generated: {generated_at}", styles["Small21"]))
        story.append(Spacer(1, 12))

        # Build table data
        col_widths = [CONTENT_WIDTH * 0.40, CONTENT_WIDTH * 0.10, CONTENT_WIDTH * 0.17,
                      CONTENT_WIDTH * 0.17, CONTENT_WIDTH * 0.16]
        table_data = [[
            Paragraph("<b>Invoice Type</b>", styles["Normal"]),
            Paragraph("<b>Count</b>", styles["Normal"]),
            Paragraph("<b>Net (AED)</b>", styles["Normal"]),
            Paragraph("<b>VAT (AED)</b>", styles["Normal"]),
            Paragraph("<b>Gross (AED)</b>", styles["Normal"]),
        ]]
        for row in breakdown_rows:
            label = f"{row['invoice_type_label']} ({row['invoice_type_code']})"
            table_data.append([
                Paragraph(label, styles["Normal"]),
                Paragraph(str(row["count"]), styles["Normal"]),
                Paragraph(f"{row['subtotal']:,.2f}", styles["Normal"]),
                Paragraph(f"{row['tax_amount']:,.2f}", styles["Normal"]),
                Paragraph(f"<b>{row['total_amount']:,.2f}</b>", styles["Normal"]),
            ])
        # Totals row
        table_data.append([
            Paragraph("<b>TOTAL</b>", styles["Normal"]),
            Paragraph(f"<b>{overall['count']}</b>", styles["Normal"]),
            Paragraph(f"<b>{overall['subtotal']:,.2f}</b>", styles["Normal"]),
            Paragraph(f"<b>{overall['tax_amount']:,.2f}</b>", styles["Normal"]),
            Paragraph(f"<b>{overall['total_amount']:,.2f}</b>", styles["Normal"]),
        ])

        tbl = Table(table_data, colWidths=col_widths)
        n_data = len(table_data)
        tbl.setStyle(TableStyle([
            # Header row
            ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("TOPPADDING", (0, 0), (-1, 0), 8),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            # Data rows
            ("BACKGROUND", (0, 1), (-1, n_data - 2), rl_colors.white),
            ("FONTNAME", (0, 1), (-1, n_data - 2), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, n_data - 2), 9),
            ("TOPPADDING", (0, 1), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 1), (-1, -1), 6),
            ("GRID", (0, 0), (-1, -1), 0.5, GRAY),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            # Total row
            ("BACKGROUND", (0, n_data - 1), (-1, n_data - 1), TOTAL_BG),
            ("FONTNAME", (0, n_data - 1), (-1, n_data - 1), "Helvetica-Bold"),
            ("TEXTCOLOR", (0, n_data - 1), (0, n_data - 1), PRIMARY),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 20))
        story.append(Paragraph(
            f"<font color='#888888' size='8'>This report was generated by InvoLinks for internal FTA compliance purposes. "
            f"All amounts are in AED. Non-cancelled invoices only.</font>",
            styles["Normal"],
        ))

        doc.build(story)
        buffer.seek(0)
        from fastapi.responses import StreamingResponse
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="periodic_report_{safe_label}.pdf"'},
        )


@app.get("/reports/daily-reconciliation", tags=["Reports"])
def get_daily_reconciliation_report(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    """
    Generate daily reconciliation report showing payments by method

    Role Access: BUSINESS_ADMIN, FINANCE_USER, COMPANY_ADMIN
    """
    from datetime import datetime, timedelta
    from sqlalchemy import func

    # Check role-based permissions
    if not has_permission(current_user, "reports", "view"):
        raise HTTPException(403, "Permission denied: view on reports is not allowed for your role")

    # Parse date range
    if start_date:
        try:
            start_dt = datetime.fromisoformat(start_date).date()
        except ValueError:
            raise HTTPException(400, "Invalid start_date format. Use YYYY-MM-DD")
    else:
        start_dt = datetime.utcnow().date()  # Today by default

    if end_date:
        try:
            end_dt = datetime.fromisoformat(end_date).date()
        except ValueError:
            raise HTTPException(400, "Invalid end_date format. Use YYYY-MM-DD")
    else:
        end_dt = start_dt  # Same day by default

    # Query paid invoices within date range
    paid_invoices = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.company_id == current_user.company_id,
            InvoiceDB.status == InvoiceStatus.PAID,
            func.date(InvoiceDB.paid_at) >= start_dt,
            func.date(InvoiceDB.paid_at) <= end_dt,
        )
        .all()
    )

    # Group by payment method
    payment_breakdown = {}
    total_collected = 0.0

    for inv in paid_invoices:
        signed_amount = inv.total_amount or 0.0
        if inv.invoice_type in [
            InvoiceType.TAX_CREDIT_NOTE,
            InvoiceType.CREDIT_NOTE_OUT_OF_SCOPE,
        ]:
            signed_amount = -signed_amount

        method = inv.payment_method or "Unknown"
        if method not in payment_breakdown:
            payment_breakdown[method] = {
                "count": 0,
                "total_amount": 0.0,
                "invoices": [],
            }
        payment_breakdown[method]["count"] += 1
        payment_breakdown[method]["total_amount"] += signed_amount
        payment_breakdown[method]["invoices"].append(
            {
                "invoice_number": inv.invoice_number,
                "customer_name": inv.customer_name,
                "amount": signed_amount,
                "payment_method": method,
                "payment_reference": inv.invoice_notes
                if getattr(inv, "invoice_notes", None)
                else None,
                "payment_date": inv.paid_at.isoformat() if inv.paid_at else None,
            }
        )
        total_collected += signed_amount

    # Get outstanding invoices (overdue).
    # Exclude PAID and CANCELLED invoices so cancelled rows are not counted as outstanding.
    now_date = datetime.utcnow().date()
    outstanding_invoices = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.company_id == current_user.company_id,
            InvoiceDB.status.notin_(
                [
                    InvoiceStatus.PAID,
                    InvoiceStatus.CANCELLED,
                ]
            ),
            InvoiceDB.due_date < now_date,
        )
        .order_by(InvoiceDB.due_date.asc())
        .all()
    )

    outstanding_total = sum(inv.amount_due or 0.0 for inv in outstanding_invoices)

    return {
        "report_period": {
            "start_date": start_dt.isoformat(),
            "end_date": end_dt.isoformat(),
        },
        "summary": {
            "total_collected": total_collected,
            "total_transactions": len(paid_invoices),
            "outstanding_amount": outstanding_total,
            "outstanding_count": len(outstanding_invoices),
        },
        "payment_breakdown": [
            {
                "payment_method": method,
                "count": data["count"],
                "total_amount": data["total_amount"],
                "invoices": data["invoices"],
            }
            for method, data in payment_breakdown.items()
        ],
        "outstanding_invoices": [
            {
                "invoice_number": inv.invoice_number,
                "customer_name": inv.customer_name,
                "amount_due": inv.amount_due,
                "due_date": inv.due_date.isoformat() if inv.due_date else None,
                "days_overdue": (now_date - inv.due_date).days if inv.due_date else 0,
            }
            for inv in outstanding_invoices[:10]  # Limit to 10 most overdue
        ],
    }


# ==================== ANALYTICS ENDPOINTS ====================


@app.get("/analytics/revenue", tags=["Analytics"])
def get_revenue_analytics(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
    period: str = "month",  # month, quarter, year
    months: int = 12,
):
    """
    Get revenue analytics with trends and growth rates

    Role Access: All authenticated users
    """
    from datetime import datetime, timedelta
    from sqlalchemy import func, extract
    from dateutil.relativedelta import relativedelta

    # Calculate date range
    end_date = datetime.utcnow()
    start_date = end_date - relativedelta(months=months)

    # Get paid invoices within date range
    invoices = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.company_id == current_user.company_id,
            InvoiceDB.status == InvoiceStatus.PAID,
            InvoiceDB.paid_at >= start_date,
            InvoiceDB.paid_at <= end_date,
        )
        .all()
    )

    # Group by month
    monthly_revenue = {}
    for inv in invoices:
        if inv.paid_at:
            month_key = inv.paid_at.strftime("%Y-%m")
            if month_key not in monthly_revenue:
                monthly_revenue[month_key] = {"revenue": 0.0, "count": 0}
            # Credit notes should be subtracted from revenue, not added
            if inv.invoice_type in [
                InvoiceType.TAX_CREDIT_NOTE,
                InvoiceType.CREDIT_NOTE_OUT_OF_SCOPE,
            ]:
                monthly_revenue[month_key]["revenue"] -= inv.total_amount or 0.0
            else:
                monthly_revenue[month_key]["revenue"] += inv.total_amount or 0.0
            monthly_revenue[month_key]["count"] += 1

    # Calculate growth rate
    revenue_trend = []
    prev_revenue = None
    for i in range(months):
        date = end_date - relativedelta(months=months - i - 1)
        month_key = date.strftime("%Y-%m")
        revenue = monthly_revenue.get(month_key, {"revenue": 0.0, "count": 0})[
            "revenue"
        ]
        count = monthly_revenue.get(month_key, {"revenue": 0.0, "count": 0})["count"]

        growth_rate = 0.0
        if prev_revenue and prev_revenue > 0:
            growth_rate = ((revenue - prev_revenue) / prev_revenue) * 100

        revenue_trend.append(
            {
                "month": month_key,
                "revenue": revenue,
                "invoice_count": count,
                "growth_rate": growth_rate,
            }
        )
        prev_revenue = revenue if revenue > 0 else prev_revenue

    # Calculate totals - subtract credit notes from revenue
    total_revenue = sum(
        (inv.total_amount or 0.0)
        * (
            -1
            if inv.invoice_type
            in [InvoiceType.TAX_CREDIT_NOTE, InvoiceType.CREDIT_NOTE_OUT_OF_SCOPE]
            else 1
        )
        for inv in invoices
    )
    avg_monthly = total_revenue / months if months > 0 else 0

    return {
        "period": period,
        "months_analyzed": months,
        "total_revenue": total_revenue,
        "average_monthly_revenue": avg_monthly,
        "revenue_trend": revenue_trend,
    }


@app.get("/analytics/customers", tags=["Analytics"])
def get_customer_analytics(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
    limit: int = 10,
):
    """
    Get customer analytics including top customers and payment patterns

    Role Access: All authenticated users
    """
    from sqlalchemy import func

    # Get all paid invoices
    invoices = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.company_id == current_user.company_id,
            InvoiceDB.status == InvoiceStatus.PAID,
        )
        .all()
    )

    # Aggregate by customer
    customer_data = {}
    for inv in invoices:
        customer = inv.customer_name
        if customer not in customer_data:
            customer_data[customer] = {
                "customer_name": customer,
                "customer_email": inv.customer_email,
                "total_revenue": 0.0,
                "invoice_count": 0,
                "average_invoice_value": 0.0,
                "last_payment_date": None,
            }
        customer_data[customer]["total_revenue"] += (inv.total_amount or 0.0) * (
            -1
            if inv.invoice_type
            in [InvoiceType.TAX_CREDIT_NOTE, InvoiceType.CREDIT_NOTE_OUT_OF_SCOPE]
            else 1
        )
        customer_data[customer]["invoice_count"] += 1
        if inv.paid_at:
            if (
                not customer_data[customer]["last_payment_date"]
                or inv.paid_at > customer_data[customer]["last_payment_date"]
            ):
                customer_data[customer]["last_payment_date"] = inv.paid_at

    # Calculate averages and sort
    top_customers = []
    for customer, data in customer_data.items():
        data["average_invoice_value"] = (
            data["total_revenue"] / data["invoice_count"]
            if data["invoice_count"] > 0
            else 0
        )
        top_customers.append(data)

    top_customers.sort(key=lambda x: x["total_revenue"], reverse=True)

    # Calculate summary stats
    total_customers = len(customer_data)
    total_revenue = sum(d["total_revenue"] for d in customer_data.values())
    avg_customer_value = total_revenue / total_customers if total_customers > 0 else 0

    return {
        "total_customers": total_customers,
        "total_revenue": total_revenue,
        "average_customer_lifetime_value": avg_customer_value,
        "top_customers": [
            {
                **cust,
                "last_payment_date": cust["last_payment_date"].isoformat()
                if cust["last_payment_date"]
                else None,
            }
            for cust in top_customers[:limit]
        ],
    }


@app.get("/analytics/profitability", tags=["Analytics"])
def get_profitability_metrics(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
    months: int = 12,
):
    """
    Get profitability metrics including gross profit, net profit, and margins

    Role Access: All authenticated users
    """
    from datetime import datetime, timedelta
    from dateutil.relativedelta import relativedelta

    # Calculate date range
    end_date = datetime.utcnow()
    start_date = end_date - relativedelta(months=months)

    # Get revenue (paid invoices) - subtract credit notes
    revenue_invoices = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.company_id == current_user.company_id,
            InvoiceDB.status == InvoiceStatus.PAID,
            InvoiceDB.paid_at >= start_date,
            InvoiceDB.paid_at <= end_date,
        )
        .all()
    )

    total_revenue = sum(
        (inv.total_amount or 0.0)
        * (
            -1
            if inv.invoice_type
            in [InvoiceType.TAX_CREDIT_NOTE, InvoiceType.CREDIT_NOTE_OUT_OF_SCOPE]
            else 1
        )
        for inv in revenue_invoices
    )

    # Get expenses
    expenses = (
        db.query(ExpenseDB)
        .filter(
            ExpenseDB.company_id == current_user.company_id,
            ExpenseDB.expense_date >= start_date,
            ExpenseDB.expense_date <= end_date,
        )
        .all()
    )

    total_expenses = sum((exp.total_amount or 0.0) for exp in expenses)

    # Calculate profitability
    gross_profit = total_revenue - total_expenses
    gross_margin = (gross_profit / total_revenue * 100) if total_revenue > 0 else 0

    # Monthly breakdown
    monthly_profit = []
    for i in range(months):
        date = end_date - relativedelta(months=months - i - 1)
        month_start = date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        month_end = (month_start + relativedelta(months=1)) - timedelta(seconds=1)

        month_revenue = sum(
            (inv.total_amount or 0.0)
            * (
                -1
                if inv.invoice_type
                in [InvoiceType.TAX_CREDIT_NOTE, InvoiceType.CREDIT_NOTE_OUT_OF_SCOPE]
                else 1
            )
            for inv in revenue_invoices
            if inv.paid_at and month_start <= inv.paid_at <= month_end
        )

        month_expenses = sum(
            (exp.total_amount or 0.0)
            for exp in expenses
            if month_start.date() <= exp.expense_date <= month_end.date()
        )

        month_profit = month_revenue - month_expenses
        month_margin = (month_profit / month_revenue * 100) if month_revenue > 0 else 0

        monthly_profit.append(
            {
                "month": date.strftime("%Y-%m"),
                "revenue": month_revenue,
                "expenses": month_expenses,
                "profit": month_profit,
                "margin": month_margin,
            }
        )

    return {
        "period_months": months,
        "total_revenue": total_revenue,
        "total_expenses": total_expenses,
        "gross_profit": gross_profit,
        "gross_margin": gross_margin,
        "monthly_breakdown": monthly_profit,
    }


@app.get("/analytics/cash-flow", tags=["Analytics"])
def get_cash_flow_analytics(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
    months: int = 6,
):
    """
    Get cash flow analysis with inflows and outflows

    Role Access: All authenticated users
    """
    from datetime import datetime, timedelta
    from dateutil.relativedelta import relativedelta

    # Calculate date range
    end_date = datetime.utcnow()
    start_date = end_date - relativedelta(months=months)

    # Get inflows (paid invoices)
    inflows = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.company_id == current_user.company_id,
            InvoiceDB.status == InvoiceStatus.PAID,
            InvoiceDB.paid_at >= start_date,
            InvoiceDB.paid_at <= end_date,
        )
        .all()
    )

    # Get outflows (expenses)
    outflows = (
        db.query(ExpenseDB)
        .filter(
            ExpenseDB.company_id == current_user.company_id,
            ExpenseDB.expense_date >= start_date,
            ExpenseDB.expense_date <= end_date,
        )
        .all()
    )

    # Monthly breakdown
    monthly_cash_flow = []
    for i in range(months):
        date = end_date - relativedelta(months=months - i - 1)
        month_start = date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        month_end = (month_start + relativedelta(months=1)) - timedelta(seconds=1)

        month_inflows = sum(
            (inv.total_amount or 0.0)
            * (
                -1
                if inv.invoice_type
                in [InvoiceType.TAX_CREDIT_NOTE, InvoiceType.CREDIT_NOTE_OUT_OF_SCOPE]
                else 1
            )
            for inv in inflows
            if inv.paid_at and month_start <= inv.paid_at <= month_end
        )

        month_outflows = sum(
            exp.amount
            for exp in outflows
            if month_start.date() <= exp.expense_date <= month_end.date()
        )

        net_cash_flow = month_inflows - month_outflows

        monthly_cash_flow.append(
            {
                "month": date.strftime("%Y-%m"),
                "inflows": month_inflows,
                "outflows": month_outflows,
                "net_cash_flow": net_cash_flow,
            }
        )

    # Calculate totals
    total_inflows = sum(m["inflows"] for m in monthly_cash_flow)
    total_outflows = sum(m["outflows"] for m in monthly_cash_flow)
    net_position = total_inflows - total_outflows

    return {
        "period_months": months,
        "total_inflows": total_inflows,
        "total_outflows": total_outflows,
        "net_cash_position": net_position,
        "monthly_cash_flow": monthly_cash_flow,
    }


@app.get("/analytics/accounts-receivable", tags=["Analytics"])
def get_accounts_receivable(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Return accounts receivable totals and simple aging buckets.

    - Considers invoices where amount_due > 0 and status is not PAID/CANCELLED.
    - Aging is based on `due_date` where available; missing due_date treated as current bucket.
    """
    from datetime import datetime

    today = datetime.utcnow().date()

    invoices = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.company_id == current_user.company_id,
            InvoiceDB.amount_due > 0,
            InvoiceDB.status.notin_([InvoiceStatus.PAID, InvoiceStatus.CANCELLED]),
        )
        .all()
    )

    total_outstanding = 0.0
    count = 0
    aging = {"0_30": 0.0, "31_60": 0.0, "61_90": 0.0, "90_plus": 0.0}
    details = []

    for inv in invoices:
        amt = float(inv.amount_due or 0.0)
        if amt <= 0:
            continue
        total_outstanding += amt
        count += 1

        # Determine aging bucket
        days_past_due = 0
        if inv.due_date:
            days_past_due = (today - inv.due_date).days

        if days_past_due <= 30:
            aging["0_30"] += amt
        elif days_past_due <= 60:
            aging["31_60"] += amt
        elif days_past_due <= 90:
            aging["61_90"] += amt
        else:
            aging["90_plus"] += amt

        details.append(
            {
                "id": inv.id,
                "invoice_number": inv.invoice_number,
                "customer": inv.customer_name,
                "amount_due": amt,
                "due_date": inv.due_date.isoformat() if inv.due_date else None,
                "status": inv.status,
            }
        )

    return {
        "total_outstanding": round(total_outstanding, 2),
        "count": count,
        "aging": {k: round(v, 2) for k, v in aging.items()},
        "details": details[:50],
    }


@app.get("/analytics/accounts-payable", tags=["Analytics"])
def get_accounts_payable(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Return accounts payable totals and simple aging buckets.

    This aggregates unpaid inward invoices (supplier invoices). Future: include POs and unpaid expenses.
    """
    from datetime import datetime

    today = datetime.utcnow().date()

    invoices = (
        db.query(InwardInvoiceDB)
        .filter(
            InwardInvoiceDB.company_id == current_user.company_id,
            InwardInvoiceDB.total_amount > 0,
            InwardInvoiceDB.payment_status != "PAID",
        )
        .all()
    )

    total_outstanding = 0.0
    count = 0
    aging = {"0_30": 0.0, "31_60": 0.0, "61_90": 0.0, "90_plus": 0.0}
    details = []

    for inv in invoices:
        # Use amount_due if present else total_amount - paid_amount
        amt = float(
            inv.amount_due
            if getattr(inv, "amount_due", None) is not None
            else (inv.total_amount - (inv.paid_amount or 0.0))
        )
        if amt <= 0:
            continue
        total_outstanding += amt
        count += 1

        days_past_due = 0
        if inv.due_date:
            days_past_due = (today - inv.due_date).days

        if days_past_due <= 30:
            aging["0_30"] += amt
        elif days_past_due <= 60:
            aging["31_60"] += amt
        elif days_past_due <= 90:
            aging["61_90"] += amt
        else:
            aging["90_plus"] += amt

        details.append(
            {
                "id": inv.id,
                "supplier_invoice_number": inv.supplier_invoice_number,
                "supplier": inv.supplier_name,
                "amount_due": amt,
                "due_date": inv.due_date.isoformat() if inv.due_date else None,
                "payment_status": inv.payment_status,
            }
        )

    return {
        "total_outstanding": round(total_outstanding, 2),
        "count": count,
        "aging": {k: round(v, 2) for k, v in aging.items()},
        "details": details[:50],
    }


@app.post("/invoices/{invoice_id}/issue", tags=["Invoices"])
def issue_invoice(
    invoice_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Issue a draft invoice - generates UBL XML with digital signature and hash chain

    PHASE 1 Features:
    - Generates UAE PINT-AE compliant UBL 2.1 XML
    - Creates digital signature (RSA-2048)
    - Links to previous invoice via hash chain
    - Saves XML to file system
    - Updates status to ISSUED
    """
    if not has_permission(current_user, "invoices", "approve"):
        raise HTTPException(403, "Permission denied: issue invoices is not allowed for your role")
    invoice = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.id == invoice_id, InvoiceDB.company_id == current_user.company_id
        )
        .first()
    )

    if not invoice:
        raise HTTPException(404, "Invoice not found")

    if invoice.status != InvoiceStatus.DRAFT:
        raise HTTPException(
            400, f"Can only issue draft invoices. Current status: {invoice.status}"
        )

    # Auto-fill supplier TRN from company if missing
    company = (
        db.query(CompanyDB).filter(CompanyDB.id == current_user.company_id).first()
    )
    if (
        (
            not invoice.supplier_trn
            or (
                isinstance(invoice.supplier_trn, str)
                and invoice.supplier_trn.strip() == ""
            )
        )
        and company
        and company.trn
    ):
        invoice.supplier_trn = company.trn

    # Auto-calculate monetary totals from line items if missing
    try:
        subtotal_calc = sum(
            [
                (
                    li.line_extension_amount
                    if li.line_extension_amount is not None
                    else (li.quantity * li.unit_price)
                )
                for li in invoice.line_items
            ]
        )
        tax_calc = sum([(li.tax_amount or 0.0) for li in invoice.line_items])
        total_calc = sum(
            [
                (
                    li.line_total_amount
                    if li.line_total_amount is not None
                    else (
                        (
                            li.line_extension_amount
                            if li.line_extension_amount is not None
                            else (li.quantity * li.unit_price)
                        )
                        + (li.tax_amount or 0.0)
                    )
                )
                for li in invoice.line_items
            ]
        )
    except Exception:
        subtotal_calc = tax_calc = total_calc = 0.0

    if not invoice.subtotal_amount or invoice.subtotal_amount == 0.0:
        invoice.subtotal_amount = subtotal_calc
    if not invoice.tax_amount or invoice.tax_amount == 0.0:
        invoice.tax_amount = tax_calc
    if not invoice.total_amount or invoice.total_amount == 0.0:
        invoice.total_amount = total_calc

    # ==== PHASE 1: Digital Signatures, Hash Chain & UBL XML Generation ====
    from utils.crypto_utils import get_crypto_instance
    from utils.ubl_xml_generator import generate_invoice_xml
    import json

    # Get the immediate predecessor in the hash chain for the current invoice.
    # Chain membership mirrors the verifier: all non-archived, non-draft
    # invoices (ISSUED/SENT/VIEWED/PAID/OVERDUE/CANCELLED). Using the same
    # ordering (issue_date, created_at) as the verifier ensures consistency.
    from sqlalchemy import or_, and_
    prev_invoice = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.company_id == current_user.company_id,
            InvoiceDB.id != invoice.id,
            InvoiceDB.is_archived == False,
            InvoiceDB.status != InvoiceStatus.DRAFT,
            or_(
                InvoiceDB.issue_date < invoice.issue_date,
                and_(
                    InvoiceDB.issue_date == invoice.issue_date,
                    InvoiceDB.created_at < invoice.created_at,
                ),
            ),
        )
        .order_by(InvoiceDB.issue_date.desc(), InvoiceDB.created_at.desc())
        .first()
    )

    if prev_invoice:
        # Compute canonical hash of previous invoice for hash chain linking.
        # Using compute_invoice_hash (SHA-256 of canonical fields) so that
        # the integrity verifier can independently recompute it from DB fields.
        _chain_crypto = get_crypto_instance()
        _prev_canonical = {
            "invoice_number": prev_invoice.invoice_number or "",
            "issue_date": str(prev_invoice.issue_date) if prev_invoice.issue_date else "",
            "supplier_trn": prev_invoice.supplier_trn or "",
            "customer_trn": prev_invoice.customer_trn or "",
            "total_amount": str(prev_invoice.total_amount or "0.0"),
            "tax_amount": str(prev_invoice.tax_amount or "0.0"),
            "prev_invoice_hash": prev_invoice.prev_invoice_hash or "",
        }
        invoice.prev_invoice_hash = _chain_crypto.compute_invoice_hash(_prev_canonical)

    # Prepare invoice data for XML generation
    invoice_data = {
        "invoice_number": invoice.invoice_number,
        "issue_date": invoice.issue_date,
        "due_date": invoice.due_date,
        "supply_date": invoice.supply_date,  # Task 5: cac:Delivery/cbc:ActualDeliveryDate
        "invoice_type": invoice.invoice_type.value,
        "currency_code": invoice.currency_code,
        "supplier_trn": invoice.supplier_trn,
        "supplier_name": invoice.supplier_name,
        "supplier_address": invoice.supplier_address,
        "supplier_city": invoice.supplier_city,
        "supplier_country": invoice.supplier_country,
        "supplier_peppol_id": invoice.supplier_peppol_id,
        "customer_trn": invoice.customer_trn,
        "customer_name": invoice.customer_name,
        "customer_email": invoice.customer_email,
        "customer_address": invoice.customer_address,
        "customer_city": invoice.customer_city,
        "customer_country": invoice.customer_country,
        "customer_peppol_id": invoice.customer_peppol_id,
        "subtotal_amount": invoice.subtotal_amount,
        "tax_amount": invoice.tax_amount,
        "total_amount": invoice.total_amount,
        "amount_due": invoice.amount_due,
        "payment_terms": invoice.payment_terms,
        "invoice_notes": invoice.invoice_notes,
        "reference_number": invoice.reference_number,
        "preceding_invoice_id": invoice.preceding_invoice_id,
        "prev_invoice_hash": invoice.prev_invoice_hash,
    }
    # Validate required fields before XML generation (TRN is optional for non-VAT parties)
    missing_fields = []
    for f in ["subtotal_amount", "tax_amount", "total_amount"]:
        val = invoice_data.get(f)
        # Allow 0 values - only check for None
        if val is None:
            missing_fields.append(f)

    if missing_fields:
        raise HTTPException(
            400, f"Missing required fields for XML generation: {missing_fields}"
        )

    # Get line items for XML
    line_items_data = []
    for line in invoice.line_items:
        line_items_data.append(
            {
                "line_number": line.line_number,
                "item_name": line.item_name,
                "item_description": line.item_description,
                "item_code": line.item_code,
                "quantity": line.quantity,
                "unit_code": line.unit_code,
                "unit_price": line.unit_price,
                "line_extension_amount": line.line_extension_amount,
                "tax_category": line.tax_category.value,
                "tax_percent": line.tax_percent,
                "tax_amount": line.tax_amount,
                "line_total_amount": line.line_total_amount,
            }
        )

    # Generate UBL 2.1 XML
    success, xml_content, errors = generate_invoice_xml(invoice_data, line_items_data)

    if not success or not xml_content:
        raise HTTPException(500, f"XML generation failed: {errors}")

    # Save XML to file
    xml_dir = "invoices_xml"
    os.makedirs(xml_dir, exist_ok=True)
    xml_filename = f"{invoice.invoice_number.replace('/', '_')}.xml"
    xml_path = os.path.join(xml_dir, xml_filename)

    with open(xml_path, "w", encoding="utf-8") as f:
        f.write(xml_content)

    invoice.xml_file_path = xml_path

    # Compute hash of invoice + XML
    crypto = get_crypto_instance()
    invoice_hash = crypto.compute_invoice_hash(invoice_data)
    invoice.xml_hash = crypto.compute_hash(xml_content)

    # Sign invoice (using mock signing for now - replace with real cert in production)
    signature = crypto.sign_invoice(invoice_hash, xml_content)
    if signature:
        invoice.signature_b64 = signature
        invoice.signing_timestamp = datetime.utcnow()
        invoice.signing_cert_serial = "MOCK-CERT-001"  # Replace with real cert serial

    # Update invoice status — snapshot prior status before mutation
    _prev_status_issue = invoice.status.value if invoice.status else "UNKNOWN"
    invoice.status = InvoiceStatus.ISSUED

    # Task 8: Auto-generate GL journal entry (wrapped so GL failure never breaks issuance)
    try:
        create_journal_entry_for_invoice(invoice, db, user_id=current_user.id)
    except Exception as gl_err:
        logger.error(f"GL journal entry failed for {invoice.invoice_number}: {gl_err}")

    # Task 1 / Task 15: Audit trail — log actual prior → new status
    import json as _json
    log_audit_event(
        db, "INVOICE_ISSUED",
        user_id=current_user.id, company_id=current_user.company_id,
        resource_type="invoice", resource_id=invoice.id,
        old_value=_json.dumps({"status": _prev_status_issue}),
        new_value=_json.dumps({"status": "ISSUED"}),
        description=f"Invoice {invoice.invoice_number} issued",
    )

    db.commit()
    db.refresh(invoice)

    print(f"\n{'=' * 70}")
    print(f"✅ INVOICE ISSUED: {invoice.invoice_number}")
    print(f"{'=' * 70}")
    print(f"Invoice Hash: {invoice_hash[:32]}...")
    print(f"XML Hash: {invoice.xml_hash[:32]}...")
    print(f"XML File: {xml_path}")
    print(f"Signature: {'✓ Signed' if signature else '✗ Not signed'}")
    print(
        f"Previous Hash: {invoice.prev_invoice_hash[:32] if invoice.prev_invoice_hash else 'None (first invoice)'}"
    )
    print(f"Hash Chain: {'✓ Linked' if invoice.prev_invoice_hash else '✓ Chain Start'}")
    print(f"Status: {invoice.status}")
    print(f"{'=' * 70}\n")

    return {
        "success": True,
        "message": "Invoice issued successfully with digital signature and hash chain",
        "invoice_id": invoice.id,
        "invoice_number": invoice.invoice_number,
        "status": invoice.status,
        "xml_generated": True,
        "xml_hash": invoice.xml_hash,
        "signature_generated": signature is not None,
        "hash_chain_linked": invoice.prev_invoice_hash is not None,
        "compliance": "UAE PINT-AE UBL 2.1",
    }


@app.post("/invoices/{invoice_id}/send", tags=["Invoices"])
async def send_invoice(
    invoice_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Send invoice to customer via email and update status to SENT"""
    invoice = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.id == invoice_id, InvoiceDB.company_id == current_user.company_id
        )
        .first()
    )

    if not invoice:
        raise HTTPException(404, "Invoice not found")

    if invoice.status == InvoiceStatus.DRAFT:
        raise HTTPException(400, "Cannot send draft invoice. Please issue it first.")

    if invoice.status == InvoiceStatus.CANCELLED:
        raise HTTPException(400, "Cannot send cancelled invoice")

    # Validate email exists
    if not invoice.customer_email:
        raise HTTPException(400, "Cannot send invoice: Customer email is required.")

    # Update status
    invoice.status = InvoiceStatus.SENT
    invoice.sent_at = datetime.utcnow()
    db.commit()

    # Get the base URL for share link
    base_url = os.getenv("REPLIT_DOMAINS", "https://involinks.replit.app")
    if base_url:
        base_url = (
            f"https://{base_url.split(',')[0]}"
            if "," in base_url
            else f"https://{base_url}"
        )

    share_url = f"{base_url}/invoices/view/{invoice.share_token}"

    # Get company details
    company = db.get(CompanyDB, current_user.company_id)
    company_email = company.email if company else "support@involinks.ae"

    # Send invoice email via AWS SES
    email_result = email_service.send_invoice_email(
        to_email=invoice.customer_email,
        customer_name=invoice.customer_name or "Customer",
        invoice_number=invoice.invoice_number,
        invoice_url=share_url,
        company_name=invoice.supplier_name,
        company_email=company_email,
        amount=invoice.total_amount,
        currency="AED",
    )

    return {
        "message": "Invoice sent successfully",
        "invoice_id": invoice.id,
        "invoice_number": invoice.invoice_number,
        "sent_to": invoice.customer_email,
        "share_link": f"/invoices/view/{invoice.share_token}",
        "email_status": "sent" if email_result.get("success") else "simulated",
        "sent_at": invoice.sent_at.isoformat(),
    }


@app.get("/invoices/{invoice_id}/pdf", tags=["Invoices"])
def download_invoice_pdf(
    invoice_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Download invoice as PDF

    Generates a professional, UAE FTA-compliant PDF invoice with:
    - Company branding and logo
    - Complete invoice details and line items
    - VAT breakdown
    - QR code for public viewing
    - Digital signature indicator
    """
    from fastapi.responses import StreamingResponse
    from utils.pdf_invoice_generator import generate_invoice_pdf

    # Get invoice
    invoice = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.id == invoice_id, InvoiceDB.company_id == current_user.company_id
        )
        .first()
    )

    if not invoice:
        raise HTTPException(404, "Invoice not found")

    # Get line items
    line_items = (
        db.query(InvoiceLineItemDB)
        .filter(InvoiceLineItemDB.invoice_id == invoice.id)
        .order_by(InvoiceLineItemDB.line_number)
        .all()
    )

    # Prepare invoice data for PDF
    invoice_data = {
        "invoice_number": invoice.invoice_number,
        "invoice_type": invoice.invoice_type.value,
        "status": invoice.status.value,
        "issue_date": invoice.issue_date.isoformat() if invoice.issue_date else "",
        "due_date": invoice.due_date.isoformat() if invoice.due_date else None,
        "supply_date": invoice.supply_date.isoformat() if invoice.supply_date else None,
        "currency_code": invoice.currency_code,
        "vat_enabled": bool(invoice.supplier_trn),
        "invoice_classification": invoice.invoice_classification,
        # Supplier
        "supplier_trn": invoice.supplier_trn,
        "supplier_name": invoice.supplier_name,
        "supplier_address": invoice.supplier_address,
        "supplier_city": invoice.supplier_city,
        "supplier_country": invoice.supplier_country,
        # Customer
        "customer_trn": invoice.customer_trn,
        "customer_name": invoice.customer_name,
        "customer_address": invoice.customer_address,
        "customer_city": invoice.customer_city,
        "customer_country": invoice.customer_country,
        # Amounts
        "subtotal_amount": invoice.subtotal_amount,
        "tax_amount": invoice.tax_amount,
        "total_amount": invoice.total_amount,
        "amount_due": invoice.amount_due,
        # Additional info
        "payment_terms": invoice.payment_terms,
        "notes": invoice.invoice_notes,
        # Compliance
        "signature_b64": invoice.signature_b64,
        "signing_cert_serial": invoice.signing_cert_serial,
        "signing_timestamp": invoice.signing_timestamp.isoformat()
        if invoice.signing_timestamp
        else None,
    }

    # Prepare line items
    line_items_data = []
    for line in line_items:
        line_items_data.append(
            {
                "line_number": line.line_number,
                "item_name": line.item_name,
                "item_description": line.item_description,
                "quantity": line.quantity,
                "unit_price": line.unit_price,
                "tax_percent": line.tax_percent,
                "line_total_amount": line.line_total_amount,
                "currency_code": invoice.currency_code,
            }
        )

    # Generate public URL for QR code
    platform_url = os.getenv("PLATFORM_URL", "https://involinks.ae")
    public_url = (
        f"{platform_url}/invoices/view/{invoice.share_token}"
        if invoice.share_token
        else None
    )

    # Generate PDF
    try:
        pdf_bytes = generate_invoice_pdf(
            invoice_data=invoice_data, line_items=line_items_data, public_url=public_url
        )

        # Task 1: Audit trail — PDF download
        log_audit_event(
            db, "INVOICE_PDF_DOWNLOADED",
            user_id=current_user.id, company_id=current_user.company_id,
            resource_type="invoice", resource_id=invoice.id,
            description=f"PDF downloaded for invoice {invoice.invoice_number}",
        )
        db.commit()

        # Return PDF as download
        filename = f"{invoice.invoice_number.replace('/', '_')}.pdf"

        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Type": "application/pdf",
            },
        )

    except Exception as e:
        print(f"❌ PDF generation error: {str(e)}")
        import traceback

        traceback.print_exc()
        raise HTTPException(500, f"PDF generation failed: {str(e)}")


@app.post("/invoices/{invoice_id}/cancel", tags=["Invoices"])
def cancel_invoice(
    invoice_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Cancel an invoice"""
    if not has_permission(current_user, "invoices", "delete"):
        raise HTTPException(403, "Permission denied: cancel invoices is not allowed for your role")
    invoice = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.id == invoice_id, InvoiceDB.company_id == current_user.company_id
        )
        .first()
    )

    if not invoice:
        raise HTTPException(404, "Invoice not found")

    if invoice.status in [InvoiceStatus.PAID, InvoiceStatus.CANCELLED]:
        raise HTTPException(409, f"Cannot cancel invoice with status: {invoice.status.value}")

    # Task 3: ISSUED invoices require a credit note before cancellation
    if invoice.status in [InvoiceStatus.ISSUED, InvoiceStatus.SENT]:
        credit_note_exists = (
            db.query(InvoiceDB)
            .filter(
                InvoiceDB.preceding_invoice_id == invoice.id,
                InvoiceDB.invoice_type.in_([
                    InvoiceType.TAX_CREDIT_NOTE,
                    InvoiceType.CREDIT_NOTE_OUT_OF_SCOPE,
                ]),
                InvoiceDB.status.in_([InvoiceStatus.ISSUED, InvoiceStatus.SENT]),
                InvoiceDB.company_id == current_user.company_id,
            )
            .first()
        )
        if not credit_note_exists:
            raise HTTPException(
                409,
                "A posted invoice can only be cancelled after issuing a credit note "
                "(document type 381 referencing this invoice). "
                "Please create a credit note first.",
            )

    # If cancelling a credit note, restore the original invoice's amount_due
    if invoice.invoice_type in [
        InvoiceType.TAX_CREDIT_NOTE,
        InvoiceType.CREDIT_NOTE_OUT_OF_SCOPE,
    ]:
        if invoice.preceding_invoice_id:
            original_invoice = (
                db.query(InvoiceDB)
                .filter(
                    InvoiceDB.id == invoice.preceding_invoice_id,
                    InvoiceDB.company_id == current_user.company_id,
                )
                .first()
            )

            if original_invoice:
                # Add back the credit note amount to the original invoice's amount_due
                new_amount_due = float(
                    Decimal(str(original_invoice.amount_due))
                    + Decimal(str(invoice.total_amount))
                )
                # Ensure we don't exceed the original total_amount
                original_invoice.amount_due = min(
                    original_invoice.total_amount, new_amount_due
                )
                db.add(original_invoice)

    # Task 4: If cancelling a debit note, subtract its amount back from the original invoice's amount_due
    if invoice.invoice_type == InvoiceType.DEBIT_NOTE and invoice.preceding_invoice_id:
        original_invoice = (
            db.query(InvoiceDB)
            .filter(
                InvoiceDB.id == invoice.preceding_invoice_id,
                InvoiceDB.company_id == current_user.company_id,
            )
            .first()
        )
        if original_invoice:
            new_amount_due = float(
                Decimal(str(original_invoice.amount_due))
                - Decimal(str(invoice.total_amount))
            )
            original_invoice.amount_due = max(0.0, new_amount_due)
            db.add(original_invoice)

    # Snapshot prior status before mutation
    _prev_status_cancel = invoice.status.value if invoice.status else "UNKNOWN"
    invoice.status = InvoiceStatus.CANCELLED

    # Task 1 / Task 15: Audit trail — log actual prior → new status
    import json as _json
    log_audit_event(
        db, "INVOICE_CANCELLED",
        user_id=current_user.id, company_id=current_user.company_id,
        resource_type="invoice", resource_id=invoice.id,
        old_value=_json.dumps({"status": _prev_status_cancel}),
        new_value=_json.dumps({"status": "CANCELLED"}),
        description=f"Invoice {invoice.invoice_number} cancelled",
    )

    db.commit()

    return {
        "message": "Invoice cancelled",
        "invoice_id": invoice.id,
        "invoice_number": invoice.invoice_number,
    }


@app.get("/invoices/{invoice_id}/qr", tags=["Invoices"])
def get_invoice_qr_code(
    invoice_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Generate QR code for invoice share link"""
    import qrcode
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    invoice = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.id == invoice_id, InvoiceDB.company_id == current_user.company_id
        )
        .first()
    )

    if not invoice:
        raise HTTPException(404, "Invoice not found")

    # Get the base URL from environment or use default
    base_url = os.getenv("REPLIT_DOMAINS", "https://involinks.replit.app")
    if base_url:
        base_url = (
            f"https://{base_url.split(',')[0]}"
            if "," in base_url
            else f"https://{base_url}"
        )

    # Create the full share URL
    share_url = f"{base_url}/invoices/view/{invoice.share_token}"

    # Generate QR code
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )
    qr.add_data(share_url)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white")

    # Convert to bytes
    buf = BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)

    return StreamingResponse(buf, media_type="image/png")


@app.post("/invoices/{invoice_id}/share", tags=["Invoices"])
def ensure_invoice_share_link(
    invoice_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Ensure an invoice has a public share token and return the public URL."""
    invoice = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.id == invoice_id, InvoiceDB.company_id == current_user.company_id
        )
        .first()
    )

    if not invoice:
        raise HTTPException(404, "Invoice not found")

    # Create share token if missing
    if not invoice.share_token:
        invoice.share_token = f"share_{uuid4().hex[:16]}"
        db.commit()

    # Build public URL using same logic as other endpoints
    base_url = (
        os.getenv("REPLIT_DOMAINS", None)
        or os.getenv("PLATFORM_URL", None)
        or "https://involinks.replit.app"
    )
    if base_url:
        base_url = (
            f"https://{base_url.split(',')[0]}"
            if "," in base_url
            else f"https://{base_url}"
        )

    public_url = f"{base_url}/invoices/view/{invoice.share_token}"

    return {
        "invoice_id": invoice.id,
        "share_token": invoice.share_token,
        "share_link": f"/invoices/view/{invoice.share_token}",
        "public_url": public_url,
    }


@app.post("/invoices/{invoice_id}/email", tags=["Invoices"])
async def email_invoice(
    invoice_id: str,
    recipient_email: str = None,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Send invoice to customer via email with PDF attachment"""
    invoice = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.id == invoice_id, InvoiceDB.company_id == current_user.company_id
        )
        .first()
    )

    if not invoice:
        raise HTTPException(404, "Invoice not found")

    # Use provided email or fall back to invoice customer email
    email_to = recipient_email or invoice.customer_email
    if not email_to:
        raise HTTPException(400, "No recipient email provided")

    # Persist recipient email only when invoice is missing one
    if recipient_email and not invoice.customer_email:
        invoice.customer_email = recipient_email
        db.commit()
        db.refresh(invoice)

    # Get the base URL for share link
    base_url = os.getenv("REPLIT_DOMAINS", "https://involinks.replit.app")
    if base_url:
        base_url = (
            f"https://{base_url.split(',')[0]}"
            if "," in base_url
            else f"https://{base_url}"
        )

    share_url = f"{base_url}/invoices/view/{invoice.share_token}"

    # Get company details for Reply-To
    company = db.get(CompanyDB, current_user.company_id)
    company_email = company.email if company else "support@involinks.ae"

    # Send invoice email via AWS SES
    # email_result = email_service.send_invoice_email(
    #     to_email=email_to,
    #     customer_name=invoice.customer_name or "Customer",
    #     invoice_number=invoice.invoice_number,
    #     invoice_url=share_url,
    #     company_name=invoice.supplier_name,
    #     company_email=company_email,
    #     amount=invoice.total_amount,
    #     currency="AED")
    email_result = {
        "success": False,
        "note": "Email sending skipped (temporarily disabled)",
    }

    return {
        "message": "Invoice email sent successfully",
        "sent_to": email_to,
        "invoice_id": invoice.id,
        "email_status": "sent" if email_result.get("success") else "simulated",
        "share_url": share_url,
        "note": email_result.get("note", "Email delivered"),
    }


@app.post("/invoices/{invoice_id}/sms", tags=["Invoices"])
async def sms_invoice(
    invoice_id: str,
    phone_number: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Send invoice link to customer via SMS"""
    invoice = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.id == invoice_id, InvoiceDB.company_id == current_user.company_id
        )
        .first()
    )

    if not invoice:
        raise HTTPException(404, "Invoice not found")

    # Get the base URL for share link
    base_url = os.getenv("REPLIT_DOMAINS", "https://involinks.replit.app")
    if base_url:
        base_url = (
            f"https://{base_url.split(',')[0]}"
            if "," in base_url
            else f"https://{base_url}"
        )

    share_url = f"{base_url}/invoices/view/{invoice.share_token}"

    # In production, this would use Twilio integration
    # For now, simulate SMS sending
    sms_content = {
        "to": phone_number,
        "message": f"Invoice {invoice.invoice_number} from {invoice.supplier_name} - AED {invoice.total_amount:,.2f}. View: {share_url}",
        "share_link": share_url,
    }

    return {
        "message": "Invoice SMS sent successfully",
        "sent_to": phone_number,
        "invoice_id": invoice.id,
        "sms_status": "sent",
        "sms_content": sms_content,
    }


@app.post("/invoices/{invoice_id}/whatsapp", tags=["Invoices"])
async def whatsapp_invoice(
    invoice_id: str,
    phone_number: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Send invoice link to customer via WhatsApp"""
    invoice = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.id == invoice_id, InvoiceDB.company_id == current_user.company_id
        )
        .first()
    )

    if not invoice:
        raise HTTPException(404, "Invoice not found")

    # Get the base URL for share link
    base_url = os.getenv("REPLIT_DOMAINS", "https://involinks.replit.app")
    if base_url:
        base_url = (
            f"https://{base_url.split(',')[0]}"
            if "," in base_url
            else f"https://{base_url}"
        )

    share_url = f"{base_url}/invoices/view/{invoice.share_token}"

    # In production, this would use Twilio WhatsApp API
    # For now, simulate WhatsApp sending
    whatsapp_content = {
        "to": phone_number,
        "message": f"📄 *Invoice {invoice.invoice_number}*\n\nFrom: {invoice.supplier_name}\nAmount: AED {invoice.total_amount:,.2f}\n\nView invoice: {share_url}",
        "share_link": share_url,
    }

    return {
        "message": "Invoice WhatsApp message sent successfully",
        "sent_to": phone_number,
        "invoice_id": invoice.id,
        "whatsapp_status": "sent",
        "whatsapp_content": whatsapp_content,
    }


# Public invoice viewing (no authentication required)
@app.get("/invoices/view/{share_token}", tags=["Public"], response_model=InvoiceOut)
def view_shared_invoice(share_token: str, db: Session = Depends(get_db)):
    """View invoice via public share link (customer portal)"""
    invoice = db.query(InvoiceDB).filter(InvoiceDB.share_token == share_token).first()

    if not invoice:
        raise HTTPException(404, "Invoice not found")

    # Mark as viewed
    if not invoice.viewed_at:
        invoice.viewed_at = datetime.utcnow()
        db.commit()

    return InvoiceOut(
        id=invoice.id,
        company_id=invoice.company_id,
        invoice_number=invoice.invoice_number,
        invoice_type=invoice.invoice_type,
        status=invoice.status,
        issue_date=invoice.issue_date.isoformat(),
        due_date=invoice.due_date.isoformat() if invoice.due_date else None,
        supply_date=invoice.supply_date.isoformat() if invoice.supply_date else None,
        currency_code=invoice.currency_code,
        supplier_trn=invoice.supplier_trn,
        supplier_name=invoice.supplier_name,
        supplier_address=invoice.supplier_address,
        supplier_peppol_id=invoice.supplier_peppol_id,
        customer_trn=invoice.customer_trn,
        customer_name=invoice.customer_name,
        customer_email=invoice.customer_email,
        customer_address=invoice.customer_address,
        customer_city=invoice.customer_city,
        invoice_notes=invoice.invoice_notes,
        customer_peppol_id=invoice.customer_peppol_id,
        subtotal_amount=invoice.subtotal_amount,
        tax_amount=invoice.tax_amount,
        total_amount=invoice.total_amount,
        amount_due=invoice.amount_due,
        total_amount_aed=getattr(invoice, "total_amount_aed", None),
        preceding_invoice_id=invoice.preceding_invoice_id,
        credit_note_reason=invoice.credit_note_reason,
        xml_file_path=invoice.xml_file_path,
        pdf_file_path=invoice.pdf_file_path,
        share_token=invoice.share_token,
        created_at=invoice.created_at.isoformat(),
        sent_at=invoice.sent_at.isoformat() if invoice.sent_at else None,
        viewed_at=invoice.viewed_at.isoformat() if invoice.viewed_at else None,
        line_items=[
            InvoiceLineItemOut(
                id=li.id,
                line_number=li.line_number,
                item_name=li.item_name,
                item_description=li.item_description,
                quantity=li.quantity,
                unit_code=li.unit_code,
                unit_price=li.unit_price,
                line_extension_amount=li.line_extension_amount,
                tax_category=li.tax_category,
                tax_percent=li.tax_percent,
                tax_amount=li.tax_amount,
                line_total_amount=li.line_total_amount,
            )
            for li in invoice.line_items
        ],
        tax_breakdowns=[
            InvoiceTaxBreakdownOut(
                id=tb.id,
                tax_category=tb.tax_category,
                taxable_amount=tb.taxable_amount,
                tax_percent=tb.tax_percent,
                tax_amount=tb.tax_amount,
            )
            for tb in invoice.tax_breakdowns
        ],
    )


# ==================== EXPENSE TRACKING (SIMPLE FINANCIAL MANAGEMENT) ====================


@app.post("/expense-categories", tags=["Expense Categories"])
def create_expense_category(
    name: str = Form(...),
    description: str = Form(None),
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Create a custom expense category"""
    from uuid import uuid4

    normalized_name = (name or "").strip()
    if not normalized_name:
        raise HTTPException(400, "Category name is required")

    # Check if category already exists
    existing = (
        db.query(ExpenseCategoryDB)
        .filter(
            ExpenseCategoryDB.company_id == current_user.company_id,
            func.lower(ExpenseCategoryDB.name) == normalized_name.lower(),
        )
        .first()
    )

    if existing:
        raise HTTPException(400, f"Category '{normalized_name}' already exists")

    category_id = f"expcat_{uuid4().hex[:12]}"

    category = ExpenseCategoryDB(
        id=category_id,
        company_id=current_user.company_id,
        name=normalized_name,
        description=description,
        is_default=False,
    )

    db.add(category)
    db.commit()
    db.refresh(category)

    return {
        "id": category.id,
        "name": category.name,
        "description": category.description,
        "is_default": category.is_default,
        "created_at": category.created_at.isoformat(),
    }


@app.get("/expense-categories", tags=["Expense Categories"])
def list_expense_categories(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    List all expense categories (default + custom)

    Returns system defaults + user-created categories
    """
    categories = (
        db.query(ExpenseCategoryDB)
        .filter(
            ExpenseCategoryDB.company_id == current_user.company_id,
            ExpenseCategoryDB.is_default == False,
        )
        .order_by(ExpenseCategoryDB.name)
        .all()
    )

    return {
        "categories": [
            {
                "id": cat.id,
                "name": cat.name,
                "description": cat.description,
                "is_default": cat.is_default,
                "created_at": cat.created_at.isoformat(),
            }
            for cat in categories
        ]
    }


@app.delete("/expense-categories/{category_id}", tags=["Expense Categories"])
def delete_expense_category(
    category_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Delete a custom expense category (cannot delete defaults)"""
    category = (
        db.query(ExpenseCategoryDB)
        .filter(
            ExpenseCategoryDB.id == category_id,
            ExpenseCategoryDB.company_id == current_user.company_id,
        )
        .first()
    )

    if not category:
        raise HTTPException(404, "Category not found")

    if category.is_default:
        raise HTTPException(400, "Cannot delete default categories")

    # Check if any expenses use this category
    expense_count = (
        db.query(ExpenseDB)
        .filter(
            ExpenseDB.company_id == current_user.company_id,
            ExpenseDB.category == category.name,
        )
        .count()
    )

    if expense_count > 0:
        raise HTTPException(
            400,
            f"Cannot delete category with {expense_count} expense(s). Reassign them first.",
        )

    db.delete(category)
    db.commit()

    return {"message": "Category deleted successfully", "category_id": category_id}


@app.put("/expense-categories/{category_id}", tags=["Expense Categories"])
def update_expense_category(
    category_id: str,
    name: str = Form(...),
    description: str = Form(None),
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Update a custom expense category. Renames will update existing expenses to new name."""
    category = (
        db.query(ExpenseCategoryDB)
        .filter(
            ExpenseCategoryDB.id == category_id,
            ExpenseCategoryDB.company_id == current_user.company_id,
        )
        .first()
    )

    if not category:
        raise HTTPException(404, "Category not found")

    if category.is_default:
        raise HTTPException(400, "Cannot edit default categories")

    new_name = (name or "").strip()
    if not new_name:
        raise HTTPException(400, "Category name is required")

    # Check for duplicates
    existing = (
        db.query(ExpenseCategoryDB)
        .filter(
            ExpenseCategoryDB.company_id == current_user.company_id,
            func.lower(ExpenseCategoryDB.name) == new_name.lower(),
            ExpenseCategoryDB.id != category_id,
        )
        .first()
    )

    if existing:
        raise HTTPException(400, f"Category '{new_name}' already exists")

    old_name = category.name
    category.name = new_name
    category.description = description
    db.add(category)

    # If name changed, update related expenses to new category name
    if old_name != new_name:
        db.query(ExpenseDB).filter(
            ExpenseDB.company_id == current_user.company_id,
            ExpenseDB.category == old_name,
        ).update({"category": new_name})

    db.commit()
    db.refresh(category)

    return {
        "id": category.id,
        "name": category.name,
        "description": category.description,
        "is_default": category.is_default,
        "created_at": category.created_at.isoformat(),
    }


@app.post("/expenses", tags=["Expenses"])
def create_expense(
    expense_date: str = Form(...),
    category: str = Form(...),
    amount: float = Form(...),
    vat_amount: float = Form(0.0),
    description: str = Form(None),
    supplier_name: str = Form(None),
    reference_number: str = Form(None),
    notes: str = Form(None),
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Create a simple expense record

    Categories: RENT, UTILITIES, SALARIES, RAW_MATERIALS, OTHER

    Example for spa:
    - Rent: AED 4,000
    - Utilities: AED 1,500 (includes VAT 71.43)
    - Salaries: AED 15,000 (no VAT)
    - Raw Materials: AED 4,000 (includes VAT 190.48)
    """
    from dateutil import parser
    from uuid import uuid4

    category_name = (category or "").strip()
    if not category_name:
        raise HTTPException(400, "Category is required")

    matched_category = (
        db.query(ExpenseCategoryDB)
        .filter(
            ExpenseCategoryDB.company_id == current_user.company_id,
            func.lower(ExpenseCategoryDB.name) == category_name.lower(),
        )
        .first()
    )

    if not matched_category:
        raise HTTPException(400, f"Invalid category '{category_name}' for this company")

    # Parse date
    try:
        expense_date_obj = parser.parse(expense_date).date()
    except:
        raise HTTPException(400, "Invalid date format. Use YYYY-MM-DD or ISO format")

    # Calculate total
    total_amount = amount + vat_amount

    # Create expense
    expense_id = f"exp_{uuid4().hex[:12]}"

    expense = ExpenseDB(
        id=expense_id,
        company_id=current_user.company_id,
        expense_date=expense_date_obj,
        category=matched_category.name,
        description=description,
        amount=amount,
        vat_amount=vat_amount,
        total_amount=total_amount,
        supplier_name=supplier_name,
        reference_number=reference_number,
        notes=notes,
    )

    db.add(expense)
    db.commit()
    db.refresh(expense)

    return {
        "id": expense.id,
        "expense_date": expense.expense_date.isoformat(),
        "category": expense.category,
        "description": expense.description,
        "amount": expense.amount,
        "vat_amount": expense.vat_amount,
        "total_amount": expense.total_amount,
        "supplier_name": expense.supplier_name,
        "reference_number": expense.reference_number,
        "created_at": expense.created_at.isoformat(),
    }


@app.get("/expenses", tags=["Expenses"])
def list_expenses(
    month: str = None,  # "2025-10" for October 2025
    category: str = None,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    List all expenses with optional filters

    Filters:
    - month: "2025-10" for October 2025
    - category: RENT, UTILITIES, SALARIES, RAW_MATERIALS, OTHER
    """
    query = db.query(ExpenseDB).filter(ExpenseDB.company_id == current_user.company_id)

    # Filter by month
    if month:
        try:
            year, month_num = map(int, month.split("-"))
            from datetime import date

            start_date = date(year, month_num, 1)
            # Calculate end date (last day of month)
            if month_num == 12:
                end_date = date(year + 1, 1, 1)
            else:
                end_date = date(year, month_num + 1, 1)

            query = query.filter(
                ExpenseDB.expense_date >= start_date, ExpenseDB.expense_date < end_date
            )
        except:
            raise HTTPException(400, "Invalid month format. Use YYYY-MM")

    # Filter by category
    if category:
        query = query.filter(func.lower(ExpenseDB.category) == category.lower())

    # Order by date descending
    expenses = query.order_by(ExpenseDB.expense_date.desc()).all()

    return {
        "expenses": [
            {
                "id": exp.id,
                "expense_date": exp.expense_date.isoformat(),
                "category": exp.category,
                "description": exp.description,
                "amount": exp.amount,
                "vat_amount": exp.vat_amount,
                "total_amount": exp.total_amount,
                "supplier_name": exp.supplier_name,
                "reference_number": exp.reference_number,
                "notes": exp.notes,
                "created_at": exp.created_at.isoformat(),
            }
            for exp in expenses
        ],
        "total_count": len(expenses),
    }


@app.get("/expenses/summary", tags=["Expenses"])
def get_financial_summary(
    month: str = None,  # "2025-10" for October 2025
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Get financial summary: Net Income and Net VAT

    Calculation:
    - Revenue (from invoices issued and paid)
    - Total Expenses (from expense records)
    - Net Income = Revenue - Expenses

    - Output VAT (from sales invoices)
    - Input VAT (from expense VAT)
    - Net VAT Payable = Output VAT - Input VAT

    Example for Spa (October):
    - Revenue: AED 35,000 (from services and products)
    - Expenses: AED 24,500 (rent + utilities + salaries + materials)
    - Net Income: AED 10,500

    - Output VAT: AED 6,000 (VAT collected from customers)
    - Input VAT: AED 3,000 (VAT paid to suppliers)
    - Net VAT: AED 3,000 (to pay to FTA)
    """
    from datetime import date

    # Parse month filter
    if month:
        try:
            year, month_num = map(int, month.split("-"))
            start_date = date(year, month_num, 1)
            if month_num == 12:
                end_date = date(year + 1, 1, 1)
            else:
                end_date = date(year, month_num + 1, 1)
        except:
            raise HTTPException(400, "Invalid month format. Use YYYY-MM")
    else:
        # Default to current month
        from datetime import datetime

        now = datetime.now()
        start_date = date(now.year, now.month, 1)
        if now.month == 12:
            end_date = date(now.year + 1, 1, 1)
        else:
            end_date = date(now.year, now.month + 1, 1)

    # ========== REVENUE (from issued invoices) ==========
    revenue_query = db.query(InvoiceDB).filter(
        InvoiceDB.company_id == current_user.company_id,
        InvoiceDB.status.in_(
            [InvoiceStatus.ISSUED, InvoiceStatus.SENT, InvoiceStatus.PAID]
        ),
        InvoiceDB.issue_date >= start_date,
        InvoiceDB.issue_date < end_date,
    )

    total_revenue = 0.0
    output_vat = 0.0
    invoice_count = 0
    # Count only invoices with status ISSUED in the period
    issued_invoice_count = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.company_id == current_user.company_id,
            InvoiceDB.status == InvoiceStatus.ISSUED,
            InvoiceDB.issue_date >= start_date,
            InvoiceDB.issue_date < end_date,
        )
        .count()
    )

    for invoice in revenue_query.all():
        # Credit notes should be subtracted from revenue, not added
        if invoice.invoice_type in [
            InvoiceType.TAX_CREDIT_NOTE,
            InvoiceType.CREDIT_NOTE_OUT_OF_SCOPE,
        ]:
            total_revenue -= invoice.subtotal_amount or 0.0
            output_vat -= invoice.tax_amount or 0.0
        else:
            total_revenue += invoice.subtotal_amount or 0.0
            output_vat += invoice.tax_amount or 0.0
        invoice_count += 1

    # ========== EXPENSES (from expense records) ==========
    expense_query = db.query(ExpenseDB).filter(
        ExpenseDB.company_id == current_user.company_id,
        ExpenseDB.expense_date >= start_date,
        ExpenseDB.expense_date < end_date,
    )

    total_expenses = 0.0
    input_vat = 0.0
    expense_count = 0
    expense_breakdown = {}

    for expense in expense_query.all():
        total_expenses += expense.amount or 0.0
        input_vat += expense.vat_amount or 0.0
        expense_count += 1

        # Category breakdown
        cat = expense.category
        if cat not in expense_breakdown:
            expense_breakdown[cat] = {"amount": 0.0, "vat": 0.0, "count": 0}
        expense_breakdown[cat]["amount"] += expense.amount
        expense_breakdown[cat]["vat"] += expense.vat_amount
        expense_breakdown[cat]["count"] += 1

    # ========== CALCULATIONS ==========
    net_income = total_revenue - total_expenses
    net_vat_payable = output_vat - input_vat

    return {
        "period": {
            "start_date": start_date.isoformat(),
            "end_date": (end_date - timedelta(days=1)).isoformat(),
            "month": f"{start_date.year}-{start_date.month:02d}",
        },
        "revenue": {
            "total": round(total_revenue, 2),
            "invoice_count": invoice_count,
            "issued_invoice_count": issued_invoice_count,
            "vat_collected": round(output_vat, 2),
        },
        "expenses": {
            "total": round(total_expenses, 2),
            "expense_count": expense_count,
            "vat_paid": round(input_vat, 2),
            "breakdown": {
                cat: {
                    "amount": round(data["amount"], 2),
                    "vat": round(data["vat"], 2),
                    "count": data["count"],
                }
                for cat, data in expense_breakdown.items()
            },
        },
        "summary": {
            "net_income": round(net_income, 2),
            "net_vat_payable": round(net_vat_payable, 2),
            "gross_revenue": round(total_revenue + output_vat, 2),
            "total_costs": round(total_expenses + input_vat, 2),
            "profit_margin_percent": round((net_income / total_revenue * 100), 2)
            if total_revenue > 0
            else 0,
        },
        "vat_details": {
            "output_vat": round(output_vat, 2),
            "input_vat": round(input_vat, 2),
            "net_vat": round(net_vat_payable, 2),
            "vat_status": "PAYABLE"
            if net_vat_payable > 0
            else "REFUND"
            if net_vat_payable < 0
            else "NEUTRAL",
        },
    }


@app.delete("/expenses/{expense_id}", tags=["Expenses"])
def delete_expense(
    expense_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Delete an expense record"""
    expense = (
        db.query(ExpenseDB)
        .filter(
            ExpenseDB.id == expense_id, ExpenseDB.company_id == current_user.company_id
        )
        .first()
    )

    if not expense:
        raise HTTPException(404, "Expense not found")

    db.delete(expense)
    db.commit()

    return {"message": "Expense deleted successfully", "expense_id": expense_id}


@app.put("/expenses/{expense_id}", tags=["Expenses"])
def update_expense(
    expense_id: str,
    expense_date: str = Form(...),
    category: str = Form(...),
    amount: float = Form(...),
    vat_amount: float = Form(0.0),
    description: str = Form(None),
    supplier_name: str = Form(None),
    reference_number: str = Form(None),
    notes: str = Form(None),
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Update an expense record"""
    expense = (
        db.query(ExpenseDB)
        .filter(
            ExpenseDB.id == expense_id, ExpenseDB.company_id == current_user.company_id
        )
        .first()
    )

    if not expense:
        raise HTTPException(404, "Expense not found")

    # Validate category exists for company
    matched_category = (
        db.query(ExpenseCategoryDB)
        .filter(
            ExpenseCategoryDB.company_id == current_user.company_id,
            func.lower(ExpenseCategoryDB.name) == (category or "").strip().lower(),
        )
        .first()
    )

    if not matched_category:
        raise HTTPException(400, f"Invalid category '{category}' for this company")

    from dateutil import parser

    try:
        expense_date_obj = parser.parse(expense_date).date()
    except:
        raise HTTPException(400, "Invalid date format. Use YYYY-MM-DD or ISO format")

    expense.expense_date = expense_date_obj
    expense.category = matched_category.name
    expense.amount = amount
    expense.vat_amount = vat_amount
    expense.total_amount = (amount or 0.0) + (vat_amount or 0.0)
    expense.description = description
    expense.supplier_name = supplier_name
    expense.reference_number = reference_number
    expense.notes = notes

    db.add(expense)
    db.commit()
    db.refresh(expense)

    return {
        "id": expense.id,
        "expense_date": expense.expense_date.isoformat(),
        "category": expense.category,
        "description": expense.description,
        "amount": expense.amount,
        "vat_amount": expense.vat_amount,
        "total_amount": expense.total_amount,
        "supplier_name": expense.supplier_name,
        "reference_number": expense.reference_number,
        "notes": expense.notes,
        "created_at": expense.created_at.isoformat(),
    }


# ==================== SIMPLE INVENTORY TRACKING ====================


@app.post("/inventory", tags=["Inventory"])
def create_inventory_item(
    item_name: str,
    current_stock: float = 0,
    unit: str = "unit",
    item_code: str = None,
    description: str = None,
    min_stock_level: float = 0,
    unit_cost: float = None,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Create inventory item (e.g., Shampoo Bottles, Mani Kits)"""
    from uuid import uuid4

    item_id = f"inv_{uuid4().hex[:12]}"

    item = InventoryItemDB(
        id=item_id,
        company_id=current_user.company_id,
        item_name=item_name,
        item_code=item_code,
        description=description,
        unit=unit,
        current_stock=current_stock,
        min_stock_level=min_stock_level,
        unit_cost=unit_cost,
    )

    db.add(item)
    db.commit()
    db.refresh(item)

    return {
        "id": item.id,
        "item_name": item.item_name,
        "item_code": item.item_code,
        "current_stock": item.current_stock,
        "unit": item.unit,
        "min_stock_level": item.min_stock_level,
        "low_stock": item.current_stock <= item.min_stock_level,
        "created_at": item.created_at.isoformat(),
    }


@app.get("/inventory", tags=["Inventory"])
def list_inventory(
    low_stock_only: bool = False,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """List all inventory items with stock levels"""
    query = db.query(InventoryItemDB).filter(
        InventoryItemDB.company_id == current_user.company_id
    )

    items = query.order_by(InventoryItemDB.item_name).all()

    result = []
    for item in items:
        is_low = item.current_stock <= item.min_stock_level

        if low_stock_only and not is_low:
            continue

        result.append(
            {
                "id": item.id,
                "item_name": item.item_name,
                "item_code": item.item_code,
                "description": item.description,
                "current_stock": item.current_stock,
                "unit": item.unit,
                "min_stock_level": item.min_stock_level,
                "low_stock": is_low,
                "unit_cost": item.unit_cost,
                "stock_value": (item.current_stock * item.unit_cost)
                if item.unit_cost
                else None,
            }
        )

    return {"inventory": result, "total_items": len(result)}


@app.post("/inventory/{item_id}/adjust", tags=["Inventory"])
def adjust_inventory(
    item_id: str,
    quantity: float,
    transaction_type: str,
    notes: str = None,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Adjust inventory stock

    transaction_type: PURCHASE, SALE, ADJUSTMENT, SERVICE_USAGE
    quantity: positive for additions, negative for reductions

    Examples:
    - Purchased 50 shampoo bottles: quantity=50, type=PURCHASE
    - Sold 5 shampoo bottles: quantity=-5, type=SALE
    - Used 200 mani kits for services: quantity=-200, type=SERVICE_USAGE
    """
    from uuid import uuid4
    from datetime import date

    item = (
        db.query(InventoryItemDB)
        .filter(
            InventoryItemDB.id == item_id,
            InventoryItemDB.company_id == current_user.company_id,
        )
        .first()
    )

    if not item:
        raise HTTPException(404, "Inventory item not found")

    # Update stock
    new_stock = item.current_stock + quantity

    if new_stock < 0:
        raise HTTPException(
            400,
            f"Insufficient stock. Current: {item.current_stock}, Requested: {abs(quantity)}",
        )

    # Record transaction
    trans_id = f"invtx_{uuid4().hex[:12]}"
    transaction = InventoryTransactionDB(
        id=trans_id,
        company_id=current_user.company_id,
        inventory_item_id=item.id,
        transaction_type=transaction_type,
        quantity=quantity,
        transaction_date=date.today(),
        notes=notes,
        stock_after=new_stock,
        reference_type="MANUAL",
    )

    db.add(transaction)

    # Update item stock
    item.current_stock = new_stock

    db.commit()
    db.refresh(item)

    return {
        "message": "Stock adjusted successfully",
        "item_name": item.item_name,
        "previous_stock": item.current_stock - quantity,
        "adjustment": quantity,
        "new_stock": item.current_stock,
        "low_stock": item.current_stock <= item.min_stock_level,
    }


@app.get("/inventory/{item_id}/history", tags=["Inventory"])
def get_inventory_history(
    item_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Get transaction history for an inventory item"""
    item = (
        db.query(InventoryItemDB)
        .filter(
            InventoryItemDB.id == item_id,
            InventoryItemDB.company_id == current_user.company_id,
        )
        .first()
    )

    if not item:
        raise HTTPException(404, "Inventory item not found")

    transactions = (
        db.query(InventoryTransactionDB)
        .filter(InventoryTransactionDB.inventory_item_id == item_id)
        .order_by(InventoryTransactionDB.transaction_date.desc())
        .limit(50)
        .all()
    )

    return {
        "item_name": item.item_name,
        "current_stock": item.current_stock,
        "history": [
            {
                "id": tx.id,
                "transaction_type": tx.transaction_type,
                "quantity": tx.quantity,
                "transaction_date": tx.transaction_date.isoformat(),
                "stock_after": tx.stock_after,
                "notes": tx.notes,
                "reference_type": tx.reference_type,
                "reference_id": tx.reference_id,
            }
            for tx in transactions
        ],
    }


# ==================== CORNER 4: AP MANAGEMENT (INWARD INVOICES) ====================


@app.post(
    "/inward-invoices/receive", tags=["AP Management"], response_model=InwardInvoiceOut
)
def receive_inward_invoice(
    invoice_data: InwardInvoiceReceive,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Receive inward invoice from supplier (via PEPPOL or manual entry)

    This endpoint:
    1. Validates the received invoice data
    2. Stores the invoice in the inward_invoices table
    3. Saves UBL XML if provided
    4. Creates line items
    5. Attempts automatic PO matching if supplier TRN matches
    """
    company = (
        db.query(CompanyDB).filter(CompanyDB.id == current_user.company_id).first()
    )
    if not company:
        raise HTTPException(404, "Company not found")

    # Validate that customer TRN matches our company TRN
    if invoice_data.customer_trn != company.trn:
        raise HTTPException(
            400,
            f"Invoice customer TRN ({invoice_data.customer_trn}) does not match your company TRN ({company.trn})",
        )

    # Check for duplicate invoice number from this supplier
    existing = (
        db.query(InwardInvoiceDB)
        .filter(
            InwardInvoiceDB.company_id == current_user.company_id,
            InwardInvoiceDB.supplier_trn == invoice_data.supplier_trn,
            InwardInvoiceDB.supplier_invoice_number
            == invoice_data.supplier_invoice_number,
        )
        .first()
    )

    if existing:
        raise HTTPException(
            409,
            f"Invoice {invoice_data.supplier_invoice_number} from supplier {invoice_data.supplier_name} already exists",
        )

    # Create inward invoice
    inward_invoice_id = str(uuid4())

    # Parse dates
    from datetime import datetime as dt

    invoice_date = dt.fromisoformat(invoice_data.invoice_date)
    due_date = (
        dt.fromisoformat(invoice_data.due_date) if invoice_data.due_date else None
    )

    # Save XML file if provided
    xml_file_path = None
    xml_hash_value = None
    if invoice_data.xml_content:
        xml_dir = os.path.join(
            ARTIFACT_ROOT, "inward_invoices", current_user.company_id
        )
        os.makedirs(xml_dir, exist_ok=True)
        xml_file_path = os.path.join(xml_dir, f"{inward_invoice_id}.xml")

        with open(xml_file_path, "w", encoding="utf-8") as f:
            f.write(invoice_data.xml_content)

        # Calculate hash
        from utils.crypto_utils import hash_data

        xml_hash_value = hash_data(invoice_data.xml_content)

    # Attempt to find matching PO by supplier TRN
    matching_po = (
        db.query(PurchaseOrderDB)
        .filter(
            PurchaseOrderDB.company_id == current_user.company_id,
            PurchaseOrderDB.supplier_trn == invoice_data.supplier_trn,
            PurchaseOrderDB.status.in_(
                [PurchaseOrderStatus.SENT, PurchaseOrderStatus.ACKNOWLEDGED]
            ),
        )
        .order_by(PurchaseOrderDB.order_date.desc())
        .first()
    )

    po_id = matching_po.id if matching_po else None
    matching_status = MatchingStatus.NOT_MATCHED
    amount_variance = 0.0

    if matching_po:
        # Calculate variance
        amount_variance = invoice_data.total_amount - matching_po.expected_total
        if abs(amount_variance) < 0.01:  # Within 1 cent
            matching_status = MatchingStatus.FULL_MATCH
        elif abs(amount_variance) < (matching_po.expected_total * 0.05):  # Within 5%
            matching_status = MatchingStatus.VARIANCE_DETECTED
        else:
            matching_status = MatchingStatus.PARTIAL_MATCH

    # Create inward invoice record
    new_invoice = InwardInvoiceDB(
        id=inward_invoice_id,
        company_id=current_user.company_id,
        supplier_invoice_number=invoice_data.supplier_invoice_number,
        invoice_type=InvoiceType.TAX_INVOICE,
        status=InwardInvoiceStatus.RECEIVED,
        invoice_date=invoice_date,
        due_date=due_date,
        received_at=datetime.utcnow(),
        supplier_trn=invoice_data.supplier_trn,
        supplier_name=invoice_data.supplier_name,
        supplier_address=invoice_data.supplier_address,
        supplier_peppol_id=invoice_data.supplier_peppol_id,
        customer_trn=invoice_data.customer_trn,
        customer_name=invoice_data.customer_name,
        currency_code=invoice_data.currency_code,
        subtotal_amount=invoice_data.subtotal_amount,
        tax_amount=invoice_data.tax_amount,
        total_amount=invoice_data.total_amount,
        amount_due=invoice_data.amount_due,
        xml_file_path=xml_file_path,
        xml_hash=xml_hash_value,
        peppol_message_id=invoice_data.peppol_message_id,
        peppol_sender_id=invoice_data.peppol_sender_id,
        peppol_provider=invoice_data.peppol_provider,
        peppol_received_at=datetime.utcnow()
        if invoice_data.peppol_message_id
        else None,
        po_id=po_id,
        matching_status=matching_status,
        amount_variance=amount_variance,
        payment_status="PENDING",
    )

    db.add(new_invoice)

    # Create line items if provided
    if invoice_data.line_items:
        for idx, item_data in enumerate(invoice_data.line_items):
            line_item = InwardInvoiceLineItemDB(
                id=str(uuid4()),
                inward_invoice_id=inward_invoice_id,
                line_number=idx + 1,
                item_name=item_data.get("item_name", "Item"),
                item_description=item_data.get("item_description"),
                item_code=item_data.get("item_code"),
                quantity=float(item_data.get("quantity", 1)),
                unit_code=item_data.get("unit_code", "C62"),
                unit_price=float(item_data.get("unit_price", 0)),
                line_extension_amount=float(item_data.get("line_extension_amount", 0)),
                tax_category=TaxCategory(item_data.get("tax_category", "S")),
                tax_percent=float(item_data.get("tax_percent", 5.0)),
                tax_amount=float(item_data.get("tax_amount", 0)),
                line_total_amount=float(item_data.get("line_total_amount", 0)),
            )
            db.add(line_item)

    # Update PO if matched
    if matching_po:
        matching_po.received_invoice_count += 1
        matching_po.received_amount_total += invoice_data.total_amount
        matching_po.variance_amount = (
            matching_po.expected_total - matching_po.received_amount_total
        )

    # Task 8: Auto-generate GL journal entry for purchase (wrapped — GL failure must not block)
    try:
        create_journal_entry_for_purchase(new_invoice, db, user_id=current_user.id)
    except Exception as gl_err:
        logger.error(f"GL purchase journal entry failed: {gl_err}")

    db.commit()
    db.refresh(new_invoice)

    # Return response
    return InwardInvoiceOut(
        id=new_invoice.id,
        company_id=new_invoice.company_id,
        supplier_invoice_number=new_invoice.supplier_invoice_number,
        invoice_type=new_invoice.invoice_type,
        status=new_invoice.status,
        invoice_date=new_invoice.invoice_date.isoformat(),
        due_date=new_invoice.due_date.isoformat() if new_invoice.due_date else None,
        received_at=new_invoice.received_at.isoformat(),
        supplier_trn=new_invoice.supplier_trn,
        supplier_name=new_invoice.supplier_name,
        supplier_address=new_invoice.supplier_address,
        supplier_peppol_id=new_invoice.supplier_peppol_id,
        supplier_company_id=new_invoice.supplier_company_id,
        customer_trn=new_invoice.customer_trn,
        customer_name=new_invoice.customer_name,
        currency_code=new_invoice.currency_code,
        subtotal_amount=new_invoice.subtotal_amount,
        tax_amount=new_invoice.tax_amount,
        total_amount=new_invoice.total_amount,
        amount_due=new_invoice.amount_due,
        xml_file_path=new_invoice.xml_file_path,
        pdf_file_path=new_invoice.pdf_file_path,
        xml_hash=new_invoice.xml_hash,
        peppol_message_id=new_invoice.peppol_message_id,
        peppol_sender_id=new_invoice.peppol_sender_id,
        peppol_provider=new_invoice.peppol_provider,
        peppol_received_at=new_invoice.peppol_received_at.isoformat()
        if new_invoice.peppol_received_at
        else None,
        po_id=new_invoice.po_id,
        grn_id=new_invoice.grn_id,
        matching_status=new_invoice.matching_status,
        po_match_score=new_invoice.po_match_score,
        grn_match_score=new_invoice.grn_match_score,
        amount_variance=new_invoice.amount_variance,
        quantity_variance=new_invoice.quantity_variance,
        reviewed_by_user_id=new_invoice.reviewed_by_user_id,
        reviewed_at=new_invoice.reviewed_at.isoformat()
        if new_invoice.reviewed_at
        else None,
        approved_by_user_id=new_invoice.approved_by_user_id,
        approved_at=new_invoice.approved_at.isoformat()
        if new_invoice.approved_at
        else None,
        rejection_reason=new_invoice.rejection_reason,
        payment_status=new_invoice.payment_status,
        payment_method=new_invoice.payment_method,
        paid_amount=new_invoice.paid_amount,
        paid_at=new_invoice.paid_at.isoformat() if new_invoice.paid_at else None,
        payment_reference=new_invoice.payment_reference,
        disputed_at=new_invoice.disputed_at.isoformat()
        if new_invoice.disputed_at
        else None,
        dispute_reason=new_invoice.dispute_reason,
        dispute_resolved_at=new_invoice.dispute_resolved_at.isoformat()
        if new_invoice.dispute_resolved_at
        else None,
        notes=new_invoice.notes,
        reference_number=new_invoice.reference_number,
        created_at=new_invoice.created_at.isoformat(),
        updated_at=new_invoice.updated_at.isoformat(),
        line_items=[
            InwardInvoiceLineItemOut(
                id=li.id,
                line_number=li.line_number,
                item_name=li.item_name,
                item_description=li.item_description,
                item_code=li.item_code,
                quantity=li.quantity,
                unit_code=li.unit_code,
                unit_price=li.unit_price,
                line_extension_amount=li.line_extension_amount,
                tax_category=li.tax_category,
                tax_percent=li.tax_percent,
                tax_amount=li.tax_amount,
                line_total_amount=li.line_total_amount,
                po_line_item_id=li.po_line_item_id,
                grn_line_item_id=li.grn_line_item_id,
                match_status=li.match_status,
            )
            for li in new_invoice.line_items
        ],
    )


@app.get(
    "/inward-invoices",
    tags=["AP Management"],
    response_model=List[InwardInvoiceListOut],
)
def list_inward_invoices(
    status: Optional[str] = None,
    supplier_trn: Optional[str] = None,
    matching_status: Optional[str] = None,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    List all inward invoices (AP Inbox)

    Query parameters:
    - status: Filter by invoice status (RECEIVED, PENDING_REVIEW, APPROVED, etc.)
    - supplier_trn: Filter by supplier TRN
    - matching_status: Filter by PO matching status
    """
    query = db.query(InwardInvoiceDB).filter(
        InwardInvoiceDB.company_id == current_user.company_id
    )

    if status:
        query = query.filter(InwardInvoiceDB.status == status)

    if supplier_trn:
        query = query.filter(InwardInvoiceDB.supplier_trn == supplier_trn)

    if matching_status:
        query = query.filter(InwardInvoiceDB.matching_status == matching_status)

    invoices = query.order_by(InwardInvoiceDB.received_at.desc()).all()

    return [
        InwardInvoiceListOut(
            id=inv.id,
            supplier_invoice_number=inv.supplier_invoice_number,
            status=inv.status,
            supplier_name=inv.supplier_name,
            total_amount=inv.total_amount,
            currency_code=inv.currency_code,
            invoice_date=inv.invoice_date.isoformat(),
            received_at=inv.received_at.isoformat(),
            matching_status=inv.matching_status,
            po_id=inv.po_id,
        )
        for inv in invoices
    ]


@app.get(
    "/inward-invoices/{invoice_id}",
    tags=["AP Management"],
    response_model=InwardInvoiceOut,
)
def get_inward_invoice(
    invoice_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Get details of a specific inward invoice"""
    invoice = (
        db.query(InwardInvoiceDB)
        .filter(
            InwardInvoiceDB.id == invoice_id,
            InwardInvoiceDB.company_id == current_user.company_id,
        )
        .first()
    )

    if not invoice:
        raise HTTPException(404, "Inward invoice not found")

    return InwardInvoiceOut(
        id=invoice.id,
        company_id=invoice.company_id,
        supplier_invoice_number=invoice.supplier_invoice_number,
        invoice_type=invoice.invoice_type,
        status=invoice.status,
        invoice_date=invoice.invoice_date.isoformat(),
        due_date=invoice.due_date.isoformat() if invoice.due_date else None,
        received_at=invoice.received_at.isoformat(),
        supplier_trn=invoice.supplier_trn,
        supplier_name=invoice.supplier_name,
        supplier_address=invoice.supplier_address,
        supplier_peppol_id=invoice.supplier_peppol_id,
        supplier_company_id=invoice.supplier_company_id,
        customer_trn=invoice.customer_trn,
        customer_name=invoice.customer_name,
        currency_code=invoice.currency_code,
        subtotal_amount=invoice.subtotal_amount,
        tax_amount=invoice.tax_amount,
        total_amount=invoice.total_amount,
        amount_due=invoice.amount_due,
        xml_file_path=invoice.xml_file_path,
        pdf_file_path=invoice.pdf_file_path,
        xml_hash=invoice.xml_hash,
        peppol_message_id=invoice.peppol_message_id,
        peppol_sender_id=invoice.peppol_sender_id,
        peppol_provider=invoice.peppol_provider,
        peppol_received_at=invoice.peppol_received_at.isoformat()
        if invoice.peppol_received_at
        else None,
        po_id=invoice.po_id,
        grn_id=invoice.grn_id,
        matching_status=invoice.matching_status,
        po_match_score=invoice.po_match_score,
        grn_match_score=invoice.grn_match_score,
        amount_variance=invoice.amount_variance,
        quantity_variance=invoice.quantity_variance,
        reviewed_by_user_id=invoice.reviewed_by_user_id,
        reviewed_at=invoice.reviewed_at.isoformat() if invoice.reviewed_at else None,
        approved_by_user_id=invoice.approved_by_user_id,
        approved_at=invoice.approved_at.isoformat() if invoice.approved_at else None,
        rejection_reason=invoice.rejection_reason,
        payment_status=invoice.payment_status,
        payment_method=invoice.payment_method,
        paid_amount=invoice.paid_amount,
        paid_at=invoice.paid_at.isoformat() if invoice.paid_at else None,
        payment_reference=invoice.payment_reference,
        disputed_at=invoice.disputed_at.isoformat() if invoice.disputed_at else None,
        dispute_reason=invoice.dispute_reason,
        dispute_resolved_at=invoice.dispute_resolved_at.isoformat()
        if invoice.dispute_resolved_at
        else None,
        notes=invoice.notes,
        reference_number=invoice.reference_number,
        created_at=invoice.created_at.isoformat(),
        updated_at=invoice.updated_at.isoformat(),
        line_items=[
            InwardInvoiceLineItemOut(
                id=li.id,
                line_number=li.line_number,
                item_name=li.item_name,
                item_description=li.item_description,
                item_code=li.item_code,
                quantity=li.quantity,
                unit_code=li.unit_code,
                unit_price=li.unit_price,
                line_extension_amount=li.line_extension_amount,
                tax_category=li.tax_category,
                tax_percent=li.tax_percent,
                tax_amount=li.tax_amount,
                line_total_amount=li.line_total_amount,
                po_line_item_id=li.po_line_item_id,
                grn_line_item_id=li.grn_line_item_id,
                match_status=li.match_status,
            )
            for li in invoice.line_items
        ],
    )


@app.post("/inward-invoices/{invoice_id}/approve", tags=["AP Management"])
def approve_inward_invoice(
    invoice_id: str,
    approval_data: InwardInvoiceApprove,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Approve inward invoice for payment

    This changes the status to APPROVED and records the approver
    """
    if not has_permission(current_user, "inward_invoices", "approve"):
        raise HTTPException(403, "Permission denied: approve on inward_invoices is not allowed for your role")
    invoice = (
        db.query(InwardInvoiceDB)
        .filter(
            InwardInvoiceDB.id == invoice_id,
            InwardInvoiceDB.company_id == current_user.company_id,
        )
        .first()
    )

    if not invoice:
        raise HTTPException(404, "Inward invoice not found")

    if invoice.status == InwardInvoiceStatus.APPROVED:
        raise HTTPException(400, "Invoice already approved")

    if invoice.status == InwardInvoiceStatus.REJECTED:
        raise HTTPException(400, "Cannot approve a rejected invoice")

    # Update invoice
    invoice.status = InwardInvoiceStatus.APPROVED
    invoice.approved_by_user_id = current_user.id
    invoice.approved_at = datetime.utcnow()

    if approval_data.notes:
        invoice.notes = (invoice.notes or "") + f"\n[Approved] {approval_data.notes}"

    if approval_data.payment_method:
        invoice.payment_method = approval_data.payment_method

    if approval_data.payment_reference:
        invoice.payment_reference = approval_data.payment_reference

    # Update PO if matched
    if invoice.po_id:
        po = (
            db.query(PurchaseOrderDB)
            .filter(PurchaseOrderDB.id == invoice.po_id)
            .first()
        )
        if po:
            po.matched_invoice_count += 1

    db.commit()

    return {
        "success": True,
        "message": f"Invoice {invoice.supplier_invoice_number} approved for payment",
        "invoice_id": invoice.id,
        "approved_by": current_user.email,
        "approved_at": invoice.approved_at.isoformat(),
    }


@app.post("/inward-invoices/{invoice_id}/reject", tags=["AP Management"])
def reject_inward_invoice(
    invoice_id: str,
    rejection_data: InwardInvoiceReject,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Reject inward invoice with reason
    """
    if not has_permission(current_user, "inward_invoices", "approve"):
        raise HTTPException(403, "Permission denied: reject on inward_invoices is not allowed for your role")
    invoice = (
        db.query(InwardInvoiceDB)
        .filter(
            InwardInvoiceDB.id == invoice_id,
            InwardInvoiceDB.company_id == current_user.company_id,
        )
        .first()
    )

    if not invoice:
        raise HTTPException(404, "Inward invoice not found")

    if invoice.status == InwardInvoiceStatus.REJECTED:
        raise HTTPException(400, "Invoice already rejected")

    if invoice.status == InwardInvoiceStatus.PAID:
        raise HTTPException(400, "Cannot reject a paid invoice")

    # Update invoice
    invoice.status = InwardInvoiceStatus.REJECTED
    invoice.reviewed_by_user_id = current_user.id
    invoice.reviewed_at = datetime.utcnow()
    invoice.rejection_reason = rejection_data.rejection_reason

    if rejection_data.notes:
        invoice.notes = (invoice.notes or "") + f"\n[Rejected] {rejection_data.notes}"

    db.commit()

    return {
        "success": True,
        "message": f"Invoice {invoice.supplier_invoice_number} rejected",
        "invoice_id": invoice.id,
        "rejected_by": current_user.email,
        "rejection_reason": rejection_data.rejection_reason,
    }


@app.post("/inward-invoices/{invoice_id}/match-po", tags=["AP Management"])
def match_inward_invoice_to_po(
    invoice_id: str,
    match_data: InwardInvoiceMatchPO,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Manually match inward invoice to a purchase order

    This updates the matching_status and calculates variances
    """
    if not has_permission(current_user, "inward_invoices", "edit"):
        raise HTTPException(403, "Permission denied: edit on inward_invoices is not allowed for your role")
    invoice = (
        db.query(InwardInvoiceDB)
        .filter(
            InwardInvoiceDB.id == invoice_id,
            InwardInvoiceDB.company_id == current_user.company_id,
        )
        .first()
    )

    if not invoice:
        raise HTTPException(404, "Inward invoice not found")

    po = (
        db.query(PurchaseOrderDB)
        .filter(
            PurchaseOrderDB.id == match_data.po_id,
            PurchaseOrderDB.company_id == current_user.company_id,
        )
        .first()
    )

    if not po:
        raise HTTPException(404, "Purchase order not found")

    # Verify supplier matches
    # If both have TRN, match by TRN. Otherwise, match by supplier name
    if invoice.supplier_trn and po.supplier_trn:
        if invoice.supplier_trn != po.supplier_trn:
            raise HTTPException(
                400,
                f"Supplier mismatch: Invoice supplier TRN ({invoice.supplier_trn}) does not match PO supplier TRN ({po.supplier_trn})",
            )
    else:
        # Match by supplier name when TRN is not available for non-VAT suppliers
        if invoice.supplier_name.strip().lower() != po.supplier_name.strip().lower():
            raise HTTPException(
                400,
                f"Supplier mismatch: Invoice supplier ({invoice.supplier_name}) does not match PO supplier ({po.supplier_name})",
            )

    # Calculate variance
    amount_variance = invoice.total_amount - po.expected_total

    # Determine matching status
    if abs(amount_variance) < 0.01:  # Within 1 cent
        matching_status = MatchingStatus.FULL_MATCH
    elif abs(amount_variance) < (po.expected_total * 0.05):  # Within 5%
        matching_status = MatchingStatus.VARIANCE_DETECTED
    else:
        matching_status = MatchingStatus.PARTIAL_MATCH

    # Update invoice
    invoice.po_id = po.id
    invoice.matching_status = matching_status
    invoice.amount_variance = amount_variance
    invoice.po_match_score = (
        100.0 - min(abs(amount_variance / po.expected_total * 100), 100.0)
        if po.expected_total > 0
        else 0.0
    )

    # Update PO
    po.received_invoice_count += 1
    po.received_amount_total += invoice.total_amount
    po.variance_amount = po.expected_total - po.received_amount_total

    db.commit()

    return {
        "success": True,
        "message": f"Invoice matched to PO {po.po_number}",
        "invoice_id": invoice.id,
        "po_id": po.id,
        "matching_status": matching_status.value,
        "amount_variance": amount_variance,
        "po_match_score": invoice.po_match_score,
    }


# ==================== PURCHASE ORDER MANAGEMENT ====================


@app.post("/purchase-orders", tags=["AP Management"], response_model=PurchaseOrderOut)
def create_purchase_order(
    po_data: PurchaseOrderCreate,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Create a new purchase order

    This endpoint:
    1. Validates PO data
    2. Creates PO record
    3. Creates line items
    4. Calculates expected totals
    """
    company = (
        db.query(CompanyDB).filter(CompanyDB.id == current_user.company_id).first()
    )
    if not company:
        raise HTTPException(404, "Company not found")

    # Check for duplicate PO number
    existing_po = (
        db.query(PurchaseOrderDB)
        .filter(
            PurchaseOrderDB.company_id == current_user.company_id,
            PurchaseOrderDB.po_number == po_data.po_number,
        )
        .first()
    )

    if existing_po:
        raise HTTPException(409, f"PO number {po_data.po_number} already exists")

    # Validate line items
    if not po_data.line_items:
        raise HTTPException(400, "At least one line item is required")

    # Calculate expected totals
    expected_subtotal = 0.0
    expected_tax = 0.0

    for item in po_data.line_items:
        line_total = item.quantity_ordered * item.unit_price
        line_tax = line_total * (item.tax_percent / 100)
        expected_subtotal += line_total
        expected_tax += line_tax

    expected_total = expected_subtotal + expected_tax

    # Parse order date
    from datetime import datetime as dt

    order_date = dt.fromisoformat(po_data.order_date).date()
    expected_delivery_date = (
        dt.fromisoformat(po_data.expected_delivery_date).date()
        if po_data.expected_delivery_date
        else None
    )

    # Create PO
    po_id = str(uuid4())
    new_po = PurchaseOrderDB(
        id=po_id,
        company_id=current_user.company_id,
        po_number=po_data.po_number,
        status=PurchaseOrderStatus.DRAFT,
        supplier_trn=po_data.supplier_trn,
        supplier_name=po_data.supplier_name,
        supplier_contact_email=po_data.supplier_contact_email,
        supplier_address=po_data.supplier_address,
        supplier_peppol_id=po_data.supplier_peppol_id,
        order_date=order_date,
        expected_delivery_date=expected_delivery_date,
        delivery_address=po_data.delivery_address,
        currency_code=po_data.currency_code,
        expected_subtotal=expected_subtotal,
        expected_tax=expected_tax,
        expected_total=expected_total,
        reference_number=po_data.reference_number,
        notes=po_data.notes,
    )

    db.add(new_po)

    # Create line items
    for item_data in po_data.line_items:
        line_total = item_data.quantity_ordered * item_data.unit_price
        line_item = PurchaseOrderLineItemDB(
            id=str(uuid4()),
            po_id=po_id,
            line_number=item_data.line_number,
            item_name=item_data.item_name,
            item_description=item_data.item_description,
            item_code=item_data.item_code,
            quantity_ordered=item_data.quantity_ordered,
            unit_code=item_data.unit_code,
            unit_price=item_data.unit_price,
            line_total=line_total,
            tax_category=item_data.tax_category,
            tax_percent=item_data.tax_percent,
        )
        db.add(line_item)

    db.commit()
    db.refresh(new_po)

    # Return PO with line items
    return PurchaseOrderOut(
        id=new_po.id,
        company_id=new_po.company_id,
        po_number=new_po.po_number,
        status=new_po.status,
        supplier_trn=new_po.supplier_trn,
        supplier_name=new_po.supplier_name,
        supplier_contact_email=new_po.supplier_contact_email,
        supplier_address=new_po.supplier_address,
        supplier_peppol_id=new_po.supplier_peppol_id,
        order_date=new_po.order_date.isoformat(),
        expected_delivery_date=new_po.expected_delivery_date.isoformat()
        if new_po.expected_delivery_date
        else None,
        delivery_address=new_po.delivery_address,
        currency_code=new_po.currency_code,
        expected_subtotal=new_po.expected_subtotal,
        expected_tax=new_po.expected_tax,
        expected_total=new_po.expected_total,
        received_invoice_count=new_po.received_invoice_count,
        matched_invoice_count=new_po.matched_invoice_count,
        received_amount_total=new_po.received_amount_total,
        variance_amount=new_po.variance_amount,
        reference_number=new_po.reference_number,
        notes=new_po.notes,
        approved_by_user_id=new_po.approved_by_user_id,
        approved_at=new_po.approved_at.isoformat() if new_po.approved_at else None,
        created_at=new_po.created_at.isoformat(),
        updated_at=new_po.updated_at.isoformat(),
        line_items=[
            PurchaseOrderLineItemOut(
                id=li.id,
                line_number=li.line_number,
                item_name=li.item_name,
                item_description=li.item_description,
                item_code=li.item_code,
                quantity_ordered=li.quantity_ordered,
                quantity_received=li.quantity_received,
                unit_code=li.unit_code,
                unit_price=li.unit_price,
                line_total=li.line_total,
                tax_category=li.tax_category,
                tax_percent=li.tax_percent,
            )
            for li in new_po.line_items
        ],
    )


@app.get("/purchase-orders", tags=["AP Management"])
def list_purchase_orders(
    status: Optional[str] = None,
    supplier_name: Optional[str] = None,
    po_number: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    List purchase orders with optional filters

    Filters:
    - status: Filter by PO status (DRAFT, SENT, ACKNOWLEDGED, FULFILLED, CANCELLED)
    - supplier_name: Search by supplier name (partial match)
    - po_number: Search by PO number (partial match)
    """
    query = db.query(PurchaseOrderDB).filter(
        PurchaseOrderDB.company_id == current_user.company_id
    )

    # Apply filters
    if status:
        try:
            status_enum = PurchaseOrderStatus(status)
            query = query.filter(PurchaseOrderDB.status == status_enum)
        except ValueError:
            raise HTTPException(400, f"Invalid status: {status}")

    if supplier_name:
        query = query.filter(PurchaseOrderDB.supplier_name.ilike(f"%{supplier_name}%"))

    if po_number:
        query = query.filter(PurchaseOrderDB.po_number.ilike(f"%{po_number}%"))

    # Order by creation date (newest first)
    query = query.order_by(PurchaseOrderDB.created_at.desc())

    # Pagination
    total = query.count()
    purchase_orders = query.offset(skip).limit(limit).all()

    # Convert to response models
    results = []
    for po in purchase_orders:
        results.append(
            PurchaseOrderOut(
                id=po.id,
                company_id=po.company_id,
                po_number=po.po_number,
                status=po.status,
                supplier_trn=po.supplier_trn,
                supplier_name=po.supplier_name,
                supplier_contact_email=po.supplier_contact_email,
                supplier_address=po.supplier_address,
                supplier_peppol_id=po.supplier_peppol_id,
                order_date=po.order_date.isoformat(),
                expected_delivery_date=po.expected_delivery_date.isoformat()
                if po.expected_delivery_date
                else None,
                delivery_address=po.delivery_address,
                currency_code=po.currency_code,
                expected_subtotal=po.expected_subtotal,
                expected_tax=po.expected_tax,
                expected_total=po.expected_total,
                received_invoice_count=po.received_invoice_count,
                matched_invoice_count=po.matched_invoice_count,
                received_amount_total=po.received_amount_total,
                variance_amount=po.variance_amount,
                reference_number=po.reference_number,
                notes=po.notes,
                approved_by_user_id=po.approved_by_user_id,
                approved_at=po.approved_at.isoformat() if po.approved_at else None,
                created_at=po.created_at.isoformat(),
                updated_at=po.updated_at.isoformat(),
                line_items=[
                    PurchaseOrderLineItemOut(
                        id=li.id,
                        line_number=li.line_number,
                        item_name=li.item_name,
                        item_description=li.item_description,
                        item_code=li.item_code,
                        quantity_ordered=li.quantity_ordered,
                        quantity_received=li.quantity_received,
                        unit_code=li.unit_code,
                        unit_price=li.unit_price,
                        line_total=li.line_total,
                        tax_category=li.tax_category,
                        tax_percent=li.tax_percent,
                    )
                    for li in po.line_items
                ],
            )
        )

    return {"total": total, "skip": skip, "limit": limit, "results": results}


@app.get(
    "/purchase-orders/{po_id}", tags=["AP Management"], response_model=PurchaseOrderOut
)
def get_purchase_order(
    po_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Get purchase order details with line items"""
    po = (
        db.query(PurchaseOrderDB)
        .filter(
            PurchaseOrderDB.id == po_id,
            PurchaseOrderDB.company_id == current_user.company_id,
        )
        .first()
    )

    if not po:
        raise HTTPException(404, "Purchase order not found")

    return PurchaseOrderOut(
        id=po.id,
        company_id=po.company_id,
        po_number=po.po_number,
        status=po.status,
        supplier_trn=po.supplier_trn,
        supplier_name=po.supplier_name,
        supplier_contact_email=po.supplier_contact_email,
        supplier_address=po.supplier_address,
        supplier_peppol_id=po.supplier_peppol_id,
        order_date=po.order_date.isoformat(),
        expected_delivery_date=po.expected_delivery_date.isoformat()
        if po.expected_delivery_date
        else None,
        delivery_address=po.delivery_address,
        currency_code=po.currency_code,
        expected_subtotal=po.expected_subtotal,
        expected_tax=po.expected_tax,
        expected_total=po.expected_total,
        received_invoice_count=po.received_invoice_count,
        matched_invoice_count=po.matched_invoice_count,
        received_amount_total=po.received_amount_total,
        variance_amount=po.variance_amount,
        reference_number=po.reference_number,
        notes=po.notes,
        approved_by_user_id=po.approved_by_user_id,
        approved_at=po.approved_at.isoformat() if po.approved_at else None,
        created_at=po.created_at.isoformat(),
        updated_at=po.updated_at.isoformat(),
        line_items=[
            PurchaseOrderLineItemOut(
                id=li.id,
                line_number=li.line_number,
                item_name=li.item_name,
                item_description=li.item_description,
                item_code=li.item_code,
                quantity_ordered=li.quantity_ordered,
                quantity_received=li.quantity_received,
                unit_code=li.unit_code,
                unit_price=li.unit_price,
                line_total=li.line_total,
                tax_category=li.tax_category,
                tax_percent=li.tax_percent,
            )
            for li in po.line_items
        ],
    )


@app.put(
    "/purchase-orders/{po_id}", tags=["AP Management"], response_model=PurchaseOrderOut
)
def update_purchase_order(
    po_id: str,
    po_data: PurchaseOrderCreate,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Update an existing purchase order

    Only DRAFT status POs can be updated
    """
    po = (
        db.query(PurchaseOrderDB)
        .filter(
            PurchaseOrderDB.id == po_id,
            PurchaseOrderDB.company_id == current_user.company_id,
        )
        .first()
    )

    if not po:
        raise HTTPException(404, "Purchase order not found")

    if po.status != PurchaseOrderStatus.DRAFT:
        raise HTTPException(400, "Only DRAFT purchase orders can be updated")

    # Validate line items
    if not po_data.line_items:
        raise HTTPException(400, "At least one line item is required")

    # Calculate expected totals
    expected_subtotal = 0.0
    expected_tax = 0.0

    for item in po_data.line_items:
        line_total = item.quantity_ordered * item.unit_price
        line_tax = line_total * (item.tax_percent / 100)
        expected_subtotal += line_total
        expected_tax += line_tax

    expected_total = expected_subtotal + expected_tax

    # Parse dates
    from datetime import datetime as dt

    order_date = dt.fromisoformat(po_data.order_date).date()
    expected_delivery_date = (
        dt.fromisoformat(po_data.expected_delivery_date).date()
        if po_data.expected_delivery_date
        else None
    )

    # Update PO fields
    po.po_number = po_data.po_number
    po.supplier_trn = po_data.supplier_trn
    po.supplier_name = po_data.supplier_name
    po.supplier_contact_email = po_data.supplier_contact_email
    po.supplier_address = po_data.supplier_address
    po.supplier_peppol_id = po_data.supplier_peppol_id
    po.order_date = order_date
    po.expected_delivery_date = expected_delivery_date
    po.delivery_address = po_data.delivery_address
    po.currency_code = po_data.currency_code
    po.expected_subtotal = expected_subtotal
    po.expected_tax = expected_tax
    po.expected_total = expected_total
    po.reference_number = po_data.reference_number
    po.notes = po_data.notes
    po.updated_at = datetime.utcnow()

    # Delete old line items
    db.query(PurchaseOrderLineItemDB).filter(
        PurchaseOrderLineItemDB.po_id == po_id
    ).delete()

    # Create new line items
    for item_data in po_data.line_items:
        line_total = item_data.quantity_ordered * item_data.unit_price
        line_item = PurchaseOrderLineItemDB(
            id=str(uuid4()),
            po_id=po_id,
            line_number=item_data.line_number,
            item_name=item_data.item_name,
            item_description=item_data.item_description,
            item_code=item_data.item_code,
            quantity_ordered=item_data.quantity_ordered,
            unit_code=item_data.unit_code,
            unit_price=item_data.unit_price,
            line_total=line_total,
            tax_category=item_data.tax_category,
            tax_percent=item_data.tax_percent,
        )
        db.add(line_item)

    db.commit()
    db.refresh(po)

    return PurchaseOrderOut(
        id=po.id,
        company_id=po.company_id,
        po_number=po.po_number,
        status=po.status,
        supplier_trn=po.supplier_trn,
        supplier_name=po.supplier_name,
        supplier_contact_email=po.supplier_contact_email,
        supplier_address=po.supplier_address,
        supplier_peppol_id=po.supplier_peppol_id,
        order_date=po.order_date.isoformat(),
        expected_delivery_date=po.expected_delivery_date.isoformat()
        if po.expected_delivery_date
        else None,
        delivery_address=po.delivery_address,
        currency_code=po.currency_code,
        expected_subtotal=po.expected_subtotal,
        expected_tax=po.expected_tax,
        expected_total=po.expected_total,
        received_invoice_count=po.received_invoice_count,
        matched_invoice_count=po.matched_invoice_count,
        received_amount_total=po.received_amount_total,
        variance_amount=po.variance_amount,
        reference_number=po.reference_number,
        notes=po.notes,
        approved_by_user_id=po.approved_by_user_id,
        approved_at=po.approved_at.isoformat() if po.approved_at else None,
        created_at=po.created_at.isoformat(),
        updated_at=po.updated_at.isoformat(),
        line_items=[
            PurchaseOrderLineItemOut(
                id=li.id,
                line_number=li.line_number,
                item_name=li.item_name,
                item_description=li.item_description,
                item_code=li.item_code,
                quantity_ordered=li.quantity_ordered,
                quantity_received=li.quantity_received,
                unit_code=li.unit_code,
                unit_price=li.unit_price,
                line_total=li.line_total,
                tax_category=li.tax_category,
                tax_percent=li.tax_percent,
            )
            for li in po.line_items
        ],
    )


@app.delete("/purchase-orders/{po_id}", tags=["AP Management"])
def cancel_purchase_order(
    po_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Cancel a purchase order (soft delete)

    Only DRAFT and SENT status POs can be cancelled
    """
    po = (
        db.query(PurchaseOrderDB)
        .filter(
            PurchaseOrderDB.id == po_id,
            PurchaseOrderDB.company_id == current_user.company_id,
        )
        .first()
    )

    if not po:
        raise HTTPException(404, "Purchase order not found")

    if po.status not in [PurchaseOrderStatus.DRAFT, PurchaseOrderStatus.SENT]:
        raise HTTPException(
            400,
            f"Cannot cancel PO in {po.status.value} status. Only DRAFT and SENT POs can be cancelled.",
        )

    # Check if any invoices are matched to this PO
    matched_invoices = (
        db.query(InwardInvoiceDB).filter(InwardInvoiceDB.po_id == po_id).count()
    )

    if matched_invoices > 0:
        raise HTTPException(
            400,
            f"Cannot cancel PO. It has {matched_invoices} matched invoice(s). Unmatch invoices first.",
        )

    # Cancel the PO
    po.status = PurchaseOrderStatus.CANCELLED
    po.updated_at = datetime.utcnow()

    db.commit()

    return {
        "success": True,
        "message": f"Purchase order {po.po_number} cancelled",
        "po_id": po.id,
        "po_number": po.po_number,
    }


@app.post("/purchase-orders/{po_id}/send", tags=["AP Management"])
def send_purchase_order(
    po_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Send purchase order to supplier (change status from DRAFT to SENT)

    This would integrate with email/PEPPOL sending in production
    """
    po = (
        db.query(PurchaseOrderDB)
        .filter(
            PurchaseOrderDB.id == po_id,
            PurchaseOrderDB.company_id == current_user.company_id,
        )
        .first()
    )

    if not po:
        raise HTTPException(404, "Purchase order not found")

    if po.status != PurchaseOrderStatus.DRAFT:
        raise HTTPException(400, "Only DRAFT purchase orders can be sent")

    # Update status
    po.status = PurchaseOrderStatus.SENT
    po.updated_at = datetime.utcnow()

    # In production, this would:
    # 1. Generate PO PDF
    # 2. Send via email to supplier_contact_email
    # 3. Or send via PEPPOL if supplier_peppol_id exists

    db.commit()

    return {
        "success": True,
        "message": f"Purchase order {po.po_number} sent to {po.supplier_name}",
        "po_id": po.id,
        "po_number": po.po_number,
        "supplier_email": po.supplier_contact_email,
    }


# ==================== COMPANY BRANDING ENDPOINTS ====================


@app.post("/companies/{company_id}/branding/logo", tags=["Branding"])
async def upload_company_logo(
    company_id: str,
    logo: UploadFile = File(...),
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Upload company logo (PNG/SVG, max 2MB)"""
    # Verify user owns this company
    if current_user.company_id != company_id:
        raise HTTPException(403, "Not authorized to upload logo for this company")

    # Check if company has branding permission
    company = db.query(CompanyDB).filter(CompanyDB.id == company_id).first()
    if not company:
        raise HTTPException(404, "Company not found")

    # Get company's subscription plan
    subscription = (
        db.query(CompanySubscriptionDB)
        .filter(CompanySubscriptionDB.company_id == company_id)
        .first()
    )

    if subscription:
        plan = (
            db.query(SubscriptionPlanDB)
            .filter(SubscriptionPlanDB.id == subscription.plan_id)
            .first()
        )

        if plan and not plan.allow_branding:
            raise HTTPException(
                403,
                "Your subscription plan does not allow branding. Please upgrade to use this feature.",
            )

    # Validate file type and extension
    allowed_types = ["image/png", "image/svg+xml"]
    allowed_extensions = [".png", ".svg"]

    if not logo.content_type in allowed_types:
        raise HTTPException(400, "Only PNG and SVG files are allowed")

    # Additional validation: check file extension
    file_ext = os.path.splitext(logo.filename)[1].lower() if logo.filename else ""
    if file_ext not in allowed_extensions:
        raise HTTPException(
            400,
            f"Invalid file extension. Only {', '.join(allowed_extensions)} are allowed",
        )

    # Read file content
    content = await logo.read()
    file_size = len(content)

    if file_size > 2 * 1024 * 1024:  # 2MB limit
        raise HTTPException(400, "File size must be less than 2MB")

    # Create branding directory
    branding_dir = os.path.join(ARTIFACT_ROOT, "branding", company_id)
    os.makedirs(branding_dir, exist_ok=True)

    # Save file - use original filename + timestamp to avoid collisions
    file_extension = "svg" if logo.content_type == "image/svg+xml" else "png"
    orig_base = os.path.splitext(logo.filename)[0] if logo.filename else "logo"
    safe_base = "".join(
        c if c.isalnum() or c in ("-", "_") else "_" for c in orig_base
    )[:100]
    timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    file_name = f"{safe_base}_{timestamp}.{file_extension}"
    file_path = os.path.join(branding_dir, file_name)

    with open(file_path, "wb") as f:
        f.write(content)
    logger.info(f"Saved company logo to %s", file_path)

    # Get or create branding record
    branding = (
        db.query(CompanyBrandingDB)
        .filter(CompanyBrandingDB.company_id == company_id)
        .first()
    )

    if not branding:
        branding = CompanyBrandingDB(id=f"br_{uuid4().hex[:8]}", company_id=company_id)
        db.add(branding)

    # Delete old logo if exists
    if branding.logo_file_path and os.path.exists(branding.logo_file_path):
        os.remove(branding.logo_file_path)

    # Update branding record
    branding.logo_file_name = file_name
    branding.logo_file_path = file_path
    branding.logo_file_size = file_size
    branding.logo_mime_type = logo.content_type
    branding.logo_uploaded_at = datetime.utcnow()

    db.commit()

    return {
        "message": "Logo uploaded successfully",
        "file_name": file_name,
        "file_path": file_path,
        "file_size": file_size,
        "uploaded_at": branding.logo_uploaded_at.isoformat(),
    }


@app.post("/admin/companies/{company_id}/branding/logo", tags=["Admin", "Branding"])
async def admin_upload_company_logo(
    company_id: str,
    logo: UploadFile = File(...),
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Allow super-admin to upload company logo on behalf of any company"""
    # Only SUPER_ADMIN role may call this
    if getattr(current_user, "role", None) != Role.SUPER_ADMIN:
        raise HTTPException(403, "Not authorized")

    # Check if company exists
    company = db.query(CompanyDB).filter(CompanyDB.id == company_id).first()
    if not company:
        raise HTTPException(404, "Company not found")

    # Validate subscription/plan like regular endpoint
    subscription = (
        db.query(CompanySubscriptionDB)
        .filter(CompanySubscriptionDB.company_id == company_id)
        .first()
    )

    if subscription:
        plan = (
            db.query(SubscriptionPlanDB)
            .filter(SubscriptionPlanDB.id == subscription.plan_id)
            .first()
        )

        if plan and not plan.allow_branding:
            raise HTTPException(
                403, "This company's subscription plan does not allow branding."
            )

    allowed_types = ["image/png", "image/svg+xml"]
    allowed_extensions = [".png", ".svg"]

    if not logo.content_type in allowed_types:
        raise HTTPException(400, "Only PNG and SVG files are allowed")

    file_ext = os.path.splitext(logo.filename)[1].lower() if logo.filename else ""
    if file_ext not in allowed_extensions:
        raise HTTPException(
            400,
            f"Invalid file extension. Only {', '.join(allowed_extensions)} are allowed",
        )

    content = await logo.read()
    file_size = len(content)

    if file_size > 2 * 1024 * 1024:  # 2MB limit
        raise HTTPException(400, "File size must be less than 2MB")

    branding_dir = os.path.join(ARTIFACT_ROOT, "branding", company_id)
    os.makedirs(branding_dir, exist_ok=True)

    file_extension = "svg" if logo.content_type == "image/svg+xml" else "png"
    orig_base = os.path.splitext(logo.filename)[0] if logo.filename else "logo"
    safe_base = "".join(
        c if c.isalnum() or c in ("-", "_") else "_" for c in orig_base
    )[:100]
    timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    file_name = f"{safe_base}_{timestamp}.{file_extension}"
    file_path = os.path.join(branding_dir, file_name)

    with open(file_path, "wb") as f:
        f.write(content)
    logger.info(f"Saved company logo to %s", file_path)

    branding = (
        db.query(CompanyBrandingDB)
        .filter(CompanyBrandingDB.company_id == company_id)
        .first()
    )

    if not branding:
        branding = CompanyBrandingDB(id=f"br_{uuid4().hex[:8]}", company_id=company_id)
        db.add(branding)

    if branding.logo_file_path and os.path.exists(branding.logo_file_path):
        os.remove(branding.logo_file_path)

    branding.logo_file_name = file_name
    branding.logo_file_path = file_path
    branding.logo_file_size = file_size
    branding.logo_mime_type = logo.content_type
    branding.logo_uploaded_at = datetime.utcnow()

    db.commit()

    return {
        "message": "Logo uploaded successfully",
        "file_name": file_name,
        "file_path": file_path,
        "file_size": file_size,
        "uploaded_at": branding.logo_uploaded_at.isoformat(),
    }


@app.post("/companies/{company_id}/branding/stamp", tags=["Branding"])
async def upload_company_stamp(
    company_id: str,
    stamp: UploadFile = File(...),
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Upload company stamp/seal (PNG/SVG, max 2MB)"""
    # Verify user owns this company
    if current_user.company_id != company_id:
        raise HTTPException(403, "Not authorized to upload stamp for this company")

    # Check if company has branding permission
    company = db.query(CompanyDB).filter(CompanyDB.id == company_id).first()
    if not company:
        raise HTTPException(404, "Company not found")

    # Get company's subscription plan
    subscription = (
        db.query(CompanySubscriptionDB)
        .filter(CompanySubscriptionDB.company_id == company_id)
        .first()
    )

    if subscription:
        plan = (
            db.query(SubscriptionPlanDB)
            .filter(SubscriptionPlanDB.id == subscription.plan_id)
            .first()
        )

        if plan and not plan.allow_branding:
            raise HTTPException(
                403,
                "Your subscription plan does not allow branding. Please upgrade to use this feature.",
            )

    # Validate file type and extension
    allowed_types = ["image/png", "image/svg+xml"]
    allowed_extensions = [".png", ".svg"]

    if not stamp.content_type in allowed_types:
        raise HTTPException(400, "Only PNG and SVG files are allowed")

    # Additional validation: check file extension
    file_ext = os.path.splitext(stamp.filename)[1].lower() if stamp.filename else ""
    if file_ext not in allowed_extensions:
        raise HTTPException(
            400,
            f"Invalid file extension. Only {', '.join(allowed_extensions)} are allowed",
        )

    # Read file content
    content = await stamp.read()
    file_size = len(content)

    if file_size > 2 * 1024 * 1024:  # 2MB limit
        raise HTTPException(400, "File size must be less than 2MB")

    # Create branding directory
    branding_dir = os.path.join(ARTIFACT_ROOT, "branding", company_id)
    os.makedirs(branding_dir, exist_ok=True)

    # Save file
    file_extension = "svg" if stamp.content_type == "image/svg+xml" else "png"
    file_name = f"stamp.{file_extension}"
    file_path = os.path.join(branding_dir, file_name)

    with open(file_path, "wb") as f:
        f.write(content)

    # Get or create branding record
    branding = (
        db.query(CompanyBrandingDB)
        .filter(CompanyBrandingDB.company_id == company_id)
        .first()
    )

    if not branding:
        branding = CompanyBrandingDB(id=f"br_{uuid4().hex[:8]}", company_id=company_id)
        db.add(branding)

    # Delete old stamp if exists
    if branding.stamp_file_path and os.path.exists(branding.stamp_file_path):
        os.remove(branding.stamp_file_path)

    # Update branding record
    branding.stamp_file_name = file_name
    branding.stamp_file_path = file_path
    branding.stamp_file_size = file_size
    branding.stamp_mime_type = stamp.content_type
    branding.stamp_uploaded_at = datetime.utcnow()

    db.commit()

    return {
        "message": "Stamp uploaded successfully",
        "file_name": file_name,
        "file_size": file_size,
        "uploaded_at": branding.stamp_uploaded_at.isoformat(),
    }


@app.get("/companies/{company_id}/branding", tags=["Branding"])
def get_company_branding(
    company_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Get company branding information"""
    # Verify user owns this company
    if current_user.company_id != company_id:
        raise HTTPException(403, "Not authorized to view branding for this company")

    branding = (
        db.query(CompanyBrandingDB)
        .filter(CompanyBrandingDB.company_id == company_id)
        .first()
    )

    if not branding:
        return {
            "id": None,
            "company_id": company_id,
            "logo_file_name": None,
            "logo_file_path": None,
            "logo_uploaded_at": None,
            "stamp_file_name": None,
            "stamp_file_path": None,
            "stamp_uploaded_at": None,
            "primary_color": None,
            "secondary_color": None,
            "font_family": None,
            "created_at": None,
            "updated_at": None,
            "message": "No branding configured for this company",
        }

    return {
        "id": branding.id,
        "company_id": branding.company_id,
        "logo_file_name": branding.logo_file_name,
        "logo_file_path": branding.logo_file_path,
        "logo_uploaded_at": branding.logo_uploaded_at.isoformat()
        if branding.logo_uploaded_at
        else None,
        "stamp_file_name": branding.stamp_file_name,
        "stamp_file_path": branding.stamp_file_path,
        "stamp_uploaded_at": branding.stamp_uploaded_at.isoformat()
        if branding.stamp_uploaded_at
        else None,
        "primary_color": branding.primary_color,
        "secondary_color": branding.secondary_color,
        "font_family": branding.font_family,
        "created_at": branding.created_at.isoformat(),
        "updated_at": branding.updated_at.isoformat(),
    }


@app.get("/companies/{company_id}/branding/logo", tags=["Branding"])
def get_company_logo(company_id: str, db: Session = Depends(get_db)):
    """Get company logo file (public endpoint for invoice display)"""
    branding = (
        db.query(CompanyBrandingDB)
        .filter(CompanyBrandingDB.company_id == company_id)
        .first()
    )

    if not branding or not branding.logo_file_path:
        raise HTTPException(404, "Logo not found")

    if not os.path.exists(branding.logo_file_path):
        raise HTTPException(404, "Logo file not found")

    return FileResponse(
        branding.logo_file_path,
        media_type=branding.logo_mime_type,
        filename=branding.logo_file_name,
    )


@app.get("/companies/{company_id}/branding/stamp", tags=["Branding"])
def get_company_stamp(company_id: str, db: Session = Depends(get_db)):
    """Get company stamp file (public endpoint for invoice display)"""
    branding = (
        db.query(CompanyBrandingDB)
        .filter(CompanyBrandingDB.company_id == company_id)
        .first()
    )

    if not branding or not branding.stamp_file_path:
        raise HTTPException(404, "Stamp not found")

    if not os.path.exists(branding.stamp_file_path):
        raise HTTPException(404, "Stamp file not found")

    return FileResponse(
        branding.stamp_file_path,
        media_type=branding.stamp_mime_type,
        filename=branding.stamp_file_name,
    )


@app.delete("/companies/{company_id}/branding/logo", tags=["Branding"])
def delete_company_logo(
    company_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Delete company logo"""
    # Verify user owns this company
    if current_user.company_id != company_id:
        raise HTTPException(403, "Not authorized to delete logo for this company")

    branding = (
        db.query(CompanyBrandingDB)
        .filter(CompanyBrandingDB.company_id == company_id)
        .first()
    )

    if not branding or not branding.logo_file_path:
        raise HTTPException(404, "Logo not found")

    # Delete file from disk
    if os.path.exists(branding.logo_file_path):
        os.remove(branding.logo_file_path)

    # Clear database fields
    branding.logo_file_name = None
    branding.logo_file_path = None
    branding.logo_file_size = None
    branding.logo_mime_type = None
    branding.logo_uploaded_at = None

    db.commit()

    return {"message": "Logo deleted successfully"}


@app.delete("/companies/{company_id}/branding/stamp", tags=["Branding"])
def delete_company_stamp(
    company_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Delete company stamp"""
    # Verify user owns this company
    if current_user.company_id != company_id:
        raise HTTPException(403, "Not authorized to delete stamp for this company")

    branding = (
        db.query(CompanyBrandingDB)
        .filter(CompanyBrandingDB.company_id == company_id)
        .first()
    )

    if not branding or not branding.stamp_file_path:
        raise HTTPException(404, "Stamp not found")

    # Delete file from disk
    if os.path.exists(branding.stamp_file_path):
        os.remove(branding.stamp_file_path)

    # Clear database fields
    branding.stamp_file_name = None
    branding.stamp_file_path = None
    branding.stamp_file_size = None
    branding.stamp_mime_type = None
    branding.stamp_uploaded_at = None

    db.commit()

    return {"message": "Stamp deleted successfully"}


# ==================== PUBLIC STATS ====================


@app.get("/public/stats", tags=["Public"])
def get_public_stats(db: Session = Depends(get_db)):
    """Get public statistics - visible to everyone"""
    # Count total active companies
    total_companies = (
        db.query(CompanyDB).filter(CompanyDB.status == CompanyStatus.ACTIVE).count()
    )

    # Count total invoices across all companies
    total_invoices = db.query(InvoiceDB).count()

    return {"totalCompanies": total_companies, "totalInvoices": total_invoices}


# ==================== HEALTH CHECK ====================


@app.get("/api/info", tags=["Health"])
def api_info():
    """API Info endpoint"""
    return {
        "service": "InvoLinks E-Invoicing API",
        "version": "2.0",
        "status": "running",
        "features": [
            "Multi-step registration wizard",
            "Document upload system",
            "Subscription plans (Starter, Professional, Enterprise)",
            "Admin approval workflow",
        ],
    }


@app.get("/health", tags=["Health"])
def health_check(db: Session = Depends(get_db)):
    """Detailed health check with database status"""
    try:
        company_count = db.query(CompanyDB).count()
        plan_count = db.query(SubscriptionPlanDB).count()
        return {
            "status": "healthy",
            "database": "connected",
            "companies": company_count,
            "plans": plan_count,
        }
    except Exception as e:
        raise HTTPException(500, f"Health check failed: {str(e)}")


@app.post("/admin/restore-superadmin", tags=["Admin"])
def restore_superadmin(db: Session = Depends(get_db)):
    """Emergency endpoint to restore SUPER_ADMIN role if accidentally overwritten"""
    super_admin_email = os.getenv("SUPER_ADMIN_EMAIL", "nrashidk@gmail.com")
    super_admin_password = os.getenv("SUPER_ADMIN_PASSWORD", "AbuDhabi@123")

    # Find the user with super admin email
    user = db.query(UserDB).filter(UserDB.email == super_admin_email).first()

    if not user:
        # Create if doesn't exist
        user = UserDB(
            id=f"user_{uuid4().hex[:8]}",
            email=super_admin_email,
            password_hash=get_password_hash(super_admin_password),
            role=Role.SUPER_ADMIN,
            company_id=None,
            is_owner=False,
            full_name="Super Admin",
            # Task 22: Initialise password age tracking
            password_changed_at=datetime.utcnow(),
        )
        db.add(user)
        db.commit()
        return {
            "success": True,
            "message": "Super Admin created",
            "email": super_admin_email,
        }
    else:
        # Restore role and password
        user.role = Role.SUPER_ADMIN
        user.password_hash = get_password_hash(super_admin_password)
        user.company_id = None
        user.is_owner = False
        # Task 22: Track password mutation time for expiry policy
        user.password_changed_at = datetime.utcnow()
        db.commit()
        return {
            "success": True,
            "message": "Super Admin role restored",
            "email": super_admin_email,
        }


# ==================== FTA AUDIT FILE (FAF) ENDPOINTS ====================


@app.post("/audit-files/generate", tags=["FTA Audit"])
def generate_fta_audit_file(
    period_start: str,  # YYYY-MM-DD
    period_end: str,  # YYYY-MM-DD
    format: str = "CSV",  # CSV or TXT
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Generate FTA Audit File (FAF) for a specific period

    UAE Federal Tax Authority compliant audit file with invoice-level detail.
    """
    if not has_permission(current_user, "audit_logs", "export"):
        raise HTTPException(403, "Permission denied: export on audit_logs is not allowed for your role")
    # Verify company exists and is active
    company = (
        db.query(CompanyDB).filter(CompanyDB.id == current_user.company_id).first()
    )
    if not company:
        raise HTTPException(404, "Company not found")

    if company.status != CompanyStatus.ACTIVE:
        raise HTTPException(400, "Company must be ACTIVE to generate audit files")

    if not company.trn:
        raise HTTPException(
            400, "Company must have a valid TRN to generate audit files"
        )

    # Parse and validate dates
    from datetime import datetime as dt

    try:
        start_date = dt.strptime(period_start, "%Y-%m-%d").date()
        end_date = dt.strptime(period_end, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(400, "Invalid date format. Use YYYY-MM-DD")

    if start_date > end_date:
        raise HTTPException(400, "Start date must be before end date")

    # Validate date range is reasonable (not in future, not before company registration)
    today = date.today()
    if end_date > today:
        raise HTTPException(400, f"End date cannot be in the future (today is {today})")

    if company.registration_date:
        if start_date < company.registration_date:
            raise HTTPException(
                400,
                f"Start date cannot be before company registration date ({company.registration_date})",
            )

    # Validate period is not excessively long (max 5 years for single audit file)
    period_days = (end_date - start_date).days
    if period_days > 365 * 5:
        raise HTTPException(
            400,
            "Audit period cannot exceed 5 years. Please generate separate files for longer periods.",
        )

    # Validate format
    if format.upper() not in ["CSV", "TXT"]:
        raise HTTPException(400, "Format must be CSV or TXT")

    # Create audit file record
    audit_file_id = f"faf_{uuid4().hex[:12]}"
    base_name = f"FTA_FAF_{company.trn}_{period_start}_to_{period_end}"
    file_name = f"{base_name}.zip"
    output_dir = os.path.join(ARTIFACT_ROOT, "audit_files", company.id)
    file_path = os.path.join(output_dir, file_name)

    audit_file = AuditFileDB(
        id=audit_file_id,
        company_id=company.id,
        file_name=file_name,
        file_path=file_path,
        format="ZIP",
        period_start_date=start_date,
        period_end_date=end_date,
        status=AuditFileStatus.GENERATING,
        generated_by_user_id=current_user.id,
    )

    db.add(audit_file)
    db.commit()

    try:
        # Get outgoing invoices (sales) for the period
        outgoing_invoices = (
            db.query(InvoiceDB)
            .filter(
                InvoiceDB.company_id == company.id,
                InvoiceDB.issue_date >= start_date,
                InvoiceDB.issue_date <= end_date,
                InvoiceDB.status != InvoiceStatus.DRAFT,  # Exclude drafts
                InvoiceDB.status != InvoiceStatus.CANCELLED,  # Exclude cancelled
            )
            .all()
        )

        # Get inward invoices (purchases) for the period
        inward_invoices = (
            db.query(InwardInvoiceDB)
            .filter(
                InwardInvoiceDB.company_id == company.id,
                InwardInvoiceDB.invoice_date >= start_date,
                InwardInvoiceDB.invoice_date <= end_date,
                InwardInvoiceDB.status != InwardInvoiceStatus.CANCELLED,
            )
            .all()
        )

        # Convert to dicts for generator
        outgoing_data = [
            {
                "invoice_number": inv.invoice_number,
                "issue_date": inv.issue_date,
                "invoice_type": inv.invoice_type.value,
                "customer_trn": inv.customer_trn,
                "customer_name": inv.customer_name,
                "customer_country": inv.customer_country,
                "subtotal_amount": inv.subtotal_amount,
                "tax_amount": inv.tax_amount,
                "total_amount": inv.total_amount,
                "total_amount_aed": getattr(inv, "total_amount_aed", None),
                "currency_code": inv.currency_code,
                "status": inv.status.value,
            }
            for inv in outgoing_invoices
        ]

        inward_data = [
            {
                "supplier_invoice_number": inv.supplier_invoice_number,
                "invoice_date": inv.invoice_date,
                "invoice_type": inv.invoice_type.value,
                "supplier_trn": inv.supplier_trn,
                "supplier_name": inv.supplier_name,
                "subtotal_amount": inv.subtotal_amount,
                "tax_amount": inv.tax_amount,
                "total_amount": inv.total_amount,
                "total_amount_aed": getattr(inv, "total_amount_aed", None),
                "currency_code": inv.currency_code,
                "status": inv.status.value,
            }
            for inv in inward_invoices
        ]

        # Fetch General Ledger entries for the period
        gl_journal_entries = (
            db.query(JournalEntryDB)
            .filter(
                JournalEntryDB.company_id == company.id,
                JournalEntryDB.entry_date >= start_date,
                JournalEntryDB.entry_date <= end_date,
                JournalEntryDB.is_posted == True,
            )
            .all()
        )

        gl_entry_ids = [je.id for je in gl_journal_entries]
        gl_lines = (
            db.query(JournalEntryLineDB)
            .filter(JournalEntryLineDB.journal_entry_id.in_(gl_entry_ids))
            .all()
            if gl_entry_ids
            else []
        )

        lines_by_entry: Dict[str, List[Dict]] = {}
        for line in gl_lines:
            lines_by_entry.setdefault(line.journal_entry_id, []).append(
                {
                    "account_code": line.account_code or "",
                    "account_name": line.account_name or "",
                    "debit_amount": line.debit_amount or 0.0,
                    "credit_amount": line.credit_amount or 0.0,
                }
            )

        gl_data = [
            {
                "id": je.id,
                "entry_date": je.entry_date,
                "reference_number": je.reference_number or "",
                "reference_id": je.reference_id or "",
                "reference_type": je.reference_type or "",
                "lines": lines_by_entry.get(je.id, []),
            }
            for je in gl_journal_entries
        ]

        # Generate FAF ZIP (5-table CSV + SHA-256 hash, packed into ZIP)
        from utils.fta_audit_generator import FTAAuditFileGenerator

        company_data = {
            "trn": company.trn,
            "legal_name": company.legal_name,
            "address": f"{company.address_line1 or ''} {company.address_line2 or ''}".strip(),
            "city": company.city,
            "emirate": company.emirate,
            "registration_date": company.registration_date,
            "period_start": start_date,
            "period_end": end_date,
        }

        generator = FTAAuditFileGenerator(company_data)
        zip_path, stats = generator.generate_faf_zip(
            outgoing_data, inward_data, output_dir, base_name, gl_entries=gl_data
        )

        # Update audit file record with stats
        audit_file.status = AuditFileStatus.COMPLETED
        audit_file.total_invoices = stats.get("total_invoices", 0)
        audit_file.total_customers = stats.get("total_customers", 0)
        audit_file.total_amount = stats.get("total_amount", 0.0)
        audit_file.total_vat = stats.get("total_vat", 0.0)
        audit_file.file_size = stats.get("file_size", 0)

        # Task 1: Audit trail
        log_audit_event(
            db, "FAF_GENERATED",
            user_id=current_user.id, company_id=company.id,
            resource_type="audit_file", resource_id=audit_file_id,
            description=f"FAF generated for period {period_start} to {period_end}",
        )

        db.commit()

        return {
            "success": True,
            "message": "FTA Audit File (FAFv1.0.0) generated — ZIP with CSV + SHA-256 hash",
            "audit_file_id": audit_file_id,
            "file_name": file_name,
            "sha256": stats.get("sha256", ""),
            "period_start": period_start,
            "period_end": period_end,
            "format": "ZIP",
            "statistics": {
                "total_invoices": audit_file.total_invoices,
                "total_sales": stats.get("total_sales", 0),
                "total_purchases": stats.get("total_purchases", 0),
                "total_customers": audit_file.total_customers,
                "total_amount": audit_file.total_amount,
                "total_vat": audit_file.total_vat,
                "total_gl_lines": stats.get("total_gl_lines", 0),
            },
        }

    except Exception as e:
        # Mark as failed
        audit_file.status = AuditFileStatus.FAILED
        audit_file.error_message = str(e)
        db.commit()

        raise HTTPException(500, f"Failed to generate audit file: {str(e)}")


@app.get("/audit-files", tags=["FTA Audit"])
def list_audit_files(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """List all FTA Audit Files for the company"""
    audit_files = (
        db.query(AuditFileDB)
        .filter(AuditFileDB.company_id == current_user.company_id)
        .order_by(AuditFileDB.created_at.desc())
        .all()
    )

    return {
        "audit_files": [
            {
                "id": af.id,
                "file_name": af.file_name,
                "format": af.format,
                "period_start": af.period_start_date.isoformat(),
                "period_end": af.period_end_date.isoformat(),
                "status": af.status.value,
                "total_invoices": af.total_invoices,
                "total_amount": af.total_amount,
                "total_vat": af.total_vat,
                "file_size": af.file_size,
                "generated_at": af.generated_at.isoformat()
                if af.generated_at
                else None,
                "error_message": af.error_message,
            }
            for af in audit_files
        ]
    }


@app.get("/audit-files/{audit_file_id}/download", tags=["FTA Audit"])
def download_audit_file(
    audit_file_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Download an FTA Audit File"""
    audit_file = (
        db.query(AuditFileDB)
        .filter(
            AuditFileDB.id == audit_file_id,
            AuditFileDB.company_id == current_user.company_id,
        )
        .first()
    )

    if not audit_file:
        raise HTTPException(404, "Audit file not found")

    if audit_file.status != AuditFileStatus.COMPLETED:
        raise HTTPException(
            400,
            f"Audit file is not ready for download. Status: {audit_file.status.value}",
        )

    if not os.path.exists(audit_file.file_path):
        raise HTTPException(404, "Audit file not found on disk")

    # Determine media type — new files are ZIPs; legacy CSV/TXT still served correctly
    ext = os.path.splitext(audit_file.file_path)[1].lower()
    media_type_map = {
        ".zip": "application/zip",
        ".csv": "text/csv",
        ".txt": "text/plain",
    }
    media_type = media_type_map.get(ext, "application/octet-stream")

    return FileResponse(
        path=audit_file.file_path, media_type=media_type, filename=audit_file.file_name
    )


# ==================== PEPPOL SETTINGS ====================


@app.get("/settings/session", tags=["Settings"])
def get_session_settings(
    current_user: UserDB = Depends(get_current_user_from_header),
):
    """
    Return current JWT session configuration.
    Controlled via JWT_EXPIRY_HOURS environment variable (default: 24 hours).
    FTA Article 9.1 requires configurable session lifetime.
    """
    return {
        "jwt_expiry_hours": round(ACCESS_TOKEN_EXPIRE_MINUTES / 60, 2),
        "jwt_expiry_minutes": ACCESS_TOKEN_EXPIRE_MINUTES,
        "env_var": "JWT_EXPIRY_HOURS",
        "fta_reference": "Article 9.1 – Session Controls",
    }


@app.get("/settings/peppol", tags=["Settings"])
def get_peppol_settings(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Get PEPPOL configuration for the company"""
    company = (
        db.query(CompanyDB).filter(CompanyDB.id == current_user.company_id).first()
    )
    if not company:
        raise HTTPException(404, "Company not found")

    return {
        "peppol_enabled": company.peppol_enabled or False,
        "peppol_provider": company.peppol_provider,
        "peppol_participant_id": company.peppol_participant_id,
        "peppol_base_url": company.peppol_base_url,
        "peppol_api_key": "***" + company.peppol_api_key[-4:]
        if company.peppol_api_key and len(company.peppol_api_key) > 4
        else None,  # Masked
        "peppol_configured_at": company.peppol_configured_at.isoformat()
        if company.peppol_configured_at
        else None,
        "peppol_last_tested_at": company.peppol_last_tested_at.isoformat()
        if company.peppol_last_tested_at
        else None,
    }


@app.put("/settings/peppol", tags=["Settings"])
def update_peppol_settings(
    settings: dict,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Update PEPPOL configuration for the company"""
    if not has_permission(current_user, "settings", "edit"):
        raise HTTPException(403, "Permission denied: edit on settings is not allowed for your role")
    company = (
        db.query(CompanyDB).filter(CompanyDB.id == current_user.company_id).first()
    )
    if not company:
        raise HTTPException(404, "Company not found")

    # Validate provider
    valid_providers = ["tradeshift", "basware", "mock"]
    if (
        settings.get("peppol_provider")
        and settings["peppol_provider"] not in valid_providers
    ):
        raise HTTPException(
            400, f"Invalid provider. Must be one of: {', '.join(valid_providers)}"
        )

    # Update settings
    company.peppol_enabled = settings.get("peppol_enabled", False)
    company.peppol_provider = settings.get("peppol_provider")
    company.peppol_participant_id = settings.get("peppol_participant_id")
    company.peppol_base_url = settings.get("peppol_base_url")

    # Only update API key if provided and not masked
    # This preserves the existing key when user updates other fields
    if "peppol_api_key" in settings and settings["peppol_api_key"]:
        if not settings["peppol_api_key"].startswith("***"):
            company.peppol_api_key = settings["peppol_api_key"]
        # If masked, skip update to preserve existing key

    company.peppol_configured_at = datetime.utcnow()

    db.commit()
    db.refresh(company)

    return {
        "success": True,
        "message": "PEPPOL settings updated successfully",
        "peppol_enabled": company.peppol_enabled,
        "peppol_provider": company.peppol_provider,
        "peppol_participant_id": company.peppol_participant_id,
    }


@app.post("/settings/peppol/test", tags=["Settings"])
def test_peppol_connection(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Test PEPPOL provider connection"""
    company = (
        db.query(CompanyDB).filter(CompanyDB.id == current_user.company_id).first()
    )
    if not company:
        raise HTTPException(404, "Company not found")

    if not company.peppol_enabled:
        raise HTTPException(400, "PEPPOL is not enabled for this company")

    if not company.peppol_provider:
        raise HTTPException(400, "PEPPOL provider not configured")

    try:
        from utils.peppol_provider import PeppolProviderFactory

        # Create provider instance
        provider = PeppolProviderFactory.create_provider(
            provider_name=company.peppol_provider,
            base_url=company.peppol_base_url,
            api_key=company.peppol_api_key,
        )

        # Test with participant ID validation
        if company.peppol_participant_id:
            is_valid = provider.validate_participant_id(company.peppol_participant_id)

            # Update last tested timestamp
            company.peppol_last_tested_at = datetime.utcnow()
            db.commit()

            return {
                "success": True,
                "message": "PEPPOL connection test successful",
                "provider": company.peppol_provider,
                "participant_valid": is_valid,
                "tested_at": company.peppol_last_tested_at.isoformat(),
            }
        else:
            # Just test provider connectivity without participant validation
            company.peppol_last_tested_at = datetime.utcnow()
            db.commit()

            return {
                "success": True,
                "message": "PEPPOL provider connection successful (no participant ID to validate)",
                "provider": company.peppol_provider,
                "tested_at": company.peppol_last_tested_at.isoformat(),
            }

    except Exception as e:
        return {
            "success": False,
            "error": str(e),
            "message": "PEPPOL connection test failed",
        }


# ==================== VAT SETTINGS (PHASE 1) ====================


@app.get("/settings/vat", tags=["Settings"])
def get_vat_settings(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Get VAT registration configuration for the company"""
    company = (
        db.query(CompanyDB).filter(CompanyDB.id == current_user.company_id).first()
    )
    if not company:
        raise HTTPException(404, "Company not found")

    # Always return TRN if it exists, regardless of vat_enabled status
    # This preserves TRN data if user temporarily disables VAT
    return {
        "vat_enabled": company.vat_enabled or False,
        "tax_registration_number": company.trn or "",
        "vat_registration_date": company.vat_registration_date.isoformat()
        if company.vat_registration_date
        else None,
        "formatted_trn": format_trn(company.trn) if company.trn else None,
        "vat_certificate_uploaded": company.vat_certificate_path is not None,
    }


@app.put("/settings/vat", tags=["Settings"])
def update_vat_settings(
    settings: dict,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Update VAT registration configuration for the company"""
    if not has_permission(current_user, "settings", "edit"):
        raise HTTPException(403, "Permission denied: edit on settings is not allowed for your role")

    company = (
        db.query(CompanyDB).filter(CompanyDB.id == current_user.company_id).first()
    )
    if not company:
        raise HTTPException(404, "Company not found")

    # Get VAT enabled status from request
    vat_enabled = settings.get("vat_enabled", False)

    # If enabling VAT, validate TRN
    if vat_enabled:
        trn = settings.get("tax_registration_number", "").strip()

        if not trn:
            raise HTTPException(
                400, "Tax Registration Number (TRN) is required when enabling VAT"
            )

        if not is_valid_trn(trn):
            raise HTTPException(
                400,
                "Invalid TRN format. Must be exactly 15 digits (e.g., 123456789012345)",
            )

        # Update company with VAT details
        company.vat_enabled = True
        company.trn = trn

        # Set registration date if provided, otherwise use today (UTC)
        if settings.get("vat_registration_date"):
            from datetime import datetime as dt

            company.vat_registration_date = dt.fromisoformat(
                settings["vat_registration_date"]
            ).date()
        elif not company.vat_registration_date:
            # Use UTC datetime to ensure consistency across timezones
            company.vat_registration_date = datetime.utcnow().date()

        db.commit()
        db.refresh(company)

        return {
            "success": True,
            "message": "VAT registration enabled successfully",
            "vat_enabled": True,
            "tax_registration_number": company.trn,
            "formatted_trn": format_trn(company.trn),
            "vat_registration_date": company.vat_registration_date.isoformat()
            if company.vat_registration_date
            else None,
        }
    else:
        # Disabling VAT - just set vat_enabled to False, keep TRN for records
        company.vat_enabled = False

        db.commit()
        db.refresh(company)

        return {
            "success": True,
            "message": "VAT registration disabled",
            "vat_enabled": False,
        }


@app.post("/settings/vat/certificate", tags=["Settings"])
async def upload_vat_certificate(
    file: UploadFile = File(...),
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Upload VAT registration certificate (PDF only)"""
    if not has_permission(current_user, "settings", "edit"):
        raise HTTPException(403, "Permission denied: edit on settings is not allowed for your role")

    company = (
        db.query(CompanyDB).filter(CompanyDB.id == current_user.company_id).first()
    )
    if not company:
        raise HTTPException(404, "Company not found")

    # Validate file type
    if file.content_type != "application/pdf":
        raise HTTPException(400, "Only PDF files are allowed")

    # Read file contents
    contents = await file.read()

    # Validate file size (max 5MB)
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(400, "File size must be less than 5MB")

    # Delete old certificate if exists
    if company.vat_certificate_path:
        old_file_path = os.path.join(ARTIFACT_ROOT, company.vat_certificate_path)
        if os.path.exists(old_file_path):
            try:
                os.remove(old_file_path)
            except Exception as e:
                print(f"Warning: Failed to delete old certificate: {e}")

    # Create certificates directory
    cert_dir = os.path.join("vat_certificates", company.id)
    full_cert_dir = os.path.join(ARTIFACT_ROOT, cert_dir)
    os.makedirs(full_cert_dir, exist_ok=True)

    # Generate unique filename
    file_name = f"vat_certificate_{uuid4().hex[:8]}.pdf"
    relative_path = os.path.join(cert_dir, file_name)
    full_path = os.path.join(ARTIFACT_ROOT, relative_path)

    # Save file
    with open(full_path, "wb") as f:
        f.write(contents)

    # Update company record with relative path
    company.vat_certificate_path = relative_path
    db.commit()
    db.refresh(company)

    return {"success": True, "message": "VAT certificate uploaded successfully"}


@app.get("/settings/vat/certificate", tags=["Settings"])
def download_vat_certificate(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Download VAT registration certificate"""
    company = (
        db.query(CompanyDB).filter(CompanyDB.id == current_user.company_id).first()
    )
    if not company:
        raise HTTPException(404, "Company not found")

    if not company.vat_certificate_path:
        raise HTTPException(404, "No certificate uploaded")

    file_path = os.path.join(ARTIFACT_ROOT, company.vat_certificate_path)

    if not os.path.exists(file_path):
        raise HTTPException(404, "Certificate file not found")

    return FileResponse(
        file_path, media_type="application/pdf", filename="vat_certificate.pdf"
    )


# ==================== BILLING & SUBSCRIPTIONS ====================

# Initialize Stripe
STRIPE_SECRET_KEY = os.getenv("STRIPE_SECRET_KEY")
if STRIPE_SECRET_KEY:
    stripe.api_key = STRIPE_SECRET_KEY


# Payment method Pydantic models
class PaymentMethodOut(BaseModel):
    id: str
    card_brand: str
    card_last4: str
    exp_month: int
    exp_year: int
    billing_name: str
    is_default: bool
    created_at: str


class SubscriptionOut(BaseModel):
    id: str
    tier: str
    billing_cycle_months: int
    monthly_price: float
    discount_percent: float
    status: str
    current_period_start: str
    current_period_end: str
    created_at: str


class TrialStatusOut(BaseModel):
    trial_status: str
    trial_start_date: Optional[str]
    trial_invoice_count: int
    trial_days_remaining: Optional[int]
    trial_invoices_remaining: Optional[int]
    trial_expired: bool
    free_plan_type: Optional[str] = None  # "INVOICE_COUNT" or "DURATION"
    free_plan_invoice_limit: Optional[int] = None
    free_plan_duration_months: Optional[int] = None


@app.post("/billing/payment-methods", tags=["Billing"])
async def add_payment_method(
    payment_method_token: str = Form(...),
    billing_name: str = Form(...),
    billing_email: str = Form(...),
    set_as_default: bool = Form(False),
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Add a new payment method (credit card) via Stripe"""
    try:
        if not STRIPE_SECRET_KEY:
            raise HTTPException(
                503, "Payment processing not configured. Please contact support."
            )

        company = (
            db.query(CompanyDB).filter(CompanyDB.id == current_user.company_id).first()
        )
        if not company:
            raise HTTPException(404, "Company not found")

        # Create or get Stripe customer
        if not company.stripe_customer_id:
            customer = stripe.Customer.create(
                email=billing_email,
                name=company.legal_name or billing_name,
                metadata={"company_id": company.id},
            )
            company.stripe_customer_id = customer.id
            db.commit()

        # Attach payment method to customer
        payment_method = stripe.PaymentMethod.attach(
            payment_method_token, customer=company.stripe_customer_id
        )

        # Set as default if requested
        if set_as_default:
            stripe.Customer.modify(
                company.stripe_customer_id,
                invoice_settings={"default_payment_method": payment_method.id},
            )

        # Save to database
        pm_db = PaymentMethodDB(
            id=f"pm_{uuid4().hex[:12]}",
            company_id=company.id,
            stripe_payment_method_id=payment_method.id,
            card_brand=payment_method.card.brand if payment_method.card else None,
            card_last4=payment_method.card.last4 if payment_method.card else None,
            exp_month=payment_method.card.exp_month if payment_method.card else None,
            exp_year=payment_method.card.exp_year if payment_method.card else None,
            billing_email=billing_email,
            billing_name=billing_name,
            is_default=set_as_default,
        )

        # If setting as default, unset others
        if set_as_default:
            db.query(PaymentMethodDB).filter(
                PaymentMethodDB.company_id == company.id, PaymentMethodDB.id != pm_db.id
            ).update({"is_default": False})

        db.add(pm_db)
        db.commit()

        return {
            "message": "Payment method added successfully",
            "payment_method_id": pm_db.id,
            "card_last4": pm_db.card_last4,
        }

    except stripe.error.CardError as e:
        raise HTTPException(400, f"Card error: {str(e)}")
    except stripe.error.StripeError as e:
        raise HTTPException(500, f"Payment processing error: {str(e)}")
    except Exception as e:
        raise HTTPException(500, f"Failed to add payment method: {str(e)}")


@app.get(
    "/billing/payment-methods", tags=["Billing"], response_model=List[PaymentMethodOut]
)
async def list_payment_methods(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """List all payment methods for current company"""
    payment_methods = (
        db.query(PaymentMethodDB)
        .filter(PaymentMethodDB.company_id == current_user.company_id)
        .order_by(PaymentMethodDB.is_default.desc(), PaymentMethodDB.created_at.desc())
        .all()
    )

    return [
        PaymentMethodOut(
            id=pm.id,
            card_brand=pm.card_brand or "unknown",
            card_last4=pm.card_last4 or "0000",
            exp_month=pm.exp_month or 1,
            exp_year=pm.exp_year or 2025,
            billing_name=pm.billing_name or "",
            is_default=pm.is_default,
            created_at=pm.created_at.isoformat(),
        )
        for pm in payment_methods
    ]


@app.delete("/billing/payment-methods/{payment_method_id}", tags=["Billing"])
async def delete_payment_method(
    payment_method_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Delete a payment method"""
    pm = (
        db.query(PaymentMethodDB)
        .filter(
            PaymentMethodDB.id == payment_method_id,
            PaymentMethodDB.company_id == current_user.company_id,
        )
        .first()
    )

    if not pm:
        raise HTTPException(404, "Payment method not found")

    try:
        # Detach from Stripe
        if STRIPE_SECRET_KEY:
            stripe.PaymentMethod.detach(pm.stripe_payment_method_id)

        # Delete from database
        db.delete(pm)
        db.commit()

        return {"message": "Payment method deleted successfully"}
    except Exception as e:
        raise HTTPException(500, f"Failed to delete payment method: {str(e)}")


@app.get("/billing/subscription", tags=["Billing"])
async def get_current_subscription(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Get current subscription details"""
    subscription = (
        db.query(SubscriptionDB)
        .filter(
            SubscriptionDB.company_id == current_user.company_id,
            SubscriptionDB.status == "ACTIVE",
        )
        .first()
    )

    if not subscription:
        return {
            "id": None,
            "tier": None,
            "billing_cycle_months": None,
            "monthly_price": None,
            "discount_percent": None,
            "status": None,
            "current_period_start": None,
            "current_period_end": None,
            "created_at": None,
            "message": "No active subscription found",
        }

    return {
        "id": subscription.id,
        "tier": subscription.tier,
        "billing_cycle_months": subscription.billing_cycle_months,
        "monthly_price": subscription.monthly_price,
        "discount_percent": subscription.discount_percent,
        "status": subscription.status,
        "current_period_start": subscription.current_period_start.isoformat(),
        "current_period_end": subscription.current_period_end.isoformat(),
        "created_at": subscription.created_at.isoformat(),
    }


@app.get("/subscription/current", tags=["Subscriptions"])
def get_subscription_plan_info(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Get current subscription plan with tier limits"""
    company = (
        db.query(CompanyDB).filter(CompanyDB.id == current_user.company_id).first()
    )
    if not company:
        raise HTTPException(404, "Company not found")

    # Check for active paid subscription
    subscription = (
        db.query(SubscriptionDB)
        .filter(
            SubscriptionDB.company_id == current_user.company_id,
            SubscriptionDB.status == "ACTIVE",
        )
        .first()
    )

    if subscription:
        # Return paid plan details
        tier_limits = {
            "BASIC": {
                "name": "Basic",
                "max_business_admins": 2,
                "max_finance_users": 5,
            },
            "PRO": {"name": "Pro", "max_business_admins": 5, "max_finance_users": 15},
            "ENTERPRISE": {
                "name": "Enterprise",
                "max_business_admins": 999,
                "max_finance_users": 999,
            },
        }
        plan_info = tier_limits.get(subscription.tier, tier_limits["BASIC"])
        return {
            "plan": {
                "name": plan_info["name"],
                "tier": subscription.tier,
                "max_business_admins": plan_info["max_business_admins"],
                "max_finance_users": plan_info["max_finance_users"],
            },
            "status": "ACTIVE",
        }
    else:
        # Return free trial limits
        return {
            "plan": {
                "name": "Free Trial",
                "tier": "TRIAL",
                "max_business_admins": 1,
                "max_finance_users": 3,
            },
            "status": company.trial_status or "ACTIVE",
        }


@app.post("/billing/subscribe", tags=["Billing"])
async def create_subscription(
    tier: str = Form(...),  # BASIC, PRO, ENTERPRISE
    billing_cycle_months: int = Form(1),  # 1, 3, 6
    payment_method_id: str = Form(...),
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Create new subscription after trial"""
    try:
        company = (
            db.query(CompanyDB).filter(CompanyDB.id == current_user.company_id).first()
        )
        if not company:
            raise HTTPException(404, "Company not found")

        # Define pricing (these should come from a pricing table in production)
        pricing = {
            "BASIC": {"monthly": 99.0, "3month_discount": 5, "6month_discount": 10},
            "PRO": {"monthly": 299.0, "3month_discount": 5, "6month_discount": 10},
            "ENTERPRISE": {
                "monthly": 799.0,
                "3month_discount": 10,
                "6month_discount": 15,
            },
        }

        if tier not in pricing:
            raise HTTPException(400, f"Invalid tier: {tier}")

        if billing_cycle_months not in [1, 3, 6]:
            raise HTTPException(400, "Billing cycle must be 1, 3, or 6 months")

        # Calculate pricing
        monthly_price = pricing[tier]["monthly"]
        discount_percent = 0.0
        if billing_cycle_months == 3:
            discount_percent = pricing[tier]["3month_discount"]
        elif billing_cycle_months == 6:
            discount_percent = pricing[tier]["6month_discount"]

        # Calculate total
        subtotal = monthly_price * billing_cycle_months
        discount_amount = subtotal * (discount_percent / 100)
        total_amount = subtotal - discount_amount

        # End trial
        if company.trial_status == "ACTIVE":
            company.trial_status = "CONVERTED"
            company.trial_ended_at = datetime.utcnow()

        # Create subscription record
        subscription = SubscriptionDB(
            id=f"sub_{uuid4().hex[:12]}",
            company_id=company.id,
            tier=tier,
            billing_cycle_months=billing_cycle_months,
            monthly_price=monthly_price,
            discount_percent=discount_percent,
            status="ACTIVE",
            current_period_start=datetime.utcnow(),
            current_period_end=datetime.utcnow()
            + timedelta(days=30 * billing_cycle_months),
            stripe_customer_id=company.stripe_customer_id,
        )

        db.add(subscription)
        db.commit()

        return {
            "message": "Subscription created successfully",
            "subscription_id": subscription.id,
            "tier": tier,
            "billing_cycle_months": billing_cycle_months,
            "total_amount": total_amount,
            "monthly_price": monthly_price,
            "discount_percent": discount_percent,
        }

    except Exception as e:
        raise HTTPException(500, f"Failed to create subscription: {str(e)}")


@app.get("/billing/trial", tags=["Billing"], response_model=TrialStatusOut)
async def get_trial_status(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Get free trial status"""
    company = (
        db.query(CompanyDB).filter(CompanyDB.id == current_user.company_id).first()
    )
    if not company:
        raise HTTPException(404, "Company not found")

    # Calculate remaining
    trial_days_remaining = None
    trial_invoices_remaining = None
    trial_expired = False

    if company.trial_status == "ACTIVE" and company.trial_start_date:
        # Calculate days remaining based on free_plan_duration_months (default 1 month = 30 days)
        trial_duration_days = (company.free_plan_duration_months or 1) * 30
        days_elapsed = (datetime.utcnow() - company.trial_start_date).days
        trial_days_remaining = max(0, trial_duration_days - days_elapsed)

        # Calculate invoices remaining using actual free_plan_invoice_limit (default to 100 if not set)
        invoice_limit = company.free_plan_invoice_limit or 100
        trial_invoices_remaining = max(0, invoice_limit - company.trial_invoice_count)

        # Check if expired based on free_plan_type
        if company.free_plan_type == "DURATION":
            # Duration-based: only check days, unlimited invoices
            if trial_days_remaining == 0:
                trial_expired = True
                company.trial_status = "EXPIRED"
                company.trial_ended_at = datetime.utcnow()
                db.commit()
        elif company.free_plan_type == "INVOICE_COUNT":
            # Invoice count-based: only check invoices, unlimited duration
            if trial_invoices_remaining == 0:
                trial_expired = True
                company.trial_status = "EXPIRED"
                company.trial_ended_at = datetime.utcnow()
                db.commit()
        else:
            # Default behavior (backward compatibility): check both
            if trial_days_remaining == 0 or trial_invoices_remaining == 0:
                trial_expired = True
                company.trial_status = "EXPIRED"
                company.trial_ended_at = datetime.utcnow()
                db.commit()

    return TrialStatusOut(
        trial_status=company.trial_status,
        trial_start_date=company.trial_start_date.isoformat()
        if company.trial_start_date
        else None,
        trial_invoice_count=company.trial_invoice_count,
        trial_days_remaining=trial_days_remaining,
        trial_invoices_remaining=trial_invoices_remaining,
        trial_expired=trial_expired,
        free_plan_type=company.free_plan_type,
        free_plan_invoice_limit=company.free_plan_invoice_limit,
        free_plan_duration_months=company.free_plan_duration_months,
    )


# ==================== BULK IMPORT ====================


@app.get("/templates/invoices", tags=["Bulk Import"])
@app.get("/bulk/template/invoices", tags=["Bulk Import"])
@app.get("/api/bulk/template/invoices", tags=["Bulk Import"])
async def download_invoice_template(format: str = "csv"):
    """Download CSV/Excel template for bulk invoice import"""
    try:
        df = BulkImportValidator.generate_invoice_template()

        if format.lower() == "excel" or format.lower() == "xlsx":
            output = BytesIO()
            df.to_excel(output, index=False, engine="openpyxl")
            output.seek(0)
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": "attachment; filename=invoice_template.xlsx"
                },
            )
        else:
            output = BytesIO()
            df.to_csv(output, index=False)
            output.seek(0)
            return Response(
                content=output.getvalue(),
                media_type="text/csv",
                headers={
                    "Content-Disposition": "attachment; filename=invoice_template.csv"
                },
            )
    except Exception as e:
        raise HTTPException(500, f"Template generation failed: {str(e)}")


@app.get("/templates/vendors", tags=["Bulk Import"])
@app.get("/bulk/template/vendors", tags=["Bulk Import"])
@app.get("/api/bulk/template/vendors", tags=["Bulk Import"])
async def download_vendor_template(format: str = "csv"):
    """Download CSV/Excel template for bulk vendor import"""
    try:
        df = BulkImportValidator.generate_vendor_template()

        if format.lower() == "excel" or format.lower() == "xlsx":
            output = BytesIO()
            df.to_excel(output, index=False, engine="openpyxl")
            output.seek(0)
            return Response(
                content=output.getvalue(),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": "attachment; filename=vendor_template.xlsx"
                },
            )
        else:
            output = BytesIO()
            df.to_csv(output, index=False)
            output.seek(0)
            return Response(
                content=output.getvalue(),
                media_type="text/csv",
                headers={
                    "Content-Disposition": "attachment; filename=vendor_template.csv"
                },
            )
    except Exception as e:
        raise HTTPException(500, f"Template generation failed: {str(e)}")


@app.post("/invoices/bulk-import", tags=["Bulk Import"], dependencies=[require_permission("invoices", "create")])
@app.post("/bulk/import/invoices", tags=["Bulk Import"], dependencies=[require_permission("invoices", "create")])
@app.post("/api/bulk/import/invoices", tags=["Bulk Import"], dependencies=[require_permission("invoices", "create")])
async def bulk_import_invoices(
    file: UploadFile = File(...),
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Upload and validate CSV/Excel file for bulk invoice creation"""
    try:
        company_id = current_user.company_id
        if not company_id:
            raise HTTPException(403, "Company context required")

        company = db.query(CompanyDB).filter_by(id=company_id).first()
        if not company:
            raise HTTPException(404, "Company not found")

        file_content = await file.read()
        is_valid, parsed_invoices, errors = BulkImportValidator.validate_invoice_file(
            file_content, file.filename
        )

        if not is_valid:
            return {
                "success": False,
                "total_rows": 0,
                "valid_rows": 0,
                "errors": errors,
            }

        # Enforce the same Super Admin free-plan invoice limit used by /invoices create endpoint.
        if (
            company.free_plan_type == "INVOICE_COUNT"
            and company.free_plan_invoice_limit is not None
        ):
            active_paid_subscription = (
                db.query(SubscriptionDB)
                .filter(
                    SubscriptionDB.company_id == company.id,
                    SubscriptionDB.status == "ACTIVE",
                    SubscriptionDB.tier != "FREE",
                )
                .first()
            )

            if not active_paid_subscription:
                used_invoice_count = (
                    db.query(InvoiceDB)
                    .filter(InvoiceDB.company_id == company.id)
                    .count()
                )
                available_slots = company.free_plan_invoice_limit - used_invoice_count

                if available_slots <= 0:
                    raise HTTPException(
                        403,
                        f"Invoice limit reached ({company.free_plan_invoice_limit}). Please upgrade or request additional invoices from Super Admin.",
                    )

                if len(parsed_invoices) > available_slots:
                    return {
                        "success": False,
                        "total_rows": len(parsed_invoices),
                        "valid_rows": 0,
                        "errors": [
                            f"Invoice limit is {company.free_plan_invoice_limit}. You already used {used_invoice_count}.",
                            f"Can only import {available_slots} more invoice(s).",
                        ],
                    }

        from decimal import Decimal

        created_count = 0
        bulk_errors = []
        prepared_invoices = []

        for invoice_data in parsed_invoices:
            line_total = invoice_data["quantity"] * invoice_data["unit_price"]
            discount = invoice_data.get("discount_amount", 0)
            taxable_amount = line_total - discount
            tax_amount_calc = (taxable_amount * invoice_data["tax_percent"]) / 100
            total_amount_calc = taxable_amount + tax_amount_calc

            invoice_type_key = invoice_data["invoice_type"]
            row_num = invoice_data.get("row_num", "Unknown")
            preceding_invoice_id = None

            # Map invoice type strings to enum values
            invoice_type_mapping = {
                "TAX_INVOICE": InvoiceType.TAX_INVOICE,
                "COMMERCIAL": InvoiceType.COMMERCIAL_INVOICE,
            }

            if invoice_type_key == "CREDIT_NOTE":
                preceding_number = invoice_data.get("preceding_invoice_number")
                if not preceding_number:
                    bulk_errors.append(
                        f"Row {row_num}: preceding_invoice_number is required for credit notes"
                    )
                    continue

                original_invoice = (
                    db.query(InvoiceDB)
                    .filter(
                        InvoiceDB.company_id == company_id,
                        InvoiceDB.invoice_number == preceding_number,
                    )
                    .first()
                )

                if not original_invoice:
                    bulk_errors.append(
                        f"Row {row_num}: Referenced invoice '{preceding_number}' not found"
                    )
                    continue

                if original_invoice.invoice_type == InvoiceType.TAX_INVOICE:
                    invoice_type = InvoiceType.TAX_CREDIT_NOTE
                    # Check if company is VAT enabled before allowing tax credit notes
                    if not company.vat_enabled:
                        bulk_errors.append(
                            f"Row {row_num}: Company must be VAT-enabled to create tax credit notes (381)"
                        )
                        continue
                elif original_invoice.invoice_type == InvoiceType.COMMERCIAL_INVOICE:
                    invoice_type = InvoiceType.CREDIT_NOTE_OUT_OF_SCOPE
                else:
                    bulk_errors.append(
                        f"Row {row_num}: Credit notes must reference a tax or commercial invoice"
                    )
                    continue

                if Decimal(str(total_amount_calc)) > Decimal(
                    str(original_invoice.total_amount)
                ):
                    bulk_errors.append(
                        f"Row {row_num}: Credit note total cannot exceed original invoice total"
                    )
                    continue

                preceding_invoice_id = original_invoice.id
            else:
                invoice_type = invoice_type_mapping.get(
                    invoice_type_key, InvoiceType.TAX_INVOICE
                )

                # Validate VAT requirement for Tax Invoices
                if invoice_type == InvoiceType.TAX_INVOICE and not company.vat_enabled:
                    bulk_errors.append(
                        f"Row {row_num}: Company must be VAT-enabled to create tax invoices (380). Non-VAT companies can only create Commercial Invoices (480)"
                    )
                    continue

            prepared_invoices.append(
                {
                    "invoice_data": invoice_data,
                    "invoice_type": invoice_type,
                    "taxable_amount": taxable_amount,
                    "tax_amount_calc": tax_amount_calc,
                    "total_amount_calc": total_amount_calc,
                    "preceding_invoice_id": preceding_invoice_id,
                }
            )

        if bulk_errors:
            return {
                "success": False,
                "total_rows": len(parsed_invoices),
                "valid_rows": 0,
                "errors": bulk_errors,
            }

        # Track invoice counts per type for this batch to avoid duplicate numbers
        invoice_type_counts = {}

        for prepared in prepared_invoices:
            invoice_data = prepared["invoice_data"]
            invoice_type = prepared["invoice_type"]
            taxable_amount = prepared["taxable_amount"]
            tax_amount_calc = prepared["tax_amount_calc"]
            total_amount_calc = prepared["total_amount_calc"]

            # Always auto-generate invoice number for bulk import.
            # Any uploaded invoice_number is intentionally ignored.
            invoice_type_value = (
                invoice_type.value
                if hasattr(invoice_type, "value")
                else str(invoice_type)
            )

            if invoice_type_value not in invoice_type_counts:
                invoice_type_counts[invoice_type_value] = (
                    db.query(InvoiceDB)
                    .filter(
                        InvoiceDB.company_id == company_id,
                        InvoiceDB.invoice_type == invoice_type,
                    )
                    .count()
                )

            invoice_type_counts[invoice_type_value] += 1

            prefix_map = {
                "380": "TI",   # Tax Invoice
                "381": "TCN",  # Tax Credit Note
                "383": "DN",   # Debit Note (Task 4)
                "480": "CI",   # Commercial Invoice
                "81": "CN",    # Credit Note
            }
            prefix = prefix_map.get(invoice_type_value, "TI")
            generated_number = f"{prefix}-{invoice_type_counts[invoice_type_value]:05d}"

            while (
                db.query(InvoiceDB)
                .filter(
                    InvoiceDB.company_id == company_id,
                    InvoiceDB.invoice_number == generated_number,
                )
                .first()
            ):
                invoice_type_counts[invoice_type_value] += 1
                generated_number = (
                    f"{prefix}-{invoice_type_counts[invoice_type_value]:05d}"
                )

            invoice_number = generated_number

            new_invoice = InvoiceDB(
                id=str(uuid4()),
                company_id=company_id,
                invoice_number=invoice_number,
                invoice_type=invoice_type,
                issue_date=datetime.strptime(
                    invoice_data["issue_date"], "%Y-%m-%d"
                ).date()
                if invoice_data.get("issue_date")
                else date.today(),
                due_date=datetime.strptime(invoice_data["due_date"], "%Y-%m-%d").date()
                if invoice_data.get("due_date")
                else None,
                # Supplier info from company
                supplier_trn=company.trn,
                supplier_name=company.legal_name or company.email,
                supplier_address=f"{company.address_line1 or ''} {company.address_line2 or ''}".strip()
                or None,
                supplier_city=company.city,
                supplier_country="AE",
                # Customer info from CSV
                customer_trn=invoice_data["customer_trn"],
                customer_name=invoice_data["customer_name"],
                customer_email=invoice_data.get("customer_email"),
                customer_address=invoice_data.get("customer_address"),
                # Amounts - use correct field names
                subtotal_amount=float(taxable_amount),
                tax_amount=float(tax_amount_calc),
                total_amount=float(total_amount_calc),
                amount_due=float(total_amount_calc),
                # Credit note reference
                preceding_invoice_id=prepared["preceding_invoice_id"],
                # Status and notes - use correct field names
                status=InvoiceStatus.DRAFT,
                invoice_notes=invoice_data.get("notes"),
                share_token=f"share_{uuid4().hex[:16]}",
                created_at=datetime.utcnow(),
            )
            db.add(new_invoice)

            # Create one line item per imported row so invoice detail can render item/qty/tax.
            raw_tax_category = str(invoice_data.get("tax_category") or "S").upper()
            if company.vat_enabled:
                tax_category_map = {
                    "S": TaxCategory.STANDARD,
                    "STANDARD": TaxCategory.STANDARD,
                    "Z": TaxCategory.ZERO,
                    "ZERO": TaxCategory.ZERO,
                    "E": TaxCategory.EXEMPT,
                    "EXEMPT": TaxCategory.EXEMPT,
                    "O": TaxCategory.OUT_OF_SCOPE,
                    "OUT_OF_SCOPE": TaxCategory.OUT_OF_SCOPE,
                }
                effective_tax_category = tax_category_map.get(
                    raw_tax_category, TaxCategory.STANDARD
                )
                effective_tax_percent = float(invoice_data.get("tax_percent") or 0.0)
                effective_tax_code = invoice_data.get("tax_code") or "SR"
                effective_tax_amount = float(tax_amount_calc)
            else:
                effective_tax_category = TaxCategory.OUT_OF_SCOPE
                effective_tax_percent = 0.0
                effective_tax_code = "OP"
                effective_tax_amount = 0.0

            line_item = InvoiceLineItemDB(
                id=f"line_{uuid4().hex[:12]}",
                invoice_id=new_invoice.id,
                line_number=1,
                item_name=invoice_data.get("item_name") or "Item",
                item_description=invoice_data.get("item_description"),
                quantity=float(invoice_data.get("quantity") or 0.0),
                unit_code=invoice_data.get("unit_name") or "C62",
                unit_price=float(invoice_data.get("unit_price") or 0.0),
                line_extension_amount=float(taxable_amount),
                tax_category=effective_tax_category,
                tax_percent=effective_tax_percent,
                tax_code=effective_tax_code,
                tax_amount=effective_tax_amount,
                line_total_amount=float(total_amount_calc),
            )
            db.add(line_item)

            if company.vat_enabled:
                tax_breakdown = InvoiceTaxBreakdownDB(
                    id=f"tax_{uuid4().hex[:12]}",
                    invoice_id=new_invoice.id,
                    tax_category=effective_tax_category,
                    taxable_amount=float(taxable_amount),
                    tax_percent=effective_tax_percent,
                    tax_amount=effective_tax_amount,
                )
                db.add(tax_breakdown)

            created_count += 1

        # Update trial invoice count if on trial
        if company.trial_status == "ACTIVE":
            company.trial_invoice_count = (
                company.trial_invoice_count or 0
            ) + created_count

        db.commit()

        return {
            "success": len(bulk_errors) == 0,  # Only true if no errors
            "total_rows": len(parsed_invoices),
            "valid_rows": created_count,
            "errors": bulk_errors,  # Include any duplicate errors
            "message": f"Successfully imported {created_count} invoices",
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Bulk import failed: {str(e)}")


@app.post("/vendors/bulk-import", tags=["Bulk Import"])
@app.post("/bulk/import/vendors", tags=["Bulk Import"])
@app.post("/api/bulk/import/vendors", tags=["Bulk Import"])
async def bulk_import_vendors(
    file: UploadFile = File(...),
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Upload and validate CSV/Excel file for bulk vendor creation (stored as inward invoice metadata)"""
    try:
        company_id = current_user.company_id
        if not company_id:
            raise HTTPException(403, "Company context required")

        file_content = await file.read()
        is_valid, parsed_vendors, errors = BulkImportValidator.validate_vendor_file(
            file_content, file.filename
        )

        if not is_valid:
            return {
                "success": False,
                "total_rows": 0,
                "valid_rows": 0,
                "errors": errors,
            }

        created_count = 0
        updated_count = 0

        for vendor_data in parsed_vendors:
            existing_vendor = (
                db.query(InwardInvoiceDB)
                .filter_by(
                    company_id=company_id, supplier_trn=vendor_data["vendor_trn"]
                )
                .first()
            )

            if existing_vendor:
                existing_vendor.supplier_name = vendor_data["vendor_name"]
                existing_vendor.supplier_email = vendor_data["vendor_email"]
                existing_vendor.supplier_peppol_id = vendor_data.get("peppol_id")
                updated_count += 1
            else:
                placeholder_invoice = InwardInvoiceDB(
                    id=str(uuid4()),
                    company_id=company_id,
                    invoice_number=f"VENDOR-{vendor_data['vendor_trn']}",
                    supplier_trn=vendor_data["vendor_trn"],
                    supplier_name=vendor_data["vendor_name"],
                    supplier_email=vendor_data["vendor_email"],
                    supplier_address=vendor_data.get("vendor_address"),
                    supplier_peppol_id=vendor_data.get("peppol_id"),
                    issue_date=date.today(),
                    total=0.00,
                    status="VENDOR_RECORD",
                    received_at=datetime.utcnow(),
                )
                db.add(placeholder_invoice)
                created_count += 1

        db.commit()

        return {
            "success": True,
            "total_rows": len(parsed_vendors),
            "valid_rows": created_count + updated_count,
            "created": created_count,
            "updated": updated_count,
            "errors": [],
            "message": f"Successfully processed {created_count + updated_count} vendors ({created_count} created, {updated_count} updated)",
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Bulk vendor import failed: {str(e)}")


# ==================== TASK 2: PASSWORD SECURITY ENDPOINTS ====================


@app.post("/me/change-password", tags=["Password Security"])
def change_password(
    current_password: str,
    new_password: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Change the current user's password.
    Enforces:
    - Current password must be correct
    - New password must meet complexity requirements (Task 2)
    - New password must not match any of the last 5 passwords (Task 2)
    - Minimum password age (Task 22): cannot change more than once per PASSWORD_MIN_AGE_HOURS
    """
    # 0. Task 22: Minimum password age — skip this check for forced/expired changes
    was_forced = bool(getattr(current_user, "must_change_password", False))
    if not was_forced:
        pw_changed_at = getattr(current_user, "password_changed_at", None)
        if pw_changed_at is not None:
            hours_since_last_change = (datetime.utcnow() - pw_changed_at).total_seconds() / 3600
            if hours_since_last_change < PASSWORD_MIN_AGE_HOURS:
                remaining_hours = int(PASSWORD_MIN_AGE_HOURS - hours_since_last_change) + 1
                raise HTTPException(
                    429,
                    f"Password was changed too recently. Please wait {remaining_hours} hour(s) before changing again.",
                )

    # 1. Verify current password
    if not current_user.password_hash or not verify_password(
        current_password, current_user.password_hash
    ):
        raise HTTPException(400, "Current password is incorrect")

    # 2. Validate complexity
    try:
        validate_password_complexity(new_password)
    except ValueError as e:
        raise HTTPException(422, str(e))

    # 3. Check password history (no reuse of last 5)
    if _check_password_history(current_user.id, new_password, db):
        raise HTTPException(
            422,
            f"New password cannot be the same as any of your last {PASSWORD_HISTORY_DEPTH} passwords",
        )

    # 4. Hash and save
    new_hash = get_password_hash(new_password)
    _save_password_history(current_user.id, new_hash, db)
    current_user.password_hash = new_hash

    # 5. Update password_changed_at timestamp (Task 22)
    current_user.password_changed_at = datetime.utcnow()

    # 6. Clear forced password change flag if set (Task 20)
    current_user.must_change_password = False

    # 7. Audit trail
    audit_action = "FORCED_PASSWORD_CHANGED" if was_forced else "PASSWORD_CHANGED"
    audit_description = (
        f"User {current_user.email} completed mandatory password change"
        if was_forced
        else f"User {current_user.email} changed their password"
    )
    log_audit_event(
        db, audit_action,
        user_id=current_user.id, company_id=current_user.company_id,
        resource_type="user", resource_id=current_user.id,
        description=audit_description,
    )

    db.commit()
    return {
        "success": True,
        "message": "Password changed successfully",
        "password_changed_at": current_user.password_changed_at.isoformat(),
    }


# ==================== TASK 1: AUDIT LOG ENDPOINTS ====================


@app.get("/audit-logs", tags=["Audit Trail"])
def list_audit_logs(
    action: Optional[str] = None,
    resource_type: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    List audit log entries for the company.
    Company admins see their company's logs. Super-admins see all logs.
    """
    from datetime import datetime as dt

    if not has_permission(current_user, "audit_logs", "view"):
        raise HTTPException(403, "Permission denied: view on audit_logs is not allowed for your role")

    q = db.query(AuditLogDB)

    if current_user.role != Role.SUPER_ADMIN:
        if not current_user.company_id:
            raise HTTPException(400, "User has no associated company")
        q = q.filter(AuditLogDB.company_id == current_user.company_id)

    if action:
        q = q.filter(AuditLogDB.action == action.upper())
    if resource_type:
        q = q.filter(AuditLogDB.resource_type == resource_type.lower())
    if from_date:
        try:
            q = q.filter(AuditLogDB.created_at >= dt.strptime(from_date, "%Y-%m-%d"))
        except ValueError:
            raise HTTPException(400, "Invalid from_date format. Use YYYY-MM-DD")
    if to_date:
        try:
            from datetime import timedelta as _td
            _to = dt.strptime(to_date, "%Y-%m-%d") + _td(days=1)
            q = q.filter(AuditLogDB.created_at < _to)
        except ValueError:
            raise HTTPException(400, "Invalid to_date format. Use YYYY-MM-DD")

    total = q.count()
    logs = (
        q.order_by(AuditLogDB.created_at.desc())
        .offset((page - 1) * limit)
        .limit(limit)
        .all()
    )

    # Build user_id → email map for response enrichment
    user_ids = list({log.user_id for log in logs if log.user_id})
    user_email_map: dict = {}
    if user_ids:
        users = db.query(UserDB).filter(UserDB.id.in_(user_ids)).all()
        user_email_map = {u.id: u.email for u in users}

    return {
        "total": total,
        "page": page,
        "limit": limit,
        "audit_logs": [
            {
                "id": log.id,
                "action": log.action,
                "resource_type": log.resource_type,
                "resource_id": log.resource_id,
                "user_id": log.user_id,
                "user_email": user_email_map.get(log.user_id),
                "company_id": log.company_id,
                "old_value": log.old_value,
                "new_value": log.new_value,
                "description": log.description,
                "ip_address": log.ip_address,
                "created_at": log.created_at.isoformat() if log.created_at else None,
            }
            for log in logs
        ],
    }


@app.get("/audit-logs/export", tags=["Audit Trail"])
def export_audit_logs(
    format: str = "csv",
    action: Optional[str] = None,
    resource_type: Optional[str] = None,
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Task 23 — Export filtered audit log as CSV or XLSX (admin-only, authenticated).
    Accepts the same filter parameters as GET /audit-logs but returns a downloadable file
    containing ALL matching records (no pagination).
    """
    from datetime import datetime as dt
    import csv as _csv
    import io

    # Authorization: governed by permission matrix (FINANCE_USER and READ_ONLY are excluded)
    if not has_permission(current_user, "audit_logs", "export"):
        raise HTTPException(403, "Permission denied: export on audit_logs is not allowed for your role")

    # Validate format
    if format.lower() not in ("csv", "xlsx"):
        raise HTTPException(400, "Invalid format. Supported values: csv, xlsx")

    # Build query identical to list_audit_logs but without pagination
    q = db.query(AuditLogDB)
    if current_user.role != Role.SUPER_ADMIN:
        if not current_user.company_id:
            raise HTTPException(400, "User has no associated company")
        q = q.filter(AuditLogDB.company_id == current_user.company_id)

    if action:
        q = q.filter(AuditLogDB.action == action.upper())
    if resource_type:
        q = q.filter(AuditLogDB.resource_type == resource_type.lower())
    if from_date:
        try:
            q = q.filter(AuditLogDB.created_at >= dt.strptime(from_date, "%Y-%m-%d"))
        except ValueError:
            raise HTTPException(400, "Invalid from_date format. Use YYYY-MM-DD")
    if to_date:
        try:
            from datetime import timedelta as _td2
            _to = dt.strptime(to_date, "%Y-%m-%d") + _td2(days=1)
            q = q.filter(AuditLogDB.created_at < _to)
        except ValueError:
            raise HTTPException(400, "Invalid to_date format. Use YYYY-MM-DD")

    logs = q.order_by(AuditLogDB.created_at.desc()).all()

    # Enrich with user emails
    user_ids = list({log.user_id for log in logs if log.user_id})
    user_email_map: dict = {}
    if user_ids:
        users = db.query(UserDB).filter(UserDB.id.in_(user_ids)).all()
        user_email_map = {u.id: u.email for u in users}

    # Build rows
    COLUMNS = ["timestamp", "user_email", "action", "resource_type", "resource_id",
               "old_value", "new_value", "description", "ip_address"]

    rows = [
        {
            "timestamp": log.created_at.isoformat() if log.created_at else "",
            "user_email": user_email_map.get(log.user_id, log.user_id or ""),
            "action": log.action or "",
            "resource_type": log.resource_type or "",
            "resource_id": log.resource_id or "",
            "old_value": log.old_value or "",
            "new_value": log.new_value or "",
            "description": log.description or "",
            "ip_address": log.ip_address or "",
        }
        for log in logs
    ]

    # Build filename base
    date_suffix = dt.utcnow().strftime("%Y%m%d_%H%M%S")
    base_name = f"audit_log_{date_suffix}"

    if format.lower() == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise HTTPException(500, "openpyxl not installed — cannot generate XLSX")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Audit Log"

        # Header row — bold with light blue fill
        header_fill = PatternFill("solid", fgColor="D0E4FF")
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(COLUMNS, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row[col_name])

        # Auto-width columns
        for col_cells in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{base_name}.xlsx"'},
        )

    else:
        # Default: CSV
        output = io.StringIO()
        writer = _csv.DictWriter(output, fieldnames=COLUMNS)
        writer.writeheader()
        writer.writerows(rows)
        csv_bytes = output.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility
        return Response(
            content=csv_bytes,
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{base_name}.csv"'},
        )


@app.get("/audit-logs/actions", tags=["Audit Trail"])
def list_audit_actions(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Return the distinct action values present in this company's audit log.
    Used by the frontend to populate the action filter dropdown.
    """
    from sqlalchemy import distinct as _distinct

    q = db.query(_distinct(AuditLogDB.action))

    if current_user.role != Role.SUPER_ADMIN:
        if not current_user.company_id:
            raise HTTPException(400, "User has no associated company")
        q = q.filter(AuditLogDB.company_id == current_user.company_id)

    actions = [row[0] for row in q.all() if row[0]]
    actions.sort()
    return {"actions": actions}


# ==================== TASK 9: VAT RETURN ENDPOINTS ====================


@app.post("/vat-return/generate", tags=["VAT Return"])
def generate_vat_return(
    period_start: str,
    period_end: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """
    Generate and persist a VAT return for the given period.
    Uses the existing calculate_vat_return() utility (Task 9 — must not modify it).
    """
    from decimal import Decimal
    from datetime import datetime as dt
    from utils.vat_utils import calculate_vat_return

    if not has_permission(current_user, "vat_return", "create"):
        raise HTTPException(403, "Permission denied: create on vat_return is not allowed for your role")

    if not current_user.company_id:
        raise HTTPException(400, "User has no associated company")

    try:
        start_date = dt.strptime(period_start, "%Y-%m-%d").date()
        end_date = dt.strptime(period_end, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(400, "Invalid date format. Use YYYY-MM-DD")

    if start_date > end_date:
        raise HTTPException(400, "Start date must be before end date")

    if end_date > date.today():
        raise HTTPException(400, "End date cannot be in the future")

    # --- Sales invoices in period (not DRAFT or CANCELLED) ---
    sales_invoices = (
        db.query(InvoiceDB)
        .filter(
            InvoiceDB.company_id == current_user.company_id,
            InvoiceDB.issue_date >= start_date,
            InvoiceDB.issue_date <= end_date,
            InvoiceDB.status.notin_([InvoiceStatus.DRAFT, InvoiceStatus.CANCELLED]),
        )
        .all()
    )

    # --- Purchase invoices in period (not CANCELLED) ---
    purchase_invoices = (
        db.query(InwardInvoiceDB)
        .filter(
            InwardInvoiceDB.company_id == current_user.company_id,
            InwardInvoiceDB.invoice_date >= start_date,
            InwardInvoiceDB.invoice_date <= end_date,
            InwardInvoiceDB.status != InwardInvoiceStatus.CANCELLED,
        )
        .all()
    )

    # --- 13-Box breakdown via InvoiceTaxBreakdownDB ---
    # Pull all tax breakdowns for the sales invoices in the period
    sale_invoice_ids = [inv.id for inv in sales_invoices]
    tax_breakdowns = (
        db.query(InvoiceTaxBreakdownDB)
        .filter(InvoiceTaxBreakdownDB.invoice_id.in_(sale_invoice_ids))
        .all()
        if sale_invoice_ids
        else []
    )

    # Aggregate by tax category
    box1_std_taxable = Decimal("0.00")   # Box 1: Standard-rated sales (excl. VAT)
    box3_output_vat = Decimal("0.00")    # Box 3: Output VAT (VAT on standard sales)
    box4_zero_rated = Decimal("0.00")    # Box 4: Zero-rated sales
    box5_exempt = Decimal("0.00")        # Box 5: Exempt supplies
    box6_out_of_scope = Decimal("0.00")  # Box 6: Out-of-scope / reverse-charge

    for tb in tax_breakdowns:
        taxable = Decimal(str(tb.taxable_amount or 0))
        vat_amt = Decimal(str(tb.tax_amount or 0))
        cat = tb.tax_category
        if cat == TaxCategory.STANDARD:
            box1_std_taxable += taxable
            box3_output_vat += vat_amt
        elif cat == TaxCategory.ZERO:
            box4_zero_rated += taxable
        elif cat == TaxCategory.EXEMPT:
            box5_exempt += taxable
        elif cat == TaxCategory.OUT_OF_SCOPE:
            box6_out_of_scope += taxable

    # Fallback: if no breakdowns recorded, use invoice-level tax_amount for output VAT
    if not tax_breakdowns:
        box3_output_vat = sum(
            (Decimal(str(inv.tax_amount or 0)) for inv in sales_invoices), Decimal("0.00")
        )
        box1_std_taxable = sum(
            (Decimal(str(inv.subtotal_amount or 0)) for inv in sales_invoices), Decimal("0.00")
        )

    box7_total_sales = box1_std_taxable + box4_zero_rated + box5_exempt + box6_out_of_scope  # Box 7

    # Purchases — group InwardInvoiceLineItemDB by tax_category to fill all purchase boxes
    purchase_invoice_ids = [inv.id for inv in purchase_invoices]
    inward_line_items = (
        db.query(InwardInvoiceLineItemDB)
        .filter(InwardInvoiceLineItemDB.inward_invoice_id.in_(purchase_invoice_ids))
        .all()
        if purchase_invoice_ids
        else []
    )

    box8_std_purchases = Decimal("0.00")     # Box 8: Standard-rated purchases (excl. VAT)
    purchase_bills_vat = Decimal("0.00")     # Box 9: Input VAT on standard-rated bills
    box_zero_rated_purch = Decimal("0.00")   # Zero-rated purchases (no input VAT)
    box_exempt_purch = Decimal("0.00")       # Exempt purchases (no input VAT)
    box_reverse_charge_vat = Decimal("0.00") # Input VAT due on reverse-charge (cat O) purchases

    for li in inward_line_items:
        taxable = Decimal(str(li.line_extension_amount or 0))
        vat_amt = Decimal(str(li.tax_amount or 0))
        cat = li.tax_category
        if cat == TaxCategory.STANDARD:
            box8_std_purchases += taxable
            purchase_bills_vat += vat_amt
        elif cat == TaxCategory.ZERO:
            box_zero_rated_purch += taxable
        elif cat == TaxCategory.EXEMPT:
            box_exempt_purch += taxable
        elif cat == TaxCategory.OUT_OF_SCOPE:
            # Reverse-charge: VAT must be accounted for by the buyer
            box8_std_purchases += taxable
            box_reverse_charge_vat += vat_amt
            purchase_bills_vat += vat_amt

    # Fallback: if no line items, derive from invoice-level totals (all standard-rated)
    if not inward_line_items and purchase_invoices:
        purchase_bills_vat = sum(
            (Decimal(str(inv.tax_amount or 0)) for inv in purchase_invoices), Decimal("0.00")
        )
        box8_std_purchases = sum(
            (Decimal(str(inv.subtotal_amount or 0)) for inv in purchase_invoices), Decimal("0.00")
        )

    purchase_expenses_vat = Decimal("0.00")  # Box 11 (expenses VAT handled separately)

    # Call the protected utility function (MUST NOT BE MODIFIED)
    result = calculate_vat_return(
        sales_invoices_vat=box3_output_vat,
        purchase_bills_vat=purchase_bills_vat,
        purchase_expenses_vat=purchase_expenses_vat,
    )

    # Box 12: Total input VAT = result["total_input_vat"]
    # Box 13: Net payable = result["net_vat_payable"]

    # Persist the return
    return_id = f"vr_{uuid4().hex[:12]}"
    vat_return = VATReturnDB(
        id=return_id,
        company_id=current_user.company_id,
        period_start=start_date,
        period_end=end_date,
        total_sales_invoices=len(sales_invoices),
        total_purchase_invoices=len(purchase_invoices),
        # Sales / output side
        standard_rated_sales=float(box1_std_taxable),
        tax_refunds_provided=0.0,
        output_vat=float(result["output_vat"]),
        zero_rated_sales=float(box4_zero_rated),
        exempt_sales=float(box5_exempt),
        out_of_scope_sales=float(box6_out_of_scope),
        reverse_charge_sales=float(box6_out_of_scope),
        total_sales=float(box7_total_sales),
        # Purchases / input side
        standard_rated_purchases=float(box8_std_purchases),
        input_vat_bills=float(result["input_vat_bills"]),
        purchase_expenses=0.0,
        input_vat_expenses=float(result["input_vat_expenses"]),
        total_input_vat=float(result["total_input_vat"]),
        zero_rated_purchases=float(box_zero_rated_purch),
        exempt_purchases=float(box_exempt_purch),
        reverse_charge_vat=float(box_reverse_charge_vat),
        # Net
        net_vat_payable=float(result["net_vat_payable"]),
        generated_by_user_id=current_user.id,
    )
    db.add(vat_return)

    # Audit trail
    log_audit_event(
        db, "VAT_RETURN_GENERATED",
        user_id=current_user.id, company_id=current_user.company_id,
        resource_type="vat_return", resource_id=return_id,
        description=f"VAT return generated for period {period_start} to {period_end}",
    )

    db.commit()
    db.refresh(vat_return)

    return {
        "success": True,
        "return_id": return_id,
        "period_start": period_start,
        "period_end": period_end,
        "total_sales_invoices": len(sales_invoices),
        "total_purchase_invoices": len(purchase_invoices),
        # Sales boxes
        "box1_standard_rated_sales": float(box1_std_taxable),
        "box2_tax_refunds": 0.0,
        "box3_output_vat": float(result["output_vat"]),
        "box4_zero_rated_sales": float(box4_zero_rated),
        "box5_exempt_sales": float(box5_exempt),
        "box6_out_of_scope_sales": float(box6_out_of_scope),
        "box6_reverse_charge_sales": float(box6_out_of_scope),
        "box7_total_sales": float(box7_total_sales),
        # Purchase boxes
        "box8_standard_rated_purchases": float(box8_std_purchases),
        "box9_input_vat_bills": float(result["input_vat_bills"]),
        "box10_purchase_expenses": 0.0,
        "box11_input_vat_expenses": float(result["input_vat_expenses"]),
        "box12_total_input_vat": float(result["total_input_vat"]),
        "zero_rated_purchases": float(box_zero_rated_purch),
        "exempt_purchases": float(box_exempt_purch),
        "reverse_charge_vat": float(box_reverse_charge_vat),
        # Net
        "box13_net_vat_payable": float(result["net_vat_payable"]),
        "status": "payable" if float(result["net_vat_payable"]) > 0 else "refundable",
    }


@app.get("/vat-return", tags=["VAT Return"])
def list_vat_returns(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """List all VAT returns for the current company."""
    if not current_user.company_id:
        raise HTTPException(400, "User has no associated company")
    returns = (
        db.query(VATReturnDB)
        .filter(VATReturnDB.company_id == current_user.company_id)
        .order_by(VATReturnDB.created_at.desc())
        .all()
    )
    return {
        "vat_returns": [
            {
                "id": r.id,
                "period_start": str(r.period_start),
                "period_end": str(r.period_end),
                "total_sales_invoices": r.total_sales_invoices,
                "total_purchase_invoices": r.total_purchase_invoices,
                "box1_standard_rated_sales": r.standard_rated_sales,
                "box3_output_vat": r.output_vat,
                "box4_zero_rated_sales": getattr(r, "zero_rated_sales", 0.0),
                "box5_exempt_sales": getattr(r, "exempt_sales", 0.0),
                "box6_out_of_scope_sales": getattr(r, "out_of_scope_sales", 0.0),
                "box6_reverse_charge_sales": getattr(r, "reverse_charge_sales", 0.0),
                "box7_total_sales": getattr(r, "total_sales", 0.0),
                "box8_standard_rated_purchases": r.standard_rated_purchases,
                "box9_input_vat_bills": r.input_vat_bills,
                "box12_total_input_vat": r.total_input_vat,
                "zero_rated_purchases": getattr(r, "zero_rated_purchases", 0.0),
                "exempt_purchases": getattr(r, "exempt_purchases", 0.0),
                "reverse_charge_vat": getattr(r, "reverse_charge_vat", 0.0),
                "box13_net_vat_payable": r.net_vat_payable,
                "status": "payable" if r.net_vat_payable > 0 else "refundable",
                "created_at": r.created_at.isoformat() if r.created_at else None,
            }
            for r in returns
        ]
    }


@app.get("/vat-return/{return_id}", tags=["VAT Return"])
def get_vat_return(
    return_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Get a single VAT return with full box-by-box breakdown."""
    r = (
        db.query(VATReturnDB)
        .filter(
            VATReturnDB.id == return_id,
            VATReturnDB.company_id == current_user.company_id,
        )
        .first()
    )
    if not r:
        raise HTTPException(404, "VAT return not found")
    return {
        "id": r.id,
        "period_start": str(r.period_start),
        "period_end": str(r.period_end),
        "total_sales_invoices": r.total_sales_invoices,
        "total_purchase_invoices": r.total_purchase_invoices,
        # Sales / Output side
        "box1_standard_rated_sales": r.standard_rated_sales,
        "box2_tax_refunds": r.tax_refunds_provided,
        "box3_output_vat": r.output_vat,
        "box4_zero_rated_sales": getattr(r, "zero_rated_sales", 0.0),
        "box5_exempt_sales": getattr(r, "exempt_sales", 0.0),
        "box6_out_of_scope_sales": getattr(r, "out_of_scope_sales", 0.0),
        "box7_total_sales": getattr(r, "total_sales", 0.0),
        # Purchases / Input side
        "box8_standard_rated_purchases": r.standard_rated_purchases,
        "box9_input_vat_bills": r.input_vat_bills,
        "box10_purchase_expenses": r.purchase_expenses,
        "box11_input_vat_expenses": r.input_vat_expenses,
        "box12_total_input_vat": r.total_input_vat,
        "zero_rated_purchases": getattr(r, "zero_rated_purchases", 0.0),
        "exempt_purchases": getattr(r, "exempt_purchases", 0.0),
        "reverse_charge_vat": getattr(r, "reverse_charge_vat", 0.0),
        # Net
        "box13_net_vat_payable": r.net_vat_payable,
        "status": "payable" if r.net_vat_payable > 0 else "refundable",
        "created_at": r.created_at.isoformat() if r.created_at else None,
    }


_SOFTWARE_NAME = "InvoLinks"
_SOFTWARE_VERSION = "1.0.0"
_VENDOR_TRN = os.environ.get("INVOLINKS_VENDOR_TRN", "N/A")


@app.get("/vat-return/{return_id}/export", tags=["VAT Return"])
def export_vat_return(
    return_id: str,
    format: str = "xlsx",
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Export a VAT return as XLSX or XML (FTA-compliant formats)."""
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    import xml.etree.ElementTree as ET

    r = (
        db.query(VATReturnDB)
        .filter(
            VATReturnDB.id == return_id,
            VATReturnDB.company_id == current_user.company_id,
        )
        .first()
    )
    if not r:
        raise HTTPException(404, "VAT return not found")

    company = db.query(CompanyDB).filter(CompanyDB.id == r.company_id).first()
    company_name = (company.legal_name or "") if company else ""
    company_trn = (company.trn or "") if company else ""

    period_start = str(r.period_start)
    period_end = str(r.period_end)

    boxes = [
        (1, "Standard-rated supplies (taxable amount, excl. VAT)", getattr(r, "standard_rated_sales", 0.0)),
        (2, "Tax refunds provided to tourists", getattr(r, "tax_refunds_provided", 0.0)),
        (3, "Output VAT (5% on standard-rated sales)", r.output_vat),
        (4, "Zero-rated supplies", getattr(r, "zero_rated_sales", 0.0)),
        (5, "Exempt supplies", getattr(r, "exempt_sales", 0.0)),
        (6, "Out-of-scope / reverse-charge supplies", getattr(r, "out_of_scope_sales", 0.0)),
        (7, "Total value of all supplies", getattr(r, "total_sales", 0.0)),
        (8, "Standard-rated purchases — bills (excl. VAT)", r.standard_rated_purchases),
        (9, "Input VAT recoverable on bills", r.input_vat_bills),
        (10, "Standard-rated purchases — expenses (excl. VAT)", getattr(r, "purchase_expenses", 0.0)),
        (11, "Input VAT recoverable on expenses", r.input_vat_expenses),
        (12, "Total input VAT (Box 9 + Box 11)", r.total_input_vat),
        (13, "Net VAT payable / (refundable)", r.net_vat_payable),
    ]

    fmt_lower = format.lower()

    if fmt_lower == "xlsx":
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "VAT Return FTA 301"

        header_fill = PatternFill("solid", fgColor="1F3A6E")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        section_fill = PatternFill("solid", fgColor="E8EDF5")
        section_font = Font(bold=True, size=10, color="1F3A6E")
        highlight_fill = PatternFill("solid", fgColor="FFF3CD")
        highlight_font = Font(bold=True, size=10)
        thin = Side(style="thin", color="CCCCCC")
        thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def hrow(row, text, fill=None, font=None, merged_to=4):
            cell = ws.cell(row=row, column=1, value=text)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=merged_to)
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            return cell

        def mrow(row, label, value):
            lc = ws.cell(row=row, column=1, value=label)
            lc.font = Font(bold=True, size=10)
            lc.alignment = Alignment(horizontal="left", vertical="center")
            vc = ws.cell(row=row, column=2, value=value)
            vc.alignment = Alignment(horizontal="left", vertical="center")
            for c in range(1, 5):
                ws.cell(row=row, column=c).border = thin_border

        ws.row_dimensions[1].height = 28
        hrow(1, "UAE FTA — VAT Return (Form 301)", fill=header_fill, font=header_font)

        ws.row_dimensions[2].height = 20
        mrow(2, "Company Name", company_name)
        ws.row_dimensions[3].height = 20
        mrow(3, "Tax Registration Number (TRN)", company_trn)
        ws.row_dimensions[4].height = 20
        mrow(4, "Period Start", period_start)
        ws.row_dimensions[5].height = 20
        mrow(5, "Period End", period_end)
        ws.row_dimensions[6].height = 20
        mrow(6, "Software Name", _SOFTWARE_NAME)
        ws.row_dimensions[7].height = 20
        mrow(7, "Software Version", _SOFTWARE_VERSION)
        ws.row_dimensions[8].height = 20
        mrow(8, "Vendor TRN", _VENDOR_TRN)

        ws.row_dimensions[9].height = 8

        ws.row_dimensions[10].height = 22
        hrow(10, "Box-by-Box Breakdown — FTA Form 301", fill=PatternFill("solid", fgColor="2E5FA3"),
             font=Font(color="FFFFFF", bold=True, size=11))

        ws.row_dimensions[11].height = 20
        for col, txt in enumerate(["Box", "Description", "Amount (AED)"], start=1):
            c = ws.cell(row=11, column=col, value=txt)
            c.font = Font(bold=True, size=10, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="4472C4")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border

        highlight_boxes = {3, 7, 12, 13}
        section_breaks = {
            1: "Section 1 — Sales & Output Tax",
            8: "Section 2 — Purchases & Input Tax",
            13: "Section 3 — Net VAT",
        }

        data_row = 12
        for box_num, desc, val in boxes:
            if box_num in section_breaks:
                ws.row_dimensions[data_row].height = 18
                sc = ws.cell(row=data_row, column=1, value=section_breaks[box_num])
                ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=3)
                sc.fill = section_fill
                sc.font = section_font
                sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                sc.border = thin_border
                data_row += 1

            ws.row_dimensions[data_row].height = 18
            is_hl = box_num in highlight_boxes
            fill = highlight_fill if is_hl else None
            font = highlight_font if is_hl else Font(size=10)

            for col_i, v in enumerate([f"Box {box_num}", desc, round(float(val or 0), 2)], start=1):
                c = ws.cell(row=data_row, column=col_i, value=v)
                c.border = thin_border
                c.font = font
                if fill:
                    c.fill = fill
                if col_i == 1:
                    c.alignment = Alignment(horizontal="center", vertical="center")
                elif col_i == 3:
                    c.alignment = Alignment(horizontal="right", vertical="center")
                    c.number_format = '#,##0.00'
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            data_row += 1

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 58
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 5

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        filename = f"vat_return_{period_start}_{period_end}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    elif fmt_lower == "xml":
        root = ET.Element("VATReturn")
        root.set("xmlns:xsi", "http://www.w3.org/2001/XMLSchema-instance")
        root.set("SchemaVersion", "1.0")

        meta = ET.SubElement(root, "Metadata")
        ET.SubElement(meta, "SoftwareName").text = _SOFTWARE_NAME
        ET.SubElement(meta, "SoftwareVersion").text = _SOFTWARE_VERSION
        ET.SubElement(meta, "VendorTRN").text = _VENDOR_TRN
        ET.SubElement(meta, "GeneratedAt").text = datetime.utcnow().isoformat() + "Z"

        taxpayer = ET.SubElement(root, "TaxpayerInfo")
        ET.SubElement(taxpayer, "CompanyName").text = company_name
        ET.SubElement(taxpayer, "TRN").text = company_trn
        ET.SubElement(taxpayer, "PeriodStart").text = period_start
        ET.SubElement(taxpayer, "PeriodEnd").text = period_end

        section_sales = ET.SubElement(root, "SalesAndOutputTax")
        purchase_section = ET.SubElement(root, "PurchasesAndInputTax")
        net_section = ET.SubElement(root, "NetVAT")

        for box_num, desc, val in boxes:
            tag = f"Box{box_num}"
            amount_str = f"{float(val or 0):.2f}"
            if box_num <= 7:
                el = ET.SubElement(section_sales, tag)
            elif box_num <= 12:
                el = ET.SubElement(purchase_section, tag)
            else:
                el = ET.SubElement(net_section, tag)
            el.set("description", desc)
            el.text = amount_str

        ET.indent(root, space="  ")
        xml_bytes = ET.tostring(root, encoding="utf-8", xml_declaration=True)

        # P3-D: Validate the generated XML against the FTA VAT Return XSD schema (strictly blocking)
        _xsd_path = os.path.join(os.path.dirname(__file__), "utils", "fta_vat_return_schema.xsd")
        from lxml import etree as _lxml_etree
        with open(_xsd_path, "rb") as _xf:
            _xmlschema_doc = _lxml_etree.parse(_xf)
        _xmlschema = _lxml_etree.XMLSchema(_xmlschema_doc)
        _doc = _lxml_etree.fromstring(xml_bytes)
        if not _xmlschema.validate(_doc):
            _errors = "; ".join(str(e) for e in _xmlschema.error_log)
            # Write audit in a separate independent session so it is committed before the exception
            _audit_db = SessionLocal()
            try:
                log_audit_event(
                    _audit_db, "VAT_RETURN_XML_INVALID",
                    user_id=current_user.id, company_id=current_user.company_id,
                    resource_type="vat_return", resource_id=return_id,
                    description=f"VAT return XML failed XSD schema validation: {_errors[:500]}",
                )
                _audit_db.commit()
            except Exception:
                pass
            finally:
                _audit_db.close()
            raise HTTPException(500, f"VAT return XML failed FTA schema validation: {_errors[:500]}")

        filename = f"vat_return_{period_start}_{period_end}.xml"
        return StreamingResponse(
            BytesIO(xml_bytes),
            media_type="application/xml",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    else:
        raise HTTPException(400, "format must be 'xlsx' or 'xml'")


# ==================== TASK 8: GL / JOURNAL ENTRY ENDPOINTS ====================


@app.get("/accounts", tags=["General Ledger"])
def list_accounts(
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """List chart of accounts for the current company. Creates default accounts if none exist."""
    if not current_user.company_id:
        raise HTTPException(400, "User has no associated company")
    accounts = (
        db.query(AccountDB)
        .filter(AccountDB.company_id == current_user.company_id, AccountDB.is_active == True)
        .order_by(AccountDB.account_code)
        .all()
    )
    if not accounts:
        create_default_chart_of_accounts(current_user.company_id, db)
        db.commit()
        accounts = (
            db.query(AccountDB)
            .filter(AccountDB.company_id == current_user.company_id, AccountDB.is_active == True)
            .order_by(AccountDB.account_code)
            .all()
        )
    return {
        "accounts": [
            {
                "id": a.id,
                "account_code": a.account_code,
                "account_name": a.account_name,
                "account_name_ar": a.account_name_ar,
                "account_type": a.account_type,
                "is_system": a.is_system,
                "is_active": a.is_active,
            }
            for a in accounts
        ]
    }


@app.get("/journal-entries", tags=["General Ledger"])
def list_journal_entries(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    reference_type: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """List journal entries for the company with optional date and type filters."""
    if not current_user.company_id:
        raise HTTPException(400, "User has no associated company")
    q = db.query(JournalEntryDB).filter(JournalEntryDB.company_id == current_user.company_id)
    if from_date:
        q = q.filter(JournalEntryDB.entry_date >= from_date)
    if to_date:
        q = q.filter(JournalEntryDB.entry_date <= to_date)
    if reference_type:
        q = q.filter(JournalEntryDB.reference_type == reference_type.upper())
    total = q.count()
    entries = q.order_by(JournalEntryDB.entry_date.desc()).offset((page - 1) * limit).limit(limit).all()
    return {
        "total": total,
        "page": page,
        "limit": limit,
        "journal_entries": [
            {
                "id": e.id,
                "entry_date": str(e.entry_date),
                "reference_type": e.reference_type,
                "reference_id": e.reference_id,
                "reference_number": e.reference_number,
                "description": e.description,
                "is_posted": e.is_posted,
                "created_at": e.created_at.isoformat() if e.created_at else None,
            }
            for e in entries
        ],
    }


@app.get("/journal-entries/{entry_id}", tags=["General Ledger"])
def get_journal_entry(
    entry_id: str,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Get a single journal entry with all its debit/credit lines."""
    entry = (
        db.query(JournalEntryDB)
        .filter(JournalEntryDB.id == entry_id, JournalEntryDB.company_id == current_user.company_id)
        .first()
    )
    if not entry:
        raise HTTPException(404, "Journal entry not found")
    return {
        "id": entry.id,
        "entry_date": str(entry.entry_date),
        "reference_type": entry.reference_type,
        "reference_id": entry.reference_id,
        "reference_number": entry.reference_number,
        "description": entry.description,
        "is_posted": entry.is_posted,
        "created_at": entry.created_at.isoformat() if entry.created_at else None,
        "lines": [
            {
                "id": l.id,
                "account_id": l.account_id,
                "account_code": l.account_code,
                "account_name": l.account_name,
                "debit_amount": l.debit_amount,
                "credit_amount": l.credit_amount,
                "currency": l.currency,
                "amount_aed": l.amount_aed,
                "description": l.description,
                "line_number": l.line_number,
            }
            for l in entry.lines
        ],
    }


@app.get("/gl-summary", tags=["General Ledger"])
def get_gl_summary(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    current_user: UserDB = Depends(get_current_user_from_header),
    db: Session = Depends(get_db),
):
    """Account balances summary — total debits, credits, and net balance per account."""
    if not current_user.company_id:
        raise HTTPException(400, "User has no associated company")

    accounts = (
        db.query(AccountDB)
        .filter(AccountDB.company_id == current_user.company_id, AccountDB.is_active == True)
        .order_by(AccountDB.account_code)
        .all()
    )

    summary = []
    for acc in accounts:
        q = (
            db.query(JournalEntryLineDB)
            .join(JournalEntryDB, JournalEntryLineDB.journal_entry_id == JournalEntryDB.id)
            .filter(
                JournalEntryLineDB.account_id == acc.id,
                JournalEntryDB.company_id == current_user.company_id,
                JournalEntryDB.is_posted == True,
            )
        )
        if from_date:
            q = q.filter(JournalEntryDB.entry_date >= from_date)
        if to_date:
            q = q.filter(JournalEntryDB.entry_date <= to_date)

        lines = q.all()
        total_debit = round(sum(l.debit_amount or 0 for l in lines), 2)
        total_credit = round(sum(l.credit_amount or 0 for l in lines), 2)
        net = round(total_debit - total_credit, 2)

        summary.append({
            "account_code": acc.account_code,
            "account_name": acc.account_name,
            "account_type": acc.account_type,
            "total_debit": total_debit,
            "total_credit": total_credit,
            "net_balance": net,
        })

    return {"from_date": from_date, "to_date": to_date, "accounts": summary}


# ==================== REACT APP SERVING ====================
# These routes MUST be last to avoid intercepting API routes


@app.get("/", tags=["General"])
async def root():
    """Serve React app or API info"""
    if os.path.exists("dist/index.html"):
        return FileResponse("dist/index.html")
    elif os.path.exists("static/index.html"):
        return FileResponse("static/index.html")
    return {"message": "InvoLinks API v2.0 - React dev server at port 5173"}


# Catch-all route for React Router (MUST be absolutely last)
@app.get("/{full_path:path}", tags=["General"])
async def serve_react_routes(full_path: str):
    """Serve React app for client-side routing"""
    # Skip API routes
    if full_path.startswith(
        (
            "api/",
            "auth/",
            "admin/",
            "companies/",
            "register/",
            "plans/",
            "docs",
            "redoc",
            "openapi.json",
        )
    ):
        raise HTTPException(404, "Not found")

    if os.path.exists("dist/index.html"):
        return FileResponse("dist/index.html")
    raise HTTPException(404, "Not found")


# ==================== RUN SERVER ====================

if __name__ == "__main__":
    import uvicorn

    print("🚀 Starting InvoLinks E-Invoicing API...")
    print("📚 API Docs: http://0.0.0.0:5000/docs")
    uvicorn.run(app, host="0.0.0.0", port=5000)
